import math
import time
import datetime
import traceback

import openpyxl

import select_ors
from EHOSP.ehospital_proper.start_page_ehosp import selector_effector
from TMS_new.sync_tms.selector_xp import searchIcon
from dkbssy.utils.colour_prints import ColourPrint
from playwright.async_api import async_playwright, Page, TimeoutError, expect
from select_ors import JSON_FILE_PATH, EXCEL_FILE_PATH
from pathlib import Path
from typing import Iterable, Union, Sequence
from openpyxl import load_workbook
from openpyxl.workbook.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet
async def type_list_generate(page:Page, case_type_button, patient_status_field_xpath, current_status):


    body_texts = await page.locator("//body").text_content()
    # ColourPrint.print_yellow('=========================>>>>==>')
    # checking the body has text "Your Hospital Dashboard"
    if "Your Hospital Dashboard" not in body_texts:
        await page.locator(select_ors.homeSVG).click()
        # await page.wait_for_timeout(1000)
        await page.wait_for_load_state("networkidle")
        await expect(page.locator(select_ors.yourHospitalDashboard)).to_be_visible()

    # selecting the "Under Treatment" in PATIENT STATUS by typing
    await page.locator(select_ors.patientStatusInput).type(current_status)  # current_status = "Under Treatment"
    # page.wait_for_selector(selector["patientStatusInput"]).type(current_status)  # current_status = "Under Treatment"
    await page.keyboard.press("Enter")


    # selecting the "50" items per page "10" is done to refresh the page count for "50"
    await page.select_option(select_ors.rowsPerPageSelect, '50')

    # getting number of pending
    total_under_treatment_element = await page.locator(patient_status_field_xpath).text_content()  # Under Treatment(173), patient_status_field_xpath="underTreatmentWith_("

    total_treat_pending = int(total_under_treatment_element.split("(")[1].split(")")[0])

    ColourPrint.print_yellow('===========================>')
    print(total_treat_pending)
    ColourPrint.print_yellow('===========================>')

    # # clicking view more text_button
    try:
        await (await page.wait_for_selector(select_ors.viewMoreButton, timeout=5000)).click()
    except:
        ColourPrint.print_turquoise('View More Skipped')
    #     pass

    # # selecting the "Under Treatment" in PATIENT STATUS by clicking button
    await page.locator(case_type_button).click()
    await page.wait_for_load_state("networkidle")


    # getting number of pages of pending
    total_pages = math.ceil(total_treat_pending /50)
    ColourPrint.print_blue("Total Page", current_status, total_pages)
    #
    #
    # getting the master table
    table_list = []
    # # page.wait_for_timeout(5000)
    for page_num in range(1, total_pages +1):
        last_reg_id = None
        each_page_all_table = await page.query_selector_all("//div[@class='row pl-3 pt-2 ']//small[text()='Registration ID: ']/strong")
        # ColourPrint.print_pink(len(each_page_all_table))
        for each_table in each_page_all_table:
            registration_id = await each_table.text_content()
            print(registration_id)
            table_list.append(registration_id)
            last_reg_id = registration_id

        # ColourPrint.print_blue('last reg id:', last_reg_id)

        # clicking the next page
        if page_num != total_pages:
            # ColourPrint.print_yellow("Next Page")
            await page.locator("//a[contains(@aria-label,'Next page')]//img").click()
            # await page.wait_for_timeout(3500)

            # '//strong[normalize-space()='1008236224']'
            await page.locator(f"//strong[normalize-space()='{last_reg_id}']").wait_for(state="detached", timeout=5000)
            await page.wait_for_load_state("networkidle")


    # table_list = set(table_list)
    ColourPrint.print_pink(current_status, len(table_list))
    return table_list

'excelmethods delete than add gspread'

def load_excel_sheet_and_return_list(EXCEL_FILE_PATH, excel_sheet_name) -> list:
    wb = openpyxl.load_workbook(EXCEL_FILE_PATH)
    ws = wb[excel_sheet_name]
    row_data = list(ws.iter_rows(values_only=True, min_row=2))
    all_reg_id = []
    for each_patient_data in row_data:
        reg_id_row = each_patient_data[2]
        all_reg_id.append(reg_id_row)
    # print(row_data)
    wb.close()
    return all_reg_id


def compare_excel_and_json(excel_sheet_data_list, json_retrieved_list):
    # print(len(excel_sheet_data_list))
    # deletable, additional = [],[]

    'getting additional - present in json but not in excel'
    # for each_reg_id_json in json_retrieved_list:
    #     if each_reg_id_json not in excel_sheet_data_list:
    #         additional.append(each_reg_id_json)
    # print('to add:-', additional)

    additional_2 = list(set(json_retrieved_list) - set(excel_sheet_data_list))
    print('add_2', len(additional_2), additional_2)

    # print('test',set(additional) == additional_2)


    "getting deletable - present in excel but not in json"
    # for each_reg_id_excel in excel_sheet_data_list:
    #     if each_reg_id_excel not in json_retrieved_list:
    #         deletable.append(each_reg_id_excel)
    # print('deletable:-', deletable)


    deletable_2 = list(set(excel_sheet_data_list) - set(json_retrieved_list))
    print('delete_2', deletable_2)
    # print('test2', set(deletable)==deletable_2)
    return additional_2, deletable_2



# ---------------------------------------------------------------------------
# Small helpers (internal use)
# ---------------------------------------------------------------------------
def _norm(values: Iterable[Union[int, str]]) -> set[str]:
    """Convert each value to '12345'‑style strings and drop Nones."""
    return {str(int(v)) for v in values if v is not None}

def _delete_rows(ws: Worksheet, delete_set: set[str], col_idx: int = 3) -> None:
    """Delete rows in *ws* where column *col_idx* matches delete_set."""
    for row in range(ws.max_row, 1, -1):           # bottom‑up
        val = ws.cell(row=row, column=col_idx).value
        if val is not None and str(int(val)) in delete_set:
            ws.delete_rows(row)

# ---------------------------------------------------------------------------
# Public API
# ---------------------------------------------------------------------------
def both_delete(
    discharge_delete_list: Sequence[Union[int, str]], query_delete_list: Sequence[Union[int, str]],
    excel_file_path: Union[str, Path], discharge_sheet_name, query_sheet_name, col_idx: int = 3) -> Workbook:
    """
    Open *excel_file_path* once, delete rows in *discharge_sheet_name* then
    *query_sheet_name*, and return the modified Workbook (unsaved).

    You can keep manipulating the workbook and decide when/where to save it.
    """
    wb: Workbook = load_workbook(excel_file_path)
    try:
        # Normalise the two deletion sets
        dis_set   = _norm(discharge_delete_list)
        query_set = _norm(query_delete_list)

        # Check sheets exist
        if discharge_sheet_name not in wb.sheetnames:
            raise ValueError(f"Sheet '{discharge_sheet_name}' not found")
        if query_sheet_name not in wb.sheetnames:
            raise ValueError(f"Sheet '{query_sheet_name}' not found")

        # Do the deletions
        _delete_rows(wb[discharge_sheet_name], dis_set,  col_idx)
        _delete_rows(wb[query_sheet_name],     query_set, col_idx)

        return wb      # still in memory, not saved
    except Exception:
        wb.close()
        raise

async def getting_more_discharge(page:Page):

    # ColourPrint.print_pink('Inside Treatment')
    amount_load = page.locator(select_ors.amount_treatment_xp)
    await amount_load.wait_for(state="visible")
    await expect(amount_load).to_be_visible()
    amount = await amount_load.text_content()

    # main treatment
    await page.locator(select_ors.mainTreatment).click()

    # treatment detail
    await page.locator(select_ors.treatmentPlanButton).click()

    # show more click
    await page.locator(select_ors.showMoreText_first).click()

    # depart
    depart = await page.locator(select_ors.depart_xp_claim_q_treatment).text_content()

    # procedure SAME FOR BOTH
    procedure = await page.locator(select_ors.procedure_xp).text_content()

    return depart, procedure, amount

async def getting_more_claim_query(page:Page):
    # ColourPrint.print_pink('Inside claim q')
    amount_load = page.locator(select_ors.amount_query_xp)
    await amount_load.wait_for(state="visible")
    await expect(amount_load).to_be_visible()
    amount = await amount_load.text_content()

    # main treatment
    await page.locator(select_ors.mainTreatment).click()

    # treatment detail
    await page.locator(select_ors.treatment_plan_for_query).click()

    # show more click
    await page.locator(select_ors.showMoreText_first).click()

    # depart
    depart = await page.locator(select_ors.depart_xp_claim_q_only).text_content()
    depart = depart.split(".")[1]  # 1.Obstetrics & Gynaecology

    # procedure SAME FOR BOTH
    procedure = await page.locator(select_ors.procedure_xp).text_content()

    return depart, procedure, amount


async def getting_more_pre_auth(page:Page):
    # ColourPrint.print_pink('Inside preaut')
    amount_load = page.locator(select_ors.amount_query_xp)
    await amount_load.wait_for(state="visible")
    await expect(amount_load).to_be_visible()
    amount = await amount_load.text_content()

    # show more click
    await page.locator(select_ors.showMoreText_first).click()

    # depart
    depart = await page.locator(select_ors.depart_xp).text_content()

    # procedure SAME FOR BOTH
    procedure = await page.locator(select_ors.procedure_xp).text_content()

    # click Finance to close it which is already open
    await page.locator(select_ors.finance_close_xp).click()

    return depart, procedure, amount


async def getting_more_details(page:Page, current_status):
    # ColourPrint.print_bg_red(current_status)
    depart = '----Department----'
    procedure = '----PROCEDURE----'
    amount = '0'
    question, ward_name, recent_query_date = None, '<<--Ward_N-->>', None

    if current_status == 'Preauthorization Queried' or current_status == 'Pre-Authorization Queried':
        depart, procedure, amount = await getting_more_pre_auth(page=page)
        question, ward_name, recent_query_date = await retrieving_query_question(page)
    elif current_status == 'Claim Queried':
        depart, procedure, amount = await getting_more_claim_query(page=page)
        question, ward_name, recent_query_date = await retrieving_query_question(page)
    elif current_status == 'Under Treatment':
        depart, procedure, amount = await getting_more_discharge(page=page)

    return depart, procedure, amount, question, ward_name, recent_query_date

async def waiting_for_enable(page:Page, selector_xp):
    loc = page.locator(selector_xp)
    await loc.wait_for(state="visible")
    await expect(loc).to_be_enabled()


async def getting_query_and_discharge_details(page:Page, case_numbers_list):
    type_wise_patient_details = []
    try:
        for case_number in case_numbers_list:
            each_patient_details = []
            try:
                await page.locator(select_ors.homeSVG).click()
            except TimeoutError:
                ColourPrint.print_pink(f'ERROR CAUSED BY {case_number}')
            your_dashboard = page.locator(select_ors.yourHospitalDashboard)
            await your_dashboard.wait_for(state="visible")

            "waiting for toggable button"
            await expect(page.locator(select_ors.self_vs_entity_toggle_xp)).to_be_enabled()
            # await page.wait_for_timeout(1500)
            time.sleep(3)
            await page.wait_for_load_state("networkidle")

            try:
                input_registration_field = page.locator(select_ors.patientStatusInput)
                await input_registration_field.wait_for(state="visible")
                await expect(input_registration_field).to_be_enabled()

                await input_registration_field.fill('All')
                # ColourPrint.print_bg_red('pass')
            except TimeoutError:
                await page.goto('https://provider.nha.gov.in/')
                await page.keyboard.press('Enter')
                ColourPrint.print_blue('Refresh done')

            await page.keyboard.press('Enter')
            await page.wait_for_timeout(250)
            await page.wait_for_load_state("networkidle")

            # search box waiting

            search_locator = page.locator(select_ors.searchBox)
            await search_locator.wait_for(state="visible")
            await expect(search_locator).to_be_enabled()

            # await page.locator(select_ors.searchBox).fill(case_number)
            await search_locator.fill(case_number)

            "waiting function"
            await waiting_for_enable(page, selector_xp=select_ors.searchIcon)
            try:
                await page.locator(select_ors.searchIcon).click()
            except TimeoutError:
                ColourPrint.print_pink(case_number, 'Caused error')
            # waiting for the one selection to be present
            while True:
                try:
                    await page.wait_for_selector(select_ors.loading_for_registration_ids, timeout=2000)  # all ids labels
                    break
                except TimeoutError:
                    ColourPrint.print_yellow("Timeout error -->")
                    await page.locator(select_ors.searchIcon).click()


            await expect(page.locator(select_ors.searchIcon)).to_be_enabled()

            await page.locator(select_ors.searchIcon).click()
            await page.locator(select_ors.searchIcon).click()

            # clicking to proceed for patient details
            await page.locator(f"//strong[text()='{case_number}']/parent::p/parent::div/parent::div//*[name()='path' and @id='Path_98789']").click()
            # print('case number', case_number)


            # getting the type of status
            current_status = (await page.locator(select_ors.current_status_xp).inner_text()).split("(")[0].strip()  # Under Treatment (1002902644)
            # print("current_status", current_status)

            await page.wait_for_load_state("networkidle")

            card = await page.locator(select_ors.card_xp).text_content()
            # print(card)

            name = await page.locator(select_ors.name_xp).text_content()
            # print(name)

            locator = page.locator(".qpz9NG90l1gorDZORgIn.OSVhhsHh74o0GLnLTmrt")
            # Wait for the element to appear (short timeout for presence check)
            await locator.wait_for(timeout=3000, state="attached")

            # Get the initial text of the element
            initial_text = await locator.inner_text()
            # ColourPrint.print_turquoise(f"Initial Text: {initial_text}")

            # Check for 'New Born' using a browser JavaScript function
            age_sex = await page.evaluate(
                """
                ({ locatorSelector, timeout }) => {
                    const start = performance.now();
                    return new Promise(resolve => {
                        const checkText = () => {
                            const element = document.querySelector(locatorSelector);
                            if (element) {
                                const text = element.innerText || element.textContent;
                                if (text.includes('New Born') || performance.now() - start > timeout) {
                                    resolve(text); // Resolve with the current text
                                } else {
                                    requestAnimationFrame(checkText); // Recheck on the next frame
                                }
                            } else {
                                resolve('<-->'); // Resolve with placeholder if element disappears
                            }
                        };
                        checkText(); // Start checking
                    });
                }
                """,
                {"locatorSelector": ".qpz9NG90l1gorDZORgIn.OSVhhsHh74o0GLnLTmrt", "timeout": 1000}
            )
            # print(status)

            date = (await page.locator(select_ors.date_xp).text_content()).split()[0]
            date_a_str = date + "a"  # to convert to string

            "getting more details"
            depart, procedure, amount, question, ward_name, recent_query_date = await getting_more_details(page=page, current_status=current_status)

            # adding remark for all
            pending_days = '--DAYS--'
            remark = 'Discharge Pending' if question is None else question

            "adding the datas"
            each_patient_details.extend([card, name, case_number, date_a_str, depart, procedure, age_sex, current_status, amount, pending_days, remark, ward_name, recent_query_date])

            "adding to master list"
            type_wise_patient_details.append(each_patient_details)

            "printing the individual details"
            print(each_patient_details)
    except Exception as ee:
        print(ee)
        traceback.print_exc()
        ColourPrint.print_pink('GOING THE SAVING FILE DUE TO ERROR')

    return type_wise_patient_details

async def retrieving_query_question(page:Page):
    try:
        # click CASE LOG button side panel
        await page.locator("//button[normalize-space()='Case log']").click()

        # TRYING TO GET THE QUERY FROM HERE - CASE LOG
        question = "----Query Question----"
        recent_query_date = '----Query Date----'

        # clicking SHOW MORE
        await page.locator(select_ors.show_more_third).click()

        # last_query_question_xp containing SHOW LESS
        question = await page.locator(select_ors.last_query_question_xp).text_content()
        # print(question)

        recent_query_date_time = await page.locator(select_ors.recent_query_date_xp).text_content()  # 22/03/2025 12:17:21
        recent_query_date_only = recent_query_date_time.split()[0]
        recent_query_date = recent_query_date_only + 'b'
        ColourPrint.print_green(recent_query_date)

        # close cross button
        await page.locator(select_ors.close_cross_xp).click()
    except TimeoutError as err:
        print(err)
        # closing the case logs
        await page.locator(select_ors.close_cross_xp).click()
        ColourPrint.print_yellow('Finance')
        # click Finance
        await page.locator("//h4[normalize-space()='Finance']/parent::button").click()
        ColourPrint.print_turquoise("Clicked Finance and opened")

        # select the query chat option and click it to open chat window
        try:
            # this is for the condition where chat is not there and query response is being asked
            # ColourPrint.print_pink('Trying to check the in next step of this - no response as -"Please add query response"')
            await page.locator("(//img[@data-tip='Chat'])[last()]").click()
            # getting question - last question
            question = await page.locator("(//div[@class=' mt-1 GoMHIRgbsMpPdNHZPHNf']/span)[last()]").text_content()
            question = question.strip()
            print('---------->',question)

            recent_query_date_xp_fin = "(//div[@class=' mt-1 GoMHIRgbsMpPdNHZPHNf']/div/p)[last()]"
            recent_query_date_time = await page.locator(recent_query_date_xp_fin).text_content()  # 03/22/2025, 12:16
            recent_query_date_only = recent_query_date_time.split(',')[0]
            # modify date
            month_date_format = datetime.datetime.strptime(recent_query_date_only, '%m/%d/%Y')
            day_date_format = datetime.datetime.strftime(month_date_format, '%d/%m/%Y')
            recent_query_date = day_date_format + 'c'
            ColourPrint.print_blue(recent_query_date)

        except TimeoutError:
            ColourPrint.print_yellow('Error Here - "Please add query response"')
            # close button
            # page.wait_for_selector("(//div[contains(@class,'react-draggable')]//img)[2]").click()
            question = "Please add query response"
            recent_query_date = "=TODAY()"
            ColourPrint.print_pink(f'{question=}, {recent_query_date=}')

    return question, 'Ward_Name_Is', recent_query_date







if __name__ == '__main__':
    discharge_row_datas = load_excel_sheet_and_return_list(EXCEL_FILE_PATH=EXCEL_FILE_PATH, excel_sheet_name='Pend Dischg2')
    compare_excel_and_json(discharge_row_datas,'9')


