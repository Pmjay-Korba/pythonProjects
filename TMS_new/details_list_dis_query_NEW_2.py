import math
import time
import datetime
import traceback
import json, os
import openpyxl
from playwright.sync_api import sync_playwright, Page, TimeoutError
from dkbssy.utils.colour_prints import ColourPrint
from old_dkbssy_folder import tms_department_wise_2
from tms_playwright.discharge_to_be_done.detail_list_getter_all import is_file_older_than_2_hours, save_to_json, \
    read_from_json
from tms_playwright.discharge_to_be_done.detail_list_getter_all import AllListGenerator as AllListGeneratorOld



# Create a directory for logs if it doesn't exist
log_dir = "play_request"
os.makedirs(log_dir, exist_ok=True)

# Generate a unique filename with a timestamp
timestamp = datetime.datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
log_file_path = os.path.join(log_dir, f"network_log_{timestamp}.txt")


def log_to_file(data):
    """Append logs to a uniquely named file."""
    with open(log_file_path, "a", encoding="utf-8") as f:
        f.write(data + "\n" + "-" * 80 + "\n")  # Separate each entry


def log_request(request):
    """Log request details."""
    request_info = f"Request: {request.url}\n"
    request_info += f"Headers: {json.dumps(request.headers, indent=2)}\n"
    if request.post_data:
        request_info += f"Payload: {request.post_data}\n"

    log_to_file(request_info)


def log_response(response):
    """Log response details."""
    response_info = f"Response: {response.url} - {response.status}\n"
    response_info += f"Headers: {json.dumps(response.headers, indent=2)}\n"
    try:
        response_info += f"Body: {response.text()[:500]}\n"  # Print first 500 chars
    except Exception:
        response_info += "Response body not available.\n"

    log_to_file(response_info)

# Attach listeners
# page.on("request", log_request)
# page.on("response", log_response)

class AllListGenerator:
    # Define the paths to Chrome executable and user data directory
    chrome_path = r"C:\Program Files (x86)\Google\Chrome\Application\chrome.exe"
    user_data_dir = r'C:\Users\RAKESH\AppData\Local\Google\Chrome\User Data'

    # Define the Google Sheets URL and download directory
    sheet_url = "https://docs.google.com/spreadsheets/d/19HHTQZe9_8hMQJDZM4aZ01RcBqVH1-xXVv3RkR2W1ls/edit?gid=0"
    sheet_2025 = 'https://docs.google.com/spreadsheets/d/1vhjV0rcODJ4lGYJBHENMnHFvqHgK25dQRt9SVpr_9N4/edit?gid=0#gid=0'
    download_dir = r"G:\My Drive\GdrivePC\Hospital\RSBY\New\down"


    def login_or_already_logged_new(self) -> Page:
        with sync_playwright() as p:
            # Connect to the running Chrome instance on the debugging port
            browser = p.chromium.connect_over_cdp('http://localhost:9222')

            # Get the active browser context and page (tab)
            context = browser.contexts[0]  # Assuming you want the first context
            page = context.pages[0]       # Assuming you want the first tab

            # Get the current URL
            current_url = page.url
            print('Current URL:', current_url)
            return page


    def type_list_generate(self, page, case_type_button, patient_status_field_xpath, current_status) -> list:
        # Load the JSON Selector file
        selector_file = r"C:\WebstormProjects\TMS_new\selectors.json"
        with open(selector_file, 'r') as file:
            selector = json.load(file)

        # checking the home page displayed or not
        # getting the "Your Hospital Dashboard" in body html

        body_texts = page.wait_for_selector("//body").text_content()
        # print(body_texts)

        # checking the body has text "Your Hospital Dashboard"
        if "Your Hospital Dashboard" not in body_texts:
            page.wait_for_selector(selector["homeSVG"]).click()
            page.wait_for_selector(selector["yourHospitalDashboard"])

        page.wait_for_timeout(2500)
        # selecting the "Under Treatment" in PATIENT STATUS by typing
        page.wait_for_selector(selector["patientStatusInput"]).type(current_status)  # current_status = "Under Treatment"
        page.keyboard.press("Enter")


        # selecting the "50" items per page "10" is done to refresh the page count for "50"
        # page.select_option(selector["rowsPerPageSelect"], "10")
        # page.wait_for_timeout(1500)
        page.select_option(selector["rowsPerPageSelect"], "50")
        page.wait_for_timeout(1500)
        page.wait_for_load_state("networkidle")

        # # add wait to count the 50 items are loaded or not
        # page.wait_for_selector("//div[@class='row pl-3 pt-2 ']")  # getting each patient box

        # getting number of pending
        total_under_treatment_element = page.wait_for_selector(selector[patient_status_field_xpath]).text_content()  # Under Treatment(173), patient_status_field_xpath="underTreatmentWith_("
        total_treat_pending = int(total_under_treatment_element.split("(")[1].split(")")[0])
        print(total_treat_pending)

        # clicking view more text_button
        try:
            page.wait_for_selector(selector["viewMoreButton"], timeout=5000).click()
        except:
            ColourPrint.print_turquoise('View More Skipped')
            pass
        # selecting the "Under Treatment" in PATIENT STATUS by clicking button
        page.wait_for_selector(selector[case_type_button]).click()
        page.wait_for_load_state("networkidle")


        # getting number of pages of pending
        total_pages = math.ceil(total_treat_pending/50)
        print("Total Page", total_pages)


        # getting the master table
        table_list = []
        # page.wait_for_timeout(5000)
        for page_num in range(1, total_pages+1):
            each_page_all_table = page.query_selector_all("//div[@class='row pl-3 pt-2 ']//small[text()='Registration ID: ']/strong")
            for each_table in each_page_all_table:
                registration_id = each_table.text_content()
                print(registration_id)
                table_list.append(registration_id)

            # clicking the next page
            if page_num != total_pages:
                ColourPrint.print_yellow("Next Page")
                page.wait_for_selector("//a[contains(@aria-label,'Next page')]//img").click()
                page.wait_for_timeout(3500)
                page.wait_for_load_state("networkidle")

                # name = each_table.wait_for_selector("//label").text_content()
                # age = each_table.query_selector("//label/following-sibling::div[1]/descendant::small[1]").text_content()
                # sex = each_table.query_selector("//label/following-sibling::div[1]/descendant::small[2]").text_content()
                # card = each_table.query_selector("//small[text()=' Program ID: ']/strong").text_content()
                # registration = each_table.query_selector("//small[text()='Registration ID: ']/strong").text_content()
                # date = each_table.query_selector("//small[text()=' Registration Date: ']/strong").text_content()
                # print(name, age, sex, card, registration, date)

            # checking the disappearance of last patient box of current page -> last item is different in successive appended list
            # last_patient_detail = table_list[-1]
            # page.wait_for_selector(f"//strong[normalize-space()='{last_patient_detail}']", state="detached")


        return table_list

    def wait_for_updated_content_direct(self, page, xpath, placeholder="NA", timeout=60):
        """
        Waits for the text content of an element to change from a placeholder value using polling.

        Args:
            page: The Playwright page instance.
            xpath: The XPath of the target element.
            placeholder: The placeholder value to wait for change (default is 'NA').
            timeout: The maximum time to wait in seconds (default is 60).

        Returns:
            str: The updated text content of the element.

        Raises:
            TimeoutError: If the text content does not change within the timeout.
        """
        end_time = time.time() + timeout
        while time.time() < end_time:
            try:
                # Retrieve the text content of the element
                element = page.query_selector(xpath)
                if element:
                    text_content = element.text_content().strip()
                    if text_content != placeholder:  # Check if it's updated
                        return text_content
            except Exception as e:
                pass  # Ignore errors and keep polling
            time.sleep(0.5)  # Wait for 500ms before retrying

        raise TimeoutError(f"Timeout: Element with XPath '{xpath}' did not update within {timeout} seconds.")

    def retrieve_date_new_claim_query(self, page: Page, case_number):
        # Load the JSON Selector file
        selector_file = r"C:\WebstormProjects\TMS_new\selectors.json"
        with open(selector_file, 'r') as file:
            selector = json.load(file)


        # checking the home page displayed or not
        # getting the "Your Hospital Dashboard" in body html

        # body_texts = page.wait_for_selector("//body").text_content()
        # print(body_texts)

        # checking the body has text "Your Hospital Dashboard"
        # if "Your Hospital Dashboard" not in body_texts:
        #     page.wait_for_selector(selector["homeSVG"]).click()
        #     page.wait_for_selector(selector["yourHospitalDashboard"])


        page.wait_for_load_state("networkidle")
        # page.wait_for_selector(selector["patientStatusInput"]).fill('')
        try:
            page.wait_for_selector(selector["patientStatusInput"], timeout=5000).fill('All')
        except TimeoutError:
            page.goto('https://provider.nha.gov.in/')
        page.keyboard.press('Enter')
        time.sleep(0.25)
        page.wait_for_load_state("networkidle")

        page.wait_for_selector(selector["searchBox"]).fill("")

        # page.on("request", log_request)
        # page.on("response", log_response)

        page.wait_for_selector(selector["searchBox"]).fill(case_number)
        page.wait_for_load_state("networkidle")

        page.wait_for_selector(selector["searchIcon"]).click()

        # waiting for the one selection to be present
        # count_registration_id = len(page.query_selector_all("//p[contains(text(),'Registration ID:')]"))  # all ids labels

        # waiting for the one selection to be present
        while True:
            try:
                page.wait_for_selector("//p[contains(text(),'Registration ID:')]", timeout=2000)  # all ids labels
                break
            except TimeoutError:
                ColourPrint.print_yellow("Timeout error -->")
                page.wait_for_selector(selector["searchIcon"]).click()


        # print(count_registration_id)
        # while True:
        #     if count_registration_id == 1:
        #         print(1)
        #         break
        #     else:
        #         time.sleep(0.25)
        #         page.wait_for_selector(selector["searchIcon"]).click()


        page.wait_for_selector(selector["searchIcon"]).click()
        # time.sleep(0.25)
        page.wait_for_selector(selector["searchIcon"]).click()
        page.wait_for_selector(selector["searchIcon"]).click()
        # time.sleep(0.25)
        # page.wait_for_selector(selector["searchIcon"]).click()

        # Attach listeners
        # page.on("request", log_request)
        # page.on("response", log_response)


        # clicking to proceed for patient details
        page.wait_for_selector(f"//strong[text()='{case_number}']/parent::p/parent::div/parent::div//*[name()='path' and @id='Path_98789']").click()

        # print('case number', case_number)
        # getting the type of status
        current_status = page.wait_for_selector("//a[normalize-space()='Home']/ancestor::ol//li[2]").inner_text().split("(")[0].strip()  # Under Treatment (1002902644)
        # print(current_status)

        page.wait_for_load_state("networkidle")
        card = page.wait_for_selector("//div[text()=' ID']/following-sibling::div").text_content()
        # print(card)
        name = page.wait_for_selector("//img[@alt='image']/parent::div/following-sibling::div/div").text_content()
        # print(name)
        # age/sex
        # status_xp = "//img[@alt='image']/parent::div/following-sibling::div/following-sibling::div/div"
        # self.wait_for_updated_content_direct(page, status_xp)
        # page.wait_for_load_state("networkidle")
        # status_element = page.wait_for_selector("//img[@alt='image']/parent::div/following-sibling::div/following-sibling::div/div")
        #
        # status = status_element.text_content()
        # # print(status)

        locator =  page.locator(".qpz9NG90l1gorDZORgIn.OSVhhsHh74o0GLnLTmrt")
        # Wait for the element to appear (short timeout for presence check)
        locator.wait_for(timeout=3000, state="attached")

        # Get the initial text of the element
        initial_text = locator.inner_text()
        # print(f"Initial Text: {initial_text}")

        # Check for 'New Born' using a browser JavaScript function
        status = page.evaluate(
            """
            ({ locatorSelector, timeout }) => {
                const start = performance.now();
                return new Promise(resolve => {
                    const checkText = () => {
                        const element = document.querySelector(locatorSelector);
                        if (element) {
                            const text = element.innerText || element.textContent;
                            if (text.includes('New Born') || performance.now() - start > timeout) {
                                resolve(text); // Resolve with the current text
                            } else {
                                requestAnimationFrame(checkText); // Recheck on the next frame
                            }
                        } else {
                            resolve('<-->'); // Resolve with placeholder if element disappears
                        }
                    };
                    checkText(); // Start checking
                });
            }
            """,
            {"locatorSelector": ".qpz9NG90l1gorDZORgIn.OSVhhsHh74o0GLnLTmrt", "timeout": 1000}
        )
        # print(status)

        date_xp = "//div[text()='Registration Date']/following-sibling::div"
        self.wait_for_updated_content_direct(page, date_xp)
        date = page.wait_for_selector("//div[text()='Registration Date']/following-sibling::div").text_content().split()[0]
        date_a_str = date + "a"  # to convert to string

        depart = '----Department----'
        procedure = '----PROCEDURE----'
        amount = '0'

        if current_status == 'Preauthorization Queried' or current_status == 'Pre-Authorization Queried':
            time.sleep(1)
            amount = page.wait_for_selector("//div[p[normalize-space()='Total payable amount (after incentives) :']]/following-sibling::div/p").text_content()
            # main treatment already opened clicking treatment plan
            page.wait_for_selector("//div[contains(text(),'Treatment Plan')]/parent::div/following-sibling::div//button").click()
            # print("clicked treat plan")
            # show more click
            page.wait_for_selector("//span[contains(text(),'Show More')]").click()
            # depart
            depart = page.wait_for_selector("//span[contains(text(),'Show Less')]/parent::p/parent::td/preceding-sibling::td[1]").text_content()
            # procedure SAME FOR BOTH
            procedure = page.wait_for_selector("//span[contains(text(),'Show Less')]/parent::p").text_content()
            # print(procedure)
            # click Finance to close it which is already open
            page.wait_for_selector("//h4[normalize-space()='Finance']/parent::button").click()

        elif current_status == 'Claim Queried':
            time.sleep(1)
            amount = page.wait_for_selector("//div[p[normalize-space()='Total payable amount (after incentives) :']]/following-sibling::div/p").text_content()
            # main treatment
            page.wait_for_selector("//h4[normalize-space()='Treatment']/parent::button").click()
            # treatment detail
            page.wait_for_selector("//span[contains(text(),'Treatment Details')]/parent::button").click()
            # show more click
            page.wait_for_selector("//span[contains(text(),'Show More')]").click()
            # department
            depart = page.wait_for_selector("//span[contains(text(),'Show Less')]/parent::p/parent::div/parent::div/preceding-sibling::div/div").text_content()
            # print(depart)
            depart = depart.split(".")[1]  # 1.Obstetrics & Gynaecology
            # print(depart)
            # procedure SAME FOR BOTH
            procedure = page.wait_for_selector("//span[contains(text(),'Show Less')]/parent::p").text_content()
            # print(procedure)

        elif current_status == 'Under Treatment':
            # wait for amount to load
            amount_xp = "//div[normalize-space()='Total Preauth Approved Amount']/following-sibling::div"
            self.wait_for_updated_content_direct(page, amount_xp)
            amount = page.wait_for_selector("//div[normalize-space()='Total Preauth Approved Amount']/following-sibling::div").text_content()

            # page.on("request", log_request)
            # page.on("response", log_response)


            # main treatment
            page.wait_for_selector("//h4[normalize-space()='Treatment']/parent::button").click()

            # page.on("request", log_request)
            # page.on("response", log_response)
            # treatment detail
            page.wait_for_selector("//span[contains(text(),'Treatment Plan')]/parent::button").click()

            # show more click
            page.wait_for_selector("//span[contains(text(),'Show More')]").click()
            # depart
            depart = page.wait_for_selector("//span[contains(text(),'Show Less')]/parent::p/parent::td/preceding-sibling::td[2]").text_content()
            # procedure SAME FOR BOTH
            procedure = page.wait_for_selector("//span[contains(text(),'Show Less')]/parent::p").text_content()
            # print("Under treat ", procedure)
            # going homepage
            page.wait_for_selector("//p[contains(text(),'Home')]/preceding-sibling::*[name()='svg']").click()
            page.wait_for_load_state("networkidle")
            time.sleep(2)


        page.wait_for_load_state("networkidle")
        time.sleep(1.5)
        # print(card, name, case_number, date, depart, procedure, status)

        date_by_datetime = datetime.datetime.strptime(date, '%d/%m/%Y').date()
        today_by_datetime = datetime.datetime.today().date()
        days_diff = today_by_datetime - date_by_datetime
        pending_days = days_diff.days
        # pending_days = str(pending_days)
        # pending_days = "=IFERROR((today()-DATE(MID(D2,7,4),mid(D2,4,2),left(D2,2))),"")"
        pending_days = '--DAYS--'
        # today_by_datetime_str = str(today_by_datetime)
        # print(type(today_by_datetime_str))
        # print((today_by_datetime_str))
        today_by_datetime_str2 = today_by_datetime.strftime('%d-%m-%Y')
        # print(today_by_datetime_str2)


        # adding remark for all
        remark = 'Discharge Pending'
        return card, name, case_number, date_a_str, depart, procedure, status, current_status, amount, pending_days, remark

        # # treatment plan
        # page.wait_for_selector("//span[contains(text(),'Treatment Plan')]/parent::button").click()
        # # show more click
        # page.wait_for_selector("//span[contains(text(),'Show More')]").click()
        #
        # depart = page.wait_for_selector("//span[contains(text(),'Show Less')]/ancestor::tr/td[2]").text_content()
        # procedure = page.wait_for_selector("//span[contains(text(),'Show Less')]/ancestor::tr/td[4]").text_content()



    def manipulating_downloaded_excel_new(self, both_query_playwright_list: list, sheet_name_to_be_compared: str,
                                      is_query_question_required: bool, page: Page):
        file_path = r"G:\My Drive\GdrivePC\Hospital\RSBY\New\down\cc.xlsx"
        workbook = openpyxl.load_workbook(filename=file_path, data_only=True)
        sheet_name = workbook[sheet_name_to_be_compared]
        case_numbers_in_cols = list(
            [row for row in sheet_name.iter_cols(min_col=3, max_col=3, min_row=2, values_only=True)][0])
        print(case_numbers_in_cols)
        case_numbers_in_cols = [str(int(i)) for i in case_numbers_in_cols if i is not None]
        # for k in case_numbers_in_cols:
        #     print((type(k)))
        print('cc-new', len(case_numbers_in_cols), case_numbers_in_cols)
        print('dd-new', len(both_query_playwright_list), both_query_playwright_list)
        to_delete_case_numbers = set(case_numbers_in_cols) - set(both_query_playwright_list)
        print('to delete', len(to_delete_case_numbers), to_delete_case_numbers)
        to_add_case_numbers = set(both_query_playwright_list) - set(case_numbers_in_cols)
        print('to_add_list', len(to_add_case_numbers), to_add_case_numbers)

        # Iterate through the rows in reverse to avoid issues with deleting rows
        for row in range(sheet_name.max_row, 1, -1):
            cell_value = sheet_name.cell(row=row, column=3).value  # Assuming 'C' is the 3rd column
            if cell_value is not None:
                cell_value = str(int(cell_value))
                # print('cellvalue ', cell_value)
            if cell_value in to_delete_case_numbers:
                sheet_name.delete_rows(row)
                # print('deleted cellvalue')

        workbook.save(r"G:\My Drive\GdrivePC\Hospital\RSBY\New\down\cc.xlsx")

        master_to_save_in_excel = []
        try:
            for idx, c_num in enumerate(to_add_case_numbers):
                if not is_query_question_required:
                    case_num_data = self.retrieve_date_new_claim_query(page, c_num)
                    # print(case_num_data)
                    for_print = "$".join(case_num_data)
                    print(idx, for_print)
                    master_to_save_in_excel.append(list(case_num_data))
                else:
                    case_num_data = self.retrieve_date_new_claim_query(page, c_num)
                    # print(case_num_data)

                    # click CASE LOG button side panel
                    page.wait_for_selector("//button[normalize-space()='Case log']").click()

                    # TRYING TO GET THE QUERY FROM HERE - CASE LOG
                    question = "----Query Question----"
                    try:
                        try:# clicking SHOW MORE
                            # ColourPrint.print_turquoise('10101010101011010101010')
                            page.wait_for_selector("//*[name()='circle' and @id='bg-icon']/ancestor::something/parent::div/parent::div/parent::div/div[2]/div[last()]/div[2]/div[3]//span[contains(text(),'...Show More')]",timeout=3000).click()
                            # ColourPrint.print_turquoise('111111111111111111')
                            # last_query_question_xp containing SHOW LESS
                            last_query_question_xp = "//*[name()='circle' and @id='bg-icon']/ancestor::something/parent::div/parent::div/parent::div/div[2]/div[last()]/div[2]/div[3]"

                            # question = "this is demo test"
                            question = page.wait_for_selector(last_query_question_xp).text_content()
                            # print(question)
                            # close cross button
                            page.wait_for_selector("//*[name()='path' and @id='cross-icon']").click()
                        except TimeoutError as err :
                            ColourPrint.print_turquoise('2222222222222222222222')
                            print(err)
                            # closing the case logs
                            page.wait_for_selector("//*[name()='path' and @id='cross-icon']").click()

                            ColourPrint.print_yellow('Finance')
                            # click Finance
                            page.wait_for_selector("//h4[normalize-space()='Finance']/parent::button").click()
                            # select the query chat option and click it to open chat window
                            page.wait_for_selector("(//img[@data-tip='Chat'])[last()]").click()
                            # getting question - last question
                            question = page.wait_for_selector(
                                "(//div[@class=' mt-1 GoMHIRgbsMpPdNHZPHNf']/span)[last()]").text_content()
                            # print('---------->',question)
                            # close button
                            page.wait_for_selector("(//div[contains(@class,'react-draggable')]//img)[2]").click()

                        # finally:
                        #     ColourPrint.print_turquoise('33333333333333333')
                        #     page.wait_for_selector("//*[name()='path' and @id='cross-icon']").click()

                    except TimeoutError as err:
                        print('Here in Exception')
                        # click Finance
                        page.wait_for_selector("//h4[normalize-space()='Finance']/parent::button").click()
                        # select the query chat option and click it to open chat window
                        page.wait_for_selector("(//img[@data-tip='Chat'])[last()]").click()
                        # getting question - last question
                        question = page.wait_for_selector("(//div[@class=' mt-1 GoMHIRgbsMpPdNHZPHNf']/span)[last()]").text_content()
                        # print(question)
                        # close button
                        page.wait_for_selector("(//div[contains(@class,'react-draggable')]//img)[2]").click()


                    case_num_data_with_remark = list(case_num_data)
                    case_num_data = case_num_data_with_remark[:-1]
                    case_num_data.append(question)
                    # print(case_num_data)
                    for_print = "$".join(case_num_data)
                    print(idx, for_print)
                    master_to_save_in_excel.append(case_num_data)
                    page.wait_for_load_state("networkidle")
                    time.sleep(2)
                    # going homepage
                    page.wait_for_selector("//p[contains(text(),'Home')]/preceding-sibling::*[name()='svg']").click()
                    page.wait_for_load_state("networkidle")
                    time.sleep(2)
        except Exception as ee:
            print(ee)
            traceback.print_exc()
            ColourPrint.print_pink('GOING THE SAVING FILE DUE TO ERROR')

        time1 = time.time()
        for case_detail_list in master_to_save_in_excel:
            sheet_name.append(case_detail_list)
        workbook.save(r"G:\My Drive\GdrivePC\Hospital\RSBY\New\down\cc.xlsx")
        # function ran
        tms_department_wise_2.department_wise_extract_for_queries_2025()
        tms_department_wise_2.department_wise_extract_for_discharge_2025()
        print('Time to save in excel:', time.time() - time1, 'seconds')



if __name__ == '__main__':
    # Call the function
    all_list_gen_obj = AllListGenerator()
    with sync_playwright() as p:
        # Connect to the running Chrome instance on the debugging port
        browser = p.chromium.connect_over_cdp('http://localhost:9222')

        # Get the active browser context and page (tab)
        context = browser.contexts[0]  # Assuming you want the first context

        # new_page = context.new_page()
        # # Navigate to the Google Sheets URL
        # AllListGeneratorOld().spreadsheet_download_and_rename(new_page,
        #                                                       'https://docs.google.com/spreadsheets/d/19HHTQZe9_8hMQJDZM4aZ01RcBqVH1-xXVv3RkR2W1ls/edit?gid=42829875#gid=42829875',
        #                                                       'AYUSHMAN REGISTRATION 2023.xlsx',
        # #                                                       "cc.xlsx")
        # new_page.close()
        # ColourPrint.print_yellow("New page closed.")

        # downloading new 2025 sheet
        # new_page_2 = context.new_page()
        # AllListGeneratorOld().spreadsheet_download_and_rename(new_page_2,
        #                                                       'https://docs.google.com/spreadsheets/d/1vhjV0rcODJ4lGYJBHENMnHFvqHgK25dQRt9SVpr_9N4/edit?gid=0#gid=0',
        #                                                       'AYUSHMAN REGISTRATION 2025.xlsx',
        #                                                       'cc_new.xlsx')
        # new_page_2.close()
        # ColourPrint.print_yellow("New page 2 closed.")

        new_page_3 = context.new_page()
        AllListGeneratorOld().spreadsheet_download_and_rename(new_page_3,
                                                              'https://docs.google.com/spreadsheets/d/1vhjV0rcODJ4lGYJBHENMnHFvqHgK25dQRt9SVpr_9N4/edit?gid=0#gid=0',
                                                              'AYUSHMAN REGISTRATION 2025.xlsx',
                                                              'cc.xlsx')  # cc_new renamed to cc also modified name from cc_new to cc at department_wise_extract_for_queries_2025() and
                                                                                                    # department_wise_extract_for_discharge_2025()
        new_page_3.close()
        ColourPrint.print_yellow("New page 3 closed.")

        for page in context.pages:
            try:
                page_title = page.title()
                print(f"Checking page title: {page_title}")
                print("Current URL: ", page.url)


                if page_title.strip() == 'PMJAY - Provider':
                    # if page_title.strip().lower() == target_title.strip().lower():
                    print(f"Found matching page: {page_title}")
                    page.set_default_timeout(20000)


                    # json file for the saving the all list getting
                    json_file_path = r"G:\My Drive\GdrivePC\Hospital\RSBY\New\down\dd2.json"
                    older_than_2_hours = is_file_older_than_2_hours(json_file_path)
                    # print(older_than_2_hours)
                    if older_than_2_hours:
                        print('Older than 2 hours:', older_than_2_hours)

                        ColourPrint.print_yellow('Start Treatment')
                        under_treatment_list_generated = all_list_gen_obj.type_list_generate(page,
                                                                                             current_status='Under Treatment',
                                                                                             patient_status_field_xpath="underTreatmentWith_(",
                                                                           case_type_button="underTreatmentOnly")
                        ColourPrint.print_pink('End Treatment')

                        ColourPrint.print_yellow('Start Pre-auth')




                        pre_auth_query_list_generated = all_list_gen_obj.type_list_generate(page,
                                                                                            current_status="Preauthorization Queried",
                                                                                            patient_status_field_xpath="PreauthorizationQueried_(",
                                                                                            case_type_button='PreauthorizationQueriedOnly')
                        ColourPrint.print_pink('End Pre-auth')





                        ColourPrint.print_yellow('Start Claim')
                        claim_query_list_generated = all_list_gen_obj.type_list_generate(page,
                                                                                         current_status="Claims Queried",
                                                                                         patient_status_field_xpath="ClaimsQueried_(",
                                                                                         case_type_button="ClaimsQueriedOnly")
                        ColourPrint.print_pink('End Claim')

                        # Save the generated lists to the CSV file
                        save_to_json(pre_auth_query_list_generated,claim_query_list_generated, under_treatment_list_generated, json_file_path)
                    else:
                        # File is less than 2 hours old, read from the file
                        pre_auth_query_list_generated,claim_query_list_generated, under_treatment_list_generated = read_from_json(json_file_path)

                    both_query_list = pre_auth_query_list_generated + claim_query_list_generated
                    print('both_query_list', both_query_list)
                    AllListGenerator().manipulating_downloaded_excel_new(both_query_list,
                                                                         'QUERY2',
                                                                         is_query_question_required=True,
                                                                         page=page)


                    # claim_query_list = claim_query_list_generated
                    # print(claim_query_list)
                    # AllListGenerator().manipulating_downloaded_excel_new(claim_query_list, 'QUERY2', is_query_question_required=True, page=page)

                    AllListGeneratorOld().upload_in_g_sheet('QUERY2')
                    ColourPrint.print_pink('Google_Spreadsheet_Query_pending_updated_NEW')

                    AllListGeneratorOld().upload_in_g_sheet_new('QUERY2')
                    ColourPrint.print_pink('Google_Spreadsheet_Query_pending_updated_NEW_2')

                    AllListGenerator().manipulating_downloaded_excel_new(under_treatment_list_generated,
                                                                         'Pend Dischg2',
                                                                         is_query_question_required=False,
                                                                         page=page)
                    AllListGeneratorOld().upload_in_g_sheet('Pend Dischg2')
                    ColourPrint.print_pink('Google_Spreadsheet_Discharge_pending_updated')

                    AllListGeneratorOld().upload_in_g_sheet_new('Pend Dischg2')
                    ColourPrint.print_pink('Google_Spreadsheet_Discharge_pending_updated_2')

            except Exception as e:
                ColourPrint.print_bg_red('Error Start Below')
                print(f"Error accessing a page: {e}")
                traceback.print_exc()
                ColourPrint.print_bg_red('Error End Above')
        # else:
        #     print('Page NOT Found')





