import asyncio
import sys
import openpyxl
from selenium.common import TimeoutException, NoSuchElementException
from selenium.webdriver.common.by import By
from dkbssy.dk_pages.dk_login_page import Page
from dkbssy.utils.colour_prints import ColourPrint
from selenium.webdriver.support import expected_conditions as EC
from dkbssy.utils.excel_utils import ExcelMethods
from playwright.async_api import async_playwright, Page as Page_play, TimeoutError

class FourthPage(Page):

    def __init__(self, driver, wait):
        super().__init__(driver, wait)

    def emp_code_old(self, code):
        self.driver.find_element(By.NAME, 'ctl00$ContentPlaceHolder1$TextBox1').send_keys(code)
        self.driver.find_element(By.NAME, 'ctl00$ContentPlaceHolder1$search').click()
        new_path = f'//*[@id="ctl00_ContentPlaceHolder1_TextBox1" and @value="{code}"]'
        self.wait.until(EC.presence_of_element_located((By.XPATH, new_path)))

        total_amount_pending = self.driver.find_element(By.XPATH, '//*[@id="ctl00_ContentPlaceHolder1_Label4"]')
        total_paid_cases = self.driver.find_element(By.XPATH, '//*[@id="ctl00_ContentPlaceHolder1_Label14"]')
        total_paid_amount = self.driver.find_element(By.XPATH, '//*[@id="ctl00_ContentPlaceHolder1_Label5"]')
        try:
            number_of_cases_pending = self.driver.find_element(By.XPATH,
                                                               '//*[@id="ctl00_ContentPlaceHolder1_GridView1"]/tbody/tr[2]/td[3]').text
        except NoSuchElementException:
            number_of_cases_pending = 0
        print(code, '\t', number_of_cases_pending, '\t', total_amount_pending.text, '\t', total_paid_cases.text, '\t',
              total_paid_amount.text)
        self.driver.find_element(By.NAME, 'ctl00$ContentPlaceHolder1$TextBox1').clear()
        return code, number_of_cases_pending, total_amount_pending.text, total_paid_cases.text, total_paid_amount.text

    def emp_code(self, code):
        return asyncio.run(self.async_emp_code(code))

    async def async_emp_code(self,code):
        code = str(code)
        play = await async_playwright().start()
        browser = await play.chromium.connect_over_cdp('http://localhost:9222')
        context = browser.contexts[0]
        page = context.pages[0]
        page.set_default_timeout(300000)
        await page.locator("//input[@id='ctl00_ContentPlaceHolder1_TextBox1']").fill(code)
        await page.locator("//input[@id='ctl00_ContentPlaceHolder1_search']").click()
        new_path = f'//*[@id="ctl00_ContentPlaceHolder1_TextBox1" and @value="{code}"]'
        await page.wait_for_selector(new_path)

        total_amount_pending = await page.locator('//*[@id="ctl00_ContentPlaceHolder1_Label4"]').text_content()
        total_paid_cases = await page.locator('//*[@id="ctl00_ContentPlaceHolder1_Label14"]').text_content()
        total_paid_amount = await page.locator('//*[@id="ctl00_ContentPlaceHolder1_Label5"]').text_content()

        try:
            number_of_cases_pending = await (await page.wait_for_selector('//*[@id="ctl00_ContentPlaceHolder1_GridView1"]/tbody/tr[2]/td[3]')).text_content()

        except TimeoutError:
            number_of_cases_pending = 0

        print(code, '\t', number_of_cases_pending, '\t', total_amount_pending, '\t', total_paid_cases, '\t', total_paid_amount)

        await play.stop()

        return code, number_of_cases_pending, total_amount_pending, total_paid_cases, total_paid_amount


    def updating_amount_use(self, final_lol):
        excel_meths_obj = ExcelMethods()
        dict_name_row_number = excel_meths_obj.retrieve_employee_id()  # function ran
        workbook_master = openpyxl.load_workbook(r'G:\My Drive\GdrivePC\Hospital\RSBY\New\Incentive_auto2.xlsx')
        worksheet_for_getting_id = workbook_master["Sheet3"]
        self.driver.get('https://dkbssy.cg.nic.in/secure/incentivemodule/IncentiveDetails_EmpCodeWiseDME.aspx')
        for inc_nam in final_lol[6:]:  # final_lol [case_num, amount,diagnosis..., names1@#,..]
            name = inc_nam.split('@')[0]
            if name in dict_name_row_number.keys():
                chosen_row_number = dict_name_row_number[name]
                chosen_emp_code = worksheet_for_getting_id.cell(row=chosen_row_number, column=3).value
                old_pending_cases = worksheet_for_getting_id.cell(row=chosen_row_number, column=9).value
                chosen_amount = worksheet_for_getting_id.cell(row=chosen_row_number, column=10).value
                print()
                print(name, chosen_emp_code, old_pending_cases, chosen_amount)
                # function ran
                e_code, pend_cases, pend_amount, total_cases, total_amount = self.emp_code(chosen_emp_code)
                worksheet_for_getting_id.cell(row=chosen_row_number, column=8).value = e_code
                worksheet_for_getting_id.cell(row=chosen_row_number, column=9).value = int(pend_cases)
                worksheet_for_getting_id.cell(row=chosen_row_number, column=10).value = float(pend_amount)
                worksheet_for_getting_id.cell(row=chosen_row_number, column=11).value = int(total_cases)
                worksheet_for_getting_id.cell(row=chosen_row_number, column=12).value = float(total_amount)
            else:
                print(name)
                sys.exit(f"THERE OCCURRED AN ERROR, Check {name} in incentive auto 2")
        ColourPrint.print_yellow('Excel Done')

        # updating from database
        results = excel_meths_obj.sqlite_process()
        for employee_name, total_count_till_now, total_incentive_till_now in results:
            chosen_row_number = dict_name_row_number[employee_name]
            worksheet_for_getting_id.cell(row=chosen_row_number, column=13).value = int(total_count_till_now)
            worksheet_for_getting_id.cell(row=chosen_row_number, column=14).value = float(total_incentive_till_now)
            # print(employee_name, total_count_till_now, total_incentive_till_now)

        # updating from database 2
        results2 = excel_meths_obj.sqlite_process_2()  # function ran
        for employee_name, total_count_till_now, total_incentive_till_now in results2:
            chosen_row_number = dict_name_row_number[employee_name]
            worksheet_for_getting_id.cell(row=chosen_row_number, column=15).value = int(total_count_till_now)
            worksheet_for_getting_id.cell(row=chosen_row_number, column=16).value = float(total_incentive_till_now)
        ColourPrint.print_yellow('Database Done 1&2')
        workbook_master.save(r'G:\My Drive\GdrivePC\Hospital\RSBY\New\Incentive_auto2.xlsx')
        workbook_master.close()
        ColourPrint.print_yellow('All Saved')




















