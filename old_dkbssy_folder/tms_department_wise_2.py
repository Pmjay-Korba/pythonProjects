import openpyxl


def department_wise_extract_for_queries():
    pending_discharge_wb = openpyxl.load_workbook(r"G:\My Drive\GdrivePC\Hospital\RSBY\New\down\cc.xlsx")
    # pending_discharge_wb = openpyxl.load_workbook(r'H:\My Drive\GdrivePC\Hospital\RSBY\New\pedia_extract2.xlsx')
    pend_dis_sheet = pending_discharge_wb['QUERY']

    # ayusman_wb = openpyxl.load_workbook(r'D:\Down\AYUSHMAN REGISTRATION 2023.xlsx')
    # # ayusman_wb = openpyxl.load_workbook(r'C:\Users\HP\Downloads\AYUSHMAN REGISTRATION 2023.xlsx')
    ayusman_sheet = pending_discharge_wb['2024']



    '''getting the list of main file card list'''
    row_values = ayusman_sheet.iter_cols(min_col=4, max_col=4)
    for col in row_values:
        main_card = [cell.value for cell in col]

    '''getting list of pending cards'''
    row_values_pend = pend_dis_sheet.iter_cols(max_col=1, min_col=1)
    for col_pend in row_values_pend:
        cards_pend = [cell.value for cell in col_pend]
        # card_row_number = [{cardx.value: cardx.row} for  cardx in col_pend]
        # print(card_row_number)
        d = {}
        for cardx in col_pend:
            d[cardx.value] = cardx.row
        print(d)
        cards_pend_list = list(filter(lambda x: x is not None, cards_pend))
        # print(cards_pend_list)

        '''finding the pending card in main file list'''
        for card_pend_scan in cards_pend_list:
            # if crds == card_pend_scan:
            idx = [idx for idx, valx in enumerate(main_card) if valx == card_pend_scan]
            # main_card.index(crds))
            # print(card_pend_scan, idx)
            y = d[card_pend_scan]
            # print(y)
            z = []
            for i in idx:
                z.append(str(ayusman_sheet.cell(row=i + 1, column=10).value))
            zz = str(' + '.join(z))
            print(card_pend_scan, zz)
            pend_dis_sheet.cell(row=y, column=9).value = zz
        pending_discharge_wb.save(r'G:\My Drive\GdrivePC\Hospital\RSBY\New\down\cc.xlsx')
        pending_discharge_wb.close()

# department_wise_extract()

def department_wise_extract_for_discharge():
    pending_discharge_wb = openpyxl.load_workbook(r"G:\My Drive\GdrivePC\Hospital\RSBY\New\down\cc.xlsx")
    pend_dis_sheet = pending_discharge_wb['Pend Dischg']
    ayusman_sheet = pending_discharge_wb['2024']

    '''getting the list of main file card list'''
    master_all_card_list = list(ayusman_sheet.iter_cols(max_col=4,
                                                        min_col=4))  # return one tuple (<Cell '2024'.D1>, <Cell '2024'.D2>, <Cell '2024'.D3>)
    # print(master_all_card_list)
    all_card_indices_dict = {}
    for card in master_all_card_list[0]:
        # print(card.value, card.row)
        card_value = card.value
        card_index = card.row
        if card_value is not None:
            corresponding_value = ayusman_sheet.cell(row=card_index, column=10).value  # Column J is the 10th column
            if card_value in all_card_indices_dict:
                all_card_indices_dict[card_value].append(corresponding_value)
            else:
                all_card_indices_dict[card_value] = [corresponding_value]

    # print(all_card_indices_dict)

    '''getting list of pending cards'''
    pending_card_list = list(pend_dis_sheet.iter_cols(max_col=1,
                                                      min_col=1))  # return one tuple (<Cell '2024'.D1>, <Cell '2024'.D2>, <Cell '2024'.D3>)
    for pend_card in pending_card_list[0]:

        # Insert the joined corresponding values into column H for matching card values
        for pend_card in pending_card_list[0]:
            pend_card_value = pend_card.value
            pend_card_row = pend_card.row
            if pend_card_value in all_card_indices_dict:  # Check if the pending card is in check_box
                corresponding_values = all_card_indices_dict[pend_card_value]
                if corresponding_values:
                    # Convert all values to strings and filter out any None values
                    joined_values = '+'.join(str(val) for val in corresponding_values if val is not None)
                    # print("______________________________________",joined_values)
                    pend_dis_sheet.cell(row=pend_card_row, column=8, value=joined_values)  # Column H is the 8th column

    # Save the changes to the workbook
    pending_discharge_wb.save(r'G:\My Drive\GdrivePC\Hospital\RSBY\New\down\cc.xlsx')
    pending_discharge_wb.close()

def department_wise_extract_for_queries_2025():
    pending_discharge_wb = openpyxl.load_workbook(r"G:\My Drive\GdrivePC\Hospital\RSBY\New\down\cc.xlsx")
    # pending_discharge_wb = openpyxl.load_workbook(r'H:\My Drive\GdrivePC\Hospital\RSBY\New\pedia_extract2.xlsx')
    pend_dis_sheet = pending_discharge_wb['QUERY2']

    # ayusman_wb = openpyxl.load_workbook(r'D:\Down\AYUSHMAN REGISTRATION 2023.xlsx')
    # # ayusman_wb = openpyxl.load_workbook(r'C:\Users\HP\Downloads\AYUSHMAN REGISTRATION 2023.xlsx')
    # ayusman_sheet_2024 = pending_discharge_wb['2024']

    # wb_2025 = openpyxl.load_workbook(r"G:\My Drive\GdrivePC\Hospital\RSBY\New\down\cc_new.xlsx")
    # modifying the name cc_new to cc
    wb_2025 = openpyxl.load_workbook(r"G:\My Drive\GdrivePC\Hospital\RSBY\New\down\cc.xlsx")
    sheet_2025 = wb_2025['2025']


    '''getting the list of main file card list'''
    row_values = sheet_2025.iter_cols(min_col=3, max_col=3)
    for col in row_values:
        main_card = [cell.value for cell in col]

    '''getting list of pending cards'''
    row_values_pend = pend_dis_sheet.iter_cols(max_col=1, min_col=1)
    for col_pend in row_values_pend:
        cards_pend = [cell.value for cell in col_pend]
        # card_row_number = [{cardx.value: cardx.row} for  cardx in col_pend]
        # print(card_row_number)
        d = {}
        for cardx in col_pend:
            d[cardx.value] = cardx.row
        print(d)
        cards_pend_list = list(filter(lambda x: x is not None, cards_pend))
        # print(cards_pend_list)

        '''finding the pending card in main file list'''
        for card_pend_scan in cards_pend_list:
            # if crds == card_pend_scan:
            idx = [idx for idx, valx in enumerate(main_card) if valx == card_pend_scan]
            # main_card.index(crds))
            # print(card_pend_scan, idx)
            y = d[card_pend_scan]
            # print(y)
            z = []
            for i in idx:
                z.append(str(sheet_2025.cell(row=i + 1, column=9).value))
            zz = str(' + '.join(z))
            print(card_pend_scan, zz)
            pend_dis_sheet.cell(row=y, column=12).value = zz
        pending_discharge_wb.save(r'G:\My Drive\GdrivePC\Hospital\RSBY\New\down\cc.xlsx')
        pending_discharge_wb.close()
        wb_2025.close()


def department_wise_extract_for_discharge_2025():
    pending_discharge_wb = openpyxl.load_workbook(r"G:\My Drive\GdrivePC\Hospital\RSBY\New\down\cc.xlsx")
    pend_dis_sheet = pending_discharge_wb['Pend Dischg2']
    # ayusman_sheet = pending_discharge_wb['2024']

    # wb_2025 = openpyxl.load_workbook(r"G:\My Drive\GdrivePC\Hospital\RSBY\New\down\cc_new.xlsx")
    # modifying the name cc_new to cc
    wb_2025 = openpyxl.load_workbook(r"G:\My Drive\GdrivePC\Hospital\RSBY\New\down\cc.xlsx")
    sheet_2025 = wb_2025['2025']

    '''getting the list of main file card list'''
    master_all_card_list = list(sheet_2025.iter_cols(max_col=3,
                                                        min_col=3))  # return one tuple (<Cell '2024'.D1>, <Cell '2024'.D2>, <Cell '2024'.D3>)
    # print(master_all_card_list)
    all_card_indices_dict = {}
    for card in master_all_card_list[0]:
        # print(card.value, card.row)
        card_value = card.value
        card_index = card.row
        if card_value is not None:
            # getting the ward name from main Ayusman google sheet
            corresponding_value = sheet_2025.cell(row=card_index, column=9).value  # Column J is the 10th column = WARD NAME
            if card_value in all_card_indices_dict:
                all_card_indices_dict[card_value].append(corresponding_value)
            else:
                all_card_indices_dict[card_value] = [corresponding_value]

    # print(all_card_indices_dict)

    '''getting list of pending cards'''
    pending_card_list = list(pend_dis_sheet.iter_cols(max_col=1,
                                                      min_col=1))  # return one tuple (<Cell '2024'.D1>, <Cell '2024'.D2>, <Cell '2024'.D3>)
    for pend_card in pending_card_list[0]:

        # Insert the joined corresponding values into column H for matching card values
        for pend_card in pending_card_list[0]:
            pend_card_value = pend_card.value
            pend_card_row = pend_card.row
            if pend_card_value in all_card_indices_dict:  # Check if the pending card is in check_box
                corresponding_values = all_card_indices_dict[pend_card_value]
                if corresponding_values:
                    # Convert all values to strings and filter out any None values
                    joined_values = '+'.join(str(val) for val in corresponding_values if val is not None)
                    # print("______________________________________",joined_values)
                    pend_dis_sheet.cell(row=pend_card_row, column=12, value=joined_values)  # Column H is the 8th column

    # Save the changes to the workbook
    pending_discharge_wb.save(r'G:\My Drive\GdrivePC\Hospital\RSBY\New\down\cc.xlsx')
    pending_discharge_wb.close()


# def department_wise_extract_for_discharge_for_all(*workbookPath_sheetName_pairs: tuple[str,str])->dict:
#     dict_all = dict()
#     name_count = 0
#     for each_workbook_name, each_sheet_name in workbookPath_sheetName_pairs:
#         each_workbook = openpyxl.load_workbook(each_workbook_name)
#         each_sheet = each_workbook[each_sheet_name]
#         sheet_cols = each_sheet.iter_rows(values_only=True)
#         dict_all[str(name_count)] = list(sheet_cols)
#         name_count += 1
#     return dict_all

