import datetime
import json
import urllib.parse
from logging import exception

import openpyxl
from playwright.sync_api import sync_playwright
from EHOSP.ehosp_2.ui_method import nextgen_ui
from dkbssy.utils.colour_prints import ColourPrint, message_box
from EHOSP.tk_ehosp.alert_boxes import error_tk_box
from TMS_new.async_tms_new.desired_page import get_desired_page_indexes_in_cdp_async_for_SYNC

WARDS = [
    {'ward_id': 4052, 'ward_name': 'New Ward', 'ward_type': 'General ward', 'active_status': 1, 'sex_id': 'B', 'location_id': None},
    {'ward_id': 4053, 'ward_name': 'LOT', 'ward_type': 'General ward', 'active_status': 1, 'sex_id': 'F', 'location_id': None},
    {'ward_id': 4054, 'ward_name': 'EYE Ward', 'ward_type': 'General ward', 'active_status': 1, 'sex_id': 'B', 'location_id': None},
    {'ward_id': 4055, 'ward_name': 'NRC,', 'ward_type': 'General ward', 'active_status': 1, 'sex_id': 'B', 'location_id': None},
    {'ward_id': 4056, 'ward_name': 'SNCU.,', 'ward_type': 'General ward', 'active_status': 1, 'sex_id': 'B', 'location_id': None},
    {'ward_id': 4057, 'ward_name': 'Female Ward', 'ward_type': 'General ward', 'active_status': 1, 'sex_id': 'F', 'location_id': '{"room_name": "21, Ground, MAIN BLOCK, MAIN BUILDING", "room_ids": "814bc1ec-71c2-11ef-9977-0050568818ce", "building_id": "81482794-71c2-11ef-9961-0050568818ce", "building_name": "MAIN BUILDING", "block_id": "8149dab2-71c2-11ef-9962-0050568818ce", "block_name": "MAIN BLOCK", "floor_name": "Ground", "floor_id": "814b7610-71c2-11ef-9963-0050568818ce", "room_number": "21"}'},
    {'ward_id': 4058, 'ward_name': 'Male Ward', 'ward_type': 'General ward', 'active_status': 1, 'sex_id': 'M', 'location_id': None},
    {'ward_id': 4059, 'ward_name': 'RESPIRATORY MEDICINE WARD', 'ward_type': 'General ward', 'active_status': 1, 'sex_id': 'B', 'location_id': '{"room_name": "201, Ground, Isolation Block, NEW BUILDING ", "room_ids": "7c9ad453-637a-4388-bfec-112701a54360", "building_id": "814b9370-71c2-11ef-9965-0050568818ce", "building_name": "NEW BUILDING ", "block_id": "e3878b59-6cfd-453d-a363-d2392fc35de6", "block_name": "Isolation Block", "floor_name": "Ground", "floor_id": "07464630-aa19-41a4-a1ab-198b0b55eade", "room_number": "201"}'},
    {'ward_id': 4060, 'ward_name': 'Emergency Ward', 'ward_type': 'General ward', 'active_status': 1, 'sex_id': 'B', 'location_id': None},
    {'ward_id': 4061, 'ward_name': 'Paying Ward,', 'ward_type': 'General ward', 'active_status': 0, 'sex_id': 'B', 'location_id': None},
    {'ward_id': 4062, 'ward_name': 'Darmatology Ward', 'ward_type': 'General ward', 'active_status': 1, 'sex_id': 'B', 'location_id': None},
    {'ward_id': 4063, 'ward_name': 'Orthopedic ward', 'ward_type': 'General ward', 'active_status': 1, 'sex_id': 'B', 'location_id': None},
    {'ward_id': 4064, 'ward_name': 'psychiatric ward', 'ward_type': 'General ward', 'active_status': 1, 'sex_id': 'B', 'location_id': None},
    {'ward_id': 4065, 'ward_name': 'ENT ward', 'ward_type': 'General ward', 'active_status': 1, 'sex_id': 'B', 'location_id': None},
    {'ward_id': 4066, 'ward_name': 'Oncology', 'ward_type': 'General ward', 'active_status': 1, 'sex_id': 'B', 'location_id': None},
    {'ward_id': 4067, 'ward_name': 'Paediatrics', 'ward_type': 'General ward', 'active_status': 1, 'sex_id': 'B', 'location_id': '{"room_name": "01, 1st, TRAUMA BLOCK, TRAUMA BUILDING", "room_ids": "814bcee4-71c2-11ef-999f-0050568818ce", "building_id": "814ba7b6-71c2-11ef-9968-0050568818ce", "building_name": "TRAUMA BUILDING", "block_id": "814bae6e-71c2-11ef-9969-0050568818ce", "block_name": "TRAUMA BLOCK", "floor_name": "1st", "floor_id": "814bb49a-71c2-11ef-996a-0050568818ce", "room_number": "01"}'},
    {'ward_id': 4068, 'ward_name': 'Burn Ward', 'ward_type': 'General ward', 'active_status': 1, 'sex_id': 'B', 'location_id': None},
    {'ward_id': 4069, 'ward_name': 'Surgical Ward', 'ward_type': 'General ward', 'active_status': 1, 'sex_id': 'B', 'location_id': None},
    {'ward_id': 8558, 'ward_name': 'Casualty ward', 'ward_type': 'Casuality Ward', 'active_status': 1, 'sex_id': 'B', 'location_id': '{"room_name": "16, Ground, MAIN BLOCK, MAIN BUILDING", "room_ids": "814bc02a-71c2-11ef-9973-0050568818ce", "building_id": "81482794-71c2-11ef-9961-0050568818ce", "building_name": "MAIN BUILDING", "block_id": "8149dab2-71c2-11ef-9962-0050568818ce", "block_name": "MAIN BLOCK", "floor_name": "Ground", "floor_id": "814b7610-71c2-11ef-9963-0050568818ce", "room_number": "16"}'}
]

def validate_ipd_as_correct_in_excel(ipd_number):
    # print('Type of the ipd data is: ', type(ipd_number))
    try:
        ipd_number_integer = int(ipd_number)
        return ipd_number_integer
    except ValueError:
        invalid_ipd_number_error = f'The IPD number -> {ipd_number} is not correct format. Only numerical digits are allowed in IPD number'
        error_tk_box(
            error_title='Invalid IPD Number',
            error_message=invalid_ipd_number_error)
        raise ValueError(invalid_ipd_number_error)

def automate_discharge_with_manual_token(token, single_dischargeable_ipd:list, whole_active_patients_data, user_id, context):

    # single_dischargeable_ipd = one of the dischargeable_ipd
    # [
    # [201717, datetime.datetime(2023, 1, 1, 0, 0), 'Bilateral cataracts(H26.9),Senile cataract(H25.9),Cataracts(H26.9)', ('COMPLAIN OF FEVER FOR 2-3 DAYS', 'IV PCM, IV DNS, IV CEFTRIAXONE', 'STABLE', 'FEEL WELL', 'BED REST'), None]
    # [3334242, datetime.datetime(2023, 11, 30, 0, 0), 'Acute gastroenteritis(A09.9)', ('COMPLAIN OF FEVER FOR 2-3 DAYS', 'IV PCM, IV DNS, IV CEFTRIAXONE', 'STABLE', 'FEEL WELL', 'BED REST'), None]
    # [2423211, datetime.datetime(2025, 6, 26, 8, 52, 55, 177491), 'Pyrexia of unknown origin(R50.9),Acute gastroenteritis(A09.9)', ('COMPLAIN OF FEVER FOR 2-3 DAYS', 'IV PCM, IV DNS, IV CEFTRIAXONE', 'STABLE', 'FEEL WELL', 'BED REST'), None]
    # ]


    ipd_number_integer:int = single_dischargeable_ipd[0]

    if not token or "your_token_here" in token:
        print("❌ No token provided. Please paste your Bearer token.")
        return


    # Setup context request object
    request = context.request

    # patient_current_ward_id = select_matching_ipd_from_active_patient_list_return_ward_id(ipd_number_integer=ipd_number_integer, whole_active_patient_data=whole_active_patients_data)  Not required now
    'check weather discharged or not'
    check_discharged_or_not(context=context, token=token, ipd_no=ipd_number_integer)

    patient_current_ward_id, patient_list_resp = get_ward_id_by_ipd(context=context, token=token, ipd_no=ipd_number_integer)
    print('current ward id', patient_current_ward_id)




    headers = {
        "Authorization": token,
        "usertype": "5",  # Add this!
        # "userdepartmentarray": ""  # Keep it empty if unsure — still better than missing
    }
    'getting the names and details of all patient in a particular ward'

    # headers2 = {  # headers added for the changes
    #     "Authorization": token,
    #     "usertype": "5",  # Add this!
    #     "wardid": f"{patient_current_ward_id}",  # new added
    #     "healthfacilityid": "7013"
    #     # "userdepartmentarray": ""  # Keep it empty if unsure — still better than missing
    # }

    # patient_list_resp = request.get(f'https://nextgen.ehospital.gov.in/api/ipd/common/patientByWard/7013/{patient_current_ward_id}', headers=headers)  # previously working now modified and id goes in headers
    # patient_list_resp = request.get(f'https://nextgen.ehospital.gov.in/api/ipd/common/patientByWard', headers=headers2)  # previously working now modified
    # print('---------------', patient_list_resp)

    # if patient_list_resp.ok:
    if patient_list_resp:
        # patient_details = patient_list_resp.json()['result']
        patient_details = patient_list_resp
        print(patient_details)  # PRINTS ALL PATIENT DATA OF THE WARD
    #     '''EACH patient_detail = {
    #         "appellation_code": 2, "appellation_value": "Mrs.", "ipd_id": 11712791, "uhid": "20250068196",
    #         "admno": "250009529", "f_name": "GANGA",
    #         "m_name": null, "l_name": "JANGDE", "admission_date": "06/05/2025", "admission_dept_code": 21,
    #         "admission_dept_name": "Medicine", "attendant_addr": null, "attendant_mobile": null, "attendant_name": "SELF",
    #         "attendant_relation_code": null, "bed_id": null, "bedno": null, "billing_cat": "Exempted", "billing_cat_code": 141,
    #         "billing_flag": null, "date_of_birth": "1983-05-06T00:00:00.000Z", "discharge_status": null, "dob_of_baby": null,
    #         "dob_time_baby": null, "duration_of_pregnancy": null, "gender_code": "F", "gravida": null,
    #         "health_facility_id": 7013, "ipd_doctor_id": "LTB_VG_6134", "mlc": 0, "mlc_number": "",
    #         "mobile": "9999999999", "mother_cr_no": null, "no_of_infants": null, "parity": null,
    #         "patientclass": 2, "pin_code": null, "pregnancy_indicator": null, "visit_id": "221819450",
    #         "unit_id": 20625, "unit_name": "Medicine", "ward_id": 4057, "ward_name": "Female Ward"
    #     }'''

        # ColourPrint.print_pink('-=' * 25)
        # print('Count of Names:-', len(patient_details))
        # ColourPrint.print_pink('-=' * 25)

        # print(patient_details)
        for each_patient_details in patient_details:  # each patient details from all patient list in ward pending discharges
            # ColourPrint.print_turquoise('/..If discharge completed "error" will be display'*2)
            # print(each_patient_details)
            # ColourPrint.print_turquoise('/..If discharge completed "error" will be display'*2)
            if ipd_number_integer == int(each_patient_details["admno"].strip()):
                ipd_id_web = each_patient_details["ipd_id"]
                # ColourPrint.print_pink('='*50)
                # ColourPrint.print_pink('Patient IPD WEB ID')
                # print(ipd_id_web)
                # ColourPrint.print_pink('='*50)
                # ColourPrint.print_turquoise('-'*50)
                # ColourPrint.print_turquoise("Each Patient Data")
                x = dict(sorted(each_patient_details.items(), key=lambda xx: xx[0]))
                # print(x)  # each_patient_details
                # ColourPrint.print_turquoise('-'*50)
                get_patient_status = request.get(f'https://nextgen.ehospital.gov.in/api/ipd/doc/patientInitStatus/7013/{ipd_id_web}',
                                                headers=headers)
                if get_patient_status.ok:
                    patient_data = get_patient_status.json()
                    # ColourPrint.print_yellow('.'*50)
                    # ColourPrint.print_yellow('Patient Initiation Status')
                    # print(patient_data)
                    # ColourPrint.print_yellow('.'*50)

                    discharge_type_code_is = single_dischargeable_ipd[-1]

                    if not patient_data['result']:
                        payload = {
                            "uhid": each_patient_details['uhid'],
                            "ipd_id": each_patient_details['ipd_id'],
                            "health_facility_id": each_patient_details["health_facility_id"],
                            "ward_id": each_patient_details["ward_id"],
                            "bed_id": each_patient_details["bed_id"],
                            "dis_initiated_by": user_id,
                            "dis_type": discharge_type_code_is,
                            "ipd_doc": each_patient_details["ipd_doctor_id"]
                        }
                        print('Payload = ', payload)

                        headers = {
                            "Authorization": token,
                            "Content-Type": "application/json",
                            "usertype": "5",  # Must be included based on previous discovery
                            "userdepartmentarray": ""}

                        resp = request.post(
                            "https://nextgen.ehospital.gov.in/api/ipd/doc/saveDischargeInit",
                            headers=headers,
                            data=json.dumps(payload)
                        )

                        if resp.ok:
                            ColourPrint.print_green("Discharge INITIATED successfully")
                            print(resp.json())
                        else:
                            print("❌ Failed to initiate discharge")
                            print(resp.status, resp.json())
                'Above till here the payload is sent to update the type of discharge'

                headers3 = {  # headers added for the changes
                    "Authorization": token,
                    "usertype": "5",  # Add this!
                    "healthfacilityid": "7013",
                    "ipdid":f"{ipd_id_web}"  # new added
                    # "userdepartmentarray": ""  # Keep it empty if unsure — still better than missing
                }
                'getting all discharges pending of a ward --- This is from PREPARE SUMMARY PAGE after INITIATION PAGE discharge type update'
                # discharge_list_of_ward_web = request.get(f'https://nextgen.ehospital.gov.in/api/ipd/doc/getDisSummary?ipd_id={ipd_id_web}&health_facility_id=7013', headers=headers)
                discharge_list_of_ward_web = request.get(f'https://nextgen.ehospital.gov.in/api/ipd/doc/getDisSummary', headers=headers3)  # modified for the new pattern change in website
                print('========', discharge_list_of_ward_web)

                # 'https://nextgen.ehospital.gov.in/api/ipd/doc/getDisSummary?ipd_id=6368234&health_facility_id=7013'
                '''This below each patient are the list of patient whose discharge summary need to be prepare after discharge type initiation in initiation page'''
                each_patient_summary_json = discharge_list_of_ward_web.json()['result']["summary"][0]
                # each_patient_summary_json =
                #{ "recordid": 6022261,     "health_facility_name": null, "health_facility_id": 7013,
                # "ipd_id": "6368235",    "uhid": "20170004506",    "admno": "2017566",    "appellation_code": 3,  "appellation_value": "Miss.",
                # "f_name": "PHOOL ", "m_name": "",  "l_name": "KUNWAR", "guardian_name": "0~0~00~0~",    "guardian_rel_code": null, "gender_code": "F",  "date_of_birth": "2002-11-17T00:00:00.000Z", "mobile": "",
                # "admission_dept_code": 21,  "billing_flag": 1, "disease": null, "ward_id": 4057,  "ward_name": "Female Ward",    "bed_id": null, "bedno": null,  "admission_date": "17/11/2017 12:42:34", "dis_datetime": "01/12/2017 00:00:00",  "discharge_status": null,
                # "billing_cat": "Student",  "mlc": 0, "mlc_number": null,  "dis_type_code": 1, "dis_type_name": "Cured",    "treatment_details": "medications",  "pat_condition_on_adm": "sick", "diagnosis": "Fever(R50.9)", "pat_condition_on_dis": "stable",  "follow_up": null,
                # "brief_summary": "fell well", "dis_summary_status": 1,  "ipd_doc": "0", "dis_prepared_by": "Neha_701336",    "dis_prepared_datetime": "21/06/2025 06:57 PM",  "dis_verified_datetime": null, "dis_verified_by": null, "is_pending_summ": null,  "health_id": "",
                # "health_id_number": "8337", "pat_address": "{\"address_line\": \"PRE MATRIC HOSTEL KORKOMA \", \"state_code\": 22, \"district_code\": null, \"pincode\": \"\"}",
                # "admission_dept_name": "Medicine"

                # ColourPrint.print_blue('-'*50)
                # ColourPrint.print_blue('Patient Summary Page Details - Whose discharge summery need to be filled')
                # y = dict(sorted(each_patient_summary_json.items(), key=lambda yy:yy[0])) # print(each_patient_summary_json)
                # print(json.dumps(each_patient_summary_json, indent=4))
                # ColourPrint.print_blue('-'*50)

                'checking the range of discharge date'
                verified_discharge_date_range_dt = check_discharge_date_range(each_patient_summary_json, single_dischargeable_ipd)


                'entry the payload data in summary'
                summary_entry(each_patient_summary_json=each_patient_summary_json,
                              single_dischargeable_ipd_entry=single_dischargeable_ipd,
                              verified_discharge_date_is=verified_discharge_date_range_dt,  # verified for the range checking
                              request=request,
                              user_id=user_id,
                              token=token)

def validate_discharge_type(single_dischargeable_ipd):
    discharge_type_text = single_dischargeable_ipd[4]
    # ColourPrint.print_bg_red(discharge_type_text)
    if discharge_type_text is None:
        dis_err_msg = f'The discharge type in IPD -> {single_dischargeable_ipd[0]} is NOT mentioned. It must be any of following:\n"D : Discharge",\n"L : LAMA",\n"M : DAMA",\n"T : Transfer",\n"R : Refer",\n"A : Abscond"'
        error_tk_box(error_title="Discharge Type Error", error_message=dis_err_msg)
        raise ValueError(dis_err_msg)
    else:
        strip_discharge_type_text = str(discharge_type_text).strip().upper()
        dis_code = ['D', 'L', 'M', 'T', 'R', 'A']
        if strip_discharge_type_text not in dis_code:
            dis_err_msg = f'The discharge type in IPD -> {single_dischargeable_ipd[0]} is NOT correct. It must be any of following:\n"D : Discharge",\n"L : LAMA",\n"M : DAMA",\n"T : Transfer",\n"R : Refer",\n"A : Abscond"'
            error_tk_box(error_title="Discharge Type Error", error_message=dis_err_msg)
            raise ValueError(dis_err_msg)
        '''     "dischargetypecode": 1,
                "dischargetypename": "Cured"
                "dischargetypecode": 3,
                "dischargetypename": "Transferred"
                "dischargetypecode": 4,
                "dischargetypename": "DAMA Discharge against Medical Advice"
                "dischargetypecode": 5,
                "dischargetypename": "LAMA Leave against Medical Advice"
                "dischargetypecode": 6,
                "dischargetypename": "Absconded"
                "dischargetypecode": 7,
                "dischargetypename": "Referral"

        '''
        mapper = {'D':"1" , 'L':"5" , 'M':"4" , 'T':"3" , 'R':"7" , 'A':"6" }
        dis_code_is = mapper[strip_discharge_type_text]
        print('Discharge Code: ', strip_discharge_type_text, dis_code_is)
        return dis_code_is



def summary_entry(each_patient_summary_json:dict, single_dischargeable_ipd_entry, user_id, verified_discharge_date_is, request, token):
    # single_dischargeable_ipd_entry = [2017287, datetime.datetime(2017, 10, 15, 0, 0), ['Pyrexia of unknown origin, Anemia'], ('C/O VOMITING AND LOOSE STOOLS, FEVER', 'IV PANTOP, IV PCM, IV ANTIBIOTICS', 'VITAL STABLE', 'FEEL WELL', 'BED REST'), None]

    processed_diagnosis = single_dischargeable_ipd_entry[2]
    processed_treatment = single_dischargeable_ipd_entry[3][1]
    processed_discharge_condition = single_dischargeable_ipd_entry[3][2]
    processed_brief_summary = single_dischargeable_ipd_entry[3][3]
    processed_follow_up = single_dischargeable_ipd_entry[3][4]
    processed_discharge_date_is = verified_discharge_date_is

    # modifying for the diagnosis entry in condition on admission->if blank, else the value provided
    processed_admission_condition = single_dischargeable_ipd_entry[3][0]
    if processed_admission_condition is None:
        processed_admission_condition = processed_diagnosis
    elif str(processed_admission_condition).strip() == "":
        processed_admission_condition = processed_diagnosis

    if not each_patient_summary_json["brief_summary"] :
        payload = {
            "summary": {
                "ipd_id": each_patient_summary_json["ipd_id"],
                "health_facility_id": each_patient_summary_json["health_facility_id"],
                "recordid": each_patient_summary_json["recordid"],
                "diagnosis": f"{processed_diagnosis}",
                "treatment_details": f"{processed_treatment}",
                "pat_condition_on_adm": f"{processed_admission_condition}",
                "pat_condition_on_dis": f"{processed_discharge_condition}",
                "brief_summary": f"{processed_brief_summary}",
                "follow_up": f"{processed_follow_up}",
                "dis_type_code": each_patient_summary_json[ "dis_type_code"],
                "discharge_date": f"{processed_discharge_date_is}",  # "dis_datetime": "23/06/2025 09:32:00", "admission_date": "22/06/2025 22:24:47",
                "dis_prepared_by": user_id
            },
            "med": []
        }

        '''NEW FORMAT FOR PAYLOAD'''
        payload3 = {
            "summary": {
                "asst_surgeon": None,
                "brief_summary": f"{processed_brief_summary}",
                "diagnosis": f"{processed_diagnosis}",
                "dis_prepared_by": user_id,
                "dis_type_code": each_patient_summary_json[ "dis_type_code"],
                "discharge_date": f"{processed_discharge_date_is}",
                "follow_up": f"{processed_follow_up}",
                "health_facility_id": each_patient_summary_json["health_facility_id"],
                "ipd_id": each_patient_summary_json["ipd_id"],
                "operation_date": None,
                "operative_finding": None,
                "pat_condition_on_adm": f"{processed_admission_condition}",
                "pat_condition_on_dis": f"{processed_discharge_condition}",
                "procedure_name": None,
                "recordid": each_patient_summary_json["recordid"],
                "senior_resident": None,
                "surgeon": None,
                "treatment_details": f"{processed_treatment}"
            },
            "med": []
        }

        print(payload3)

        headers = {
            "Authorization": token,
            "Content-Type": "application/json",
            "usertype": "5",  # Must be included based on previous discovery
            "userdepartmentarray": ""}

        add_summary_post = request.post('https://nextgen.ehospital.gov.in/api/ipd/doc/addDischargeSummary',
                                       headers=headers,
                                        data = json.dumps(payload3)
                                        )
        if add_summary_post.ok:
            ColourPrint.print_green('=.'*50)
            ColourPrint.print_green('Discharge Completed')
            print(add_summary_post.json())
            ColourPrint.print_green('=.' * 50)
        else:
            ColourPrint.print_bg_red('Not able to complete discharge')
            print(add_summary_post.status, add_summary_post.json())

def check_discharge_date_range(each_patient_summary_json:dict, single_dischargeable_ipd:list): #  one of dischargeable_ipds =
    # [
    # [2017237, datetime.datetime(2017, 1, 20, 0, 0), ['Acute gastroenteritis', 'Pyrexia of unknown origin'], ('C/O VOMITING AND LOOSE STOOLS, FEVER', 'IV PANTOP, IV PCM, IV ANTIBIOTICS', 'VITAL STABLE', 'FEEL WELL', 'BED REST')]
    # [2017287, datetime.datetime(2017, 10, 15, 0, 0), ['Pyrexia of unknown origin'], ('C/O VOMITING AND LOOSE STOOLS, FEVER', 'IV PANTOP, IV PCM, IV ANTIBIOTICS', 'VITAL STABLE', 'FEEL WELL', 'BED REST')]
    # ]
    """
    Checking the range of discharge date
    :param each_patient_summary_json: json for parsing the admission date
    :param single_dischargeable_ipd: list from Excel to get discharge date
    :return: modified discharge date if modified or same discharge date
    """
    datetime_discharge_date = single_dischargeable_ipd[1]
    admission_date = each_patient_summary_json["admission_date"]  # "admission_date": "17/11/2017 12:42:34"
    datetime_admission_date = datetime.datetime.strptime(admission_date, "%d/%m/%Y %H:%M:%S")


    if datetime_discharge_date.date() == datetime_admission_date.date():
        ColourPrint.print_bg_red('Admission and Discharge dates are SAME')

        if datetime_discharge_date.date() == datetime.datetime.now().date():
            # print(']]]]]]]]]]]]', datetime_discharge_date.date() == datetime.datetime.now().date())
            # print(']]]]]]]]]]]]', datetime_discharge_date.date() )
            # print(']]]]]]]]]]]]', datetime.datetime.now().date())
            same_date_err = f'The admission and discharge date in IPD -> {single_dischargeable_ipd[0]} is TODAY. Kindly discharge tomorrow or do it manually for now'
            error_tk_box(error_title="discharge Date Error", error_message= same_date_err)
            raise ValueError(same_date_err)

        # print(datetime.datetime.combine(datetime_discharge_date.date(), datetime.time(23, 59, 59)))
        return datetime.datetime.combine(datetime_discharge_date.date(), datetime.time(23, 59, 59))

    elif datetime_discharge_date < datetime_admission_date:
        dis_before_adm_date_error = f"The discharge date in IPD: {single_dischargeable_ipd[0]} -> {datetime_discharge_date.strftime('%d/%m/%Y %H:%M:%S')} is BEFORE the admission date -> {admission_date}. Discharge date must be later than admission date."
        error_tk_box(
            error_title="Discharge Date Error",
            error_message=dis_before_adm_date_error)
        raise ValueError(dis_before_adm_date_error)

    elif datetime_discharge_date >= datetime.datetime.now():
        dis_mor_than_today_error = f"The discharge date in IPD: {single_dischargeable_ipd[0]} -> {datetime_discharge_date.strftime('%d/%m/%Y %H:%M:%S')} is AFTER the TODAY date. Discharge date must not be more than today."
        error_tk_box(
            error_title='Discharge Date Error',
            error_message=dis_mor_than_today_error)
        raise ValueError(dis_mor_than_today_error)
    # return datetime_discharge_date

    final_datetime_discharge_date = datetime.datetime.combine(datetime_discharge_date.date(), datetime.datetime.now().time().replace(microsecond=0)) - datetime.timedelta(minutes=30) # subtracted to prevent timestamp server check error

    ColourPrint.print_yellow('<>' * 50)
    print('admission_time', datetime_admission_date, 'discharge_date', final_datetime_discharge_date)
    ColourPrint.print_yellow('<>' * 50)

    return final_datetime_discharge_date

def get_all_active_patients_list(context, token: str):
    url = "https://nextgen.ehospital.gov.in/api/ipd/common/reportPatientList/7013"

    headers = {
            "Authorization": token,
            "Content-Type": "application/json",
            "usertype": "5",  # Must be included based on previous discovery
            "userdepartmentarray": ""
    }

    response = context.request.get(url, headers=headers, timeout= 120_000)
    # print('>>>>>>>> ALL PATIENT DATA', response)

    if response.ok:
        '''searching the ipd'''
        all_patient_list = response.json()
        _active_patients_data = all_patient_list["result"]
        # print('active patient data', _active_patients_data)
        return _active_patients_data
    else:
        print(f"❌ Error {response.status}: {response.status_text}")
        return None

def get_ward_id_by_ipd(context, token: str, ipd_no):
    url = 'https://nextgen.ehospital.gov.in/api/ipd/common/patientbyIdMN'
    headers4 = {
        "Authorization": token,
        "Content-Type": "application/json",
        "usertype": "5",  # Must be included based on previous discovery
        "userdepartmentarray": ""
    }

    payload = {"pat_ipdid": f"{ipd_no}",
               "SearchCri": "IPDID",
               "health_facility_id": 7013
    }

    response = context.request.post(url=url, data=payload, headers=headers4)
    if response.ok:
        patient_data = response.json()
        print("==>>>>>>", patient_data)

        result = patient_data.get("result", [])
        if result:  # check if list is not empty

            # if result[0]['dis_summary_status'] == 2:
            #     error_tk_box(error_title='Error',
            #                  error_message=f"The IPD:{ipd_no}' might be already discharged. Please Check")
            #     raise ValueError(f"The IPD:{ipd_no}' might be already discharged. Please Check")

            ward_id = result[0]["ward_id"]
            return ward_id, result
        else:
            error_tk_box(error_title='Error',
                         error_message=f'The ward name and ID could not be fetched from website for IPD:{ipd_no}. Check IPD number is correct.')
            raise ValueError(f'The ward name and ID could not be fetched from website for IPD:{ipd_no}')

    else:
        print(f"Request failed: {response.status}")
        return None






def check_discharged_or_not(context, token: str, ipd_no):
    url = f"https://nextgen.ehospital.gov.in/api/ipd/doc/printSummStatus/7013/{ipd_no}"
    headers = {
        "Authorization": token,
        "Content-Type": "application/json",
        "usertype": "5",  # Must be included based on previous discovery
        "userdepartmentarray": ""
    }

    response = context.request.get(url, headers=headers, timeout=120_000)
    if response.ok:
        patient_data = response.json()
        print("=======>", patient_data)

        result = patient_data.get("result", [])
        if result:  # check if list is not empty

            if result[0]['dis_summary_status']==2:
                error_tk_box(error_title='Error',
                             error_message=f"The IPD:{ipd_no}' might be already discharged. Please Check")
                raise ValueError(f"The IPD:{ipd_no}' might be already discharged. Please Check")
        return None

    else:
        print(f"Request failed: {response.status}")
        return None


def get_fresh_token_and_userid(context):
    page = context.new_page()
    try:

        token_data = {}

        def capture_request(request):
            if "api/user_mgmt/v1/user_project_menus" in request.url:
                token_data["headers"] = request.headers
                # ColourPrint.print_yellow(token_data)
                # for k, v in request.headers.items():
                    # ColourPrint.print_blue(f"{k}: {v}")

        page.on("request", capture_request)

        # Go to the page that triggers the API call
        page.goto("https://nextgen.ehospital.gov.in/adminHome")

        page.wait_for_selector("//span[normalize-space()='IPD']").click()

        # Wait until that specific request is made
        context.wait_for_event(
            "request",
            lambda req: "api/user_mgmt/v1/user_project_menus" in req.url
        )
        return token_data['headers']['authorization'], token_data['headers']['userid']
    except Exception as e:
        print(e)
    finally:
        page.close()

# NOT USED
def select_matching_ipd_from_active_patient_list_return_ward_id(ipd_number_integer:int, whole_active_patient_data ):
    for individual_patient_data in whole_active_patient_data:
        '''individual_patient_data = 
            {
                "appellation_value": "Ms.",
                "uhid": "20230073318",
                "ipdid": "202310428",
                "full_name": "USHA SONI ",
                "gender_code": "F",
                "date_of_birth": "1981-06-26T00:00:00.000Z",
                "admission_dept_code": 26,
                "ward_id": 4057,
                "ward_name": "Female Ward",
                "unit_name": "Obs AND Gyne",
                "bedno": null,
                "admission_date": "26/06/2023 05:09 PM",
                "attendant_name": "test",
                "discharge_status": null,
                "health_facility_id": 7013,
                "ipd_doctor_id": "0",
                "mlc": 0,
                "mobile": ""
            },
            '''
        if int(individual_patient_data["ipdid"]) == ipd_number_integer:
            patient_ward_id_is = individual_patient_data["ward_id"]
            ColourPrint.print_turquoise('()'*50)
            print('Name:', individual_patient_data['full_name'],'|', 'IPD:', individual_patient_data['ipdid'], '|', 'ward name: ', individual_patient_data["ward_name"], '|', 'Admission Date:', individual_patient_data["admission_date"])
            ColourPrint.print_turquoise('()'*50)
            return patient_ward_id_is
    error_tk_box(error_title='Error',
                 error_message=f'The ward name and ID could not be fetched from website for IPD:{ipd_number_integer}. Check IPD number is correct, or the IPD might be already discharged. ')
    raise ValueError(f'The ward name and ID could not be fetched from website for IPD:{ipd_number_integer}')


def fetch_excel(workbook_path):
    wb = openpyxl.load_workbook(workbook_path, data_only=True)
    main_sheet = wb['MAIN']
    main_sheet_rows = main_sheet.iter_rows(min_row=2, max_col=5, values_only=True)
    # print(main_sheet_rows)  # <generator object Worksheet._cells_by_row at 0x000001ED0B697640>

    "not required now"
    # """getting data from USER sheet"""
    # user_sheet = wb['USER']
    # user_id_of_excel = user_sheet.cell(row=3, column=2).value.strip()
    # # ward_of_excel = user_sheet.cell(row=21, column=28).value
    # # ward_id_of_excel = user_sheet.cell(row=22, column=28).value
    # print(user_id_of_excel)

    """getting data from DIAG_CODE sheet"""
    diagnosis_sheet = wb['DIAG_CODE']
    diagnosis_rows = list(diagnosis_sheet.iter_rows(min_row=2, values_only=True))
    # print(diagnosis_rows)   # [('aa', 'Acute gastroenteritis'),
    diagnosis_dict = {i[0].lower(): i[1].strip() for i in diagnosis_rows if i[1] and i[1].strip()}
    ColourPrint.print_yellow('--'*50)
    print(diagnosis_dict)
    ColourPrint.print_yellow('--'*50)

    # """getting data from TREATMENT sheet"""
    # treatment_sheet = wb['TREATMENT']
    # treatment_rows = list(treatment_sheet.iter_rows(min_row=2, max_col=6, values_only=True))
    # treatment_dict = {i[0]:i[1:] for i in treatment_rows}
    # ColourPrint.print_blue('-=' * 50)
    # print(treatment_dict)
    # ColourPrint.print_blue('-=' * 50)

    from itertools import islice  # helps keep the row number editable if you ever change min_row

    """ modified getting data from TREATMENT sheet for checking the None Values"""
    treatment_sheet = wb['TREATMENT']
    treatment_rows = list(treatment_sheet.iter_rows(min_row=2, max_col=6, values_only=True))

    # treatment_dict = {}
    # for excel_row_num, row in zip(range(2, 2 + len(treatment_rows)), treatment_rows):
    #     key, *values = row
    #
    #     # ── 1. Blank row: key is None and *every* value is None ────────────────────────
    #     if key is None and all(v is None for v in values):
    #         continue  # skip silently ─ it's just an empty row
    #
    #     # ── 2. Key missing but some value(s) filled in ────────────────────────────────
    #     if key is None:
    #         error_tk_box(error_title='Excel Sheet Entry Error',
    #                      error_message=f"TREATMENT sheet ‑ row {excel_row_num}: Code (first cell) is empty but one or more value cells are not → {values}"
    #                      )
    #         raise KeyError(
    #             f"TREATMENT sheet ‑ row {excel_row_num}: Code (first cell) is empty "
    #             f"but one or more value cells are not → {values}"
    #         )
    #
    #     # ── 3. Key present but at least one value is missing ──────────────────────────
    #     if any(v is None for v in values):
    #         error_tk_box(error_title='Excel Sheet Entry Error',
    #                      error_message=f"TREATMENT sheet ‑ row no. {excel_row_num}: Code number -> '{key}' has missing data in one or more value cells → {values}")
    #         raise ValueError(
    #             f"TREATMENT sheet ‑ row no. {excel_row_num}: Code number -> '{key}' has missing "
    #             f"data in one or more value cells → {values}"
    #         )
    #
    #     # ── 4. Fully valid row ────────────────────────────────────────────────────────
    #     treatment_dict[key] = tuple(values)  # or keep them as list if you prefer

    treatment_dict = {}
    for excel_row_num, row in zip(range(2, 2 + len(treatment_rows)), treatment_rows):
        key, *values = row

        # ── 1. Blank row: key is None and *every* value is None ────────────────────────
        if key is None and all(v is None for v in values):
            continue  # skip silently ─ it's just an empty row

        # ── 2. Key missing but some value(s) filled in ────────────────────────────────
        if key is None:
            error_tk_box(
                error_title='Excel Sheet Entry Error',
                error_message=f"TREATMENT sheet - row {excel_row_num}: Code (first cell) is empty but one or more value cells are not → {values}"
            )
            raise KeyError(
                f"TREATMENT sheet - row {excel_row_num}: Code (first cell) is empty "
                f"but one or more value cells are not → {values}"
            )

        # ── 3. Key present but missing values ──────────────────────────────────────────
        # allow 2nd column (values[0]) to be None, but others (values[1:]) must not be None
        if any(v is None for v in values[1:]):
            error_tk_box(
                error_title='Excel Sheet Entry Error',
                error_message=f"TREATMENT sheet - row no. {excel_row_num}: Code number -> '{key}' has missing data in one or more mandatory value cells → {values}"
            )
            raise ValueError(
                f"TREATMENT sheet - row no. {excel_row_num}: Code number -> '{key}' has missing "
                f"data in one or more mandatory value cells → {values}"
            )

        # ── 4. Fully valid row ────────────────────────────────────────────────────────
        treatment_dict[key] = tuple(values)

    ColourPrint.print_blue('-=' * 50)
    print(treatment_dict)

    dischargeable_ipd = []
    # ColourPrint.print_green('-:' * 50)
    for each_row in main_sheet_rows:
        if each_row[0] is not None:
            each_row_list = list(each_row)
            """each row list [250015223, '26.07.2025', 'bb,aa', 2, None]"""
            new_each_row_list =[each_row[0]]  # each_row[0] = ipd no # new list for fresh entry of modified each data
            print('each row list', each_row_list)

            """making changes to discharge date"""
            discharge_date_verified = validate_discharge_date(each_row_list)
            new_each_row_list.append(discharge_date_verified)

            """making changes to the diagnosis from diagnosis codes"""
            # checking the diagnosis mentioned or not, function run
            many_diagnosis = validate_and_split_multi_diagnosis_new(each_row_list=each_row_list)  # 'aa', 'bb', 'aa,bb', AA, Bb
            multi_diag_list = []  # in order to achieve multi diagnosis
            for each_diag_codes in many_diagnosis:  # 'aa', 'bb', 'aa,bb', AA, Bb
                try:
                    multi_diag_list.append(diagnosis_dict[each_diag_codes.lower().strip()])
                except KeyError:
                    diagnosis_key_error_message = f'The DIAGNOSIS code -> "{each_diag_codes}" in IPD: {each_row_list[0]} is not present in DIAG_CODE EXCEL SHEET. Please Check'
                    error_tk_box(error_title="Diagnosis Code Error",
                                 error_message=diagnosis_key_error_message)
                    raise KeyError(diagnosis_key_error_message)

            """Place here the ICD code update function"""

            new_each_row_list.append(multi_diag_list)
            """Making changes to receive the treatment list"""
            treatment_verified_code = validate_treatment_codes_new(each_row_list)
            try:
                new_each_row_list.append(treatment_dict[treatment_verified_code])
            except KeyError:
                raise KeyError(f'The TREATMENT code -> "{treatment_verified_code}" in IPD: {each_row_list[0]} is not present in TREATMENT EXCEL SHEET. Please Check')

            # adding the discharge type
            new_each_row_list.append(each_row_list[-1])

            dischargeable_ipd.append(new_each_row_list)
    print('------>>>>', dischargeable_ipd)
    return dischargeable_ipd



def validate_and_split_multi_diagnosis_new(each_row_list:list):
    diagnosis_code = each_row_list[2]  # 'aa', 'bb', 'aa,bb', AA, Bb
    # print(diagnosis_code)
    if diagnosis_code:
        if type(diagnosis_code) == str:
            return diagnosis_code.split(',')  # non lower containing
        else:
            err_msg= f'The DIAGNOSIS code -> "{diagnosis_code}" is invalid in IPD: {each_row_list[0]}. It must be alphabet ot two character like aa, bb'
            error_tk_box(error_title='Excel Diagnosis Error',
                         error_message=err_msg)
            raise ValueError(err_msg)
    else:
        no_diagnosis_err_msg = f'No DIAGNOSIS code provided in MAIN Sheet in IPD: {each_row_list[0]}'
        error_tk_box(error_title="Main Sheet Diagnosis Error",
                     error_message=no_diagnosis_err_msg)
        raise ValueError(no_diagnosis_err_msg)


def validate_treatment_codes_new(each_row_list:list):
    treatment_code = each_row_list[3]
    if treatment_code:
        if type(treatment_code) == int:
            return treatment_code
        elif type(treatment_code) == str and treatment_code.isdigit():
            return int(treatment_code)
        else:
            incorrect_treat_code_error = f'The Provided TREATMENT code is not correct in IPD: {each_row_list[0]}. Provide the digits only in codes'
            error_tk_box(error_title="Treatment Code Error",
                         error_message=incorrect_treat_code_error)
            raise  ValueError(incorrect_treat_code_error)
    else:
        no_treatment_provided_error = f'No TREATMENT code provided in IPD: {each_row_list[0]}'
        error_tk_box(error_title="Treatment Code Error",
                     error_message=no_treatment_provided_error)
        raise ValueError(no_treatment_provided_error)


def yyyy_date_from_yy(yy_char_only:str, each_row_list)->str:  # ie only '25 or '2025'
    year = None
    if len(yy_char_only) == 2:
        year =  '20'+yy_char_only
    elif len(yy_char_only) == 4:
        year =  yy_char_only
    else:
        year_format_error = f'The DISCHARGE DATE "YEAR" -> "{yy_char_only}" is invalid in IPD: {each_row_list[0]}. Must be like either "25" or "2025"'
        error_tk_box(error_title="Discharge Date Error",
                     error_message=year_format_error)
        raise ValueError(year_format_error)

    current_year = datetime.datetime.now().year
    current_year_int = int(current_year)
    if int(year) >  current_year_int:
        over_year_date_error = f'The "YEAR" in Discharge date in IPD: {each_row_list[0]} -> {yy_char_only} is greater then current year, which is not possible. Please Check'
        error_tk_box(error_title="Discharge Date Error",
                     error_message= over_year_date_error)
        raise ValueError(over_year_date_error)

    return year

def validate_discharge_date(each_row_list):
    discharge_date =  each_row_list[1]
    # print('Original discharge date =', discharge_date)
    if discharge_date:
        discharge_dt = None
        # if type(discharge_date) == datetime.datetime:
        if isinstance(discharge_date, datetime.datetime):
            discharge_dt = discharge_date
            # print('Print from Validate date', discharge_dt)
            # return discharge_dt

        elif type(discharge_date) == str:
            discharge_date = discharge_date.strip()
            if discharge_date.lower() == "t":
                discharge_dt = datetime.datetime.now().replace(microsecond=0) - datetime.timedelta(minutes=30) # subtracted to prevent timestamp server check error
                # print(discharge_dt)

            elif discharge_date.count('-') == 2:  # 22-12-25
                discharge_dt = get_dash_date(discharge_date,each_row_list)

            elif discharge_date.count('.') == 2:  # 22/12/25
                discharge_dt =get_dot_date(discharge_date, each_row_list)

            elif discharge_date.count('/') == 2:  # 22.12.25
                discharge_dt = get_oblique_date(discharge_date, each_row_list)

            else:
                wrong_dis_date_format_error = f'The DISCHARGE DATE -> "{discharge_date}" provided is in incorrect format in IPD: {each_row_list[0]}.\nIt can be either dd-mm-yy or dd.mm.yy or dd/mm/yy or dd-mm-yyyy or dd.mm.yyyy or dd/mm/yyyy or "t" for today'
                error_tk_box(error_title="Discharge Date Error",
                             error_message=wrong_dis_date_format_error)
                raise ValueError(wrong_dis_date_format_error)


            # formatted_discharge_dt = process_date_time(discharge_dt)

            # print(formatted_discharge_dt)
            # ColourPrint.print_pink("=.-"*50)
        else:
            invalid_discharge_DATE_error = f'The DISCHARGE DATE -> "{discharge_date}" provided is invalid for ID: {each_row_list[0]}. Use proper date format.\nIt can be either dd-mm-yy or dd.mm.yy or dd/mm/yy or dd-mm-yyyy or dd.mm.yyyy or dd/mm/yyyy or "t" for today'
            error_tk_box(error_title='INVALID DISCHARGE DATE ERROR',
                         error_message=invalid_discharge_DATE_error)
            raise ValueError(invalid_discharge_DATE_error)
        # formatted_discharge_date = process_date_time(discharge_dt)

        return discharge_dt

    else:
        no_discharge_date_provided_error = f'No DISCHARGE DATE provided in IPD: {each_row_list[0]}'
        error_tk_box(error_title="No Discharge Date Error",
                     error_message=no_discharge_date_provided_error)
        raise ValueError(no_discharge_date_provided_error)

def get_oblique_date(discharge_date,each_row_list):
    if len(discharge_date.split('/')) == 3:  # 22-12-25
        d, m, y = discharge_date.split('/')
        y_str = yyyy_date_from_yy(y, each_row_list=each_row_list)
        discharge_dt = datetime.datetime(year=int(y_str), month=int(m), day=int(d))
        return discharge_dt
    else:
        date_format_error_oblique = f'The DISCHARGE DATE -> "{discharge_date}" provided is invalid for ID: {each_row_list[0]}. Use proper date format.\nIt can be either dd-mm-yy or dd.mm.yy or dd/mm/yy or dd-mm-yyyy or dd.mm.yyyy or dd/mm/yyyy or "t" for today'
        error_tk_box(error_title="Discharge Date Format Error O",
                     error_message=date_format_error_oblique)
        raise ValueError(date_format_error_oblique)


def get_dot_date(discharge_date, each_row_list):
    if len(discharge_date.split('.')) == 3:  # 22-12-25
        d, m, y = discharge_date.split('.')
        y_str = yyyy_date_from_yy(y, each_row_list=each_row_list)
        discharge_dt = datetime.datetime(year=int(y_str), month=int(m), day=int(d))
        return discharge_dt
    else:
        invalid_date_error_dot = f'The DISCHARGE DATE -> "{discharge_date}" provided is invalid for ID: {each_row_list[0]}. Use proper date format.\nIt can be either dd-mm-yy or dd.mm.yy or dd/mm/yy or dd-mm-yyyy or dd.mm.yyyy or dd/mm/yyyy or "t" for today'
        error_tk_box(error_title='Discharge Date Format Error Do',
                     error_message=invalid_date_error_dot)
        raise ValueError(invalid_date_error_dot)

def get_dash_date(discharge_date,each_row_list):
    if len(discharge_date.split('-')) == 3:  # 22-12-25
        d, m, y = discharge_date.split('-')
        y_str = yyyy_date_from_yy(y, each_row_list=each_row_list)
        discharge_dt = datetime.datetime(year=int(y_str), month=int(m), day=int(d))
        return discharge_dt
    else:
        invalid_date_format_dash_error = f'The DISCHARGE DATE -> "{discharge_date}" provided is invalid for ID: {each_row_list[0]}. Use proper date format.\nIt can be either dd-mm-yy or dd.mm.yy or dd/mm/yy or dd-mm-yyyy or dd.mm.yyyy or dd/mm/yyyy or "t" for today'
        error_tk_box(error_title="Discharge Date Format Error Ds",
                     error_message=invalid_date_format_dash_error)
        raise ValueError(invalid_date_format_dash_error)

def process_date_time(discharge_dt):
    # discharge_dt = discharge_dt.replace(microsecond=0)
    # ColourPrint.print_pink("=/-" * 50)
    # print(discharge_dt)
    formatted_discharge_dt = discharge_dt.strftime("%d/%m/%Y %H:%M:%S")
    # print(formatted_discharge_dt)
    # ColourPrint.print_blue("=/-" * 50)
    return formatted_discharge_dt


def search_snomed_js(page, term: str):
    # Properly encode the term (e.g., spaces → %20)
    encoded_term = urllib.parse.quote(term)

    # Build the full JSONP URL
    jsonp_url = (
        f"https://nextgen.ehospital.gov.in/csnoserv/api/search/search?"
        f"term={encoded_term}&state=active&semantictag=disorder++finding"
        f"&acceptability=synonyms&returnlimit=30&groupbyconcept=false"
        f"&refsetid=null&parentid=null&excludeparentconcept=false"
        f"&fullconcept=true&callback=__my_callback__"
    )

    # Inject script using JS inside the browser
    result = page.evaluate(f"""
        () => {{
            return new Promise((resolve, reject) => {{
                const script = document.createElement('script');
                script.src = "{jsonp_url}";

                window.__my_callback__ = (data) => {{
                    resolve(data);
                    delete window.__my_callback__;
                }};

                script.onerror = reject;
                document.body.appendChild(script);
            }});
        }}
    """)

    return result


def process_diagnosis_with_icd_code(page, dischargeable_ipds):  #  dischargeable_ipds =
    # [
    # [2017237, datetime.datetime(2017, 1, 20, 0, 0), ['Acute gastroenteritis'], ('C/O VOMITING AND LOOSE STOOLS, FEVER', 'IV PANTOP, IV PCM, IV ANTIBIOTICS', 'VITAL STABLE', 'FEEL WELL', 'BED REST'), None]
    # [2017287, datetime.datetime(2017, 10, 15, 0, 0), ['Pyrexia of unknown origin'], ('C/O VOMITING AND LOOSE STOOLS, FEVER', 'IV PANTOP, IV PCM, IV ANTIBIOTICS', 'VITAL STABLE', 'FEEL WELL', 'BED REST'), None]
    # [2017171, datetime.datetime(2025, 6, 25, 16, 18, 18, 478228), ['Acute gastroenteritis', 'Pyrexia of unknown origin'], ('COMPLAIN OF FEVER FOR 2-3 DAYS', 'IV PCM, IV DNS, IV CEFTRIAXONE', 'STABLE', 'FEEL WELL', 'BED REST'), None]
    # ]
    for every_dischargeable_ipd in dischargeable_ipds:

        "checking the diagnosis type"
        discharge_code_is = validate_discharge_type(single_dischargeable_ipd=every_dischargeable_ipd)

        diagnosis_are = every_dischargeable_ipd[2]  # all the diagnosis of an IPD in single list, ['Acute gastroenteritis', 'Pyrexia of unknown origin']
        # print('The diagnosis are >===============>', diagnosis_are)

        modified_diagnosis_accumulator = []
        for each_diagnosis in diagnosis_are:  # every diagnosis in an IPD, 'Acute gastroenteritis' ----> ['Acute gastroenteritis', 'Pyrexia of unknown origin']
            each_diagnosis_snomed_data = search_snomed_js(page, each_diagnosis)

            "check the snomed data is empty i.e. wrong item searched"
            if not each_diagnosis_snomed_data:

                snomed_diagnosis_unmatch_error = f'The diagnosis provided -> "{each_diagnosis}" is wrong in IPD:{every_dischargeable_ipd[0]}. The diagnosis should be matching the diagnosis present in website "EXACTLY". Please check spellings and correct in DATA_CODE SHEET'
                error_tk_box(error_title='Diagnosis Mismatch Error',
                             error_message=snomed_diagnosis_unmatch_error)
                raise ValueError(snomed_diagnosis_unmatch_error)

            # print('THE SNO-MED SEARCH', each_diagnosis_snomed_data)  #
            # [
            # {'hierarchy': 'disorder', 'isPreferredTerm': '1', 'conceptState': '1', 'conceptFsn': 'Acute gastroenteritis (disorder)', 'definitionStatus': '900000000000073002', 'conceptId': '69776003', 'languageCode': 'en', 'typeId': 'SYNONYM', 'term': 'Acute gastroenteritis', 'caseSignificanceId': 'CASE_INSENSITIVE', 'id': '115902018', 'effectiveTime': 'Jul 31, 2017 12:00:00 AM', 'activeStatus': 1, 'moduleId': '900000000000207008'},
            # {'hierarchy': 'disorder', 'isPreferredTerm': '1', 'conceptState': '1', 'conceptFsn': 'Acute infective gastroenteritis (disorder)', 'definitionStatus': '900000000000073002', 'conceptId': '36789003', 'languageCode': 'en', 'typeId': 'SYNONYM', 'term': 'Acute infective gastroenteritis', 'caseSignificanceId': 'CASE_INSENSITIVE', 'id': '61368016', 'effectiveTime': 'Jul 31, 2017 12:00:00 AM', 'activeStatus': 1, 'moduleId': '900000000000207008'},
            # {'hierarchy': 'disorder', 'isPreferredTerm': '1', 'conceptState': '1', 'conceptFsn': 'Acute infectious nonbacterial gastroenteritis (disorder)', 'definitionStatus': '900000000000074008', 'conceptId': '359613008', 'languageCode': 'en', 'typeId': 'SYNONYM', 'term': 'Acute infectious nonbacterial gastroenteritis', 'caseSignificanceId': 'CASE_INSENSITIVE', 'id': '473899018', 'effectiveTime': 'Jul 31, 2017 12:00:00 AM', 'activeStatus': 1, 'moduleId': '900000000000207008'},
            # {'hierarchy': 'disorder', 'isPreferredTerm': '1', 'conceptState': '1', 'conceptFsn': 'Acute ulcerative gastroenteritis complicating pneumonia (disorder)', 'definitionStatus': '900000000000073002', 'conceptId': '109814008', 'languageCode': 'en', 'typeId': 'SYNONYM', 'term': 'Acute ulcerative gastroenteritis complicating pneumonia', 'caseSignificanceId': 'CASE_INSENSITIVE', 'id': '174489010', 'effectiveTime': 'Jul 31, 2017 12:00:00 AM', 'activeStatus': 1, 'moduleId': '900000000000207008'}
            # ]
            for every_matching_diagnosis in each_diagnosis_snomed_data:  # all the matching diagnosis returned as list from diagnosis search in website, each is dictionary in list
                term_in_snomed = every_matching_diagnosis.get('term')
                if each_diagnosis == term_in_snomed:  # checking the Excel diagnosis in searched diagnosis
                    # ColourPrint.print_turquoise('The term is >>>>>>>',term_in_snomed)
                    concept_id = every_matching_diagnosis['conceptId']
                    concept_id_data = get_icd_map_using_page(page, concept_id)
                    # print(concept_id_data)
                    icd_code = concept_id_data["mapGroup"][0]["mappedICDCode"]
                    # ColourPrint.print_yellow('The term is >>>>>>>',term_in_snomed,'The ICD CODE is', str(icd_code))
                    modded_diagnosis = f'{term_in_snomed}({icd_code})'
                    # ColourPrint.print_pink('The Modded Diagnosis:', modded_diagnosis)
                    # print()
                    modified_diagnosis_accumulator.append(modded_diagnosis)
        "Remove duplicate diagnosis"
        modified_diagnosis_accumulator = set(modified_diagnosis_accumulator)
        ColourPrint.print_pink('*'*50)
        final_modded_diagnosis = ','.join(modified_diagnosis_accumulator)
        print(final_modded_diagnosis)
        ColourPrint.print_pink('*'*50)
        print()

        'modifying the old list to replace the diagnosis with desired string format'
        every_dischargeable_ipd[2] = final_modded_diagnosis

        """validating the ipd number as correct"""
        # converting the ipd number string to integer
        ipd_number_integer: int = validate_ipd_as_correct_in_excel(every_dischargeable_ipd[0])
        every_dischargeable_ipd[0] = ipd_number_integer
        every_dischargeable_ipd[-1] = discharge_code_is


    ColourPrint.print_green('>/>>>>>>>>> Final Discharge IPDS >>>>>>>>>>/> '*2)
    for single_ipd in dischargeable_ipds:  # exact diagnosis syntax in website = "Cataracts(H26.9),Bilateral cataracts(H26.9),Senile cataract(H25.9)"
        print(single_ipd)
    ColourPrint.print_green('>/>>>>>>>>> Final Discharge IPDS >>>>>>>>>>/> '*2)
    print()

    return dischargeable_ipds



def get_icd_map_using_page(page, concept_id: str):
    return page.evaluate(
        """(conceptId) => {
            return fetch(`https://nextgen.ehospital.gov.in/csnoserv/api/map/icdmap?id=${conceptId}`)
                .then(response => response.json())
                .then(data => data)
                .catch(error => {
                    console.error("❌ Error fetching ICD map:", error);
                    return null;
                });
        }""",
        concept_id  # Passed as argument to the evaluate function
    )

def verify_user(user_id):
    already_present_user_list = ["kiran_701303", "Jyoti_701333", "Sumitra_701378", "Surendra_701384", "Sangeeta_701358", "Yamini_701348", "Veena_701377", "Neha_701336"]
    if user_id not in already_present_user_list:
        invalid_user_error = f'User_Name -> {user_id} is not correct or not in list. Check Username in "USER" Sheet in Excel'
        error_tk_box(error_title="Invalid User ID",
                     error_message=invalid_user_error)
        raise ValueError(invalid_user_error)
    else:
        print('User is :', user_id)
        ColourPrint.print_yellow(message_box('Please wait ...'))


def main():
    # check for chromeDev Opened and the desired page present return the indexed of page
    pages_indicis = get_desired_page_indexes_in_cdp_async_for_SYNC('NextGen eHospital')  # print = [0,2,..]

    with sync_playwright() as p:
        # Connect to already running Chrome (start with --remote-debugging-port=9222)
        browser = p.chromium.connect_over_cdp("http://localhost:9222")

        # Use existing context and open a fresh new tab
        context = browser.contexts[0]

        # Set default timeout for all actions in this context
        # context.set_default_timeout(120_000)

        # Manually paste your fresh Bearer token here
        print(get_fresh_token_and_userid(context=context))
        token, user_id_from_get_request = get_fresh_token_and_userid(context=context)
        ColourPrint.print_pink('User Id of Web:', user_id_from_get_request)

        # Keep tab 1 (manual login tab) untouched
        page1 = context.pages[0]
        # print("Page 1: manual login session preserved.")

        # 🔁 Check if second tab (automation tab) already exists
        if len(context.pages) > 1:
            # page2 = context.pages[1]
            desired_url_page = pages_indicis[0]
            page2 = context.pages[desired_url_page]  # using the page indices
            # print("🔁 Reusing existing automation tab (Page 2).")
        else:
            page2 = context.new_page()
            # print("🆕 Opened fresh automation tab (Page 2).")
            page2.goto("https://nextgen.ehospital.gov.in/login")




        headers = {
            "Authorization": token,
            "usertype": "5",  # Add this!
            "Content-Type": "application/json"
            # "userdepartmentarray": ""  # Keep it empty if unsure — still better than missing
        }


        dischargeable_ipd = fetch_excel(r"..\EHOSP\ehosp_2\ward_discharge_entry.xlsx")
        formatted_final_dischargeable_ipds_except_datetime = process_diagnosis_with_icd_code(page=page2,dischargeable_ipds=dischargeable_ipd)
        # print(formatted_final_dischargeable_ipds_except_datetime)
        # print('len', len(formatted_final_dischargeable_ipds_except_datetime))

        "VERIFY USER"
        user_id_of_excel = user_id_from_get_request
        # print('=========', user_id_of_excel)
        verify_user(user_id_of_excel)

        # 'getting the ward name of the ipd number'
        # whole_active_patients_data = get_all_active_patients_list(context, token)
        # print(all_active_patients_data)

        # single_patient_data = get_single_patient_ward_id(context=context, token=token, ipd_no='250017315')
        # print(single_patient_data)


        for idx, single_dischargeable_ipd in enumerate(formatted_final_dischargeable_ipds_except_datetime, start=1):
            ColourPrint.print_pink(message_box(f'Serial No.{idx}. IPD: {single_dischargeable_ipd[0]}'))
            print(f'Single Discharge IPD: ', single_dischargeable_ipd)
            print()

            # selecting the ipd number patient from whole active data. <-- This was previously employed but not now.
            automate_discharge_with_manual_token(token=token, single_dischargeable_ipd=single_dischargeable_ipd, whole_active_patients_data=None, user_id=user_id_of_excel, context=context)

            # This has been deprecated. Because of no require screenshots
            # nextgen_ui(context=context, headers=headers, ipd_number_integer=single_dischargeable_ipd[0])

            ColourPrint.print_green(message_box(f'Completed All for IPD {single_dischargeable_ipd[0]}'))




if __name__ == "__main__":
    fetch_excel(r"..\EHOSP\ehosp_2\ward_discharge_entry.xlsx")
    main()


"""https://nextgen.ehospital.gov.in/api/ipd/doc/printSummStatus/7013/250017674 to know the dischsrge status completeed 2 is complete dis_summary_status": 1"""