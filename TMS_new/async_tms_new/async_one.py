import asyncio
import datetime
import json
import time
from playwright.async_api import async_playwright, Page, TimeoutError

import openpyxl
import pyautogui
import pygetwindow as gw
from playwright.async_api import async_playwright, Page, expect
import select_ors
from dkbssy.utils.colour_prints import ColourPrint

from list_generator_only import list_generator

def split(lis, number_of_splits):
    k, m = divmod(len(lis), number_of_splits)
    return (lis[i * k + min(i, m):(i + 1) * k + min(i + 1, m)] for i in range(number_of_splits))



class AllListGeneratorAsync:

    async def retrieve_date_new_claim_query(self, page: Page, case_number):
        """Async version of retrieve_date_new_claim_query"""

        selector_file = r"C:\WebstormProjects\TMS_new\selectors.json"
        with open(selector_file, 'r') as file:
            selector = json.load(file)

        await page.wait_for_load_state("networkidle")

        try:
            await page.wait_for_selector(selector["patientStatusInput"], timeout=5000)
            await page.fill(selector["patientStatusInput"], 'All')
        except TimeoutError:
            await page.goto('https://provider.nha.gov.in/')

        await page.keyboard.press('Enter')
        await asyncio.sleep(0.25)
        await page.wait_for_load_state("networkidle")

        await page.fill(selector["searchBox"], "")
        await page.fill(selector["searchBox"], case_number)
        await page.wait_for_load_state("networkidle")

        await page.click(selector["searchIcon"])

        while True:
            try:
                await page.wait_for_selector("//p[contains(text(),'Registration ID:')]", timeout=2000)
                break
            except TimeoutError:
                await page.click(selector["searchIcon"])

        await page.click(selector["searchIcon"])

        await page.click(
            f"//strong[text()='{case_number}']/parent::p/parent::div/parent::div//*[name()='path' and @id='Path_98789']")

        current_status = (await page.inner_text("//a[normalize-space()='Home']/ancestor::ol//li[2]")).split("(")[
            0].strip()
        card = await page.text_content("//div[text()=' ID']/following-sibling::div")
        name = await page.text_content("//img[@alt='image']/parent::div/following-sibling::div/div")

        status = await page.evaluate(
            """
            async ({ locatorSelector, timeout }) => {
                const start = performance.now();
                return new Promise(resolve => {
                    const checkText = () => {
                        const element = document.querySelector(locatorSelector);
                        if (element) {
                            const text = element.innerText || element.textContent;
                            if (text.includes('New Born') || performance.now() - start > timeout) {
                                resolve(text);
                            } else {
                                requestAnimationFrame(checkText);
                            }
                        } else {
                            resolve('<-->');
                        }
                    };
                    checkText();
                });
            }
            """,
            {"locatorSelector": ".qpz9NG90l1gorDZORgIn.OSVhhsHh74o0GLnLTmrt", "timeout": 1000}
        )

        date = (await page.text_content("//div[text()='Registration Date']/following-sibling::div")).split()[0]
        date_a_str = date + "a"

        depart = '----Department----'
        procedure = '----PROCEDURE----'
        amount = '0'

        if current_status in ['Preauthorization Queried', 'Pre-Authorization Queried']:
            await asyncio.sleep(1)
            amount = await page.text_content(
                "//div[p[normalize-space()='Total payable amount (after incentives) :']]/following-sibling::div/p")
            await page.click("//div[contains(text(),'Treatment Plan')]/parent::div/following-sibling::div//button")
            await page.click("//span[contains(text(),'Show More')]")
            depart = await page.text_content(
                "//span[contains(text(),'Show Less')]/parent::p/parent::td/preceding-sibling::td[1]")
            procedure = await page.text_content("//span[contains(text(),'Show Less')]/parent::p")
            await page.click("//h4[normalize-space()='Finance']/parent::button")

        elif current_status == 'Claim Queried':
            await asyncio.sleep(1)
            amount = await page.text_content(
                "//div[p[normalize-space()='Total payable amount (after incentives) :']]/following-sibling::div/p")
            await page.click("//h4[normalize-space()='Treatment']/parent::button")
            await page.click("//span[contains(text(),'Treatment Details')]/parent::button")
            await page.click("//span[contains(text(),'Show More')]")
            depart = await page.text_content(
                "//span[contains(text(),'Show Less')]/parent::p/parent::div/parent::div/preceding-sibling::div/div")
            depart = depart.split(".")[1]
            procedure = await page.text_content("//span[contains(text(),'Show Less')]/parent::p")

        elif current_status == 'Under Treatment':
            await asyncio.sleep(1)
            amount = await page.text_content(
                "//div[normalize-space()='Total Preauth Approved Amount']/following-sibling::div")
            await page.click("//h4[normalize-space()='Treatment']/parent::button")
            await page.click("//span[contains(text(),'Treatment Plan')]/parent::button")
            await page.click("//span[contains(text(),'Show More')]")
            depart = await page.text_content(
                "//span[contains(text(),'Show Less')]/parent::p/parent::td/preceding-sibling::td[2]")
            procedure = await page.text_content("//span[contains(text(),'Show Less')]/parent::p")
            await page.click("//p[contains(text(),'Home')]/preceding-sibling::*[name()='svg']")
            await page.wait_for_load_state("networkidle")
            await asyncio.sleep(2)

        await page.wait_for_load_state("networkidle")
        await asyncio.sleep(1.5)

        date_by_datetime = datetime.datetime.strptime(date, '%d/%m/%Y').date()
        today_by_datetime = datetime.datetime.today().date()
        pending_days = (today_by_datetime - date_by_datetime).days

        pending_days = '--DAYS--'
        today_by_datetime_str2 = today_by_datetime.strftime('%d-%m-%Y')

        remark = 'Discharge Pending'
        return card, name, case_number, date_a_str, depart, procedure, status, current_status, amount, pending_days, remark

async def right_click_duplicate(n, tab_x=111, tab_y=23, duplicate_x=171,duplicate_y=266):
    for _ in range(n):
        # Step 1: Right-click on the Chrome tab area
        # tab_x, tab_y = 111, 23  # Your captured tab position
        pyautogui.rightClick(tab_x, tab_y)  # Right-click on the tab

        await asyncio.sleep(3)  # Wait for the menu to appear

        # Step 2: Click on "Duplicate" at exact coordinates
        # duplicate_x, duplicate_y = 171, 266  # Your captured "Duplicate" position
        pyautogui.click(duplicate_x, duplicate_y)

        print("Clicked on 'Duplicate' in the right-click menu")

        await asyncio.sleep(5)  # Wait for tab duplication to complete

def filter_deletable_data(sheet_name_to_be_compare, list_to_compare)->list:
    """
    filter the case number which are to be deletable
    :param list_to_compare: the new list to compare
    :param sheet_name_to_be_compare: name of sheet
    :return: filtered list
    """


    file_path = r"G:\My Drive\GdrivePC\Hospital\RSBY\New\down\cc.xlsx"
    workbook = openpyxl.load_workbook(filename=file_path, data_only=True)
    sheet_name = workbook[sheet_name_to_be_compare]
    case_numbers_in_cols = list(
        [row for row in sheet_name.iter_cols(min_col=3, max_col=3, min_row=2, values_only=True)][0])
    print(case_numbers_in_cols)
    case_numbers_in_cols = [str(int(i)) for i in case_numbers_in_cols if i is not None]
    # for k in case_numbers_in_cols:
    #     print((type(k)))
    print('cc-new', len(case_numbers_in_cols), case_numbers_in_cols)
    print('dd-new', len(list_to_compare), list_to_compare)
    to_delete_case_numbers = set(case_numbers_in_cols) - set(list_to_compare)
    print('to delete', len(to_delete_case_numbers), to_delete_case_numbers)
    to_add_case_numbers = set(list_to_compare) - set(case_numbers_in_cols)
    ColourPrint.print_blue('-'*50)
    print('to_add_list', len(to_add_case_numbers), to_add_case_numbers)
    ColourPrint.print_blue('-' * 50)
    # Iterate through the rows in reverse to avoid issues with deleting rows
    for row in range(sheet_name.max_row, 1, -1):
        cell_value = sheet_name.cell(row=row, column=3).value  # Assuming 'C' is the 3rd column
        if cell_value is not None:
            cell_value = str(int(cell_value))
            # print('cellvalue ', cell_value)
        if cell_value in to_delete_case_numbers:
            sheet_name.delete_rows(row)
            # print('deleted cellvalue')

    workbook.save(file_path)
    workbook.close()

    return list(to_add_case_numbers)


async def scrapping_page(page:Page, case_num_list:list, storing_list:list):
    # await page.locator(select_ors.).click()
    await expect(page.get_by_placeholder('Search')).to_be_visible()
    for index_s, case_number in enumerate(case_num_list):
        all_lis_gen_obj = AllListGeneratorAsync()
        case_num_data = await all_lis_gen_obj.retrieve_date_new_claim_query(page,case_number)
        print(case_num_data)


        # await page.get_by_placeholder('Search').fill(case_number)
        # print(index_s, case_number)
    # page_one = await context.new_page()

both_query_playwright_list, under_treatment_play_list = list_generator()
print('both list printed')
    # print(f'{both_query_list_out=}')
    # print(f'{under_treatment_out=}')
both_query_list_out = filter_deletable_data('QUERY2', both_query_playwright_list)
under_treatment_list_out = filter_deletable_data('Pend Dischg2', under_treatment_play_list)
async def main():
    play = await async_playwright().start()
    browser = await play.chromium.connect_over_cdp('http://localhost:9222')

    context = browser.contexts[0]
    page = context.pages[0]
    page.set_default_timeout(5000)
    original_url = page.url
    # await page.get_by_text('Claims to be Submitted').click()
    # await page.wait_for_timeout(1000)

    # await page.pause()

    cdp_chrome_title = await page.title()
    print(f"✅ Found CDP Chrome page title: {cdp_chrome_title}")

    # Step 3: Find the matching Chrome window by title
    cdp_chrome_window = None
    for win in gw.getWindowsWithTitle(cdp_chrome_title):
        if "Chrome" in win.title:  # Ensure it's a Chrome window
            cdp_chrome_window = win
            break

    if cdp_chrome_window:
        cdp_chrome_window.activate()  # Bring the CDP Chrome to the front
        time.sleep(1)  # Allow time for focus change
        print("✅ CDP Chrome window activated")
    else:
        print("❌ No matching CDP Chrome window found!")

    # opening the pages

    num_splits = 3
    # await right_click_duplicate(num_splits)

    for index, page in enumerate(context.pages):
        print(index,page)

    pages = context.pages

    chunks_query = list(split(both_query_list_out, num_splits))
    print(f'{chunks_query=}')
    # Shared list to store all scraped data
    all_scraped_data_query = []
    tasks = [scrapping_page(pages[i], chunks_query[i], all_scraped_data_query) for i in range(num_splits)]
    await asyncio.gather(*tasks)



    chunks_dis = list(split(under_treatment_list_out, num_splits))
    print(f'{chunks_dis=}')
    all_scraped_data_dis = []
    tasks = [scrapping_page(pages[i], chunks_dis[i], all_scraped_data_dis) for i in range(num_splits)]
    await asyncio.gather(*tasks)

    # await scrapping_page(page=pages[1], case_num_list=both_query_list_out, storing_list=[])
    # await scrapping_page(page=pages[2], case_num_list=under_treatment_lost_out, storing_list=[])





    # closing the pages except 1
    # for page in context.pages[:0:-1]:
    #     time.sleep(0.5)
    #     await page.close()
    #     time.sleep(0.5)
    # await context.close()
    await play.stop()


asyncio.run(main())

