import openpyxl

from TMS_new.sync_tms.excel_and_sql_methods import is_file_older_than_specific_time, save_to_json, read_from_json, \
    ExcelMethods, SqlMethods
from TMS_new.sync_tms.sync_methods_tms import TmsProvider
from tms_playwright.discharge_to_be_done.detail_list_getter_all import AllListGenerator as AllListGeneratorOld
from dkbssy.utils.colour_prints import ColourPrint

# print(ExcelMethods('q',r"G:\My Drive\GdrivePC\Hospital\RSBY\New\down\cc.xlsx",'a',False,'page').get_excel_column_list(
#     # workbook_path = r"G:\My Drive\GdrivePC\Hospital\RSBY\New\down\cc.xlsx",
#     sheet_name = 'Pend Dischg2',
#     min_cols=3,
#     max_cols=3
#     ))


# (ExcelMethods('q',r"G:\My Drive\GdrivePC\Hospital\RSBY\New\down\cc8.xlsx",'a',False,'page')
#  .delete_rows(sheet_name='Pend Dischg2',
#               to_delete_item_values_list=[],
#               matching_column_number=3,
#               workbook_path=r"G:\My Drive\GdrivePC\Hospital\RSBY\New\down\cc.xlsx"))

database_path_is = 'temp.db'
excel_wb_is = r"G:\My Drive\GdrivePC\Hospital\RSBY\New\down\cc_new.xlsx"
# l = '''card test, name test, case_number_test_123, date test, depart test, procedure test, age_and_sex test,
#             current_status test, amount test, pending_days test, remark test'''.split(',')
# random_datas = [i.strip() + '_'+ str(random.choice([i for i in range(1,100)])) for i in l]
# # print(random_datas)
# SqlMethods(database_path=database_path_is).temp_save_data(to_save_data=random_datas, table_name='tempData')

# loaded_temp_data = SqlMethods(database_path=database_path_is).load_temp_saved_data(table_name='tempData')
# print(loaded_temp_data)
# SqlMethods(database_path=database_path_is).delete_all_data(table_name='tempData')
# print(is_file_older_than_2_hours(database_path_is))

def main_app():
    tms = TmsProvider()
    # page, context = tms.get_desired_page('Shaheed Veer Narayan Singh Ayushman Swasthya Yojna')
    page, context = tms.get_desired_page('PMJAY - Provider')

    # Navigate to the Google Sheets URL

    # new_page = context.new_page()
    # if is_file_older_than_specific_time(r"G:\My Drive\GdrivePC\Hospital\RSBY\New\down\cc.xlsx",hours=1):
    #     AllListGeneratorOld().spreadsheet_download_and_rename(new_page,
    #                                                           'https://docs.google.com/spreadsheets/d/19HHTQZe9_8hMQJDZM4aZ01RcBqVH1-xXVv3RkR2W1ls/edit?gid=42829875#gid=42829875',
    #                                                           'AYUSHMAN REGISTRATION 2023.xlsx',
    #                                                           "cc.xlsx")
    #     new_page.close()
    #     ColourPrint.print_yellow("New page closed.")
    #

    # downloading new 2025 sheet
    if is_file_older_than_specific_time(r"G:\My Drive\GdrivePC\Hospital\RSBY\New\down\cc_new.xlsx",
                                        hours=0.05):
        new_page_2 = context.new_page()
        AllListGeneratorOld().spreadsheet_download_and_rename(new_page_2,
                                                              'https://docs.google.com/spreadsheets/d/1vhjV0rcODJ4lGYJBHENMnHFvqHgK25dQRt9SVpr_9N4/edit?gid=0#gid=0',
                                                              'AYUSHMAN REGISTRATION 2025.xlsx',
                                                              'cc_new.xlsx')
        new_page_2.close()
        ColourPrint.print_yellow("New page 2 closed.")

    # getting database oldness
    if is_file_older_than_specific_time(database_path_is, hours=1000000):
        print(f'Older than ... hours: ----')
        pre_auth_query_list_generated, claim_query_list_generated, under_treatment_list_generated = tms.refresh_pending_lists(page)

        # saving data to json
        save_to_json(list_of_keys=['pre_auth_list', 'claim_query_list', 'under_treatment_list'],
                     list_of_values= [pre_auth_query_list_generated, claim_query_list_generated, under_treatment_list_generated],
                     file_path=r"G:\My Drive\GdrivePC\Hospital\RSBY\New\down\dd2.json")
    else:
        # File is less than 2 hours old, read from the file
        all_dict_data = read_from_json(file_path=r"G:\My Drive\GdrivePC\Hospital\RSBY\New\down\dd2.json")
        print(all_dict_data)
        pre_auth_query_list_generated = all_dict_data['PreAuth List']
        claim_query_list_generated = all_dict_data['Claim List']
        under_treatment_list_generated =  all_dict_data['Discharge List']

    both_query_list = pre_auth_query_list_generated + claim_query_list_generated
    print('both_query_list', both_query_list)

    """Starting the 
    discharge process"""

    # Excel works-ups
    excel_file = ExcelMethods(excel_wb_is)
    # getting the column data = [('Case Number', 1003637676.0, 1004458743.0, 1004365832.0, 1004396234.0)]

    # performing the discharge deletion and addition
    to_delete_discharges, to_add_discharges = excel_file.filtered_excel_data_with_new_web_data(
                                                sheet_name='Pend Dischg2',
                                                min_cols=3,
                                                max_cols=3,
                                                comparator_new_list=under_treatment_list_generated)

    # deleting the rows
    excel_file.delete_rows(sheet_name='Pend Dischg2',
                           to_delete_item_values_list=to_delete_discharges,
                           matching_column_number=3)

    # creating the list to save scrapped data
    master_to_save_data_dis = []
    # retrieving the data from web
    try:
        for c_num in to_add_discharges:
            case_num_data = tms.retrieve_data_from_new_site(page,c_num,)
            for_print = '$'.join(case_num_data)
            print(for_print)
            master_to_save_data_dis.append(list(case_num_data))
        sheet_name = 'Pend Dischg2'
        wb = openpyxl.load_workbook(excel_wb_is)
        ws = wb[sheet_name]
        if master_to_save_data_dis:
            for case_detail_list in master_to_save_data_dis:
                ws.append(case_detail_list)
        wb.save(excel_wb_is)

    except Exception as e:
        ColourPrint.print_pink("Error occurred. Saving in excel file")
        print(e)
        sheet_name = 'Pend Dischg2'
        wb=openpyxl.load_workbook(excel_wb_is)
        ws = wb[sheet_name]
        if master_to_save_data_dis:
            for case_detail_list in master_to_save_data_dis:
                ws.append(case_detail_list)
        wb.save(excel_wb_is)

    """Starting 
    Query"""
    # performing the discharge deletion and addition
    to_delete_query, to_add_query = excel_file.filtered_excel_data_with_new_web_data(
        sheet_name='QUERY2',
        min_cols=3,
        max_cols=3,
        comparator_new_list=both_query_list)
    # deleting the rows
    excel_file.delete_rows(sheet_name='QUERY2',
                           to_delete_item_values_list=to_delete_query,
                           matching_column_number=3)
    # creating the list to save scrapped data
    master_to_save_data_query = []
    # retrieving the data from web
    try:
        for c_num in to_add_query:
            case_num_data = tms.retrieve_data_from_new_site(page,c_num, is_query_question_required=True)
            for_print = '$'.join(case_num_data)
            print(for_print)
            master_to_save_data_query.append(list(case_num_data))
        sheet_name = 'QUERY2'
        wb = openpyxl.load_workbook(excel_wb_is)
        ws = wb[sheet_name]
        if master_to_save_data_query:
            for case_detail_list in master_to_save_data_query:
                ws.append(case_detail_list)
        wb.save(excel_wb_is)
    except Exception as e:
        ColourPrint.print_pink("Error occurred in query. Saving in excel file")
        print(e)
        sheet_name = 'QUERY2'
        wb = openpyxl.load_workbook(excel_wb_is)
        ws = wb[sheet_name]
        if master_to_save_data_query:
            for case_detail_list in master_to_save_data_query:
                ws.append(case_detail_list)
        wb.save(excel_wb_is)



# main_app()
