import openpyxl


pending_discharge_wb = openpyxl.load_workbook(r'H:\My Drive\GdrivePC\Hospital\RSBY\New\pedia_extract2.xlsx')
# pending_discharge_wb = openpyxl.load_workbook(r'H:\My Drive\GdrivePC\Hospital\RSBY\New\pedia_extract2.xlsx')
pend_dis_sheet = pending_discharge_wb['Sheet1']

ayusman_wb = openpyxl.load_workbook(r'D:\Down\AYUSHMAN REGISTRATION 2023.xlsx')
# ayusman_wb = openpyxl.load_workbook(r'C:\Users\HP\Downloads\AYUSHMAN REGISTRATION 2023.xlsx')
# ayusman_wb = openpyxl.load_workbook(r"C:\Users\RAKESH\Downloads\AYUSHMAN REGISTRATION 2023.xlsx")
ayusman_sheet = ayusman_wb['2024']


'''getting the list of main file card list'''
row_values = ayusman_sheet.iter_cols(min_col=4, max_col=4)
for col in row_values:
    main_card = [cell.value for cell in col]
# main_card = [i.strip() if i is not None and not isinstance(i, float) else i for i in main_card]



'''getting list of pending cards'''
row_values_pend = pend_dis_sheet.iter_cols(max_col=1, min_col=1 )
# print(row_values_pend)
for col_pend in row_values_pend:
    cards_pend = [cell.value for cell in col_pend]
    card_row_number = [{cardx.value: cardx.row} for  cardx in col_pend]
    print(card_row_number)
    d = {}
    for cardx in col_pend:
        # cardx_indexes = []
        d[cardx.value] = cardx.row
    print('d',d)
    cards_pend_list = list(filter(lambda x: x is not None, cards_pend))
    # print(cards_pend_list)

    '''finding the pending card in main file list'''
    for card_pend_scan in cards_pend_list:
        # if crds == card_pend_scan:
        idx = [idx for idx, valx in enumerate(main_card) if valx == card_pend_scan]
        # main_card.index(crds))
        # print(card_pend_scan, idx)
        y = d[card_pend_scan]
        # print(y)
        z = []
        for i in idx:
            z.append(str(ayusman_sheet.cell(row=i + 1, column=10).value))
        zz = str(' + '.join(z))
        print(card_pend_scan, zz)
        pend_dis_sheet.cell(row = y, column=9).value = zz
    pending_discharge_wb.save(r'H:\My Drive\GdrivePC\Hospital\RSBY\New\pedia_extract2.xlsx')








