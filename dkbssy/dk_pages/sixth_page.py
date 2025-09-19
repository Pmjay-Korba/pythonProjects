import asyncio
from pathlib import Path
from openpyxl import load_workbook, Workbook
from playwright.async_api import async_playwright, Page, TimeoutError
import time

from TMS_new.async_tms_new.desired_page import get_desired_page_indexes_in_cdp_async_for_ASYNC
from dkbssy.utils.colour_prints import ColourPrint
from dkbssy.utils.excel_utils import retrieve_emp_code_only
from dkbssy.utils.sqlite_updation_long import sql_update_bulk_optimized
from dkbssy.dk_pages.fifth_page import FifthPage
from collections import defaultdict

BROWSER_TAB_COUNT = 5
BATCH_PER_TAB = 10  # 10 cases per tab each round
OUTPUT_XLSX = Path(r"G:\My Drive\GdrivePC\Hospital\RSBY\New\INCENTIVE query.xlsx")


class IncentiveScraper:
    case_number_xpath = '//*[@id="ctl00_ContentPlaceHolder1_caseNoReqText"]'
    status_xpath = '//*[@id="ctl00_ContentPlaceHolder1_caseStatusText"]'
    procedure_xpath = '//*[@id="ctl00_ContentPlaceHolder1_procName"]'
    last_updated_xpath = '//*[@id="ctl00_ContentPlaceHolder1_caseLastUpdateDate"]'


    def modified_url(self, case_number):
        return f'https://dkbssy.cg.nic.in/secure/incentivemodule/incentivemoduleViewDME.aspx?ci={case_number}'

    async def _get_data_async(self, case_number, page):
        new_url = self.modified_url(case_number)
        await page.goto(new_url)
        await page.wait_for_load_state("networkidle")
        await page.wait_for_selector('//*[@id="incentiveTableData"]/tr[1]/td[1]')

        case_number_text = await page.locator(self.case_number_xpath).text_content()
        status_text = await page.locator(self.status_xpath).text_content()
        last_updated_text = await page.locator(self.last_updated_xpath).text_content()
        procedure_text = await page.locator(self.procedure_xpath).text_content()
        table_data_tbody = await page.locator('//*[@id="incentiveTableData"]').inner_html()

        data_of_line = []
        table_by_lines = FifthPage('demo', 'demo').bs4_parser(html_string=table_data_tbody)
        category = None

        for i in table_by_lines:
            name_amount_list = i[1].split()
            if name_amount_list[0] == 'अधिष्ठाता':
                category = 1
            elif name_amount_list[0] == 'पैथोलॉजी' and name_amount_list[8] == 'फैकल्टी':
                category = 2
            elif name_amount_list[0] == 'पैथोलॉजी' and name_amount_list[8] == '(टेक्निशियन':
                category = 3
            elif name_amount_list[0] == 'सभी' and name_amount_list[1] == 'फिजिशियन':
                category = 4
            elif name_amount_list[0] == 'सभी' and name_amount_list[1] == 'सीनियर':
                category = 5
            elif name_amount_list[0] == 'एनेस्थीसिया':
                category = 6
            elif name_amount_list[0] == 'नर्सिंग':
                category = 7
            elif name_amount_list[0] == 'चतुर्थ':
                category = 8
            elif name_amount_list[0] == 'डाटा':
                category = 9
            amount = i[7]
            emp_id = i[6]
            data_of_line.append([category, emp_id, amount])
        print(case_number_text, status_text, last_updated_text, procedure_text, data_of_line)
        return case_number_text, status_text, last_updated_text, procedure_text, data_of_line

    async def get_data_async_main(self, case_number, page):
        return await self._get_data_async(case_number, page)



    def save_to_db(self, success_list):
        # print('success list')  # Success List  # [('CASE/PS6/HOSP22G146659/CK7323852', 'Hospital Initiated', '12/07/2025 23:20:05', 'SICS with non-foldable IOL', [[1, '11170010478', '3.16'], [1, '66170010344', '3.16'], [1, '04170140656', '3.16'], [1, '09530010327', '3.16'], [1, '02530010055', '3.16'], [1, '04170140099', '3.16'], [1, '66170010023', '3.16'], [2, '05170011289', '15.47'], [2, '05170040053', '15.47'], [2, 'DME55173201', '15.47'], [2, '05530010026', '15.47'], [2, '05530010031', '15.47'], [2, '05530010065', '15.47'], [2, '05530010079', '15.47'], [2, 'DME55172439', '15.47'], [2, '04170140518', '15.47'], [2, '05530010013', '15.47'], [2, 'DME55172441', '15.47'], [2, 'DME55173188', '15.47'], [2, 'DME55173185', '15.47'], [2, '05530010029', '15.47'], [3, '04170150173', '2.49'], [3, 'J55172863', '2.49'], [3, 'N5517593', '2.49'], [3, '05170010396', '2.49'], [3, '05170020193', '2.49'], [3, '05170020054', '2.49'], [3, 'AI55170024', '2.49'], [3, 'AI55170030', '2.49'], [3, '03170010366', '2.49'], [3, '05170020117', '2.49'], [3, '05170060099', '2.49'], [3, '06170050083', '2.49'], [3, '07170021387', '2.49'], [3, '09170010257', '2.49'], [3, '09530010377', '2.49'], [3, '05170011797', '2.49'], [4, '05170020187', '198.9'], [4, '05170020043', '198.9'], [4, '05170020182', '198.9'], [4, '05530010028', '198.9'], [4, '05530010030', '198.9'], [7, '15170200084', '2.65'], [7, '19170040056', '2.65'], [7, 'AI55170022', '2.65'], [7, 'AI55170026', '2.65'], [7, 'N5517599', '2.65'], [7, '05170020059', '2.65'], [7, '06170010814', '2.65'], [7, '02170280014', '2.65'], [7, '11170010433', '2.65'], [7, 'N5517866', '2.65'], [7, '05170020063', '2.65'], [7, '66170010273', '2.65'], [7, '66170020530', '2.65'], [7, '66170020641', '2.65'], [7, '66170020668', '2.65'], [7, '66170020533', '2.65'], [7, '66170020614', '2.65'], [7, '05170020133', '2.65'], [7, '04170120033', '2.65'], [7, 'N5517809', '2.65'], [7, 'N5517860', '2.65'], [7, 'N0517353', '2.65'], [7, 'N5517515', '2.65'], [7, '04170060068', '2.65'], [7, '04170140331', '2.65'], [7, '04170140436', '2.65'], [7, '04170140447', '2.65'], [7, '04170150162', '2.65'], [7, '05170020124', '2.65'], [7, '05170020142', '2.65'], [7, '05170020143', '2.65'], [7, '05170020150', '2.65'], [7, '05170020153', '2.65'], [7, '05170020161', '2.65'], [7, '05170020178', '2.65'], [7, '05170040078', '2.65'], [7, '08170040064', '2.65'], [7, '09170160077', '2.65'], [7, '12170050188', '2.65'], [7, '14170011702', '2.65'], [7, '15170200036', '2.65'], [7, '15170200090', '2.65'], [7, '22170040038', '2.65'], [7, '05170050074', '2.65'], [7, '0617009003', '2.65'], [7, '23170020062', '2.65'], [7, 'N5517989', '2.65'], [7, 'N5517990', '2.65'], [7, '05170011626', '2.65'], [7, '04170140534', '2.65'], [8, 'N5517032', '1.51'], [8, 'N5517042', '1.51'], [8, 'N5517056', '1.51'], [8, 'N5517083', '1.51'], [8, 'N5517084', '1.51'], [8, 'N5517085', '1.51'], [8, 'N55171003', '1.51'], [8, 'N5517136', '1.51'], [8, 'N5517028', '1.51'], [8, 'N5517055', '1.51'], [8, 'N5517610', '1.51'], [8, 'N5517613', '1.51'], [8, 'N5517614', '1.51'], [8, 'N5517615', '1.51'], [8, 'N5517073', '1.51'], [8, 'N5517589', '1.51'], [8, '05170020095', '1.51'], [8, 'N5517857', '1.51'], [8, '05170020131', '1.51'], [8, '05170020137', '1.51'], [8, '05170020184', '1.51'], [8, '05170020189', '1.51'], [8, '05170020097', '1.51'], [8, '05170020098', '1.51'], [8, '05170020099', '1.51'], [8, '05170020100', '1.51'], [8, '04170150035', '1.51'], [8, '04170150084', '1.51'], [8, '05170020082', '1.51'], [8, '05170020084', '1.51'], [8, '05170020086', '1.51'], [8, '05170020087', '1.51'], [8, '05170020091', '1.51'], [8, '05170020092', '1.51'], [8, '07170250399', '1.51'], [8, '07170250453', '1.51'], [8, '04170150044', '1.51'], [8, '05170020088', '1.51'], [8, '05170020090', '1.51'], [8, '05170020093', '1.51'], [8, '05170020103', '1.51'], [8, '05170020105', '1.51'], [8, '05170020107', '1.51'], [8, '05170020110', '1.51'], [9, 'J55172798', '15.79'], [9, 'J55173309', '15.79'], [9, 'DME55173036', '15.79'], [9, 'DME55173037', '15.79'], [9, 'DME55173200', '15.79'], [9, 'N5517569', '15.79'], [9, '05170020166', '15.79']]), ('CASE/PS6/HOSP22G146659/CK7323811', 'Hospital Initiated', '12/07/2025 23:18:38', 'SICS with non-foldable IOL', [[1, '11170010478', '3.16'], [1, '66170010344', '3.16'], [1, '04170140656', '3.16'], [1, '09530010327', '3.16'], [1, '02530010055', '3.16'], [1, '04170140099', '3.16'], [1, '66170010023', '3.16'], [2, '05170011289', '15.47'], [2, '05170040053', '15.47'], [2, 'DME55173201', '15.47'], [2, '05530010026', '15.47'], [2, '05530010031', '15.47'], [2, '05530010065', '15.47'], [2, '05530010079', '15.47'], [2, 'DME55172439', '15.47'], [2, '04170140518', '15.47'], [2, '05530010013', '15.47'], [2, 'DME55172441', '15.47'], [2, 'DME55173188', '15.47'], [2, 'DME55173185', '15.47'], [2, '05530010029', '15.47'], [3, '04170150173', '2.49'], [3, 'J55172863', '2.49'], [3, 'N5517593', '2.49'], [3, '05170010396', '2.49'], [3, '05170020193', '2.49'], [3, '05170020054', '2.49'], [3, 'AI55170024', '2.49'], [3, 'AI55170030', '2.49'], [3, '03170010366', '2.49'], [3, '05170020117', '2.49'], [3, '05170060099', '2.49'], [3, '06170050083', '2.49'], [3, '07170021387', '2.49'], [3, '09170010257', '2.49'], [3, '09530010377', '2.49'], [3, '05170011797', '2.49'], [4, '05170020187', '198.9'], [4, '05170020043', '198.9'], [4, '05170020182', '198.9'], [4, '05530010028', '198.9'], [4, '05530010030', '198.9'], [7, '15170200084', '2.65'], [7, '19170040056', '2.65'], [7, 'AI55170022', '2.65'], [7, 'AI55170026', '2.65'], [7, 'N5517599', '2.65'], [7, '05170020059', '2.65'], [7, '06170010814', '2.65'], [7, '02170280014', '2.65'], [7, '11170010433', '2.65'], [7, 'N5517866', '2.65'], [7, '05170020063', '2.65'], [7, '66170010273', '2.65'], [7, '66170020530', '2.65'], [7, '66170020641', '2.65'], [7, '66170020668', '2.65'], [7, '66170020533', '2.65'], [7, '66170020614', '2.65'], [7, '05170020133', '2.65'], [7, '04170120033', '2.65'], [7, 'N5517809', '2.65'], [7, 'N5517860', '2.65'], [7, 'N0517353', '2.65'], [7, 'N5517515', '2.65'], [7, '04170060068', '2.65'], [7, '04170140331', '2.65'], [7, '04170140436', '2.65'], [7, '04170140447', '2.65'], [7, '04170150162', '2.65'], [7, '05170020124', '2.65'], [7, '05170020142', '2.65'], [7, '05170020143', '2.65'], [7, '05170020150', '2.65'], [7, '05170020153', '2.65'], [7, '05170020161', '2.65'], [7, '05170020178', '2.65'], [7, '05170040078', '2.65'], [7, '08170040064', '2.65'], [7, '09170160077', '2.65'], [7, '12170050188', '2.65'], [7, '14170011702', '2.65'], [7, '15170200036', '2.65'], [7, '15170200090', '2.65'], [7, '22170040038', '2.65'], [7, '05170050074', '2.65'], [7, '0617009003', '2.65'], [7, '23170020062', '2.65'], [7, 'N5517989', '2.65'], [7, 'N5517990', '2.65'], [7, '05170011626', '2.65'], [7, '04170140534', '2.65'], [8, 'N5517032', '1.51'], [8, 'N5517042', '1.51'], [8, 'N5517056', '1.51'], [8, 'N5517083', '1.51'], [8, 'N5517084', '1.51'], [8, 'N5517085', '1.51'], [8, 'N55171003', '1.51'], [8, 'N5517136', '1.51'], [8, 'N5517028', '1.51'], [8, 'N5517055', '1.51'], [8, 'N5517610', '1.51'], [8, 'N5517613', '1.51'], [8, 'N5517614', '1.51'], [8, 'N5517615', '1.51'], [8, 'N5517073', '1.51'], [8, 'N5517589', '1.51'], [8, '05170020095', '1.51'], [8, 'N5517857', '1.51'], [8, '05170020131', '1.51'], [8, '05170020137', '1.51'], [8, '05170020184', '1.51'], [8, '05170020189', '1.51'], [8, '05170020097', '1.51'], [8, '05170020098', '1.51'], [8, '05170020099', '1.51'], [8, '05170020100', '1.51'], [8, '04170150035', '1.51'], [8, '04170150084', '1.51'], [8, '05170020082', '1.51'], [8, '05170020084', '1.51'], [8, '05170020086', '1.51'], [8, '05170020087', '1.51'], [8, '05170020091', '1.51'], [8, '05170020092', '1.51'], [8, '07170250399', '1.51'], [8, '07170250453', '1.51'], [8, '04170150044', '1.51'], [8, '05170020088', '1.51'], [8, '05170020090', '1.51'], [8, '05170020093', '1.51'], [8, '05170020103', '1.51'], [8, '05170020105', '1.51'], [8, '05170020107', '1.51'], [8, '05170020110', '1.51'], [9, 'J55172798', '15.79'], [9, 'J55173309', '15.79'], [9, 'DME55173036', '15.79'], [9, 'DME55173037', '15.79'], [9, 'DME55173200', '15.79'], [9, 'N5517569', '15.79'], [9, '05170020166', '15.79']]), ('CASE/PS6/HOSP22G146659/CK7323761', 'Hospital Initiated', '12/07/2025 23:17:12', 'SICS with non-foldable IOL', [[1, '11170010478', '3.16'], [1, '66170010344', '3.16'], [1, '04170140656', '3.16'], [1, '09530010327', '3.16'], [1, '02530010055', '3.16'], [1, '04170140099', '3.16'], [1, '66170010023', '3.16'], [2, '05170011289', '15.47'], [2, '05170040053', '15.47'], [2, 'DME55173201', '15.47'], [2, '05530010026', '15.47'], [2, '05530010031', '15.47'], [2, '05530010065', '15.47'], [2, '05530010079', '15.47'], [2, 'DME55172439', '15.47'], [2, '04170140518', '15.47'], [2, '05530010013', '15.47'], [2, 'DME55172441', '15.47'], [2, 'DME55173188', '15.47'], [2, 'DME55173185', '15.47'], [2, '05530010029', '15.47'], [3, '04170150173', '2.49'], [3, 'J55172863', '2.49'], [3, 'N5517593', '2.49'], [3, '05170010396', '2.49'], [3, '05170020193', '2.49'], [3, '05170020054', '2.49'], [3, 'AI55170024', '2.49'], [3, 'AI55170030', '2.49'], [3, '03170010366', '2.49'], [3, '05170020117', '2.49'], [3, '05170060099', '2.49'], [3, '06170050083', '2.49'], [3, '07170021387', '2.49'], [3, '09170010257', '2.49'], [3, '09530010377', '2.49'], [3, '05170011797', '2.49'], [4, '05170020187', '198.9'], [4, '05170020043', '198.9'], [4, '05170020182', '198.9'], [4, '05530010028', '198.9'], [4, '05530010030', '198.9'], [7, '15170200084', '2.65'], [7, '19170040056', '2.65'], [7, 'AI55170022', '2.65'], [7, 'AI55170026', '2.65'], [7, 'N5517599', '2.65'], [7, '05170020059', '2.65'], [7, '06170010814', '2.65'], [7, '02170280014', '2.65'], [7, '11170010433', '2.65'], [7, 'N5517866', '2.65'], [7, '05170020063', '2.65'], [7, '66170010273', '2.65'], [7, '66170020530', '2.65'], [7, '66170020641', '2.65'], [7, '66170020668', '2.65'], [7, '66170020533', '2.65'], [7, '66170020614', '2.65'], [7, '05170020133', '2.65'], [7, '04170120033', '2.65'], [7, 'N5517809', '2.65'], [7, 'N5517860', '2.65'], [7, 'N0517353', '2.65'], [7, 'N5517515', '2.65'], [7, '04170060068', '2.65'], [7, '04170140331', '2.65'], [7, '04170140436', '2.65'], [7, '04170140447', '2.65'], [7, '04170150162', '2.65'], [7, '05170020124', '2.65'], [7, '05170020142', '2.65'], [7, '05170020143', '2.65'], [7, '05170020150', '2.65'], [7, '05170020153', '2.65'], [7, '05170020161', '2.65'], [7, '05170020178', '2.65'], [7, '05170040078', '2.65'], [7, '08170040064', '2.65'], [7, '09170160077', '2.65'], [7, '12170050188', '2.65'], [7, '14170011702', '2.65'], [7, '15170200036', '2.65'], [7, '15170200090', '2.65'], [7, '22170040038', '2.65'], [7, '05170050074', '2.65'], [7, '0617009003', '2.65'], [7, '23170020062', '2.65'], [7, 'N5517989', '2.65'], [7, 'N5517990', '2.65'], [7, '05170011626', '2.65'], [7, '04170140534', '2.65'], [8, 'N5517032', '1.51'], [8, 'N5517042', '1.51'], [8, 'N5517056', '1.51'], [8, 'N5517083', '1.51'], [8, 'N5517084', '1.51'], [8, 'N5517085', '1.51'], [8, 'N55171003', '1.51'], [8, 'N5517136', '1.51'], [8, 'N5517028', '1.51'], [8, 'N5517055', '1.51'], [8, 'N5517610', '1.51'], [8, 'N5517613', '1.51'], [8, 'N5517614', '1.51'], [8, 'N5517615', '1.51'], [8, 'N5517073', '1.51'], [8, 'N5517589', '1.51'], [8, '05170020095', '1.51'], [8, 'N5517857', '1.51'], [8, '05170020131', '1.51'], [8, '05170020137', '1.51'], [8, '05170020184', '1.51'], [8, '05170020189', '1.51'], [8, '05170020097', '1.51'], [8, '05170020098', '1.51'], [8, '05170020099', '1.51'], [8, '05170020100', '1.51'], [8, '04170150035', '1.51'], [8, '04170150084', '1.51'], [8, '05170020082', '1.51'], [8, '05170020084', '1.51'], [8, '05170020086', '1.51'], [8, '05170020087', '1.51'], [8, '05170020091', '1.51'], [8, '05170020092', '1.51'], [8, '07170250399', '1.51'], [8, '07170250453', '1.51'], [8, '04170150044', '1.51'], [8, '05170020088', '1.51'], [8, '05170020090', '1.51'], [8, '05170020093', '1.51'], [8, '05170020103', '1.51'], [8, '05170020105', '1.51'], [8, '05170020107', '1.51'], [8, '05170020110', '1.51'], [9, 'J55172798', '15.79'], [9, 'J55173309', '15.79'], [9, 'DME55173036', '15.79'], [9, 'DME55173037', '15.79'], [9, 'DME55173200', '15.79'], [9, 'N5517569', '15.79'], [9, '05170020166', '15.79']]), ('CASE/PS6/HOSP22G146659/CK7323554', 'Hospital Initiated', '12/07/2025 23:15:51', 'SICS with non-foldable IOL', [[1, '11170010478', '3.16'], [1, '66170010344', '3.16'], [1, '04170140656', '3.16'], [1, '09530010327', '3.16'], [1, '02530010055', '3.16'], [1, '04170140099', '3.16'], [1, '66170010023', '3.16'], [2, '05170011289', '15.47'], [2, '05170040053', '15.47'], [2, 'DME55173201', '15.47'], [2, '05530010026', '15.47'], [2, '05530010031', '15.47'], [2, '05530010065', '15.47'], [2, '05530010079', '15.47'], [2, 'DME55172439', '15.47'], [2, '04170140518', '15.47'], [2, '05530010013', '15.47'], [2, 'DME55172441', '15.47'], [2, 'DME55173188', '15.47'], [2, 'DME55173185', '15.47'], [2, '05530010029', '15.47'], [3, '04170150173', '2.49'], [3, 'J55172863', '2.49'], [3, 'N5517593', '2.49'], [3, '05170010396', '2.49'], [3, '05170020193', '2.49'], [3, '05170020054', '2.49'], [3, 'AI55170024', '2.49'], [3, 'AI55170030', '2.49'], [3, '03170010366', '2.49'], [3, '05170020117', '2.49'], [3, '05170060099', '2.49'], [3, '06170050083', '2.49'], [3, '07170021387', '2.49'], [3, '09170010257', '2.49'], [3, '09530010377', '2.49'], [3, '05170011797', '2.49'], [4, '05170020187', '198.9'], [4, '05170020043', '198.9'], [4, '05170020182', '198.9'], [4, '05530010028', '198.9'], [4, '05530010030', '198.9'], [7, '15170200084', '2.65'], [7, '19170040056', '2.65'], [7, 'AI55170022', '2.65'], [7, 'AI55170026', '2.65'], [7, 'N5517599', '2.65'], [7, '05170020059', '2.65'], [7, '06170010814', '2.65'], [7, '02170280014', '2.65'], [7, '11170010433', '2.65'], [7, 'N5517866', '2.65'], [7, '05170020063', '2.65'], [7, '66170010273', '2.65'], [7, '66170020530', '2.65'], [7, '66170020641', '2.65'], [7, '66170020668', '2.65'], [7, '66170020533', '2.65'], [7, '66170020614', '2.65'], [7, '05170020133', '2.65'], [7, '04170120033', '2.65'], [7, 'N5517809', '2.65'], [7, 'N5517860', '2.65'], [7, 'N0517353', '2.65'], [7, 'N5517515', '2.65'], [7, '04170060068', '2.65'], [7, '04170140331', '2.65'], [7, '04170140436', '2.65'], [7, '04170140447', '2.65'], [7, '04170150162', '2.65'], [7, '05170020124', '2.65'], [7, '05170020142', '2.65'], [7, '05170020143', '2.65'], [7, '05170020150', '2.65'], [7, '05170020153', '2.65'], [7, '05170020161', '2.65'], [7, '05170020178', '2.65'], [7, '05170040078', '2.65'], [7, '08170040064', '2.65'], [7, '09170160077', '2.65'], [7, '12170050188', '2.65'], [7, '14170011702', '2.65'], [7, '15170200036', '2.65'], [7, '15170200090', '2.65'], [7, '22170040038', '2.65'], [7, '05170050074', '2.65'], [7, '0617009003', '2.65'], [7, '23170020062', '2.65'], [7, 'N5517989', '2.65'], [7, 'N5517990', '2.65'], [7, '05170011626', '2.65'], [7, '04170140534', '2.65'], [8, 'N5517032', '1.51'], [8, 'N5517042', '1.51'], [8, 'N5517056', '1.51'], [8, 'N5517083', '1.51'], [8, 'N5517084', '1.51'], [8, 'N5517085', '1.51'], [8, 'N55171003', '1.51'], [8, 'N5517136', '1.51'], [8, 'N5517028', '1.51'], [8, 'N5517055', '1.51'], [8, 'N5517610', '1.51'], [8, 'N5517613', '1.51'], [8, 'N5517614', '1.51'], [8, 'N5517615', '1.51'], [8, 'N5517073', '1.51'], [8, 'N5517589', '1.51'], [8, '05170020095', '1.51'], [8, 'N5517857', '1.51'], [8, '05170020131', '1.51'], [8, '05170020137', '1.51'], [8, '05170020184', '1.51'], [8, '05170020189', '1.51'], [8, '05170020097', '1.51'], [8, '05170020098', '1.51'], [8, '05170020099', '1.51'], [8, '05170020100', '1.51'], [8, '04170150035', '1.51'], [8, '04170150084', '1.51'], [8, '05170020082', '1.51'], [8, '05170020084', '1.51'], [8, '05170020086', '1.51'], [8, '05170020087', '1.51'], [8, '05170020091', '1.51'], [8, '05170020092', '1.51'], [8, '07170250399', '1.51'], [8, '07170250453', '1.51'], [8, '04170150044', '1.51'], [8, '05170020088', '1.51'], [8, '05170020090', '1.51'], [8, '05170020093', '1.51'], [8, '05170020103', '1.51'], [8, '05170020105', '1.51'], [8, '05170020107', '1.51'], [8, '05170020110', '1.51'], [9, 'J55172798', '15.79'], [9, 'J55173309', '15.79'], [9, 'DME55173036', '15.79'], [9, 'DME55173037', '15.79'], [9, 'DME55173200', '15.79'], [9, 'N5517569', '15.79'], [9, '05170020166', '15.79']])]
        # print(success_list)
        # for case_num, status, last_updated, proc, data_lines in success_list:
        #     # Insert using your DB library
        #     # definition - def sql_update(case_number_text, status_text, procedure_text, last_updated_text, list_of_list):
        #     sql_update(case_number_text=case_num, status_text=status, procedure_text=proc,
        #                last_updated_text=last_updated, list_of_list=data_lines)

        sql_update_bulk_optimized(success_list)



    def append_errors_to_excel(self, error_list):
        if not error_list:
            return

        if OUTPUT_XLSX.exists():
            wb = load_workbook(OUTPUT_XLSX)
            ws = wb.active
        else:
            wb = Workbook()
            ws = wb.active
            ws.append(["case_number", "error"])  # Add headers only if new

        for case_number, error in error_list:
            ws.append([case_number, str(error)])

        wb.save(OUTPUT_XLSX)
        wb.close()

    async def process_one_batch(self, case_nums, page):
        results = []
        for case_numx in case_nums:
            try:
                # t1 = time.perf_counter()
                res = await self.get_data_async_main(case_numx, page)
                # t2 = time.perf_counter()
                # duration = t2 - t1
                # ColourPrint.print_blue(f"✅ {case_numx} processed in {duration:.2f} seconds")
            except Exception as e:
                res = e
                ColourPrint.print_yellow(case_numx, e)
            results.append((case_numx, res))
        # ColourPrint.print_pink('zip-list', results)
        return results

    async def run(self, case_list):
        play = await async_playwright().start()
        browser = await play.chromium.connect_over_cdp('http://localhost:9222')
        context = browser.contexts[0]

        pages = [await context.new_page() for _ in range(BROWSER_TAB_COUNT)]

        it = iter(case_list)
        round_num = 0


        while True:
            round_num += 1
            round_cases = [next(it, None) for _ in range(BROWSER_TAB_COUNT * BATCH_PER_TAB)]
            round_cases = [c for c in round_cases if c]
            if not round_cases:
                break

            ColourPrint.print_turquoise(f"Starting Round {round_num}, {len(round_cases)} cases")

            tab_batches = defaultdict(list)
            for idx, case in enumerate(round_cases):
                tab_index = idx % BROWSER_TAB_COUNT
                tab_batches[tab_index].append(case)

            tasks = []
            for tab_index, batch_slice in tab_batches.items():
                tasks.append(self.process_one_batch(batch_slice, pages[tab_index]))

            all_results = await asyncio.gather(*tasks)
            # ColourPrint.print_green(all_results)
            flat = [(case, res) for chunk in all_results for case, res in chunk]
            successes = [res for _, res in flat if not isinstance(res, Exception)]
            # ColourPrint.print_blue('Success List', successes)
            ColourPrint.print_green('Success Count:', len(successes))
            errors = [(case, str(res)) for case, res in flat if isinstance(res, Exception)]
            ColourPrint.print_pink("Error Count:", len(errors))
            # print(errors)
            self.save_to_db(successes)
            self.append_errors_to_excel(errors)

            print(f"✅ Round {round_num} complete: {len(successes)} saved, {len(errors)} errors")
            await asyncio.sleep(1)

        [await page.close() for page in pages]
        await play.stop()


    def start(self, case_list):
        asyncio.run(self.run(case_list))


def batch_divide(batch_count, list_to_divide):
    batch_list = [[] for _ in range(batch_count)]
    for i, v in enumerate(list_to_divide):
        target_index = i % batch_count
        # print(divided_index)
        target_list = batch_list[target_index]
        target_list.append(v)
        # print(batch_list)
    return batch_list


class AmountScrapper:
    excel_ver_3_path = r"G:\My Drive\GdrivePC\Hospital\RSBY\New\Incentive_auto_ver_3.xlsx" # for getting the names in sheet 1 and getting the name and ecode from sheet 3
    auto_2_excel_path = r"G:\My Drive\GdrivePC\Hospital\RSBY\New\Incentive_auto2.xlsx" # updating the amount and ALSO THE NAMES AND ECODE IF NOT PRESENT IN AUTO 2 AND PRESET IN VER 3

    "first matching the names nt present in auto 2"

    async def scrap_emp_amount(self, page:Page, e_code):
        await page.goto('https://dkbssy.cg.nic.in/secure/incentivemodule/IncentiveDetails_EmpCodeWiseDME.aspx')
        await page.locator('//input[@id="ctl00_ContentPlaceHolder1_TextBox1"]').fill(str(e_code))
        await page.locator('//input[@id="ctl00_ContentPlaceHolder1_search"]').click()
        web_e_code = None

        try:
            await page.wait_for_selector('//*[@id="ctl00_ContentPlaceHolder1_GridView1"]/tbody/tr[2]/td[3]')
            web_e_code = await page.locator('//a[contains(@href, "IncentiveDetails_EmpCodeWise_DetailsDME")]').text_content()

        except TimeoutError:
            ColourPrint.print_bg_red(f'The {e_code} either has Zero amount or Invalid ')

        pending_count = await page.locator('//*[@id="ctl00_ContentPlaceHolder1_Label3"]').text_content()
        pending_amount = await page.locator('//*[@id="ctl00_ContentPlaceHolder1_Label4"]').text_content()
        paid_count =  await page.locator('//*[@id="ctl00_ContentPlaceHolder1_Label14"]').text_content()
        paid_amount = await page.locator('//*[@id="ctl00_ContentPlaceHolder1_Label5"]').text_content()

        print(web_e_code, pending_count, pending_amount, paid_count, paid_amount)
        return web_e_code, pending_count, pending_amount, paid_count, paid_amount

    async def process_one_batch_list(self, page, batch_list):
        collect_batch_scrape = []
        for e_code in batch_list:
            web_e_code, pending_count, pending_amount, paid_count, paid_amount = await self.scrap_emp_amount(page, e_code)
            collect_batch_scrape.append([web_e_code, pending_count, pending_amount, paid_count, paid_amount])
        return collect_batch_scrape

    async def main_scraper(self, set_timeout_is):
        page_indexes = await get_desired_page_indexes_in_cdp_async_for_ASYNC(user_title_of_page='Shaheed Veer Narayan Singh Ayushman Swasthya Yojna')
        # print(is_chrome_dev_open)

        async with async_playwright() as p:
            cdp_for_main = 9222
            browser = await p.chromium.connect_over_cdp(f"http://localhost:{cdp_for_main}")
            context = browser.contexts[0]
            all_pages = context.pages

            page_index = page_indexes[0]  # selecting the first index of matching page
            page = all_pages[page_index]  # selecting the first PAGE of matching page

            # l = '''N5517039 N5517035 DME55173194 DME55172713 DME55173199 05530010060 05170020039 05530010074 0000'''. split()
            all_e_codes = retrieve_emp_code_only()
            batch = 5
            "creating pages"
            pages = [await context.new_page() for _ in range(batch)]
            for page in pages:
                page.set_default_timeout(set_timeout_is)
                page.set_default_navigation_timeout(set_timeout_is)

            main_list_of_list = batch_divide(batch_count=batch, list_to_divide=all_e_codes)
            print(main_list_of_list)
            print(len(main_list_of_list))

            tasks = [asyncio.create_task(self.process_one_batch_list(page, batch_list)) for page, batch_list in zip(pages, main_list_of_list)]

            try:
                result = await asyncio.gather(*tasks)

            finally:
                # await asyncio.sleep(10)
                [await page.close() for page in pages]

            print(result)





if __name__ == '__main__':
    asyncio.run(AmountScrapper().main_scraper(30000))
