import datetime
import sys

import openpyxl

from dkbssy.utils.colour_prints import ColourPrint


def working_date_extractor():
    workbook_master = openpyxl.load_workbook(r'G:\My Drive\GdrivePC\Hospital\RSBY\New\Incentive_auto_ver_3.xlsx')
    worksheet_for_getting_names = workbook_master["Sheet1"]
    # print(worksheet["B1:B5"])

    incentive_names_to_enter = []
    row_values = worksheet_for_getting_names.iter_cols(min_col=3, max_col=11, min_row=2, max_row=26)
    for col in row_values:
        row_data = [cell.value for cell in col]
        incentive_names_to_enter += row_data

    incentive_names_to_enter = [i for i in incentive_names_to_enter if i is not None]
    print(incentive_names_to_enter)

    # retrieving the employee ID--------
    worksheet_for_getting_leave = workbook_master["Sheet2"]

    # getting all names from sheet3
    column_names_in_sheet2 = worksheet_for_getting_leave.iter_rows(min_col=1, max_col=1)

    dict_name_row_number = {}
    for namez in column_names_in_sheet2:
        # print(namez[0].value)
        name_is = namez[0].value
        row_number = namez[0].row
        # print({name_is:row_number})
        dict_name_row_number[name_is] = row_number  # gives the row containing name
    # print(dict_name_row_number)

    name_and_joining_date_dict = {}
    for name in incentive_names_to_enter:
        name = name.strip()
        # print(dict_name_row_number.keys())
        join_resign_leave_dates = []
        if name in dict_name_row_number.keys():
            chosen_row_number = dict_name_row_number[name]
            # print(chosen_row_number)
            chosen_emp_join_date = worksheet_for_getting_leave.cell(row=chosen_row_number, column=5).value
            chosen_emp_resign_date = worksheet_for_getting_leave.cell(row=chosen_row_number, column=6).value
            chosen_emp_leave_1_date = worksheet_for_getting_leave.cell(row=chosen_row_number, column=7).value
            chosen_emp_return_1_date = worksheet_for_getting_leave.cell(row=chosen_row_number, column=8).value
            chosen_emp_leave_2_date = worksheet_for_getting_leave.cell(row=chosen_row_number, column=9).value
            chosen_emp_return_2_date = worksheet_for_getting_leave.cell(row=chosen_row_number, column=10).value
            chosen_emp_leave_3_date = worksheet_for_getting_leave.cell(row=chosen_row_number, column=11).value
            chosen_emp_return_3_date = worksheet_for_getting_leave.cell(row=chosen_row_number, column=12).value
            join_resign_leave_dates.extend([chosen_emp_join_date, chosen_emp_resign_date,
                                            chosen_emp_leave_1_date, chosen_emp_return_1_date,
                                            chosen_emp_leave_2_date, chosen_emp_return_2_date,
                                            chosen_emp_leave_3_date, chosen_emp_return_3_date])
            # print(name, join_resign_leave_dates)
            name_and_joining_date_dict[name] = join_resign_leave_dates
    workbook_master.close()
    return name_and_joining_date_dict


def extract_date(casear):
    # print(casear)  # list of incen
    # loading the data sheet for getting the date of incentive cases
    date_workbook = openpyxl.load_workbook(
        r'G:\My Drive\GdrivePC\Hospital\RSBY\New\DKBSSY\GMC 16.1.24 6704 Dr. Khoobchand Baghel Swastyha Sahayata Yojana.xlsx')
    date_worksheet = date_workbook['all']
    column_of_case_number = list(date_worksheet.iter_rows(min_col=2, max_col=2))

    date_all = []
    for case_nox in casear:  # casear.split('\n')
        yy = case_nox.split("\t")
        # print(yy)
        if len(yy) == 6:
            # print(yy)
            # yy_date = yy[4].split("-")
            # date_is = datetime.date(year=int(yy_date[2]), month=int(yy_date[1]), day=int(yy_date[0]))
            # print('....',date_is)
            date_str = yy[4]
            # print(date_str)
            try:
                date_is = (datetime.datetime.strptime(date_str, "%d-%m-%Y"))
            except ValueError:
                try:
                    date_is = (datetime.datetime.strptime(date_str, "%d-%m-%y"))
                except ValueError:
                    ColourPrint.print_yellow('Date format should be separated by dash ("-")')
                    ColourPrint.print_bg_red('Program Stopped. Restart after modifying date format')
                    sys.exit()


            # date_is = date_is.date()  # converts to only date strips time
            # print('mmm', date_is)
            date_all.append(date_is)
    # print(date_all)
    date_workbook.close()
    return date_all

def extract_date_only_by_case_number(casear):
    # print(casear)
    # Load the workbook and the specific worksheet
    date_workbook = openpyxl.load_workbook(
        r'G:\My Drive\GdrivePC\Hospital\RSBY\New\GMC 16.1.24 6704 Dr. Khoobchand Baghel Swastyha Sahayata Yojana.xlsx')
    date_worksheet = date_workbook['all']

    # List to store the corresponding values for each case number
    results = []

    # Iterate over each case number in the search list
    for search_case_number in casear:
        # Flag to check if the case number is found
        found = False

        # Iterate over rows to find the matching case number and get the corresponding value
        for row in date_worksheet.iter_rows(min_row=2, max_col=7, values_only=True):  # Adjust max_col if necessary
            case_number = row[1]  # Column B (index 1)

            # Check if the case number matches the search case number
            if case_number == search_case_number:
                corresponding_value = row[5]  # Assuming value of interest is in Column E (index 4)
                # print(corresponding_value)
                date_is = corresponding_value
                results.append(date_is)
                found = True
                break
    return results


def checker_with_dict_output(casear):
    dict_output = {}
    # below extracts date of workings
    working_date_extracted_dict = working_date_extractor()  # {name:[datejoin,dateresign,leave1,return1,leave2,retur2,..}]
    print()

    # use case for the name and date and amount copy and pasted
    casewise_extracted_dates_list = extract_date(casear)  # [datetime.date(2022, 8, 27), datetime.date(2022, 9, 9)]

    # use this for the process of entry only by pasting the case number
    # casewise_extracted_dates_list = extract_date_only_by_case_number(casear)  # [datetime.date(2022, 8, 27), datetime.date(2022, 9, 9)]

    # print("casewise_extracted_dates_list", casewise_extracted_dates_list)
    if not casewise_extracted_dates_list:
        print('Error - ENTER CASE NUMBER WITH AMOUNT AND PROCEDURE AND NAME')
        ColourPrint.print_bg_red('Programme Stopped')
        print()
        sys.exit()
    else:
        pass
    for date in casewise_extracted_dates_list:
        out = []
        un_updated_leaves, non_employed = [], []
        trail_dict = {}
        for name, date_periods in working_date_extracted_dict.items():
            chosen_emp_join_date, chosen_emp_resign_date, _1_leave, _1_return, _2_leave, _2_return, _3_leave, _3_return = date_periods
            leave_periods_pairs = _1_leave, _1_return, _2_leave, _2_return, _3_leave, _3_return
            # print(type(leave_periods_pairs))
            if chosen_emp_join_date <= date <= chosen_emp_resign_date:  # working period check
                # on_leave = False
                for non_work_day in range(0, len(leave_periods_pairs), 2):
                    leave_date = leave_periods_pairs[non_work_day]
                    return_date = leave_periods_pairs[non_work_day+1]
                    # print(name, leave_date, return_date)
                    if leave_date and return_date:  # non None data
                        if leave_date <= date <= return_date:
                            # on_leave = True
                            out.append(name)
                            break
                    elif (leave_date and not return_date) or (not leave_date and return_date):
                        un_updated_leaves.append(name)
                        break

            else:  # red colour print(f"\033[91m{error_message}\033[0m")
                non_employed.append(name)
                # print(f'Check the JOINING and RELIEVING Dates of ================={name}================')

        if non_employed:
            # print(f"\033[92m{date.date()} इन तिथियों में नौकरी पर नहीं रखे गए थे ================= {non_employed} ==========\033[0m")
            trail_dict['Not Employed'] = non_employed
        if un_updated_leaves:
            # print(f'\033[94m{date.date()} इनकी छुटियाँ दिनांक सही प्रविष्ट करें ......................{un_updated_leaves}............\033[0m')
            trail_dict["To Update Leaves"] = un_updated_leaves
        trail_dict['On leave']=out

        # print(f'\033[92m{trail_dict}\033[0m')
        # print('clr', clear_list)
        dict_output[str(date.date())] = trail_dict
    for k, v in dict_output.items():
        # print(f'\033[96m-------------------------------------------------------------------------------------------------------\033[0m')
        print(f'\033[41m{k}:\033[0m')
        for vk,vv in v.items():
            if vk == 'Not Employed':
                print(" "*5 + f'\033[92m{vk}: {vv}\033[0m')
            elif vk == 'To Update Leaves':
                print(" "*5 + f'\033[94m{vk}: {vv}\033[0m')
            elif vk == 'On leave':
                print(" "*5 + f'\033[93m{vk}: {vv}\033[0m')
        print(f'\033[95m-------------------------------------------------------------------------------------------------------\033[0m')
        print()
        # print()
    input('Press Enter to continue')
    # return dict_output


# checker_with_dict_output(casear)
#