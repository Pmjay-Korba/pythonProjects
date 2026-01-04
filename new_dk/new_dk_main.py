import asyncio
import json
import time
from dkbssy.utils.colour_prints import ColourPrint, message_box
import openpyxl
from aiohttp import ClientSession
from bs4 import BeautifulSoup
from playwright.async_api import async_playwright
from EHOSP.tk_ehosp.alert_boxes import error_tk_box
from TMS_new.async_tms_new.desired_page import get_desired_page_indexes_in_cdp_async_for_ASYNC
from dkbssy.utils import name_for_date_check_gmc
from dkbssy.utils.excel_utils import ExcelMethods
from dkbssy.utils.sqlite_updation_long import sql_update_for_3
from svnsssy.new_svnsssy.amount_update_request import main_aiohttp_optimized
from svnsssy.new_svnsssy.async_dkbssy import indent_json_print, split_multiline_to_list, get_cat_and_names_for_entry, \
    update_sql_in_auto_2, update_sql_in_auto_2_by_emp_code

SSL = False

async def _main(all_incentive_multiline):
    excel_path = r'G:\My Drive\GdrivePC\Hospital\RSBY\New\Incentive_auto_ver_3.xlsx'

    user_title_of_page = 'Shaheed Veer Narayan Singh Ayushman Swasthya Yojna'
    page_indexes = await get_desired_page_indexes_in_cdp_async_for_ASYNC(user_title_of_page=user_title_of_page)

    url = "https://dkbssy.cg.nic.in/secure/incentivemodule/incentivemoduledme.aspx"

    # only for checking the approver site open
    # await get_desired_page_indexes_in_cdp_async_for_ASYNC(user_title_of_page='Shaheed Veer Narayan Singh Ayushman Swasthya Yojna', cdp_port=9223)

    # from_date = '01-08-2022'
    from_date = '2022-08-01'
    # to_date = '31-12-2023'
    to_date = '2023-12-31'
    department_name_is = get_department(excel_path=excel_path)

    # splits to individual data line
    incentive_cases_list_pasted = split_multiline_to_list(all_incentive_multiline=all_incentive_multiline)  # [CASE/PS5/HOSP22G146659/CK5998243	14960	23/02/2023	Mina Bai Sahu	Burns Management	Per, ...]
    # indent_json_print(incentive_cases_list_pasted)

    "EXCEL WORKS"
    # function ran
    name_for_date_check_gmc.checker_with_dict_output(incentive_cases_list_pasted)
    "Incentive Names in NAME format"
    final_cat_and_names = get_cat_and_names_for_entry()
    print()
    ColourPrint.print_yellow('-'*150)
    print(final_cat_and_names)
    ColourPrint.print_yellow('-'*150)

    "Incentive names in Code format"
    e_coded_names = convert_names_to_codes(name_format_list_of_list=final_cat_and_names, excel_path=excel_path)
    print()
    ColourPrint.print_green('-'*150)
    print(e_coded_names)
    ColourPrint.print_green('-'*150)

    async with async_playwright() as p:
        cdp_for_main = 9222
        browser = await p.chromium.connect_over_cdp(f"http://localhost:{cdp_for_main}")
        context = browser.contexts[0]
        page = context.pages[page_indexes[0]]

        # Get cookies
        cookies = await context.cookies()
        session_cookie = next((c for c in cookies if c['name'] == 'ASP.NET_SessionId'), None)

        if not session_cookie:
            print("Session cookie not found.")
            return

        asp_net_session = session_cookie['value']
        cookie_dict = {"ASP.NET_SessionId": asp_net_session}
        print(f"Extracted ASP.NET_SessionId: {cookie_dict}")

        async with ClientSession() as session:
            all_incentive_list_html = await get_the_dept_incentive_list_of_date_range(
                url=url,
                session=session,
                cookies=cookie_dict,
                depart=department_name_is,
                from_date=from_date,
                to_date=to_date
            )
            incentive_case_list_web = extract_case_ids(all_incentive_list_html)
            # print(incentive_case_list_web)
            for each_incentive_case_detailed in incentive_cases_list_pasted:  # pasted = [CASE/PS5/HOSP22G146659/CK5998243	14960	23/02/2023	Mina Bai Sahu	Burns Management	Per, ...]
                "check the already initiated"
                time1 = time.perf_counter()
                each_incentive_case, amount_is = retrieve_case_number_and_amount(each_case_number_line_data=each_incentive_case_detailed)
                # print(each_incentive_case_detailed)
                # print(each_incentive_case)
                check_if_initiated(each_incentive_case, amount_is, incentive_case_list_web)

                "STARTING ENTRY"
                "step-1 Redirect"
                raw_case_url = f"https://dkbssy.cg.nic.in/secure/incentivemodule/incentivedetailsdme.aspx?c={each_incentive_case}&amt={amount_is}"
                redirect_url = await fetch_case_redirect_details(session=session, cookies=cookie_dict, raw_case_url=raw_case_url)

                payload_name_radio_select = {
                    "ctl00$ContentPlaceHolder1$RadioButtonList1": "1",
                    "__EVENTTARGET": "ctl00$ContentPlaceHolder1$RadioButtonList1$0",
                    "__EVENTARGUMENT": "",
                }
                response_html = await post_common(raw_case_url,session,cookie_dict,payload_name_radio_select, redirect_url)

                for i, cat_wise_names in enumerate(e_coded_names,start=1):
                    payload_select_category = {
                        "ctl00$ContentPlaceHolder1$empCategory": f"{i}",
                        "__EVENTTARGET": "ctl00$ContentPlaceHolder1$empCategory",
                        "__EVENTARGUMENT": "",
                    }
                    response_cat_select = await post_common(raw_case_url, session, cookie_dict, payload_select_category, response_html)
                    if i==1: # Checking names Once:
                        web_codename_list = name_codes_present_in_web_options(response_cat_select)

                    "Check the names during each category entry"
                    cat_wise_codenames = cat_wise_names[1:]  # 1st element is hindi cat
                    check_the_name_codes_prent_in_web_options(cat_wise_codenames, web_codename_list)

                    # print (response_cat_select)
                    payload_name_selection = {
                    "ctl00$ContentPlaceHolder1$empName":cat_wise_codenames,
                    "ctl00$ContentPlaceHolder1$Button1":"Add Staff"
                    }

                    response_name_selected = await post_common(raw_case_url,session,cookie_dict,payload_name_selection,response_cat_select)
                    # print(response_name_selected)
                    response_html = response_name_selected
                    ColourPrint.print_yellow(f"Completed category -> {i}. Number of names: {len(cat_wise_codenames)}")

                payload_submit ={
                    "ctl00$ContentPlaceHolder1$CheckBox1":"on",
                    "ctl00$ContentPlaceHolder1$CheckBox2":"on",
                    "ctl00$ContentPlaceHolder1$CheckBox3":"on",
                    "ctl00$ContentPlaceHolder1$CheckBox4":"on",
                    "ctl00$ContentPlaceHolder1$button_submit":"Submit"
                }

                # print('Now Submitting')
                response_submit = await post_common(raw_case_url,session,cookie_dict,payload_submit,response_html)
                # print(response_submit)
                time_submitting = time.perf_counter()
                ColourPrint.print_blue("Time", time_submitting-time1)


                "resubmitting the submitted"
                await reinitiate_the_initiated(case_number=each_incentive_case,session=session,cookie_dict=cookie_dict)
                time2 =time.perf_counter()
                ColourPrint.print_blue("Time", time2-time1)

                "printing in the one page. The reinitiated"
                print_url = await incentive_print_url(case_number=each_incentive_case,amount=amount_is)
                page_print = await context.new_page()
                await page_print.goto(print_url)

                master_for_db = [each_incentive_case, amount_is, "A", "B", "C", "D"]
                incentive_amount_data_json = await get_incentive_amount_data(case_number=each_incentive_case, amount=amount_is, session=session,cookie_dict=cookie_dict)
                processed_data = processing_incentive_data(incentive_amount_data_json)  # converting to detail_str = f'{name}@{amt}#{cat_num}^{e_id}'
                master_for_db.extend(processed_data)
                indent_json_print(master_for_db)

                'Excel save'
                ExcelMethods().excel_save_new(final_lol=master_for_db)
                sql_update_for_3(full_list_of_list=master_for_db)
                ColourPrint.print_green(message_box('Excel and DB saved'))
                time3 = time.perf_counter()
                ColourPrint.print_blue("Time",time3-time1)

        "Updating the incentive auto 2 from db3 -> to update the auto2 fron DB entered amount"
        update_sql_in_auto_2_by_emp_code()

        "updating auto2 globally -> to update the auto2 from web the paid unpaid"
        await main_aiohttp_optimized()


async def post_common(url,session,cookies,payload_actual,response_html):
    if response_html is not None:
        hidden = extract_hidden_fields(response_html)
    else:
        hidden = {}

    payload_with_hidden = {
        **hidden,
        **payload_actual
    }

    async with session.post(url,cookies=cookies,data=payload_with_hidden, ssl=SSL) as resp:
        text = await resp.text()
        # ColourPrint.print_turquoise("POST STATUS:", resp.status)
        # print(text)  # print first 2000 chars only
        return text

async def reinitiate_the_initiated(case_number, session, cookie_dict):
    """
    reinitiate the same case number via query modify
    :param case_number: incentive case number
    :return: html
    """
    url = f"https://dkbssy.cg.nic.in/secure/incentivemodule/incentivemoduleQuerryViewDME_Edit.aspx?ci={case_number}"
    fetching_for_hidden = await fetch_case_redirect_details(session=session, cookies=cookie_dict, raw_case_url=url)

    payload_reinitiate = {
        "ctl00$ContentPlaceHolder1$button_submit":"Submit"
    }

    reinitiating = await post_common(url=url,session=session,cookies=cookie_dict,payload_actual=payload_reinitiate,response_html=fetching_for_hidden)

    # return html for next processing the hidden
    return reinitiating


async def incentive_print_url(case_number, amount):
    # url = "https://dkbssy.cg.nic.in/secure/incentivemodule/incentivemoduleViewPrintDME.aspx?ci=CASE/PS6/HOSP22G146659/CK6377979&amt=785.4" ACTUAL IT IS
    url = f"https://dkbssy.cg.nic.in/secure/incentivemodule/incentivemoduleViewPrintDME.aspx?ci={case_number}&amt={amount}"
    return url


async def get_incentive_amount_data(case_number, amount, session, cookie_dict):
    url ="https://dkbssy.cg.nic.in/secure/incentivemodule/incentivemoduleViewPrintDME.aspx/getData"
    payload_get_data = {"caseNoReqR": f"{case_number}","incentiveAmtText": f"{amount}"}
    headers = {
        "Content-Type": "application/json"
    }

    # response = await post_common(url=url,session=session,cookies=cookie_dict,payload_actual=json_payload,response_html=None, headers=headers)
    # # print(response.status)

    async with session.post(url,cookies=cookie_dict,json=payload_get_data, headers=headers, ssl=SSL) as resp:
        response = await resp.json()

    return response

def processing_incentive_data(web_json_data):
    """    web_json_data =
    {'d': [{'__type': 'incentivemodule_incentivemoduleViewPrintDME+clsData1',
               'empCategory': 'डाटा एंट्री ऑपरेटर', 'empPost': 'Data Entry Operator(DME)',
               'empName': 'SURINDAR LAL CHANDRA',
               'empCode': 'DME55173655',
               'incentiveText': '26.5',
               'incentveWithTDSamt': '29.45',
               'incentiveTDSamt': '2.95'},
         {'__type': 'incentivemodule_incentivemoduleViewPrintDME+clsData1', 'empCategory': 'चतुर्थ वर्ग एवं सफाई कर्मचारी', 'empPost': 'Sweepers(DME)', 'empName': 'K. Damodar', 'empCode': '05170020110', 'incentiveText': '1.76', 'incentveWithTDSamt': '1.96', 'incentiveTDSamt': '0.2'}
         {'__type': 'incentivemodule_incentivemoduleViewPrintDME+clsData1', 'empCategory': 'नर्सिंग एवं पैरामेडिकल स्टाफ ', 'empPost': 'Health Assistant(DME)', 'empName': 'Narendra Kumar Kanwar', 'empCode': '04170140534', 'incentiveText': '2.71', 'incentveWithTDSamt': '3.01', 'incentiveTDSamt': '0.3'}
         {"__type": "incentivemodule_incentivemoduleViewPrintDME+clsData1", "empCategory": "एनेस्थीसिया", "empPost": "Professor(DME)", "empName": "Dr. Durga Shankar Patel", "empCode": "66170010344", "incentiveText": "212.06", "incentveWithTDSamt": "235.62", "incentiveTDSamt": "23.56"},
         {'__type': 'incentivemodule_incentivemoduleViewPrintDME+clsData1', 'empCategory': 'सभी सीनियर एवं जूनियर रेसिडेंट ', 'empPost': 'Registrar / Tutor / Senior Resident(DME)', 'empName': 'Dr. Arunima', 'empCode': 'DME55172473', 'incentiveText': '19.28', 'incentveWithTDSamt': '21.42', 'incentiveTDSamt': '2.14'}
         {'__type': 'incentivemodule_incentivemoduleViewPrintDME+clsData1', 'empCategory': 'सभी फिजिशियन / सर्जन ', 'empPost': 'Professor(DME)', 'empName': 'Dr. Gopal Singh Kanwer', 'empCode': '11170010478', 'incentiveText': '119.29', 'incentveWithTDSamt': '132.54', 'incentiveTDSamt': '13.25'}
         {'__type': 'incentivemodule_incentivemoduleViewPrintDME+clsData1', 'empCategory': 'पैथोलॉजी ,रेडियोलॉजी , माइक्रोबायोलॉजी , बायोकेमेस्ट्री , निश्चेतना (टेक्निशियन )', 'empPost': 'Technician(DME)', 'empName': 'ACHCHHE KUMAR PATLEY', 'empCode': '04170150173', 'incentiveText': '2.55', 'incentveWithTDSamt': '2.83', 'incentiveTDSamt': '0.28'}
         {'__type': 'incentivemodule_incentivemoduleViewPrintDME+clsData1', 'empCategory': 'पैथोलॉजी ,रेडियोलॉजी , माइक्रोबायोलॉजी , बायोकेमेस्ट्री , ( फैकल्टी एंड रेसीडेंट )', 'empPost': 'Senior Registrar -Radiology(DME)', 'empName': 'Dr. manoj Kumar', 'empCode': '05530010029', 'incentiveText': '13.85', 'incentveWithTDSamt': '15.39', 'incentiveTDSamt': '1.54'}
         {'__type': 'incentivemodule_incentivemoduleViewPrintDME+clsData1', 'empCategory': 'अधिष्ठाता अस्पताल अधीक्षक ,सहायक अधीक्षक नोडल अधिकारी एवं सहायक नोडल अधिकारी , अस्पताल  सलाहकार  ', 'empPost': 'Dean (DME)', 'empName': 'AVINASH MESHRAM', 'empCode': '66170010023', 'incentiveText': '3.54', 'incentveWithTDSamt': '3.93', 'incentiveTDSamt': '0.39'}
         ]
         }"""
    main_data = web_json_data['d']
    _category_all_value_pair_website_based = {
        'अधिष्ठाता अस्पताल अधीक्षक ,सहायक अधीक्षक नोडल अधिकारी एवं सहायक नोडल अधिकारी , अस्पताल  सलाहकार': '1',
        'पैथोलॉजी ,रेडियोलॉजी , माइक्रोबायोलॉजी , बायोकेमेस्ट्री , ( फैकल्टी एंड रेसीडेंट )': '2',
        'पैथोलॉजी ,रेडियोलॉजी , माइक्रोबायोलॉजी , बायोकेमेस्ट्री , निश्चेतना (टेक्निशियन )': '3',
        'सभी फिजिशियन / सर्जन': '4', 'सभी सीनियर एवं जूनियर रेसिडेंट': '5',
        'एनेस्थीसिया': '6', 'नर्सिंग एवं पैरामेडिकल स्टाफ': '7', 'चतुर्थ वर्ग एवं सफाई कर्मचारी': '8',
        'डाटा एंट्री ऑपरेटर': '9'
    }

    processed_data = []
    for individual_data in main_data:
        # print(individual_data)
        name = individual_data['empName']
        hindi_category = individual_data['empCategory']
        cat_num = _category_all_value_pair_website_based[hindi_category.strip()]
        emp_code = individual_data['empCode']
        amount = individual_data['incentveWithTDSamt']

        detail_str = f'{name}@{amount}#{cat_num}^{emp_code}'
        # print(detail_str)
        processed_data.append(detail_str)

    return processed_data[::-1]  # reversing due to operator name is at start


def retrieve_case_number_and_amount(each_case_number_line_data):
    # "CASE/PS6/HOSP22G146659/CK7008808\tOphthalmology\tSICS with non-foldable IOL\t2550\t07-10-2023\tMayamati"
    case_number_is, amount_is = each_case_number_line_data.split('\t')[0],  each_case_number_line_data.split('\t')[1]
    # print("case number is:", case_number_is)
    return case_number_is, amount_is


async def fetch_case_redirect_details(session, cookies, raw_case_url):

    async with session.get(raw_case_url, cookies=cookies, allow_redirects=True,ssl=SSL) as resp:
        html = await resp.text()
        # print("Final URL:", str(resp.url))
        print("Status:", resp.status)
        return html


async def get_hidden_fields(url, session: ClientSession, cookies):
    async with session.get(url, cookies=cookies,ssl=SSL) as resp:
        text = await resp.text()
        return extract_hidden_fields(text)


def name_codes_present_in_web_options(page_with_options_html):
    soup = BeautifulSoup(page_with_options_html, "html.parser")

    # find the SELECT element by ID
    select_tag = soup.find(id="ctl00_ContentPlaceHolder1_empName")

    # extract only option values
    values = [opt["value"] for opt in select_tag.find_all("option")]
    return values

def check_the_name_codes_prent_in_web_options(excel_entry_code_list, name_codes_web_list):
    not_present =[]
    for code in excel_entry_code_list:
        if code not in name_codes_web_list:
            not_present.append(code)
    if not_present:
        msg =f'The name(s) -> {not_present} not present in the option list. Check Manually'
        error_tk_box(error_message=msg)
        raise NameError(msg)


def extract_hidden_fields(html_text):
    """
    Extracts ASP.NET hidden form fields dynamically.
    """
    soup = BeautifulSoup(html_text, "html.parser")
    fields = {}

    # extract all <input type="hidden">
    for tag in soup.find_all("input", type="hidden"):
        name = tag.get("name")
        value = tag.get("value", "")
        if name:
            fields[name] = value

    # Extract ScriptManager hidden async fields if present
    # (These can be non-hidden but required)
    script_manager = soup.find("input", id=lambda x: x and "ScriptManager" in x)
    if script_manager:
        fields[script_manager.get("name")] = script_manager.get("value", "")

    return fields


async def get_the_dept_incentive_list_of_date_range(url, session: ClientSession, cookies, depart, from_date, to_date):

    # Step 1: GET → get fresh VIEWSTATE etc.
    hidden = await get_hidden_fields(url, session, cookies)

    data = {
        **hidden,
        "ctl00$ContentPlaceHolder1$HospID": "HOSP22G146659",
        "ctl00$ContentPlaceHolder1$speciality": depart,
        "ctl00$ContentPlaceHolder1$fdate": f"{from_date}",
        "ctl00$ContentPlaceHolder1$tdate": f'{to_date}',
        "ctl00$ContentPlaceHolder1$SearchCases": "Search",
    }

    async with session.post(url, cookies=cookies, data=data, ssl=SSL) as resp:
        text = await resp.text()
        # print("\n---- POST STATUS:", resp.status)
        # print(text)  # print first 2000 chars only
        return text


def check_if_initiated(case_number, amount, incentive_case_list_web):
    if case_number not in incentive_case_list_web:
        msg = (f"The incentive case number -> {case_number} might be already initiated. Check dates, department and then restart.\n\n"
               f"Check below link in approver site for details.")
        error_tk_box(msg)
        print("CHECK IN APPROVER SITE", f'https://dkbssy.cg.nic.in/secure/incentivemodule/incentivemoduleApViewDME.aspx?ci={case_number}')
        print("CHECK IN MAIN SITE", f"https://dkbssy.cg.nic.in/secure/incentivemodule/incentivedetailsdme.aspx?c={case_number}&amt={amount}")
        raise FileExistsError(msg)


def get_department(excel_path):
    wb_3 = openpyxl.load_workbook(excel_path)
    w_sheet_3_sh1 = wb_3["Sheet1"]
    department_choice = w_sheet_3_sh1["B2"].value
    wb_3.close()
    ColourPrint.print_pink(department_choice)
    return department_choice


def extract_case_ids(html_text):
    soup = BeautifulSoup(html_text, "html.parser")

    case_ids = []

    # Find all <a> tags with class "badge-success"
    for a in soup.find_all("a", class_="badge-success"):
        text = a.get_text(strip=True)
        if text.startswith("CASE/"):
            case_ids.append(text)
        else:
            print("Casw not stared with 'CASE/'")

    return case_ids


def name_emp_code_pair_dict(excel_path)->dict:
    """
    get the name and e_code pairs
    :param excel_path: ver_3
    :return: dict(name,code)
    :error: when name is present and code is missing in Excel
    """
    work_book = openpyxl.load_workbook(excel_path,read_only=True, data_only=True)
    sheet_3 = work_book['Sheet3']
    rows = list(sheet_3.iter_rows(min_col=2, max_col=3, values_only=True))
    # print(rows)
    name_code_dict = {}
    for name_code in rows:
        # print(name_code)
        n, e = name_code
        # print(n,e)
        if n is not None:
            if e is None or str(e).strip()=="":
                raise NameError(f'The Employee Code for name-> "{n}" is missing')
            name_is = n.strip()
            code_is =str(e).strip()
            name_code_dict[name_is] = code_is
    # print(name_code_dict)
    # indent_json_print(name_code_dict)
    work_book.close()
    return name_code_dict

def convert_names_to_codes(name_format_list_of_list,excel_path)->list:
    """
    Used to convert the list of names in master list to code list
    :param name_format_list_of_list: names in category sublist wise of list
    :param excel_path: ver_3
    :return: list of category wise codes in master list
    """

    """[
          [
            "अधिष्ठाता अस्पताल अधीक्षक ,सहायक अधीक्षक नोडल अधिकारी एवं सहायक नोडल अधिकारी , अस्पताल  सलाहकार  ",
            "DR.ANMOL MADHUR MINZ",
            "AVINASH MESHRAM",
            "RAVIKANT JATWAR",
            "Dr. Durga Shankar Patel",
            "Dr. Rakesh Kumar Verma",
            "Dr. Gopal Singh Kanwer",
            "Dr. Aditya Siodiya"
            ],
          [hindi category 2... names
          ] 
    ]"""

    name_code_dict = name_emp_code_pair_dict(excel_path=excel_path)
    # print(name_code_dict)

    master_list = []
    for all_name_in_cat in name_format_list_of_list:
        sub_list = [all_name_in_cat[0]]
        for name in all_name_in_cat[1:]:
            # print('----', name)
            e_code_of_name = name_code_dict[name]
            sub_list.append(e_code_of_name)
        master_list.append(sub_list)

    # print(master_list)
    return master_list




def main(all_incentive_multiline):
    asyncio.run(_main(all_incentive_multiline))


if __name__ == "__main__":
    name_emp_code_pair_dict(r'G:\My Drive\GdrivePC\Hospital\RSBY\New\Incentive_auto_ver_3.xlsx')
