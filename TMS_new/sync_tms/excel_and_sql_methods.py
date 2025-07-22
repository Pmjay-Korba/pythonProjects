import json
import sqlite3
import time
from datetime import datetime, timedelta
import os
import openpyxl
from playwright.sync_api import sync_playwright, Page, TimeoutError
from dkbssy.utils.colour_prints import ColourPrint
from TMS_new.sync_tms import selector_xp
from old_dkbssy_folder import tms_department_wise_2
from TMS_new.details_list_dis_query_NEW import AllListGenerator

class ExcelMethods:

    # def __init__(self, list_to_be_compared: list, workbook_path: str, sheet_name_to_be_compared: str,
    #              is_query_question_required: bool, page: Page):
    #     """
    #     :param workbook_path: path of workbook
    #     :param list_to_be_compared: Python list to be used as iterator for playwright (for latest Updating) E.g. pending discharge list, query list
    #     :param sheet_name_to_be_compared: Excel Sheet where the data's already saved which will be updated deleted etc. e.g. pend_discharge2
    #     :param is_query_question_required: check if query question is required or not. True means query will be scraped
    #     :param page: Playwright Page
    #     """
    #     self.workbook_path = workbook_path
    #     self.both_query_playwright_list = list_to_be_compared
    #     self.sheet_name_to_be_compared = sheet_name_to_be_compared
    #     self.is_query_question_required = is_query_question_required
    #     self.page = page

    def __init__(self, workbook_path: str):
        """
        Initialising the methods
        :param workbook_path: Path of Excel file
        """
        self.workbook_path = workbook_path

    def excel_difference_of_set(self, old_comparison_list: list, new_comparison_list: list) -> tuple[set, set]:
        """
        used to get the difference of the two list and output as sets of two data as tuple
        :param old_comparison_list: existing old list
        :param new_comparison_list:
        :return: tuple of 'delete set' and 'add set' respectively
        """
        to_delete_set = set(old_comparison_list) - set(new_comparison_list)
        to_add_set = set(new_comparison_list) - set(old_comparison_list)
        return to_delete_set, to_add_set

    def get_excel_column_data_list(self, sheet_name: str, min_cols: int, max_cols: int | None = None,
                                   workbook_path: str | None = None) -> list[
        tuple[str | float | datetime | None, ...]]:
        """
        Use to get the column data from the Excel. Useful to get many column at once,
        but prefer one column request and use same function to get another column data
        :param workbook_path: absolute path of workbook, if none than taken from object creation time from class
        :param sheet_name: sheet name of workbook
        :param min_cols: min_cols parameter same as openpyxl
        :param max_cols: max_cols parameter same as openpyxl
        :return: list of raw data [(data1,...), (data2,...)]
        """
        if workbook_path is None:
            wb = openpyxl.load_workbook(self.workbook_path)  # get parent workbook at object time creation fron __init__
        else:
            wb = openpyxl.load_workbook(workbook_path)  # path given for this function
        work_sheet = wb[sheet_name]
        column_datas = work_sheet.iter_cols(min_col=min_cols, max_col=max_cols, values_only=True)
        # print(type(column_datas))

        return list(column_datas)

    def filtered_excel_data_with_new_web_data(self,sheet_name:str ,min_cols:int,max_cols:int, comparator_new_list)-> tuple:
        """
        Filter the old and new scrapped web items and return to delete and to add list
        :param sheet_name: Sheet name
        :param min_cols: min cols
        :param max_cols: max cols
        :param comparator_new_list: list scrapped for comparison
        :return: tuple of delete and add list
        """
        # getting the column data = [('Case Number', 1003637676.0, 1004458743.0, 1004365832.0, 1004396234.0)]
        case_number_from_excel = self.get_excel_column_data_list(sheet_name=sheet_name,min_cols=min_cols,max_cols=max_cols)
        # print(case_number_from_excel_discharge)

        # filtering the case numbers only. data = [('Case Number', 1003637676.0, 1004458743.0, 1004365832.0, 1004396234.0)]
        filter_case_number_from_excel_str_old = [str(int(i)) for i in case_number_from_excel[0] if (type(i) is float or type(i) is int)]
        # print(case_number_from_excel_discharge_string)

        ColourPrint.print_pink('-' * 150)
        print('Downloaded Excel length', len(filter_case_number_from_excel_str_old), filter_case_number_from_excel_str_old)
        # print('Downloaded Excel Old', case_number_from_excel_discharge)
        print('JSON Length', len(comparator_new_list), comparator_new_list)

        ColourPrint.print_pink('-' * 150)
        ColourPrint.print_turquoise('-' * 150)

        to_delete_case_number_set, to_add_case_number_set = self.excel_difference_of_set(old_comparison_list=filter_case_number_from_excel_str_old,
                                                                                           new_comparison_list=comparator_new_list)
        to_delete_case_number_list = list(to_delete_case_number_set)
        to_add_case_number_list = list(to_add_case_number_set)

        print(f'Delete items len:{len(to_delete_case_number_list)} -> {to_delete_case_number_list}')
        print(f'Add items len:{len(to_add_case_number_list)} -> {to_add_case_number_list}')

        ColourPrint.print_turquoise('-' * 150)

        return to_delete_case_number_list, to_add_case_number_list



    # def delete_rows(self, sheet_name:str, to_delete_item_values_list:list[int|str], matching_column_number:int, workbook_path:str|None = None) -> None:
    #     """
    #     :param workbook_path: absolute path of workbook, if none than taken from object creation time from class
    #     :param sheet_name: sheet name of workbook
    #     :param to_delete_item_values_list: list of value which is to be searched and delete its row (int or str)
    #     :param matching_column_number: column index in which the matching value is present
    #     :return: None
    #     """
    #
    #     if not to_delete_item_values_list:
    #         print("No values provided for deletion.")
    #         return
    #
    #     to_delete_item_values_list_string = [str(int(i)) for i in to_delete_item_values_list]
    #
    #     wb_path = workbook_path if workbook_path else self.workbook_path  ## get parent workbook at object time creation from __init__
    #     wb = openpyxl.load_workbook(wb_path)  # Load workbook
    #     work_sheet = wb[sheet_name]
    #
    #     # Iterate through the rows in reverse to avoid issues with deleting rows
    #     for row in range(work_sheet.max_row, 1, -1):
    #         cell_value = work_sheet.cell(row=row, column=matching_column_number).value  # Assuming 'C' is the 3rd column
    #         if cell_value is not None:
    #             cell_value = str(int(cell_value))
    #             # print(f'Deleted Function {cell_value = }')
    #         if cell_value in to_delete_item_values_list_string:
    #             work_sheet.delete_rows(row)
    #             print(f'Deleted Value {cell_value}')
    #
    #     wb.save(wb_path)
    #     wb.close()

    def delete_rows(self, sheet_name: str, to_delete_item_values_list: list[int | str], matching_column_number: int,
                    workbook_path: str | None = None) -> None:
        """
        :param workbook_path: absolute path of workbook, if none than taken from object creation time from class
        :param sheet_name: sheet name of workbook
        :param to_delete_item_values_list: list of value which is to be searched and delete its row (int or str)
        :param matching_column_number: column index in which the matching value is present
        :return: None
        """

        if not to_delete_item_values_list:
            print("No values provided for deletion.")
            return

        start_time = time.time()

        to_delete_item_values_list_string = [str(int(i)) for i in to_delete_item_values_list]
        # print(f"Step 1: Converted delete values to strings - {time.time() - start_time:.2f} sec")

        wb_path = workbook_path if workbook_path else self.workbook_path
        wb = openpyxl.load_workbook(wb_path)
        # print(f"Step 2: Loaded workbook - {time.time() - start_time:.2f} sec")

        work_sheet = wb[sheet_name]
        # print(f"Step 3: Selected worksheet '{sheet_name}' - {time.time() - start_time:.2f} sec")

        # Iterate through rows in reverse to avoid skipping issues
        for row in range(work_sheet.max_row, 1, -1):
            cell_value = work_sheet.cell(row=row, column=matching_column_number).value  # Get cell value
            if cell_value is not None:
                cell_value = str(int(cell_value))

            if cell_value in to_delete_item_values_list_string:
                work_sheet.delete_rows(row)
                # print(f"Deleted Value {cell_value} at row {row} - {time.time() - start_time:.2f} sec")

        # print(f"Step 4: Completed row deletions - {time.time() - start_time:.2f} sec")

        wb.save(wb_path)
        # print(f"Step 5: Saved workbook - {time.time() - start_time:.2f} sec")

        wb.close()
        # print(f"Step 6: Closed workbook - {time.time() - start_time:.2f} sec")

        total_time = time.time() - start_time
        print(f"Total execution time: {total_time:.2f} sec")

    def manipulating_downloaded_excel_new_2(self, both_query_playwright_list: list, sheet_name_to_be_compared: str,
                                            is_query_question_required: bool, page: Page):
        """
        both_query_playwright_list: the retrieved updated list for those the updates process need to be run. Expected from JSON file
        sheet_name_to_be_compared: the sheet where the data is stored (an old data). Data which need to updated/deleted according
                                    to the "both_query_playwright_list" parameter
        is_query_question_required: weather the query question is required or not
        page: Playwright Page
        :return: set of difference of data
        """
        file_path = r"G:\My Drive\GdrivePC\Hospital\RSBY\New\down\cc.xlsx"
        workbook = openpyxl.load_workbook(filename=file_path, data_only=True)
        sheet_name = workbook[sheet_name_to_be_compared]
        case_numbers_in_cols = list(
            [row for row in sheet_name.iter_cols(min_col=3, max_col=3, min_row=2, values_only=True)][0])
        case_numbers_in_cols = [str(int(i)) for i in case_numbers_in_cols if i is not None]
        # for k in case_numbers_in_cols:
        #     print((type(k)))
        print('cc-new', len(case_numbers_in_cols), case_numbers_in_cols)
        print('dd-new', len(both_query_playwright_list), both_query_playwright_list)
        to_delete_case_numbers = set(case_numbers_in_cols) - set(both_query_playwright_list)
        print('to delete', len(to_delete_case_numbers), to_delete_case_numbers)
        to_add_case_numbers = set(both_query_playwright_list) - set(case_numbers_in_cols)
        print('to_add_list', len(to_add_case_numbers), to_add_case_numbers)

        # Iterate through the rows in reverse to avoid issues with deleting rows
        for row in range(sheet_name.max_row, 1, -1):
            cell_value = sheet_name.cell(row=row, column=3).value  # Assuming 'C' is the 3rd column
            if cell_value is not None:
                cell_value = str(int(cell_value))
                # print('cellvalue ', cell_value)
            if cell_value in to_delete_case_numbers:
                sheet_name.delete_rows(row)
                # print('deleted cellvalue')

        workbook.save(r"G:\My Drive\GdrivePC\Hospital\RSBY\New\down\cc.xlsx")

        master_to_save_in_excel = []
        for c_num in to_add_case_numbers:
            if not is_query_question_required:
                case_num_data = self.retrieve_date_new_claim_query(page, c_num)
                # print(case_num_data)
                for_print = "$".join(case_num_data)
                print(for_print)
                master_to_save_in_excel.append(list(case_num_data))
            else:
                case_num_data = self.retrieve_date_new_claim_query(page, c_num)
                # print(case_num_data)

                # click CASE LOG button side panel
                page.wait_for_selector("//button[normalize-space()='Case log']").click()

                # TRYING TO GET THE QUERY FROM HERE - CASE LOG
                try:
                    try:  # clicking SHOW MORE
                        page.wait_for_selector(
                            "//*[name()='circle' and @id='bg-icon']/ancestor::something/parent::div/parent::div/parent::div/div[2]/div[last()]/div[2]/div[3]//span[contains(text(),'...Show More')]").click()
                        # last_query_question_xp containing SHOW LESS
                        last_query_question_xp = "//*[name()='circle' and @id='bg-icon']/ancestor::something/parent::div/parent::div/parent::div/div[2]/div[last()]/div[2]/div[3]"

                        # question = "this is demo test"
                        question = page.wait_for_selector(last_query_question_xp).text_content()
                        # print(question)
                        # close cross button
                    except TimeoutError as err:
                        print(err)
                    finally:
                        page.wait_for_selector("//*[name()='path' and @id='cross-icon']").click()

                except TimeoutError as err:
                    # click Finance
                    page.wait_for_selector("//h4[normalize-space()='Finance']/parent::button").click()
                    # select the query chat option and click it to open chat window
                    page.wait_for_selector("(//img[@data-tip='Chat'])[last()]").click()
                    # getting question - last question
                    question = page.wait_for_selector(
                        "(//div[contains(@class,'react-draggable')]//strong)[last()]").text_content()
                    # print(question)
                    # close button
                    page.wait_for_selector("(//div[contains(@class,'react-draggable')]//img)[2]").click()

                case_num_data = list(case_num_data)
                case_num_data.append(question)
                # print(case_num_data)
                for_print = "$".join(case_num_data)
                print(for_print)
                master_to_save_in_excel.append(case_num_data)
                page.wait_for_load_state("networkidle")
                time.sleep(2)
                # going homepage
                page.wait_for_selector("//p[contains(text(),'Home')]/preceding-sibling::*[name()='svg']").click()
                page.wait_for_load_state("networkidle")
                time.sleep(2)

        time1 = time.time()
        for case_detail_list in master_to_save_in_excel:
            sheet_name.append(case_detail_list)
        workbook.save(r"G:\My Drive\GdrivePC\Hospital\RSBY\New\down\cc.xlsx")
        # function ran
        tms_department_wise_2.department_wise_extract_for_queries_2025()
        tms_department_wise_2.department_wise_extract_for_discharge_2025()
        print('Time to save in excel:', time.time() - time1, 'seconds')


class SqlMethods:
    def __init__(self, database_path):
        """
        :param database_path: database path
        """
        self.database_path = database_path

    def temp_save_data(self, table_name, to_save_data: list) -> None:
        """
        :param to_save_data: list of data to be saved
        :param table_name: name of the sqlite table
        :return: None
        """
        conn = sqlite3.connect(database=self.database_path)
        cursor = conn.cursor()

        cursor.execute(
            f'''
            CREATE TABLE IF NOT EXISTS "{table_name}"(
            card TEXT, 
            name TEXT, 
            case_number TEXT PRIMARY KEY UNIQUE, 
            date TEXT, 
            depart TEXT, 
            procedure TEXT, 
            age_and_sex TEXT, 
            current_status TEXT, 
            amount TEXT,
            pending_days TEXT,
            remark TEXT
            )
            '''
        )

        # creating the test data
        cursor.execute(
            f'''
            INSERT INTO {table_name} (
            card,
            name,
            case_number,
            date,
            depart,
            procedure,
            age_and_sex,
            current_status,
            amount,
            pending_days,
            remark)

            VALUES(
            ?,?,?,?,?,?,?,?,?,?,?
            )
            ''', to_save_data


            # [i.strip() for i in """card test,
            # name test, case_number_test_123, date test, depart test, procedure test, age_and_sex test,
            # current_status test, amount test, pending_days test, remark test""".split(',')]
        )

        conn.commit()
        conn.close()

    def load_temp_saved_data(self, table_name:str) -> list[tuple]:
        """
        Used to load the temporary data
        :param table_name: table name inside the database
        :return: list of tuple data of each row
        """

        conn = sqlite3.connect(database=self.database_path)
        cursor = conn.cursor()
        cursor.execute(
            f'''
            SELECT * FROM "{table_name}"
            '''
        )

        data_fetched = cursor.fetchall()
        data_fetched_list = list(data_fetched)

        conn.commit()
        conn.close()

        return data_fetched_list

    def delete_all_data(self, table_name:str) -> None:
        """
        Delete the all row data from the table
        :param table_name: name of the table inside the database
        :return: None
        """

        conn = sqlite3.connect(database=self.database_path)
        cursor = conn.cursor()
        cursor.execute(
            f'''
            DELETE FROM "{table_name}"
            '''
        )

        conn.commit()
        conn.close()

# Function to get the file creation time
def get_file_creation_time(file_path):
    if os.path.exists(file_path):
        print('------', os.path.getmtime(file_path))
        return os.path.getmtime(file_path)
    return None


# Check if the CSV file exists and is older than 2 hours
def is_file_older_than_specific_time(file_path, hours=2):
    creation_time = get_file_creation_time(file_path)
    if creation_time:
        file_time = datetime.fromtimestamp(creation_time)
        current_time = datetime.now()
        time_diff = current_time - file_time
        return time_diff > timedelta(hours=hours)
    return True  # file not found inaccessible thus meaning OLD

# Function to save lists to CSV (separating pre-auth, claim, and discharge)
def save_to_json(list_of_keys, list_of_values, file_path)->None:
    """
    Saving the data in json format
    :param list_of_keys: json keys
    :param list_of_values: corresponding json values for keys
    :param file_path: location of json file
    :return: None
    """
    data = {}
    for k, v in zip(list_of_keys, list_of_values):
        data[k]=v

    # Write to JSON file (overwrite existing data)
    with open(file_path, mode='w', encoding='utf-8') as file:
        json.dump(data, file, ensure_ascii=False, indent=4)


# Function to read lists from CSV
def read_from_json(file_path)->dict:
    """
    return the saved data from json as dict
    :param file_path: location of json file
    :return: dict of data
    """
    # Read from JSON file
    with open(file_path, mode='r', encoding='utf-8') as file:
        data = json.load(file)

    return data

