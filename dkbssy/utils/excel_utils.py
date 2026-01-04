import os
import sqlite3
import datetime
import time
import openpyxl
import pandas as pd
from dkbssy.utils.colour_prints import ColourPrint
import csv



class ExcelMethods:
    workbook_3 = r'G:\My Drive\GdrivePC\Hospital\RSBY\New\Incentive_auto_ver_3.xlsx'

    def entry_department(self):
        wb_3 = openpyxl.load_workbook(self.workbook_3)
        w_sheet_3_sh1 = wb_3["Sheet1"]
        department_choice = w_sheet_3_sh1["B2"].value
        wb_3.close()
        ColourPrint.print_pink(department_choice)
        return department_choice

    def get_cat_wise_inc_names(self):
        cat_and_name_lol = []
        wb_3 = openpyxl.load_workbook(self.workbook_3)
        w_sheet_3_sh1 = wb_3["Sheet1"]
        all_cate_and_names_lol = w_sheet_3_sh1.iter_cols(min_col=3, max_col=11, max_row=51)  # max_row=26
        for cols_data in all_cate_and_names_lol:
            names = []
            for each_cell in cols_data:
                if each_cell.value is not None:
                    names.append(each_cell.value)
            # print(names)
            cat_and_name_lol.append(names)
        print(cat_and_name_lol)
        wb_3.close()
        return cat_and_name_lol

    def flatten_list(self, nested_list):
        flattened = []
        for item in nested_list:
            if isinstance(item, list):
                flattened.extend(self.flatten_list(item))
            else:
                flattened.append(item)
        return flattened

    def excel_saves(self, final_lol):
        save_location = 'G:\\My Drive\\GdrivePC\\Hospital\\RSBY\\New\\Incentive_Entered_New\\'
        to_save_list_df = pd.DataFrame(final_lol).T
        case_number = final_lol[0]
        split_case_number = case_number.split('/')
        file_name_retrieved = split_case_number[-1]
        file_name = f'{file_name_retrieved}.csv'
        full_path = os.path.join(save_location, file_name)
        # saving csv
        to_save_list_df.to_csv(full_path, index=False, header=False)
        ColourPrint.print_yellow(f'File_name - {file_name} saved')

        # commented as taking very long time to update
        # saving MASTER
        t1 = time.time()


    def excel_save_new(self, final_lol):
        ''' [
            "CASE/PS6/HOSP22G146659/CB7008568",
            "2210",
            "Ophthalmology",
            "SICS with non-foldable IOL",
            "2023-10-07",
            "Lal Khan",
            "Dr. Gopal Singh Kanwer@3.68#1^11170010478",
            "DR.ANMOL MADHUR MINZ@3.68#1^04170140656",
            "Dr. Aditya Siodiya@3.68#1^09530010327",
            "Dr. Rakesh Kumar Verma@3.68#1^02530010055",
            "RAVIKANT JATWAR@3.68#1^04170140099",
            "AVINASH MESHRAM@3.68#1^66170010023",
            "GHANSHYAM SINGH JATRA@15.47#2^05170011289",
            "DR RAKESH KUMAR AGRAWAL@15.47#2^05170040053",
            "Dr. Awadh Sahu@15.47#2^DME55173201",
            "Dr. Deepa Janghel@15.47#2^05530010026",
        '''
        save_location = 'G:\\My Drive\\GdrivePC\\Hospital\\RSBY\\New\\Incentive_Entered_New\\'

        # processing final lol removing the emp id
        final_lol = [i if "^" not in i else i.split("^")[0] for i in final_lol]

        # print('final_lol', final_lol)
        case_number = final_lol[0]
        split_case_number = case_number.split('/')
        file_name_retrieved = split_case_number[-1]
        file_name = f'{file_name_retrieved}.csv'
        full_path = os.path.join(save_location, file_name)

        # Save with builtin csv module
        with open(full_path, mode='w', newline='', encoding='utf-8') as f:
            writer = csv.writer(f)
            writer.writerows([final_lol])
        ColourPrint.print_yellow(f'File_name - {file_name} saved')


    def sql_save(self, final_lol):
            conn = sqlite3.connect(r'G:\My Drive\GdrivePC\Hospital\RSBY\New\incentiveDatabase.db')
            cur = conn.cursor()
            current_timestamp = str(datetime.datetime.now())
            try:
                # Prepare the SELECT statement to check if the case number exists
                check_statement = "SELECT 1 FROM incentive_entry_T WHERE incentive_case_number = ? LIMIT 1"
                # Execute the SELECT statement
                cur.execute(check_statement, (final_lol[0],))
                result = cur.fetchone()

                # If the case number is found, proceed to delete
                if result:
                    # Prepare the DELETE statement
                    delete_statement = "DELETE FROM incentive_entry_T WHERE incentive_case_number = ?"
                    # Execute the DELETE statement
                    cur.execute(delete_statement, (final_lol[0],))

                for inc_nam in final_lol[6:]:
                    cur.execute('''
                            INSERT INTO incentive_entry_T (employee_name, incentive_case_number, incentive_amount, inc_categ, timestamp)
                            VALUES (?,?,?,?,?)''',
                                (inc_nam.split('@')[0], final_lol[0], inc_nam.split('@')[1].split('#')[0],
                                 inc_nam.split('#')[1], current_timestamp))
                conn.commit()
                conn.close()
                ColourPrint.print_yellow('DataBase Saved')
            except Exception as e:
                ColourPrint.print_bg_red("ERROR OCCURRED, ROLLBACK")
                print("Exception:", e)
                conn.rollback()
                raise FileExistsError('The Database File is corrupt. Download the latest "incentiveDatabase.db" from Google Drive backup')

    def sqlite_process(self) -> list[tuple]:
        conn = sqlite3.connect(r'G:\My Drive\GdrivePC\Hospital\RSBY\New\incentiveDatabase.db')
        cursor = conn.cursor()

        # Execute the SQL query
        cursor.execute("""
            SELECT employee_name, COUNT(incentive_amount) AS total_count, SUM(incentive_amount) AS total_incentive
            FROM incentive_entry_T
            GROUP BY employee_name
        """)

        # Fetch the result
        results = cursor.fetchall()  # [(n1,c1,a1),(n2,c2,a2)..]

        # Close the cursor and connection
        cursor.close()
        conn.close()
        # print(results)
        return results

    def sqlite_process_2(self) -> list[tuple]:
        conn = sqlite3.connect(r'G:\My Drive\GdrivePC\Hospital\RSBY\New\incentiveDatabase_2.db')
        cursor = conn.cursor()
        # Execute the SQL query
        cursor.execute("""
            SELECT e.name_emp, COUNT(d.d_inc_amt) AS total_incentive_count, SUM(d.d_inc_amt) AS total_incentive
            FROM emp_detail_t e
            JOIN distribution_t d ON e.SN = d.d_emp_code
            GROUP BY e.SN  
        """)
        # Fetch the result
        results = cursor.fetchall()  # [(n1,c1,a1),(n2,c2,a2)..]
        # Close the cursor and connection
        cursor.close()
        conn.close()
        # print(results)
        return results


    def sqlite_process_3(self):
        db = r"G:\My Drive\GdrivePC\Hospital\RSBY\New\incentiveDatabase_3.db"
        conn = sqlite3.connect(db)
        cursor = conn.cursor()

        # Execute SQL query
        cursor.execute("""
            SELECT e.name_emp, COUNT(d.distri_amount) AS total_incentive_count, SUM(d.distri_amount) AS total_incentive
            FROM emp_detail_table e
            JOIN distribution_table d ON e.id_emp = d.distri_name_id
            GROUP BY e.id_emp
        """)

        # Fetch all results: [(name1, count1, sum1), (name2, count2, sum2), ...]
        results = cursor.fetchall()

        # Close connection
        cursor.close()
        conn.close()

        return results


    def retrieve_employee_id(self) -> dict:
        workbook_master = openpyxl.load_workbook(r'G:\My Drive\GdrivePC\Hospital\RSBY\New\Incentive_auto2.xlsx')
        worksheet_for_getting_id = workbook_master["Sheet3"]
        column_names_in_sheet3 = worksheet_for_getting_id.iter_rows(min_col=2, max_col=2)
        dict_name_row_number = {}
        for namez in column_names_in_sheet3:
            # print(namez[0].value)
            name_is = namez[0].value
            row_number = namez[0].row
            # print({name_is:row_number})
            dict_name_row_number[name_is] = row_number  # gives the row containing name
        # print(dict_name_row_number)
        workbook_master.close()
        return dict_name_row_number


def get_employee_code_row_number(auto2)->dict:
    """
    Used to get the employee code row number
    :param auto2: excel path of auto2
    :return: dict
    """
    wb_auto2 = openpyxl.load_workbook(auto2)
    ws = wb_auto2['Sheet3']
    name_and_ecode_list_data = ws.iter_rows(min_col=3, max_col=3)
    # print(name_and_ecode_list)
    ecode_list = [(str(cell[0].value),cell[0].row) for cell in name_and_ecode_list_data]
    ecode_row_num_dict = dict(ecode_list)
    wb_auto2.close()
    return ecode_row_num_dict


def sqlite_process_3_by_ecode():
    db = r"G:\My Drive\GdrivePC\Hospital\RSBY\New\incentiveDatabase_3.db"
    conn = sqlite3.connect(db)
    cursor = conn.cursor()

    # Execute SQL query
    cursor.execute("""
            SELECT e.name_emp, e.code_emp, COUNT(d.distri_amount) AS total_incentive_count, 
                   SUM(d.distri_amount) AS total_incentive
            FROM emp_detail_table e
            JOIN distribution_table d ON e.id_emp = d.distri_name_id
            GROUP BY e.id_emp
            """)

    # Fetch all results: [(name1, ecode1, count1, sum1), (name2, ecode2, count2, sum2), ...]
    results = cursor.fetchall()

    # Close connection
    cursor.close()
    conn.close()

    return results


def remove_case_number_from_excel(case_number, excel_path):
    wb = openpyxl.load_workbook(excel_path)
    ws = wb['Sheet1']

    row_to_delete = None

    # Loop through rows to find the case_number
    for row in ws.iter_rows(min_row=1, values_only=False):
        cell_value = row[0].value  # Assuming case_number is in column A
        if cell_value == case_number:
            row_to_delete = row[0].row
            break

    if row_to_delete:
        ws.delete_rows(row_to_delete)
        wb.save(excel_path)
        print(f"Removed case number {case_number} from Excel.")
    else:
        print(f"Case number {case_number} not found in Excel.")

def add_in_excel_for_query_stage_pending(excel_path, case_number):
    wb = openpyxl.load_workbook(excel_path)
    ws = wb['Sheet1']
    query_modify = f'https://dkbssy.cg.nic.in/secure/incentivemodule/incentivemoduleApViewDME.aspx?ci={case_number}'
    ws.append([case_number, query_modify])
    wb.save(excel_path)
    print(f"Added case number {case_number} from Excel.")

def retrieve_emp_code_only():
    workbook_3 = openpyxl.load_workbook(r'G:\My Drive\GdrivePC\Hospital\RSBY\New\Incentive_auto_ver_3.xlsx')
    sheet = workbook_3["Sheet3"]
    e_codes = sheet.iter_cols(min_col=3, max_col=3, min_row=2, values_only=True)
    # print(list(e_codes))
    removed_none = [ i for i in list(e_codes)[0] if i is not None]
    # # print(e_codes)
    workbook_3.close()
    return removed_none


if __name__== "__main__":
    # em = ExcelMethods()
    # # print(em.retrieve_employee_id())
    # em.get_cat_wise_inc_names()
    print(retrieve_emp_code_only())