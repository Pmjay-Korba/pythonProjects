import asyncio
import datetime
import json
from pathlib import Path
import openpyxl
from numpy.ma.core import swapaxes
from playwright.async_api import async_playwright, Page
from TMS_new.async_tms_new.desired_page import get_desired_page_indexes_in_cdp_async_for_ASYNC
from dkbssy.utils.colour_prints import ColourPrint
from tms_playwright.discharge_to_be_done.detail_list_getter_all import is_file_older_than_2_hours
from tms_playwright.discharge_to_be_done.detail_list_getter_all import AllListGenerator as AllListGeneratorOld
from old_dkbssy_folder import tms_department_wise_2

class AsyncTms:
    CDP_PORT = 9222
    BENEFICIARY_URL = "https://apisprod.nha.gov.in/pmjay/provider/nproviderdashboard/V3/beneficiary/list"
    QUERY_INFO_URL = staticmethod(lambda reg_id: f'https://apisprod.nha.gov.in/pmjay/provider/provider/activity/log?registrationid={reg_id}')  # use as INFO_URL(reg_id)
    TREATMENT_INFO_URL = "https://apisprod.nha.gov.in/pmjay/provider/provider/claim/info"
    UNDER_TREATMENT_SEARCH_VALUE = '11'
    PREAUTHORIZATION_QUERY_SEARCH_VALUE ='12'
    CLAIM_QUERY_SEARCH_VALUE = '21'
    DOWNLOAD_DIR = str(Path(__file__).resolve().parent.parent / 'async_tms_new' / 'download')
    EXCEL_FILE_PATH = DOWNLOAD_DIR + "\\" + "AYUSHMAN REGISTRATION 2025.xlsx"
    INFO_OF_BENEFICIARY = "https://apisprod.nha.gov.in/pmjay/provider/provider/V3/beneficiary/info"

    async def main(self):
        async with async_playwright() as p:
            browser = await p.chromium.connect_over_cdp(f"http://localhost:{self.CDP_PORT}")
            context = browser.contexts[0]
            all_pages = context.pages

            pages_indexes = await get_desired_page_indexes_in_cdp_async_for_ASYNC(user_title_of_page='PMJAY - Provider', cdp_port=self.CDP_PORT)
            page_index = pages_indexes[0]  # selecting the first index of matching page

            page = all_pages[page_index]  # selecting the first PAGE of matching page

            # get session storage
            session_storage = await self.get_session_storage(page)

            # downloading excel
            # await spreadsheet_download_as_excel_async(browser=browser,
            #                                           download_directory_folder=self.DOWNLOAD_DIR,
            #                                           sheet_url='https://docs.google.com/spreadsheets/d/1vhjV0rcODJ4lGYJBHENMnHFvqHgK25dQRt9SVpr_9N4/edit?gid=0#gid=0')

            headers = {
                "accept": "application/json",
                "accept-encoding": "gzip, deflate, br, zstd",
                "accept-language": "en-US,en;q=0.9",
                "access-control-allow-origin": "https://provider.nha.gov.in/",
                "appname": "TMS-Provider",
                "authorization": f"Bearer {session_storage['idmToken']}",
                "uauthorization": f"Bearer {session_storage['token']}",
                "cache-control": "no-cache",
                "cid": "0",
                "content-type": "application/json; charset=UTF-8",
                "hid": "3649",
                "origin": "https://provider.nha.gov.in",
                "pid": "1935",
                "pragma": "no-cache",
                "priority": "u=1, i",
                "referer": "https://provider.nha.gov.in/",
                "scode": "22",
                "sec-ch-ua": '"Not)A;Brand";v="8", "Chromium";v="138", "Google Chrome";v="138"',
                "sec-ch-ua-mobile": "?0",
                "sec-ch-ua-platform": '"Windows"',
                "sec-fetch-dest": "empty",
                "sec-fetch-mode": "cors",
                "sec-fetch-site": "same-site",
                "tid": session_storage["transactionid"],
                "uid": "USER684607",
                "uname": "Rakesh Kumar Verma",
                "urole": "MEDCO",
                "user-agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/138.0.0.0 Safari/537.36",
                "ustate": "1935"
            }


            # await self.retrieve_logs(context=context,
            #                    headers=headers,
            #                    registration_no= '1011919559'    )

            # await self.retrieve_treatment_details(context=context,
            #                    headers=headers,
            #                    registration_no= '1011919559'  )

            # await self.retrieve_beneficiary_info(context=context,
            #                    headers=headers,
            #                    registration_no= '1003407721')

            'ALL LIST GENERATE'
            all_list_generated = await self.get_lists_of_ALL(
                url=self.BENEFICIARY_URL,
                context=context,
                headers=headers
                )

            # generating the lists
            # under_treatment_list_generated = await self.get_lists_of_pendings(
            #     url=self.BENEFICIARY_URL,
            #     search_value=self.UNDER_TREATMENT_SEARCH_VALUE,
            #     context=context,
            #     headers=headers
            #     )
            #
            # pre_auth_query_list_generated = await self.get_lists_of_pendings(
            #     url=self.BENEFICIARY_URL,
            #     search_value=self.PREAUTHORIZATION_QUERY_SEARCH_VALUE,
            #     context=context,
            #     headers=headers
            # )
            #
            # claim_query_list_generated = await self.get_lists_of_pendings(
            #     url=self.BENEFICIARY_URL,
            #     search_value=self.CLAIM_QUERY_SEARCH_VALUE,
            #     context=context,
            #     headers=headers
            # )
            #
            # both_query_list = pre_auth_query_list_generated + claim_query_list_generated
            # print('BOTH QUERY', both_query_list)

            # # deleting old registration ids from excel
            # to_add_pending_query_reg_no = self.post_deleting_in_downloaded_excel(both_query_playwright_list=both_query_list,
            #                                                                      sheet_name_to_be_compared='QUERY2')

            # query_all_data = await self.process_all(both_query_list=to_add_pending_query_reg_no,context=context,headers=headers)
            # # ColourPrint.print_green(query_all_data)
            # ColourPrint.print_yellow('Completed Query', len(query_all_data))
            # self.save_to_excel(sheet_name='QUERY2', data=query_all_data)
            #
            # # updating ward
            # tms_department_wise_2.department_wise_extract_for_queries_2025(excel_path=self.EXCEL_FILE_PATH)
            #
            # # uploading in g sheet
            # AllListGeneratorOld().upload_in_g_sheet('QUERY2', self.EXCEL_FILE_PATH)
            # ColourPrint.print_pink('Google_Spreadsheet_Query_pending_updated_NEW')
            # # updating old sheet
            # AllListGeneratorOld().upload_in_g_sheet_new('QUERY2',  self.EXCEL_FILE_PATH)
            # ColourPrint.print_pink('Google_Spreadsheet_Query_pending_updated_NEW_2')
            #
            # "UNDER TREATMENT DISCHARGE PROCESS"
            # print(under_treatment_list_generated)
            # # deleting old registration ids from excel
            # to_add_pending_dis_reg_no = self.post_deleting_in_downloaded_excel(both_query_playwright_list=under_treatment_list_generated,
            #                                                                    sheet_name_to_be_compared='Pend Dischg2')
            # under_treat_all_data = await self.process_all(both_query_list=to_add_pending_dis_reg_no, context=context,
            #                                         headers=headers)
            # ColourPrint.print_yellow('Completed Discharge', len(under_treat_all_data))
            # self.save_to_excel(sheet_name='Pend Dischg2', data=under_treat_all_data)
            # # updating discharge wards
            # tms_department_wise_2.department_wise_extract_for_discharge_2025(excel_path=self.EXCEL_FILE_PATH)
            # # uploading in g sheet
            # AllListGeneratorOld().upload_in_g_sheet('Pend Dischg2', self.EXCEL_FILE_PATH)
            # ColourPrint.print_pink('Google_Spreadsheet_Discharge_pending_updated')
            # # updating old sheet
            # AllListGeneratorOld().upload_in_g_sheet_new('Pend Dischg2', self.EXCEL_FILE_PATH)
            # ColourPrint.print_pink('Google_Spreadsheet_Discharge_pending_updated_2')

    def parse_date_of_birth(self, dob_str):
        formats = ["%Y-%m-%d", "%d-%m-%Y", "%m/%d/%Y", "%Y"]  # Add more if needed

        for fmt in formats:
            try:
                return datetime.datetime.strptime(dob_str, fmt)
            except ValueError:
                continue

        return None  # Invalid format

    def calculate_age(self, dob_str):
        dob = self.parse_date_of_birth(dob_str)
        if dob is None:
            return None  # or raise ValueError("Invalid date format")

        today = datetime.datetime.today()
        age = today.year - dob.year - ((today.month, today.day) < (dob.month, dob.day))
        return age

    async def retrieve_beneficiary_info(self, registration_no, context, headers):
        """
        Used for getting the Age and if Newborn
        :param headers: headers
        :param registration_no: registration number
        :param context: playwright context
        :return: Age, isNewborn
        """
        payload = {"registrationid": f"{registration_no}"}
        r = await context.request.post(url=self.INFO_OF_BENEFICIARY, headers=headers, data=json.dumps(payload))
        r_data = await r.json()
        date_of_birth_str = r_data['dateofbirth']
        is_newborn = r_data['newborn']
        age_is = self.calculate_age(date_of_birth_str)
        print(age_is, is_newborn)


    async def process_all(self, both_query_list, context, headers):
        semaphore = asyncio.Semaphore(10)

        async def limited_call(reg_no):
            async with semaphore:
                return await self.each_case_number_details(
                    registration_no=reg_no,
                    context=context,
                    headers=headers
                )

        tasks = [limited_call(reg_no) for reg_no in both_query_list]
        result = await asyncio.gather(*tasks)
        return result


    async def get_session_storage(self, page):
        """
        Get the session storage of indexed page
        :return: session storage dict
        """
        ColourPrint.print_green(f"✅ Connected to page: {page}")
        session_storage = await page.evaluate("sessionStorage")
        return session_storage

        # print(json.dumps(session_storage, indent=2))
        # print(session_storage["userid"],session_storage['token'])

    async def get_lists_of_ALL(self, url, context, headers):

        payload = {"hospitalid":"3649",
                   "pagenumber":"1",
                   "pagesize":"20000",
                   "searchcriteria":None,
                   "searchvalue":None}
        request_context =  context.request
        r = await request_context.post(url= url,headers=headers, data=json.dumps(payload))
        ColourPrint.print_turquoise(r)
        val = await r.json()
        patient_details = val['cases']
        # print(json.dumps(patient_details, indent=4))
        ColourPrint.print_pink(len(patient_details))
        list_generation = []
        for each_patient_details in patient_details:
            list_generation.append(each_patient_details["registrationid"])
        return list_generation


    async def get_lists_of_pendings(self, url, search_value, context, headers):

        payload = {"hospitalid":"3649",
                   "pagenumber":"1",
                   "pagesize":"20000",
                   "searchcriteria":[
                       {"key":"status",
                        "value":f"{search_value}",
                        "operation":"Equal"
                        }
                   ],
                   "searchvalue":None}
        request_context =  context.request
        r = await request_context.post(url= url,headers=headers, data=json.dumps(payload))
        ColourPrint.print_turquoise(r)
        val = await r.json()
        patient_details = val['cases']
        # print(json.dumps(patient_details, indent=4))
        ColourPrint.print_pink(len(patient_details))
        list_generation = []
        for each_patient_details in patient_details:
            list_generation.append(each_patient_details["registrationid"])
        return list_generation

    async def each_case_number_details(self, registration_no, context, headers):
        """
        Using to retrieve all data of single registration id
        :param headers: headers
        :param registration_no: registration number
        :param context: playwright context
        :return: all data of single registration id
        """
        card_is, name_is, department_is, procedure_is, sex_is, amount_is = await self.retrieve_treatment_details(registration_no=registration_no, context=context,headers=headers)
        registration_date_is, last_status, last_remark, last_remark_date = await self.retrieve_logs(registration_no=registration_no, context=context,headers=headers)

        print(registration_no, card_is, name_is, registration_no, registration_date_is, department_is, procedure_is, sex_is, amount_is, last_status,last_remark_date,last_remark)
        return card_is, name_is, registration_no, registration_date_is, department_is, procedure_is, sex_is, last_status, amount_is, '__PENDING DAYS__', last_remark,'__WARD NAME__',last_remark_date

    async def retrieve_logs(self, registration_no, context, headers) -> tuple[str, str, str, str]:
        """
        Get the logs present with raised query and raised date
        :param headers: headers
        :param registration_no: registration number
        :param context: playwright context
        :return: Query data and remark string
        """
        # request_context = context.request
        r = await context.request.get(url=self.QUERY_INFO_URL(registration_no), headers=headers)
        print('r', r)
        r_data = await r.json()
        first_json_data = r_data[0]
        last_json_data = r_data[-1]
        # print(json.dumps(last_json_data, indent=2))
        last_remark = last_json_data["remarks"]
        last_remark_date = last_json_data["raiseddate"][:10]+'b'
        last_status = last_json_data["status"]
        registration_date_is = first_json_data['raiseddate'] +'a'
        ColourPrint.print_yellow('Date: ', last_remark_date, 'Remark', last_remark)
        return  registration_date_is, last_status, last_remark, last_remark_date

    async def retrieve_treatment_details(self, registration_no, context, headers) -> tuple[str,str,str,str,str,str]:
        """
        Getting the treatment details and amount
        :param headers: headers
        :param registration_no: registration number
        :param context: playwright context
        :return: Department, Diagnosis, Amount
        """
        payload = {"registrationid":f"{registration_no}"}
        r = await context.request.post(url=self.TREATMENT_INFO_URL, headers=headers, data=json.dumps(payload))
        print(r)
        r_data = await r.json()
        # print(json.dumps(r_data, indent=2))
        card_is = r_data['encounter']['benid']
        name_is = r_data['encounter']['patientname']
        # registration_date_is = r_data["registrationdate"]
        department_is = r_data['treatments'][0]['typedesc']
        procedure_is = r_data['treatments'][0]['procedurename']
        sex_is = r_data['encounter']['patientgender']
        amount_is = r_data['amount']['claimedamount']
        ColourPrint.print_blue(card_is,name_is,department_is,procedure_is,sex_is,amount_is)
        return card_is,name_is,department_is,procedure_is,sex_is,amount_is


    def post_deleting_in_downloaded_excel(self, both_query_playwright_list: list, sheet_name_to_be_compared: str, ):
        """
        Compare the sheets after download and eliminate the non required and add new
        :param both_query_playwright_list: list newly generated by playwright
        :param sheet_name_to_be_compared: generated which are present in. Opened by openpyxl
        :return: Saving in Excel -> None
        """
        file_path = self.EXCEL_FILE_PATH
        workbook = openpyxl.load_workbook(filename=file_path, data_only=True)
        sheet_name = workbook[sheet_name_to_be_compared]
        case_numbers_in_cols = list(
            [row for row in sheet_name.iter_cols(min_col=3, max_col=3, min_row=2, values_only=True)][0])
        case_numbers_in_cols = [str(int(i)) for i in case_numbers_in_cols if i is not None]
        # print("EXCEL'S", case_numbers_in_cols)
        # for k in case_numbers_in_cols:
        #     print((type(k)))
        print("Excel's Sheets values", len(case_numbers_in_cols), case_numbers_in_cols)
        print("Playwright's Values", len(both_query_playwright_list), both_query_playwright_list)
        to_delete_case_numbers = set(case_numbers_in_cols) - set(both_query_playwright_list)
        print('to delete', len(to_delete_case_numbers), to_delete_case_numbers)
        to_add_case_numbers = set(both_query_playwright_list) - set(case_numbers_in_cols)
        ColourPrint.print_blue('to_add_list', len(to_add_case_numbers), to_add_case_numbers)

        # Iterate through the rows in reverse to avoid issues with deleting rows
        for row in range(sheet_name.max_row, 1, -1):
            cell_value = sheet_name.cell(row=row, column=3).value  # Assuming 'C' is the 3rd column
            if cell_value is not None:
                cell_value = str(int(cell_value))
                # print('cellvalue ', cell_value)
            if cell_value in to_delete_case_numbers:
                sheet_name.delete_rows(row)
                # print('deleted cellvalue')

        workbook.save(self.EXCEL_FILE_PATH)
        return to_add_case_numbers

    def save_to_excel(self, sheet_name, data):
        file_path = self.EXCEL_FILE_PATH
        workbook = openpyxl.load_workbook(filename=file_path)
        sheet = workbook[sheet_name]
        for each_data in data:
            sheet.append(each_data)
        workbook.save(file_path)


async def spreadsheet_download_as_excel_async(browser, sheet_url, download_directory_folder: str):

    new_context = await browser.new_context()
    page = await new_context.new_page()
    # Navigate to the Google Sheets URL
    await page.goto(sheet_url)
    # Click on "File" -> "Download" -> "Microsoft Excel (.xlsx)"
    await page.click("text=File")
    await asyncio.sleep(0.5)
    await page.click("text=Download")
    await asyncio.sleep(0.5)
    await page.click("text=Microsoft Excel (.xlsx)")
    # Expect the download to start and capture the download event
    async with page.expect_download() as download_info:
        pass  # The download should automatically start after clicking

    # Get the download object
    download = await download_info.value  # Waits until the download is complete
    # Get the suggested filename (e.g., "Sheet1.xlsx")
    filename = download.suggested_filename
    print(filename)
    # Save the file to your desired folder with the same name

    filepath = f"{download_directory_folder}\\{filename}"
    await download.save_as(filepath)
    await page.close()
    await new_context.close()

class Headers:
    def __init__(self, session_storage, context, sec_ch_ua, user_agent):
        self.session_storage=session_storage
        self.context=context

        print(user_agent)
        print(sec_ch_ua)

        self.header_for_tms_1 = {
        "accept": "application/json",
        "accept-encoding": "gzip, deflate, br, zstd",
        "accept-language": "en-US,en;q=0.9",
        "access-control-allow-origin": "https://provider.nha.gov.in/",
        "appname": "TMS-Provider",
        "authorization": f"Bearer {self.session_storage['idmToken']}",
        "uauthorization": f"Bearer {self.session_storage['token']}",
        "cache-control": "no-cache",
        "cid": "0",
        "content-type": "application/json; charset=UTF-8",
        "hid": "3649",
        "origin": "https://provider.nha.gov.in",
        "pid": "1935",
        "pragma": "no-cache",
        "priority": "u=1, i",
        "referer": "https://provider.nha.gov.in/",
        "scode": "22",
        # "sec-ch-ua": '"Not)A;Brand";v="8", "Chromium";v="138", "Google Chrome";v="138"',
        "sec-ch-ua": sec_ch_ua,
        "sec-ch-ua-mobile": "?0",
        "sec-ch-ua-platform": '"Windows"',
        "sec-fetch-dest": "empty",
        "sec-fetch-mode": "cors",
        "sec-fetch-site": "same-site",
        "tid": self.session_storage["transactionid"],
        "uid": self.session_storage["userid"],
        "uname": self.session_storage["username"],
        "urole": self.session_storage["userRole"],
        # "user-agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/138.0.0.0 Safari/537.36",
        "user-agent": user_agent,
        "ustate": "1935"
    }
        self.header_for_tms_2 = {
            "accept": "application/json",
            "accept-encoding": "gzip, deflate, br, zstd",
            "accept-language": "en-US,en;q=0.9",
            "access-control-allow-origin": "https://provider.nha.gov.in/",
            "appname": "TMS-Provider",
            "authorization": f"Bearer {self.session_storage['idmToken']}",
            "uauthorization": f"Bearer {self.session_storage['token']}",
            "cache-control": "no-cache",
            "cid": "0",
            "content-type": "application/json; charset=UTF-8",
            "hid": "3649",
            "origin": "https://provider.nha.gov.in",
            "pid": "1935",
            "pragma": "no-cache",
            "priority": "u=1, i",
            "referer": "https://provider.nha.gov.in/",
            "scode": "22",
            # "sec-ch-ua": '"Not)A;Brand";v="8", "Chromium";v="138", "Google Chrome";v="138"',
            "sec-ch-ua": '"Not;A=Brand";v="99", "Google Chrome";v="139", "Chromium";v="139"',
            "sec-ch-ua-mobile": "?0",
            "sec-ch-ua-platform": '"Windows"',
            "sec-fetch-dest": "empty",
            "sec-fetch-mode": "cors",
            "sec-fetch-site": "same-site",
            "tid": self.session_storage["transactionid"],
            "uid": self.session_storage["userid"],
            "uname": self.session_storage["username"],
            "urole": self.session_storage["userRole"],
            # "user-agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/138.0.0.0 Safari/537.36",
            "user-agent": '"Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/139.0.0.0 Safari/537.36',
            "ustate": "1935"
        }

if __name__ == '__main__':
    asyncio.run(AsyncTms().main())