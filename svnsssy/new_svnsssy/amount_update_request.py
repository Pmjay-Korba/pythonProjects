import asyncio
import time
import aiohttp
import openpyxl
from bs4 import BeautifulSoup
from playwright.async_api import async_playwright

# -------------------- Utility Functions --------------------

def get_name_and_ecode_from_ver_3():
    excel_path = r"G:\My Drive\GdrivePC\Hospital\RSBY\New\Incentive_auto_ver_3.xlsx"
    print('Excel Path: ', excel_path)
    wb = openpyxl.load_workbook(excel_path)
    sheet = wb['Sheet3']
    rows_datas = list(sheet.iter_rows(min_row=2, min_col=2, max_col=3, values_only=True))
    filter_name = [(n, e) for n, e in rows_datas if n is not None]
    wb.close()
    return filter_name


def parse_amount_data(html: str):
    soup = BeautifulSoup(html, "html.parser")
    data = {}
    table = soup.find("table")
    if not table:
        return data
    rows = table.find_all("tr")
    for row in rows:
        cols = row.find_all("td")
        if len(cols) == 2:
            key = cols[0].get_text(strip=True)
            value = cols[1].get_text(strip=True)
            data[key] = value
    return data


# -------------------- aiohttp Scraper --------------------

async def get_amount_data_single(session: aiohttp.ClientSession, emp_code: str, viewstate_data: dict):
    URL = "https://dkbssy.cg.nic.in/secure/incentivemodule/IncentiveDetails_EmpCodeWiseDME.aspx"

    payload = {
        "__EVENTTARGET": "",
        "__EVENTARGUMENT": "",
        "__VIEWSTATE": viewstate_data["__VIEWSTATE"],
        "__VIEWSTATEGENERATOR": viewstate_data["__VIEWSTATEGENERATOR"],
        "__VIEWSTATEENCRYPTED": "",
        "__EVENTVALIDATION": viewstate_data["__EVENTVALIDATION"],
        "ctl00$ContentPlaceHolder1$TextBox1": emp_code,
        "ctl00$ContentPlaceHolder1$search": "Search"
    }

    headers = {"Content-Type": "application/x-www-form-urlencoded"}

    async with session.post(URL, data=payload, headers=headers, ssl=False) as r:
        if r.status != 200:
            print(f"❌ POST failed {r.status} for {emp_code}")
            return {}
        html = await r.text()

    data = parse_amount_data(html)

    # Add emp_code to the dictionary
    data['emp_code'] = emp_code

    print(f"✅ Parsed {emp_code} → {data}")
    return data


async def main_aiohttp_optimized():
    start_total = time.perf_counter()

    async with async_playwright() as p:
        cdp_for_main = 9222
        browser = await p.chromium.connect_over_cdp(f"http://localhost:{cdp_for_main}")
        context = browser.contexts[0]
        pages = context.pages
        if not pages:
            print("No pages found in context.")
            return

        cookies = await context.cookies()
        session_cookie = next((c for c in cookies if c['name'] == 'ASP.NET_SessionId'), None)
        if not session_cookie:
            print("Session cookie not found.")
            return

        session_id = session_cookie['value']

        emp_data = get_name_and_ecode_from_ver_3()
        print(emp_data)  # [('Dr. Nidhi Kashyap', 'DME41172629', 2), ('Dr. Durga Shankar Patel', 66170010344, 3), ('Dr. Prabhu Dutta Sahu', 'DME41172630', 4),
        emp_codes = [e for n, e in emp_data]

        cookie_dict = {"ASP.NET_SessionId": session_id}

        URL = "https://dkbssy.cg.nic.in/secure/incentivemodule/IncentiveDetails_EmpCodeWiseDME.aspx"

        # -------------------- Extract hidden form fields ONCE --------------------
        async with aiohttp.ClientSession(cookies=cookie_dict) as session:
            async with session.get(URL, ssl=False) as resp:
                if resp.status != 200:
                    print(f"❌ Initial GET failed with {resp.status}")
                    return
                text = await resp.text()
            soup = BeautifulSoup(text, "html.parser")
            viewstate_data = {
                "__VIEWSTATE": soup.find("input", {"id": "__VIEWSTATE"})["value"],
                "__VIEWSTATEGENERATOR": soup.find("input", {"id": "__VIEWSTATEGENERATOR"})["value"],
                "__EVENTVALIDATION": soup.find("input", {"id": "__EVENTVALIDATION"})["value"]
            }

            # -------------------- Process in batches --------------------
            results = []
            batch_size = 50  # can try 100-200 if server allows
            for i in range(0, len(emp_codes), batch_size):
                batch = emp_codes[i:i + batch_size]
                batch_start = time.perf_counter()

                tasks = [get_amount_data_single(session, code, viewstate_data) for code in batch]
                batch_results = await asyncio.gather(*tasks, return_exceptions=True)

                # Filter successful results
                cleaned = [r for r in batch_results if isinstance(r, dict)]
                results.extend(cleaned)

                batch_time = time.perf_counter() - batch_start
                print(f"⏱️ Batch {i//batch_size+1} ({len(batch)} codes) took {batch_time:.2f} sec")

            total_time = time.perf_counter() - start_total
            print(f"🚀 Optimized aiohttp finished in {total_time:.2f} sec for {len(emp_codes)} codes")
            # return results
            update_auto_2_globally(results)

def update_auto_2_globally(global_data):
    excel_path = r'G:\My Drive\GdrivePC\Hospital\RSBY\New\Incentive_auto2.xlsx'
    wb = openpyxl.load_workbook(excel_path)
    ws = wb['Sheet3']
    rows_data = list(ws.iter_rows())
    # print(list(rows_data))
    # ecode_dict = {}
    for data in global_data:  #data = {'Unpaid Cases': '943', 'Unpaid Total Amount': '36258', 'Total Paid Cases': '1057', 'Total Paid Amount': '52302', 'emp_code': '12345'}
        for each in rows_data:
            if data['emp_code'] == each[2].value:
                row_value_is = each[2].row
                ws.cell(row=row_value_is, column=9).value = int(data['Unpaid Cases'])
                ws.cell(row=row_value_is, column=10).value = float(data['Unpaid Total Amount'])
                ws.cell(row=row_value_is, column=11).value = int(data['Total Paid Cases'])
                ws.cell(row=row_value_is, column=12).value = float(data['Total Paid Amount'])
    wb.save(excel_path)
    wb.close()




if __name__ == '__main__':
    update_auto_2_globally('')
    print(asyncio.run(main_aiohttp_optimized()))
