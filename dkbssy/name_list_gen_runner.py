import openpyxl
from dkbssy.name_list_generator.name_list_gen import IncentiveNameGenerator, dates_of_incentive_cases, \
    retrieving_all_incentive_case_data, retrieving_all_incentive_names_data, \
    amount_distribution_of_all_case_number_of_sheet, update_incentive2_from_sql
from dkbssy.name_list_generator.sql_for_temp_db import SqlForTemp
from dkbssy.utils.colour_prints import ColourPrint


##################################################################################################################

INCENTIVE_NAME_ADDING_WORKBOOK_NAME = r"G:\My Drive\GdrivePC\Hospital\RSBY\New\Upload Formats\Med\Med 5\Med 5 1.xlsx"

required_cat_num = [1, 2, 3, 4, 5,  7, 8, 9]  # Anaesthesia removed if required add '6' with comma ',' sign


##################################################################################################################



incen_obj = IncentiveNameGenerator(r"G:\My Drive\GdrivePC\Hospital\RSBY\New\Incentive_auto_ver_3.xlsx")
# all_sheet_data = incen_obj.get_all_sheet_data('Sheet3')

depart_name_for_generated = incen_obj.depart_name_for_entry('Sheet1', 'B5')
# ColourPrint.print_yellow('DEPARTMENT NAME ===============>')
# print(depart_name_for_generated)

ColourPrint.print_pink(f'Confirm the Department of Entry in Excel B5. The department name is {depart_name_for_generated}.')
input("Press Enter to continue")
ColourPrint.print_pink(f"Confirm all REQUIRED CATEGORY are present. Categories present are: {required_cat_num}")
# print(required_cat_num)
input("Press Enter to continue")

"getting the unwanted post which can be decreased in some incentive cases -> post names must be same as in post column of Excel ver3 sheet3"
unwanted_post_list_collection = ['Counsellor', 'Psychologist', 'Health Assistant', 'Accountant']
unwanted_post_list = [] # for storing the preferences
for unwanted in unwanted_post_list_collection:
    ColourPrint.print_pink(f'Do you want to add the names under the {unwanted} category. Press "Y" or "y" for YES OR press "N" or "n" for NO.')
    response = input("Type Here: ").strip().lower()
    if response == "y":
        continue
    elif response == "n":
        unwanted_post_list.append(unwanted)
    else:
        raise KeyError('Press "Y" or "y" for YES OR press "N" for NO.')

ColourPrint.print_yellow("Unwanted", unwanted_post_list)
# input('lllllllllllllllllllllllllllllllllll')

# GETTING THE DATES OF INCENTIVES IN DEPARTMENT WISE EXCEL
incentive_date_data = dates_of_incentive_cases(workbook_name=INCENTIVE_NAME_ADDING_WORKBOOK_NAME, sheet_name='Sheet4')
# incentive_date_data = incentive_date_data[0]  # list unpacking as it has tuple of dates
# print(incentive_date_data)

number_of_inventive_cases_count = len(incentive_date_data)


all_category_names = []


all_date_data = []
for date in incentive_date_data:
    each_data_data = [date]
    ColourPrint.print_bg_red(str(date))
    for cat_num in required_cat_num:
        collect = incen_obj.collect_names_for_single_date(
            'Sheet3',
            'Sheet2',
            category_number=cat_num,
            depart_name_for_gen=depart_name_for_generated,
            sort_by_amount=True,
            inc_date=date,
            number_of_names=50,
            unwanted_post_list=unwanted_post_list)
        # print('---------------->', collect)
        each_data_data.append(collect)
    all_date_data.append(each_data_data)
    # ColourPrint.print_yellow('\nEach Date Data = ')
    # print('Each Date Data = ', each_data_data)

print(len(all_date_data))


# print(all_date_data)

def percent_wise_names(all_date_data) -> dict:
    """
    Getting the count of names for whole set of incentive cases in each Excel Sheet -> in dict form
    :param all_date_data: All names for all date in list of list format
    :return: names and counting dict
    """
    master_name_repeated_count_dict = {}
    for each_date_data in all_date_data:
        for each_category_names in each_date_data[1:]:
            for name in each_category_names:
                if type(name) is str:
                    if name not in master_name_repeated_count_dict:
                        master_name_repeated_count_dict[name] = 1
                    else:
                        master_name_repeated_count_dict[name] += 1

    return master_name_repeated_count_dict


# getting the names and the counts of same names = Unique Names
unique_names = percent_wise_names(all_date_data)
# print(unique_names)
print('Total Names: ', len(unique_names))

# sorted according the count in dict
sorted_names_count = sorted(unique_names.items(), key=lambda x: x[1])
print(sorted_names_count)

"""
gives all names generated raw"
# for name in sorted_names_count:
#     
#     print(name)
#     '''('Dr. Durga Shankar Patel', 38)'''
#     # print(name,':', count)
# print(unique_names)
"""


# percent_filter_name_list =[]

def percent_filtered_name_more_than_90_percent(sorted_names_count, number_of_inventive_cases_count, percent_filter_name_list=None):
    """
    This gives the LIMITS of names count to be at-least more than 90% in all the incentive cases of Excel sheet
    :param sorted_names_count: names and count dict in sorted manner
    :param number_of_inventive_cases_count: total incentive cases in Excel sheet
    :param percent_filter_name_list: temp list for collection and returning
    :return:list of names whose percentage score is more than 90%
    """
    if percent_filter_name_list is None:
        percent_filter_name_list = []
    cccc = 0
    for name_and_count in sorted_names_count:
        if name_and_count[1] / number_of_inventive_cases_count > 0.90:
            cccc += 1
            # print(cccc)
            # print(name_and_count)
            percent_filter_name_list.append(name_and_count)
    return percent_filter_name_list


percent_filtered_name_list_more_than_90_percent = percent_filtered_name_more_than_90_percent(sorted_names_count, number_of_inventive_cases_count)
print(percent_filtered_name_list_more_than_90_percent)
ColourPrint.print_pink('==========================')
print(f'{len(percent_filtered_name_list_more_than_90_percent)=}')
ColourPrint.print_pink('--------==========================--------')


def final_formated_names_for_excel_as_per_excel_headers(percent_filtered_name_list_more_than_the_90):
    """
    This is used to get data matching according to the Excel sheet headers
    :param percent_filtered_name_list_more_than_the_90: filtered names having the percentage more than 90 percent eligible
    in total incentive cases
    :return: formatted name for pasting in Excel
    """
    # ColourPrint.print_yellow('================================')
    all_sheet_data = incen_obj.get_all_sheet_data('Sheet3')
    '''Each_employee_data -> [('Column1', 'Employee Name', 'Employee Code', 'Category', 'Post Name', 'Posting', 'Depart', 'Cases', 'Pend Cases', 'Pend Amount', 'Paid Cases', 'Paid Amount', 'Web Tot cases', 'Web Tot Amount', 'DB1Tot Case entry', 'DB1Tot amount entry', 'DB2Tot Case 2', 'DB2Tot Amount 2', None, ' ', None, None, None, None, None, None), 
        (3, 'Dr. Durga Shankar Patel', 66170010344, '6,1', 'Professor', 'ANAES', 'ANAES', datetime.datetime(2022, 3, 3, 0, 0), 8154, 403446, 1750, 71397, 9904, 474843, 8229, 403718.3462428572, 7793, 399197.5933333333, None, None, None, None, None, None, None, None), ...]'''
    # print(all_sheet_data)

    # depart_wise_excel = openpyxl.load_workbook(incentive_sheet_path)
    # sheet_for_updating_name = depart_wise_excel['Sheet2']

    name_and_categ_collection = []
    for inc_name, inc_count in percent_filtered_name_list_more_than_the_90:
        if inc_name == "JYOTI GHRITLAHRE":
            continue
        for each_employee_data in all_sheet_data:
            master_employee_name = each_employee_data[1]
            master_employee_code = each_employee_data[2]
            master_sheet_category = each_employee_data[3]
            master_sheet_post = each_employee_data[4]

            if inc_name == master_employee_name:
                # ColourPrint.print_yellow(inc_name, '==>><<==', master_employee_name)
                # print(f'{type(emp_category)=}')
                if type(master_sheet_category) == str:
                    master_sheet_category = select_category_from_multiple_category(depart_name_for_generated,
                                                                                   each_employee_data)

                hindi_cat_name = hindi_category_name(master_sheet_category)
                name_and_categ_collection.append(
                    (master_employee_name, master_sheet_post, hindi_cat_name, master_employee_code,
                     master_sheet_category))

    return name_and_categ_collection


def select_category_from_multiple_category(depart_name_for_generated, each_employee_data):
    """
    Selection of the names of consultants from different categories according tho the current entry department
    :param depart_name_for_generated: current department for which entry to done
    :param each_employee_data: full row data of employee as same as row of incentive_ver_3
    :return: category number of the consultant for the current entry department
    """
    # all depart = ['ENT', 'EYE', 'MED', 'OBGY', 'ORTH', 'PEDIA', 'RADTHRY', 'RESP', 'SKIN', 'SURG']
    surgical_depart = ['ENT', 'OBGY', 'ORTH', 'SURG']

    name = each_employee_data[1]
    category = each_employee_data[3]

    if name == 'Dr. Durga Shankar Patel':
        if depart_name_for_generated in surgical_depart:
            category = 6
        elif 6 in required_cat_num:  # for medicine depart if willing
            category = 6
        else:
            category = 1

    if name == "Dr. Rakesh Kumar Verma":
        if depart_name_for_generated == 'PEDIA':
            category = 4
        elif depart_name_for_generated == 'OBGY':
            category = 5
        else:
            category = 1

    if name == 'Dr. Gopal Singh Kanwer':
        if depart_name_for_generated == 'MED':
            category = 4
        else:
            category = 1

    if name == 'Dr. Aditya Siodiya':
        if depart_name_for_generated == 'OBGY':
            category = 4
        else:
            category = 1
    return category


def hindi_category_name(category_number):
    """
    Used for filling the Excel post name for the final Excel Sheet using the category number
    :param category_number:
    :return: Hindi category names
    """
    hindi_name = "No Hindi Name"
    if category_number == 1:
        hindi_name = 'vfèk"Bkrk vLirky vèkh{kd ]lgk;d vèkh{kd uksMy vfèkdkjh'
    if category_number == 2:
        hindi_name = 'iSFkksy‚th ]jsfM;ksy‚th ] ekbØksck;ksy‚th ] ck;ksdsesLVªh ] ¼ QSdYVh ,aM jslhMsaV ½'
    if category_number == 3:
        hindi_name = "iSFkksy‚th ]jsfM;ksy‚th ] ekbØksck;ksy‚th ] ck;ksdsesLVªh ] fu'psruk ¼VsfDuf'k;u ½"
    if category_number == 4:
        hindi_name = "lHkh fQftf'k;u @ ltZu"
    if category_number == 5:
        hindi_name = "lHkh lhfu;j ,oa twfu;j jsflMsaV"
    if category_number == 6:
        hindi_name = "fu'psruk"
    if category_number == 7:
        hindi_name = "uflZax ,oa iSjkesfMdy LVkQ"
    if category_number == 8:
        hindi_name = "prqFkZ oxZ ,oa lQkbZ deZpkjh"
    if category_number == 9:
        hindi_name = "MkVk ,aVªh v‚ijsVj"
    return hindi_name


ColourPrint.print_green('----------------------------------------------------')
final_names_more_than_90_percent_formatted_for_excel = final_formated_names_for_excel_as_per_excel_headers(percent_filtered_name_list_more_than_90_percent)
final_names_more_than_90_percent_formatted_for_excel = sorted(final_names_more_than_90_percent_formatted_for_excel, key=lambda x: x[4])
print(final_names_more_than_90_percent_formatted_for_excel)
ColourPrint.print_green('----------------------------------------------------')


# for i in final_names:
#     print(i)


def uploading_names_in_excel_sheet(final_names_90_formatted_for_excel, workbook_path, worksheet_name):
    """
    Used for directly filling the final Excel sheet
    :param final_names_90_formatted_for_excel: final names whose names are more than 90 percent eligible in total incentive cases
    :param workbook_path: Excel workbook path of department wise Excel e.g. eye 2 1.xlsx
    :param worksheet_name:Sheet name of the above workbook. Usually sheet 2
    :return: None
    """

    # Load the workbook and worksheet
    wb = openpyxl.load_workbook(workbook_path)
    ws = wb[worksheet_name]

    # Start writing at row 2, column 2 (cell B2)
    start_row = 2
    start_col = 2

    # Clear all data except the first row (headers)
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=ws.max_column):
        for cell in row:
            cell.value = None

    # Loop through each row of data
    for row_offset, row_values in enumerate(final_names_90_formatted_for_excel):
        for col_offset, value in enumerate(row_values):
            ws.cell(row=start_row + row_offset, column=start_col + col_offset).value = value

    # Save the updated workbook
    wb.save(workbook_path)
    wb.close()


uploading_names_in_excel_sheet(final_names_more_than_90_percent_formatted_for_excel, INCENTIVE_NAME_ADDING_WORKBOOK_NAME, 'Sheet2')
print("Updated: ", INCENTIVE_NAME_ADDING_WORKBOOK_NAME)
# ColourPrint.print_pink('DELETE THE DUPLICATE NAMES IF PRESENT LIKE JYOTI GHRITLAHRE. THAN SAVE AND CLOSE EXCEL. THAN PRESS ENTER')
# input()

""" now retrieving the amount gaining by the members """

# retrieving_all_incentive_data = retrieving_all_incentive_case_data(INCENTIVE_NAME_ADDING_WORKBOOK_NAME)
ret_case_num_data = retrieving_all_incentive_case_data(INCENTIVE_NAME_ADDING_WORKBOOK_NAME)
ret_name_data = retrieving_all_incentive_names_data(INCENTIVE_NAME_ADDING_WORKBOOK_NAME)
# print(amount_distribution_of_all_case_number_of_sheet(ret_case_num_data, ret_name_data))
data_for_db_amount_of_all_case_number_and_employee_wise = amount_distribution_of_all_case_number_of_sheet(ret_case_num_data, ret_name_data)


temp_db = r"G:\My Drive\GdrivePC\Hospital\RSBY\New\down\temp_incen.db"

sql_temp_object = SqlForTemp(temp_db_path=temp_db)
sql_temp_object.insert_all_case_num_data(data_for_db_amount_of_all_case_number_and_employee_wise)
emp_id_and_amount_list = sql_temp_object.fetch_updated_amount()
print(emp_id_and_amount_list)

sql_temp_object.close()

'updating the incentive2 Excel'
update_incentive2_from_sql(emp_id_and_total_amount_list=emp_id_and_amount_list)



'counceller, health assistant, Psychologist'