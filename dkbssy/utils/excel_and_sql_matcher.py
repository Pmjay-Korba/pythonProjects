import time
import openpyxl
import sqlite3
import os
from selenium.webdriver.common.by import By
import asyncio
from dkbssy.utils.colour_prints import ColourPrint
from dkbssy.utils.file_renamer import get_default_download_dir
from bs4 import BeautifulSoup
from playwright.async_api import async_playwright


class CheckerUpdate:
    download_folder_path = get_default_download_dir()
    work_book_path = os.path.join(download_folder_path, "bb.xlsx")

    def get_all_initiated_data(self):
        return asyncio.run(self._get_scraped_initiated_data())

    async def _get_scraped_initiated_data(self):
        play = await async_playwright().start()
        browser = await play.chromium.connect_over_cdp('http://localhost:9222')
        context = browser.contexts[0]
        # page = await context.new_page()


        response = await context.request.get('https://dkbssy.cg.nic.in/secure/incentivemodule/incentivemoduleInitiatedcasesdme.aspx')
        # print(await response.text())

        # returning all text
        response_text = await response.text()

        # await page.close()
        await play.stop()

        return response_text

    def processed_all_inititated_data(self):  # return list case_text, status_text, datetime_text
        # getting unprocessed data
        html = self.get_all_initiated_data()

        soup = BeautifulSoup(html, "html.parser")

        rows = soup.find_all("tr")

        capture_all_data = []
        for row in rows:
            columns = row.find_all("td")
            if len(columns) >= 4:
                # Extract case string
                case_tag = columns[1].find("a")
                case_text = case_tag.get_text(strip=True) if case_tag else None

                # Get the status and datetime
                status_text = columns[2].get_text(strip=True)
                datetime_text = columns[3].get_text(strip=True)

                # print(case_text, status_text, datetime_text)
                capture_all_data.append([case_text, status_text, datetime_text])
                # print(capture_all_data)
        return capture_all_data


    def excel_checker_old(self):  #  This can be replaced by request module
        excel_file = openpyxl.load_workbook(self.work_book_path)
        sheet = excel_file['Sheet1']
        case_number_and_date = []
        column_values = sheet.iter_rows(min_col=2, max_col=4, min_row=3)
        for row in column_values:
            row_data = [cell.value for cell in row]
            case_number_and_date.append(row_data[0])
        ColourPrint.print_turquoise(case_number_and_date)
        return case_number_and_date  # return case numbers only

    def excel_checker(self):
        """
        Returning list of case numbers
        :return: list of case numbers
        """
        all_3_data= self.processed_all_inititated_data()  # case_text, status_text, datetime_text
        case_number_data_only = [d[0] for d in all_3_data]
        # print('only cn', case_number_data_only)
        return case_number_data_only

    def sqlite_check_1(self):
        # Connect to the database (replace 'your_database.db' with your database file)
        conn = sqlite3.connect(r'G:\My Drive\GdrivePC\Hospital\RSBY\New\incentiveDatabase.db')
        cursor = conn.cursor()
        # Execute the SQL query
        cursor.execute("SELECT DISTINCT incentive_case_number FROM incentive_entry_T")
        # Fetch all results
        results = cursor.fetchall()
        # Convert the results to a list of lists
        # results_list = [[row[0], row[1]] for row in results]
        # Close the connection
        conn.close()
        # Print the results list
        results_list_n = [row[0] for row in results]
        # print(results_list_n)
        return results_list_n

    def sqlite_check_2(self):
        # Connect to the database (replace 'your_database.db' with your database file)
        conn = sqlite3.connect(r'G:\My Drive\GdrivePC\Hospital\RSBY\New\incentiveDatabase_2.db')
        cursor = conn.cursor()
        # Execute the SQL query
        cursor.execute("SELECT DISTINCT case_number FROM case_num_t")
        # Fetch all results
        results = cursor.fetchall()
        # Convert the results to a list of lists
        # results_list = [[row[0], row[1]] for row in results]
        # Close the connection
        conn.close()
        # Print the results list
        results_list_n = [row[0] for row in results]
        # print(results_list_n)
        return results_list_n

    def check_proper_to_update_sql(self):
        excel_list = self.excel_checker_complete()
        # print('=========================================')
        sql_list = self.sqlite_check_2_complete()
        # print('...........................................')
        to_be_updated = []
        for ee in excel_list:  # exc ['CASE/PS5/HOSP22G146659/CB5767565', 'ACO Approved', '30/05/2024 05:08:51']
            if ee[1] != 'Payment Done':
                # ColourPrint.print_bg_red('Payment Done: ', ee[0])
                for ss in sql_list:  # sql [('CASE/PS6/HOSP22G146659/R6206528', '27/05/2024 05:19:06')]
                    if ee[0] == ss[0]:
                        if ee[2].strip() != ss[1].strip():
                            # print('////////////////////////////////////')
                            # print(1, ee[0], 2, ss[0], 3, repr(ee[2]), 4, repr(ss[1]))
                            # print('000000000000000000000000000000000000000000')
                            to_be_updated.append(ee[0])
        print('To be Updated list:', to_be_updated)
        # print('>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>')
        return to_be_updated

    def rename_file(self, original_name, new_name, directory):
        original_file = os.path.join(directory, original_name)
        new_file = os.path.join(directory, new_name)

        if os.path.exists(new_file):
            os.remove(new_file)  # Remove the existing file if it exists

        while not os.path.exists(original_file):
            time.sleep(10)  # Wait until the file is downloaded

        os.rename(original_file, new_file)

    def download_file(self, driver, download_dir):
        # Open the webpage
        driver.get("https://dkbssy.cg.nic.in/secure/incentivemodule/incentivemoduleInitiatedcasesdme.aspx")

        # Find the download link and click it
        download_link = driver.find_element(By.XPATH,
                                            '//*[@id="ctl00_ContentPlaceHolder1_GridView1_wrapper"]/div[1]/button[3]')  # Change this to the correct XPath
        download_link.click()
        time.sleep(10)
        # Rename the file after downloading
        self.rename_file("Shaheed Veer Narayan Singh Ayushman Swasthya Yojna.xlsx", "bb.xlsx", download_dir)

    def find_difference_SE(self):
        sql_list = set(self.sqlite_check_2())
        exc_list = set(self.excel_checker())
        diff_SE = sql_list - exc_list
        print('S-E', len(sql_list), len(exc_list), len(diff_SE))
        # print(list(diff_SE))
        # print('E-S', exc_list-sql_list)
        return list(diff_SE)

    def find_ES_difference(self):
        sql_list = set(self.sqlite_check_2())
        exc_list = set(self.excel_checker())
        diff_ES = exc_list - sql_list
        # print('E-S', len(exc_list), len(sql_list), len(diff_ES))
        # print(list(diff_ES))
        # print('E-S', exc_list-sql_lis/t)
        return list(diff_ES)

    def sql_delete_records(self, case_number):
        # Connect to the database (replace 'your_database.db' with your database file)
        conn = sqlite3.connect(r'G:\My Drive\GdrivePC\Hospital\RSBY\New\incentiveDatabase.db')
        cursor = conn.cursor()
        # SQL query to delete records with the specified case number
        sql_query = "DELETE FROM incentive_entry_T WHERE incentive_case_number = ?"
        # Execute the query
        cursor.execute(sql_query, (case_number,))

        # Commit the transaction
        conn.commit()

    def sql_delete_records_2(self, case_number):
        # Connect to the database
        conn = sqlite3.connect(r'G:\My Drive\GdrivePC\Hospital\RSBY\New\incentiveDatabase_2.db')
        cursor = conn.cursor()

        try:
            # Begin a transaction
            conn.execute('BEGIN TRANSACTION')

            # Find the id_case_num for the given case_number
            cursor.execute("SELECT id_case_num FROM case_num_t WHERE case_number = ?", (case_number,))
            id_case_num = cursor.fetchone()

            if id_case_num:
                id_case_num = id_case_num[0]

                # SQL query to delete records from distribution_t with the specified id_case_num
                sql_delete_distribution = "DELETE FROM distribution_t WHERE d_case_num = ?"
                cursor.execute(sql_delete_distribution, (id_case_num,))

                # SQL query to delete records from case_num_t with the specified case_number
                sql_delete_case_num = "DELETE FROM case_num_t WHERE case_number = ?"
                cursor.execute(sql_delete_case_num, (case_number,))

            # Commit the transaction
            conn.commit()
        except Exception as e:
            # Rollback the transaction in case of error
            conn.rollback()
            print(f"Error occurred: {e}")
        finally:
            # Close the connection
            conn.close()

    def sql_delete_cycle(self):
        to_delete_list = self.find_difference_SE()
        print('TO DELETE LIST:', to_delete_list)
        for i in to_delete_list:
            print('Deleting', i)
            self.sql_delete_records(i)

    def sql_delete_cycle_2(self):
        to_delete_list = self.find_difference_SE()
        print('TO DELETE LIST2:', to_delete_list)
        for i in to_delete_list:
            print('Deleting', i)
            self.sql_delete_records_2(i)

    def update_record(self) -> list:
        list_to_add = self.find_ES_difference()
        return list_to_add

    def excel_checker_complete_old(self):  # This can be replaced by request module
        excel_file = openpyxl.load_workbook(self.work_book_path)
        sheet = excel_file['Sheet1']
        case_number_and_date = []
        column_values = sheet.iter_rows(min_col=2, max_col=4, min_row=3)
        for row in column_values:
            row_data = [cell.value for cell in row]
            case_number_and_date.append(row_data)
        # print(case_number_and_date)
        return case_number_and_date  # RETURN CASE_NUMBER, STATUS, TIME
        # ['CASE/PS5/HOSP22G146659/CB5767565', 'ACO Approved', '30/05/2024 05:08:51']

    def excel_checker_complete(self):
        return self.processed_all_inititated_data()

    def sqlite_check_2_complete(self):
        # Connect to the database (replace 'your_database.db' with your database file)
        conn = sqlite3.connect(r'G:\My Drive\GdrivePC\Hospital\RSBY\New\incentiveDatabase_2.db')
        cursor = conn.cursor()
        # Execute the SQL query
        cursor.execute("SELECT case_number, updated_time FROM case_num_t")
        # Fetch all results
        results = cursor.fetchall()
        # Close the connection
        conn.close()
        # print(results)
        return results  # RETURN CASE_NUMBER, TIME [('CASE/PS6/HOSP22G146659/R6206528', '27/05/2024 05:19:06')]

    def add_payment_done_directly_inc_db_2(self, payment_done_data):
        """
        Add the payment done data to sql db_2 so to avoid unnecessary waiting
        :param payment_done_data: Excel download data
        :return: None
        """
        '''payment_done_data 
                    header NOT included in data  Case No	        Last Updated Date	Done Payment	Case Status
                    actual data CASE/PS6/HOSP22G146659/CK6147485	28/05/2024 07:58:19	Yes	            Payment Done
                                CASE/PS6/HOSP22G146659/CK6152981	28/05/2024 07:58:19	Yes	            Payment Done
                                CASE/PS5/HOSP22G146659/CK5907965	27/05/2024 05:13:30	Yes	            Payment Done
        Done payment is sekf created for the column match data

        '''
        # Split rows and columns
        rows = [line.split('\t') for line in payment_done_data.strip().split('\n')]

        # Connect to your SQLite DB
        conn = sqlite3.connect(r"G:\My Drive\GdrivePC\Hospital\RSBY\New\incentiveDatabase_2.db")
        cur = conn.cursor()

        # Insert data
        for case_number, updated_time, inc_procedure, inc_status in rows:
            cur.execute("""
                INSERT INTO case_num_t (case_number, updated_time, inc_procedure, inc_status)
                VALUES (?, ?, ?, ?)
            """, (case_number, updated_time, inc_procedure, inc_status))

        # Commit and close
        conn.commit()
        conn.close()



if __name__ == '__main__':

    cu = CheckerUpdate()
    #     cu.check_proper_to_update_sql()
    # print(cu.excel_checker_complete())
    # print(cu.sqlite_check_2_complete())
    # cu.excel_checker_old()
    # print("==================")
    # cu.excel_checker()
    # print('--------------------')
    cu.excel_checker_complete_old()
    print("+++++++++++++++++++")
    print(cu.excel_checker_complete())
    # cu.sqlite_check()
    # cu.find_difference_SE()
    # input('Proceed to delete')
    # input("Do you want to proceed")
    # cu.sql_delete_cycle()
