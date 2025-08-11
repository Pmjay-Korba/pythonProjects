import json
import math
import time
import os
import csv
import time
from datetime import datetime, timedelta
# user_id = 'CHH009264'
import gspread
import gspread_dataframe
import openpyxl
import pandas as pd
from gspread.utils import ValueInputOption
from playwright.sync_api import sync_playwright, TimeoutError, Page

from dkbssy.utils.colour_prints import ColourPrint
from dkbssy.utils.file_renamer import rename_file
from old_dkbssy_folder import tms_department_wise_2
from tms_playwright.discharge_to_be_done.discharge_details import DischargeGetParameters
from tms_playwright.page_objs_tms import tms_xpaths


class AllListGenerator:
    # Define the paths to Chrome executable and user data directory
    chrome_path = r"C:\Program Files (x86)\Google\Chrome\Application\chrome.exe"
    user_data_dir = r'C:\Users\RAKESH\AppData\Local\Google\Chrome\User Data'

    # Define the Google Sheets URL and download directory
    sheet_url = "https://docs.google.com/spreadsheets/d/19HHTQZe9_8hMQJDZM4aZ01RcBqVH1-xXVv3RkR2W1ls/edit?gid=0"
    download_dir = r"G:\My Drive\GdrivePC\Hospital\RSBY\New\down"

    def pre_auth_list_generate(self, page: Page):
        page.wait_for_load_state()
        try:  # pre-auth is already opened
            page.wait_for_selector(tms_xpaths.list_discharge_left_side_tab_ACTIVE_xp, timeout=3000)
            pass
        except TimeoutError:
            pre_auth_query = page.wait_for_selector(tms_xpaths.pre_auth_query_left_side_tab_xp)
            pre_auth_query.click()

        pre_auth_query_child_menu = page.wait_for_selector(tms_xpaths.pre_auth_query_left_side_tab_sub_tab_xp)
        pre_auth_query_child_menu.click()

        total_pending_query = page.wait_for_selector(tms_xpaths.pre_auth_query_count_xp)
        total_pending_query = total_pending_query.inner_text()
        pages_count = math.ceil(int(total_pending_query) / 10)
        print(total_pending_query, pages_count)

        middle_frame_element = page.wait_for_selector(tms_xpaths.middle_frame_xp)
        middle_frame_content = middle_frame_element.content_frame()

        count_individual_cases = 1
        to_save_list_pre = []
        for sheet in range(1, pages_count + 1):
            for numx in range(1, 11):
                if count_individual_cases <= int(total_pending_query):
                    middle_frame_content.wait_for_selector(tms_xpaths.serial_number_waiting(count_individual_cases))
                    idx = middle_frame_content.wait_for_selector(tms_xpaths.case_serial_num_xp(numx))
                    idx_text = idx.inner_text()
                    # print(count_individual_cases, idx_text)
                    print(idx_text)
                    count_individual_cases += 1
                    to_save_list_pre.append(idx_text)
                else:
                    break
            if sheet == 9:
                print('pq1')
                try:
                    next_sheet = middle_frame_content.wait_for_selector(tms_xpaths.list_discharge_next_xp)
                    next_sheet.click()
                    waiting = middle_frame_content.wait_for_selector(
                        tms_xpaths.list_discharge_next_sheet_waiting_xp(sheet))
                    waiting.click()
                except TimeoutError:
                    break
            elif (sheet + 1) % 10 == 0:
                print('pq2')
                try:
                    next_sheet = middle_frame_content.wait_for_selector(tms_xpaths.list_discharge_next_xp)
                    next_sheet.click()
                    waiting = middle_frame_content.wait_for_selector(
                        tms_xpaths.list_discharge_next_sheet_waiting_xp(sheet))
                    waiting.click()
                except TimeoutError:
                    break
            else:
                print('pq3')
                if sheet < pages_count:
                    sheet += 1
                    adjacent_sheet = middle_frame_content.wait_for_selector(
                        tms_xpaths.list_discharge_adjacent_page_xp(sheet))
                    adjacent_sheet.click()
                    sheet -= 1
        return to_save_list_pre

    def discharge_list_generate(self, page: Page):
        print('d1')
        page.wait_for_load_state()
        try:  # pre-auth is already opened
            page.wait_for_selector(tms_xpaths.list_discharge_left_side_tab_ACTIVE_xp, timeout=3000)
            pass
        except TimeoutError:
            discharge_tab = page.wait_for_selector(tms_xpaths.list_discharge_left_side_tab_xp)
            discharge_tab.click()

        print('d2')
        discharge_tab_button = page.wait_for_selector(tms_xpaths.list_discharge_left_side_tab_sub_tab_xp)
        discharge_tab_button.click()
        print('d3')
        total_pending = page.wait_for_selector(tms_xpaths.list_discharge_count_xp).inner_text()
        pages_count = math.ceil(int(total_pending) / 10)
        print('Discharge: ', total_pending, pages_count)

        middle_frame_element = page.wait_for_selector(tms_xpaths.middle_frame_xp)
        middle_frame_content = middle_frame_element.content_frame()

        to_save_list_dis = []
        count_individual_cases = 1
        for sheet in range(1, pages_count + 1):
            for numx in range(1, 11):
                # print('count', count,numx)
                if count_individual_cases <= int(total_pending):
                    middle_frame_content.wait_for_selector(tms_xpaths.serial_number_waiting(count_individual_cases))
                    idx = middle_frame_content.wait_for_selector(tms_xpaths.case_serial_num_xp(numx))
                    idx_text = idx.inner_text()
                    # print(count_individual_cases, idx_text)
                    print(idx_text)
                    count_individual_cases += 1
                    to_save_list_dis.append(idx_text)
                else:
                    break
            if sheet == 9:
                print('dd1')
                try:
                    next_sheet = middle_frame_content.wait_for_selector(tms_xpaths.list_discharge_next_xp)
                    next_sheet.click()
                    waiting = middle_frame_content.wait_for_selector(
                        tms_xpaths.list_discharge_next_sheet_waiting_xp(sheet))
                    waiting.click()
                except TimeoutError:
                    break
            elif (sheet + 1) % 10 == 0:
                print('dd2')
                try:
                    next_sheet = middle_frame_content.wait_for_selector(tms_xpaths.list_discharge_next_xp)
                    next_sheet.click()
                    waiting = middle_frame_content.wait_for_selector(
                        tms_xpaths.list_discharge_next_sheet_waiting_xp(sheet))
                    waiting.click()
                except TimeoutError:
                    break
            else:
                print('dd3')
                if sheet < pages_count:
                    sheet += 1
                    adjacent_sheet = middle_frame_content.wait_for_selector(
                        tms_xpaths.list_discharge_adjacent_page_xp(sheet))
                    adjacent_sheet.click()
                    sheet -= 1
        return to_save_list_dis

    def claim_query_list_generate(self, page: Page):
        page.wait_for_load_state()
        claim_menu = page.wait_for_selector(tms_xpaths.claim_list_gen_menu_xp)
        claim_menu.click()
        claim_child_menu = page.wait_for_selector(tms_xpaths.claim_list_gen_child_menu_xp)
        claim_child_menu.click()

        total_pending = page.wait_for_selector(tms_xpaths.claim_list_pending_count_xp)
        total_pending = total_pending.inner_text()
        pages_count = math.ceil(int(total_pending) / 10)
        print("Claim Query", total_pending, pages_count)

        middle_frame_element = page.wait_for_selector(tms_xpaths.middle_frame_xp)
        middle_frame_content = middle_frame_element.content_frame()

        to_save_list_claim = []
        count_individual_cases = 1
        for sheet in range(1, pages_count + 1):
            for numx in range(1, 11):
                # print('count', count,numx)
                if count_individual_cases <= int(total_pending):
                    middle_frame_content.wait_for_selector(tms_xpaths.serial_number_waiting(count_individual_cases))
                    idx = middle_frame_content.wait_for_selector(
                        tms_xpaths.case_serial_num_for_claim_query_gen_xp(numx))
                    idx_text = idx.inner_text()
                    # print(count_individual_cases, idx_text)
                    print(idx_text)
                    count_individual_cases += 1
                    to_save_list_claim.append(idx_text)
                else:
                    break
            if sheet == 9:
                print('cq1')
                try:
                    next_sheet = middle_frame_content.wait_for_selector(tms_xpaths.list_discharge_next_xp)
                    next_sheet.click()
                    waiting = middle_frame_content.wait_for_selector(
                        tms_xpaths.list_discharge_next_sheet_waiting_xp(sheet))
                    waiting.click()
                except TimeoutError:
                    break
            elif (sheet + 1) % 10 == 0:
                print('cq2')
                try:
                    next_sheet = middle_frame_content.wait_for_selector(tms_xpaths.list_discharge_next_xp)
                    next_sheet.click()
                    waiting = middle_frame_content.wait_for_selector(
                        tms_xpaths.list_discharge_next_sheet_waiting_xp(sheet))
                    waiting.click()
                except TimeoutError:
                    break
            else:
                print('cq3')
                if sheet < pages_count:
                    sheet += 1
                    adjacent_sheet = middle_frame_content.wait_for_selector(
                        tms_xpaths.list_discharge_adjacent_page_xp(sheet))
                    adjacent_sheet.click()
                    sheet -= 1
        return to_save_list_claim

    def spreadsheet_download_and_rename(self, page: Page, sheet_url, down_filename_with_extension,  save_filename_with_extension):

        # Navigate to the Google Sheets URL
        page.goto(sheet_url)
        # Wait for the page to load
        time.sleep(.5)
        # Click on "File" -> "Download" -> "Microsoft Excel (.xlsx)"
        page.click("text=File")
        time.sleep(0.05)  # Wait for the dropdown to appear
        page.click("text=Download")
        time.sleep(0.05)
        page.click("text=Microsoft Excel (.xlsx)")
        # Expect the download to start and capture the download event
        with page.expect_download() as download_info:
            pass  # The download should automatically start after clicking

        # Get the download object
        download = download_info.value

        # Save the downloaded file to the specified location
        download_path = self.download_dir + "\\" + download.suggested_filename
        download.save_as(download_path)
        # Renaming the file
        original_name = down_filename_with_extension
        new_name = save_filename_with_extension
        directory = self.download_dir
        rename_file(original_name, new_name, directory)

    def manipulating_downloaded_excel(self, both_query_playwright_list: list, sheet_name_to_be_compared: str,
                                      is_query_question_required: bool, page: Page):
        """is_query_question_required: gives the details which a query required to be cleared"""

        file_path = r"G:\My Drive\GdrivePC\Hospital\RSBY\New\down\cc.xlsx"
        workbook = openpyxl.load_workbook(filename=file_path, data_only=True)

        sheet_name = workbook[sheet_name_to_be_compared]
        case_numbers_in_cols = list(
            [row for row in sheet_name.iter_cols(min_col=3, max_col=3, min_row=2, values_only=True)][0])
        case_numbers_in_cols = [i for i in case_numbers_in_cols if i is not None]
        print('cc', len(case_numbers_in_cols), case_numbers_in_cols)
        print('bb', len(both_query_playwright_list), both_query_playwright_list)
        to_delete_case_numbers = set(case_numbers_in_cols) - set(both_query_playwright_list)
        print('to delete', len(to_delete_case_numbers), to_delete_case_numbers)
        to_add_case_numbers = set(both_query_playwright_list) - set(case_numbers_in_cols)
        print('to_add_list', len(to_add_case_numbers), to_add_case_numbers)

        # Iterate through the rows in reverse to avoid issues with deleting rows
        for row in range(sheet_name.max_row, 0, -1):
            cell_value = sheet_name.cell(row=row, column=3).value  # Assuming 'C' is the 3rd column
            if cell_value in to_delete_case_numbers:
                sheet_name.delete_rows(row)

        workbook.save(r"G:\My Drive\GdrivePC\Hospital\RSBY\New\down\cc.xlsx")

        # Step 2: Add new rows
        discharge_details_parameter_obj = DischargeGetParameters()

        master_to_save_in_excel = []
        for c_num in to_add_case_numbers:
            if not is_query_question_required:
                # detail page
                discharge_details_parameter_obj.detail_entry_page(page, c_num)
                case_num_data = discharge_details_parameter_obj.retrieve_data(page, c_num)  # returns tuple
                print(case_num_data)
                master_to_save_in_excel.append(list(case_num_data))
                # discharge_details_parameter_obj.excel_save(data, raw_excel_file_path)
                # sheet_name.append(case_num_data)
                # workbook.save(r"G:\My Drive\GdrivePC\Hospital\RSBY\New\down\cc.xlsx")
            else:
                # detail page
                discharge_details_parameter_obj.detail_entry_page(page, c_num)
                case_num_data = discharge_details_parameter_obj.retrieve_data(page, c_num)  # returns tuple
                # print(data)
                question = discharge_details_parameter_obj.query_question_data(page)
                case_num_data = list(case_num_data)
                case_num_data.append(question)
                print(case_num_data)
                master_to_save_in_excel.append(case_num_data)
                # sheet_name.append(case_num_data)
                # workbook.save(r"G:\My Drive\GdrivePC\Hospital\RSBY\New\down\cc.xlsx")
                case_search_left_panel = page.wait_for_selector(tms_xpaths.claim_query_cases_search_tab_xp)
                case_search_left_panel.click()
        time1 = time.time()
        for case_detail_list in master_to_save_in_excel:
            sheet_name.append(case_detail_list)
        workbook.save(r"G:\My Drive\GdrivePC\Hospital\RSBY\New\down\cc.xlsx")
        # function ran
        tms_department_wise_2.department_wise_extract_for_queries()
        print('Time to save in excel:', time.time() - time1, 'seconds')

    # def upload_in_g_sheet(self, sheet_name):
    #     # excel data getting
    #     excel_workbook_path = r"G:\My Drive\GdrivePC\Hospital\RSBY\New\down\cc.xlsx"
    #     df = pd.read_excel(excel_workbook_path, sheet_name=sheet_name)
    #     df_cleaned = df.dropna(how='all')
    #     df_cleaned = df_cleaned.reset_index(drop=True)
    #     # print(df_cleaned)
    #     df_cleaned = df_cleaned.sort_values(by=['Status'])
    #     df_cleaned = df_cleaned.sort_values(by=['Procedure'])
    #     df_cleaned = df_cleaned.sort_values(by=['Depart'])
    #
    #     # columns = df_cleaned.columns.tolist()
    #     # columns[-1], columns[-2] = columns[-2], columns[-1]  # Swap the last and second-to-last column names
    #     # df_rearranged = df[columns]
    #
    #     # Convert all data to strings
    #     df_cleaned = df_cleaned.astype(str)  # Convert all data to strings
    #
    #     # getting the google sheet
    #     gc = gspread.service_account(
    #         filename=r"G:\My Drive\GdrivePC\Hospital\RSBY\New\down\luminous-girder-433203-m6-aaef1468d984.json")
    #     g_workbook = gc.open_by_key('19HHTQZe9_8hMQJDZM4aZ01RcBqVH1-xXVv3RkR2W1ls')
    #
    #     g_worksheet = g_workbook.worksheet(sheet_name)
    #     # print(g_workbook.worksheets())
    #     g_worksheet.clear()
    #     gspread_dataframe.set_with_dataframe(g_worksheet, df_cleaned)
    #     print(df_cleaned)

    def upload_in_g_sheet(self, sheet_name, excel_workbook_path= r"G:\My Drive\GdrivePC\Hospital\RSBY\New\down\cc.xlsx"):
        # Excel data getting
        # excel_workbook_path = r"G:\My Drive\GdrivePC\Hospital\RSBY\New\down\cc.xlsx"
        df = pd.read_excel(excel_workbook_path, sheet_name=sheet_name,
                           dtype=str)  # Read as string to preserve empty cells

        df_cleaned = df.dropna(how='all').reset_index(drop=True)  # Remove fully empty rows
        df_cleaned = df_cleaned.sort_values(by=['Depart', 'Procedure', 'Status'])  # Sort correctly

        df_cleaned.fillna("", inplace=True)  # Replace NaN with empty strings

        # Getting the Google Sheet
        gc = gspread.service_account(
            filename=r"G:\My Drive\GdrivePC\Hospital\RSBY\New\down\luminous-girder-433203-m6-aaef1468d984.json")
        g_workbook = gc.open_by_key('19HHTQZe9_8hMQJDZM4aZ01RcBqVH1-xXVv3RkR2W1ls')

        g_worksheet = g_workbook.worksheet(sheet_name)
        g_worksheet.clear()
        gspread_dataframe.set_with_dataframe(g_worksheet, df_cleaned)

        '''INSERTING THE FORMULA IN SHEET '''
        # Find the last filled row in the reference column
        data = g_worksheet.col_values(4)  # Column D (4th column)
        max_row = len(data)
        # Create formula and update in column J

        if sheet_name == 'QUERY2':
            # ColourPrint.print_yellow('8888888888888888888888888')
            reference_column = "M"
        else:
            reference_column = "D"
        column_to_apply = "J"

        # Create formula list
        formulas = [[
                        f'=IFERROR(TODAY()-DATE(MID({reference_column}{row},7,4),MID({reference_column}{row},4,2),LEFT({reference_column}{row},2)),"")']
                    for row in range(2, max_row + 1)]

        # Define the range dynamically
        cell_range = f"{column_to_apply}2:{column_to_apply}{max_row}"

        # Batch update with chunking to avoid exceeding quota
        chunk_size = 50  # Google allows ~100 per minute, so update in safe chunks
        for i in range(0, len(formulas), chunk_size):
            chunk_range = f"{column_to_apply}{i + 2}:{column_to_apply}{min(i + chunk_size + 1, max_row)}"
            g_worksheet.update(chunk_range, formulas[i:i + chunk_size], value_input_option=ValueInputOption.user_entered)
            print(f"Updated rows {i + 2} to {min(i + chunk_size + 1, max_row)}")
            time.sleep(1)  # Prevent hitting quota

        print(df_cleaned)

    def upload_in_g_sheet_new(self, sheet_name,excel_workbook_path= r"G:\My Drive\GdrivePC\Hospital\RSBY\New\down\cc.xlsx"):
        # Excel data getting
        # excel_workbook_path = r"G:\My Drive\GdrivePC\Hospital\RSBY\New\down\cc.xlsx"
        df = pd.read_excel(excel_workbook_path, sheet_name=sheet_name,
                           dtype=str)  # Read as string to preserve empty cells

        df_cleaned = df.dropna(how='all').reset_index(drop=True)  # Remove fully empty rows
        df_cleaned = df_cleaned.sort_values(by=['Depart', 'Procedure', 'Status'])  # Sort correctly

        df_cleaned.fillna("", inplace=True)  # Replace NaN with empty strings

        # Getting the Google Sheet
        gc = gspread.service_account(
            filename=r"G:\My Drive\GdrivePC\Hospital\RSBY\New\down\luminous-girder-433203-m6-aaef1468d984.json")
        g_workbook = gc.open_by_key('1vhjV0rcODJ4lGYJBHENMnHFvqHgK25dQRt9SVpr_9N4')

        g_worksheet = g_workbook.worksheet(sheet_name)
        g_worksheet.clear()
        gspread_dataframe.set_with_dataframe(g_worksheet, df_cleaned)

        '''INSERTING THE FORMULA IN SHEET '''
        # Find the last filled row in the reference column
        data = g_worksheet.col_values(4)  # Column D (4th column)
        max_row = len(data)
        # Create formula and update in column J
        if sheet_name == 'QUERY2':
            # ColourPrint.print_yellow('999999999999999999999')
            reference_column = "M"
        else:
            reference_column = "D"
        column_to_apply = "J"

        # Create formula list
        formulas = [[
                        f'=IFERROR(TODAY()-DATE(MID({reference_column}{row},7,4),MID({reference_column}{row},4,2),LEFT({reference_column}{row},2)),"")']
                    for row in range(2, max_row + 1)]

        # Define the range dynamically
        cell_range = f"{column_to_apply}2:{column_to_apply}{max_row}"

        # Batch update with chunking to avoid exceeding quota
        chunk_size = 50  # Google allows ~100 per minute, so update in safe chunks
        for i in range(0, len(formulas), chunk_size):
            chunk_range = f"{column_to_apply}{i + 2}:{column_to_apply}{min(i + chunk_size + 1, max_row)}"
            g_worksheet.update(chunk_range, formulas[i:i + chunk_size], value_input_option=ValueInputOption.user_entered)
            print(f"Updated rows {i + 2} to {min(i + chunk_size + 1, max_row)}")
            time.sleep(1)  # Prevent hitting quota

        # print(df_cleaned)


    def discharge_manipulating_excel(self, discharge_playwright_list: list, sheet_name_to_be_compared: str,
                                     is_query_question_required: bool, page: Page):
        """is_query_question_required: gives the details which a query required to be cleared"""
        file_path = r"G:\My Drive\GdrivePC\Hospital\RSBY\New\down\cc.xlsx"
        workbook = openpyxl.load_workbook(filename=file_path, data_only=True)
        sheet_name = workbook[sheet_name_to_be_compared]
        case_numbers_in_cols = list(
            [row for row in sheet_name.iter_cols(min_col=3, max_col=3, min_row=2, values_only=True)][0])
        case_numbers_in_cols = [i for i in case_numbers_in_cols if i is not None]
        print('cc_dis', len(case_numbers_in_cols), case_numbers_in_cols)
        print('bb_dis', len(discharge_playwright_list), discharge_playwright_list)
        to_delete_case_numbers = set(case_numbers_in_cols) - set(discharge_playwright_list)
        print('to delete dis', len(to_delete_case_numbers), to_delete_case_numbers)
        to_add_case_numbers = set(discharge_playwright_list) - set(case_numbers_in_cols)
        print('to_add_list_dis', len(to_add_case_numbers), to_add_case_numbers)

        # Iterate through the rows in reverse to avoid issues with deleting rows
        for row in range(sheet_name.max_row, 0, -1):
            cell_value = sheet_name.cell(row=row, column=3).value  # Assuming 'C' is the 3rd column
            if cell_value in to_delete_case_numbers:
                sheet_name.delete_rows(row)

        workbook.save(r"G:\My Drive\GdrivePC\Hospital\RSBY\New\down\cc.xlsx")
        # Step 2: Add new rows
        discharge_details_parameter_obj = DischargeGetParameters()

        master_to_save_in_excel = []
        for c_num in to_add_case_numbers:
            if not is_query_question_required:
                # detail page
                discharge_details_parameter_obj.detail_entry_page(page, c_num)
                case_num_data = discharge_details_parameter_obj.retrieve_data(page, c_num)  # returns tuple
                print(case_num_data)
                master_to_save_in_excel.append(list(case_num_data))
                # discharge_details_parameter_obj.excel_save(data, raw_excel_file_path)
                # sheet_name.append(case_num_data)
                # workbook.save(r"G:\My Drive\GdrivePC\Hospital\RSBY\New\down\cc.xlsx")
            else:
                # detail page
                discharge_details_parameter_obj.detail_entry_page(page, c_num)
                case_num_data = discharge_details_parameter_obj.retrieve_data(page, c_num)  # returns tuple
                # print(data)
                question = discharge_details_parameter_obj.query_question_data(page)
                case_num_data = list(case_num_data)
                case_num_data.append(question)
                print(case_num_data)
                master_to_save_in_excel.append(case_num_data)
                # sheet_name.append(case_num_data)
                # workbook.save(r"G:\My Drive\GdrivePC\Hospital\RSBY\New\down\cc.xlsx")
                case_search_left_panel = page.wait_for_selector(tms_xpaths.claim_query_cases_search_tab_xp)
                case_search_left_panel.click()
        time1 = time.time()
        for case_detail_list in master_to_save_in_excel:
            case_detail_list.append('Pending Discharge')
            sheet_name.append(case_detail_list)
        workbook.save(r"G:\My Drive\GdrivePC\Hospital\RSBY\New\down\cc.xlsx")
        # function ran
        tms_department_wise_2.department_wise_extract_for_discharge()
        print('Time to save in excel:', time.time() - time1, 'seconds')

    def main_all_list_getter_tms(self, user_id, is_dis_list=False, is_pre_auth_list=False,
                                 is_claim_list=False):

        chrome_path = r"C:\Program Files (x86)\Google\Chrome\Application\chrome.exe"
        user_data_dir = r'C:\Users\RAKESH\AppData\Local\Google\Chrome\User Data'
        with sync_playwright() as p:
            # browser = p.chromium.launch_persistent_context(headless=False, user_data_dir=user_data_dir,executable_path=chrome_path)
            browser = p.chromium.launch(headless=False, args=['--start-maximized'])
            page = browser.new_page(no_viewport=True)
            # page = browser.new_page()
            page.set_default_timeout(60000)

            '''the below line is to be un-commented'''
            # download spreadsheet
            self.spreadsheet_download_and_rename(page,
                                                'https://docs.google.com/spreadsheets/d/19HHTQZe9_8hMQJDZM4aZ01RcBqVH1-xXVv3RkR2W1ls/edit?gid=42829875#gid=42829875',
                                                'AYUSHMAN REGISTRATION 2023.xlsx',
                                                "cc.xlsx")

            # starting TMS
            DischargeGetParameters().login(user_id, page)
            # print(1)
            # check for older than 2 hours the all lists
            older_than_2_hours = is_file_older_than_2_hours(json_file_path)
            if older_than_2_hours:
                print('Older than 2 hours:', older_than_2_hours)
                if is_pre_auth_list:
                    ColourPrint.print_yellow('Pre Auth Query Start ')
                    pre_auth_list_generated = self.pre_auth_list_generate(page)
                    ColourPrint.print_pink('Pre Auth Query End')

                if is_claim_list:
                    ColourPrint.print_yellow('Claim Query Start ')
                    claim_list_generated = self.claim_query_list_generate(page)
                    ColourPrint.print_pink('Claim Query End')
                if is_dis_list:
                    ColourPrint.print_yellow('Discharge Start')
                    discharge_list_generated = self.discharge_list_generate(page)
                    ColourPrint.print_pink("Discharge End")
                    # Save the generated lists to the CSV file
                    save_to_json(pre_auth_list_generated, claim_list_generated, discharge_list_generated, json_file_path)

            else:
                # File is less than 2 hours old, read from the file
                pre_auth_list_generated, claim_list_generated, discharge_list_generated = read_from_json(
                    json_file_path)
                # print(type(pre_auth_list_generated), type(claim_list_generated), type(discharge_list_generated))
            # getting to case search page
            page.wait_for_selector(tms_xpaths.claim_query_cases_search_tab_xp).click()
            page.wait_for_load_state()

            both_query_list = claim_list_generated + pre_auth_list_generated
            # both_query_list = pre_auth_list_generated
            # print(both_query_list)
            self.manipulating_downloaded_excel(both_query_list, 'QUERY', is_query_question_required=True, page=page)
            self.upload_in_g_sheet('QUERY')
            ColourPrint.print_pink('Google_Spreadsheet_Query_pending_updated')

            self.discharge_manipulating_excel(discharge_list_generated, 'Pend Dischg', is_query_question_required=False,
                                              page=page)
            self.upload_in_g_sheet('Pend Dischg')
            ColourPrint.print_pink('Google_Spreadsheet_Discharge_pending_updated')

            ColourPrint.print_yellow('Logging Out')
            down_logout_arrow = page.wait_for_selector(tms_xpaths.down_arrow_xp)
            down_logout_arrow.click()
            logout_button = page.wait_for_selector(tms_xpaths.logout_button_xp)
            logout_button.click()
            ColourPrint.print_yellow('Log out done')



# Function to get the file creation time
json_file_path = r"G:\My Drive\GdrivePC\Hospital\RSBY\New\down\dd.json"
def get_file_creation_time(file_path):
    if os.path.exists(file_path):
        print('------', os.path.getmtime(file_path))
        return os.path.getmtime(file_path)
    return None

# Function to save lists to CSV (separating pre-auth, claim, and discharge)
def save_to_json(pre_auth_list, claim_list, discharge_list, file_path):
    data = {
        'PreAuth List': pre_auth_list,
        'Claim List': claim_list,
        'Discharge List': discharge_list
    }

    # Write to JSON file (overwrite existing data)
    with open(file_path, mode='w', encoding='utf-8') as file:
        json.dump(data, file, ensure_ascii=False, indent=4)


# Function to read lists from CSV
def read_from_json(file_path):
    # Read from JSON file
    with open(file_path, mode='r', encoding='utf-8') as file:
        data = json.load(file)

    pre_auth_list = data.get('PreAuth List', [])
    claim_list = data.get('Claim List', [])
    discharge_list = data.get('Discharge List', [])

    return pre_auth_list, claim_list, discharge_list


# Check if the CSV file exists and is older than 2 hours
def is_file_older_than_2_hours(file_path):
    creation_time = get_file_creation_time(file_path)
    if creation_time:
        file_time = datetime.fromtimestamp(creation_time)
        current_time = datetime.now()
        time_diff = current_time - file_time
        return time_diff > timedelta(hours=2)
    return True  # file not found inaccessible thus meaning OLD


if __name__ == '__main__':
    t =time.time()
    AllListGenerator().main_all_list_getter_tms(user_id='CHH009264', is_pre_auth_list=True, is_claim_list=True,
                                                is_dis_list=True)
    print((time.time() - t)/60, 'min')
# AllListGenerator().manipulating_downloaded_excel()
