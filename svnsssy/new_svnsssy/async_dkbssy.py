import asyncio
import json
import sys
import time
from enum import Enum
import openpyxl
from playwright.async_api import async_playwright, TimeoutError, Page
from EHOSP.ehosp_2.async_ui_entry import apply_throttle
from dkbssy.utils.excel_utils import ExcelMethods, get_employee_code_row_number, sqlite_process_3_by_ecode
from EHOSP.tk_ehosp.alert_boxes import error_tk_box
from TMS_new.async_tms_new.desired_page import get_desired_page_indexes_in_cdp_async_for_ASYNC
from dkbssy.utils import name_for_date_check_gmc
from dkbssy.utils.colour_prints import ColourPrint, message_box
from dkbssy.utils.incen_percent import inc_percent_amt_calc
from dkbssy.utils.sqlite_updation_long import sql_update_for_3
from svnsssy.new_svnsssy.amount_update_request import main_aiohttp_optimized

DUPLICATE_NAMED_CODES:dict = {
    "Vandana Kanwar":["05170020188", "N0517466"],
    "DEMO":['01234','56789']
}

async def _main(all_incentive_multiline:str, set_timeout_is, attach_upload_file_path:str=None):
    page_indexes = await get_desired_page_indexes_in_cdp_async_for_ASYNC(user_title_of_page='Shaheed Veer Narayan Singh Ayushman Swasthya Yojna')
    # print(is_chrome_dev_open)

    # only for checking the approver site open
    # await get_desired_page_indexes_in_cdp_async_for_ASYNC(user_title_of_page='Shaheed Veer Narayan Singh Ayushman Swasthya Yojna', cdp_port=9223)

    # 'Alerting the incomplete incentives'
    # stuck_display()


    # from_date = '01-08-2022'
    from_date = '2022-08-01'
    # to_date = '31-12-2023'
    to_date = '2023-12-31'

    # check upload file -> skipping in new modified
    # checked_file_path = check_file_path(attach_file_path=attach_upload_file_path)
    print()

    # splits to individual data line
    incentive_cases_list = split_multiline_to_list(all_incentive_multiline=all_incentive_multiline)
    # indent_json_print(incentive_cases_list)

    # function ran
    name_for_date_check_gmc.checker_with_dict_output(incentive_cases_list)

    # where_start_incentive_case_index = get_start_index(incentive_cases_list)

    async with async_playwright() as p:
        cdp_for_main = 9222
        browser = await p.chromium.connect_over_cdp(f"http://localhost:{cdp_for_main}")
        context = browser.contexts[0]
        all_pages = context.pages

        page_index = page_indexes[0]  # selecting the first index of matching page
        page = all_pages[page_index]  # selecting the first PAGE of matching page

        page.set_default_timeout(set_timeout_is)
        page.set_default_navigation_timeout(set_timeout_is)

        # ⚡ Attach dialog handler exactly once
        page.on("dialog", handle_dialog)

        # await apply_throttle(page=page, mode=None)

        depart_choice = ExcelMethods().entry_department()

        # getting each incentive case number
        for idx, case_number_with_data in enumerate(incentive_cases_list, start=1):
            case_number, amount_is = retrieve_case_number(case_number_with_data)
            ColourPrint.print_pink(f"Case Number {idx} Is:", case_number)

            "QUERY MODIFY IF STUCK AND NOT SHOWING IN BOTH SITES"
            ColourPrint.print_turquoise("Query Modify")
            print("CHECK IN MAIN SITE",f"https://dkbssy.cg.nic.in/secure/incentivemodule/incentivedetailsdme.aspx?c={case_number}&amt={amount_is}")
            query_modify = f'https://dkbssy.cg.nic.in/secure/incentivemodule/incentivemoduleApViewDME.aspx?ci={case_number}'
            ColourPrint.print_blue(query_modify)

            # "CHECKING THE MANUAL QUERY WAS CLEARED METHOD INSIDE THE SELECT DEPARTMENT"
            await select_department_and_dates(page=page, depart_name=depart_choice,from_date=from_date, to_date=to_date, case_number=case_number)

            'Alerting the incomplete incentives'
            stuck_display()

            await check_already_initiated(page, case_number)  # return true

            """Scrape the details of incentive case"""
            main_master_c_num_with_name_amt_cat_e_code =  []
            details_list = list(await scrape_incentive_case_details(page=page))
            # print(details_list)
            main_master_c_num_with_name_amt_cat_e_code.extend(details_list)  # adding the case details in the main master

            await radio_click(page=page)
            # await selection_of_category()
            final_cat_and_names = get_cat_and_names_for_entry()
            # print(f'{final_cat_and_names=}')

            """Entry is doing here"""
            await entry_proper(page, final_cat_and_names)
            # main_master_c_num_with_name_amt_cat_e_code.extend(filled_entered_names_cat_amount)
            #
            # print('entered inc data master', main_master_c_num_with_name_amt_cat_e_code)
            # # saving the names in Excel and DB
            # ExcelMethods().excel_save_new(main_master_c_num_with_name_amt_cat_e_code)

            # await submit_upload(page=page,upload_file_path=checked_file_path)

            await check_box_and_submit(page=page)

            # sys.exit()

            Status().set_status(case_number=case_number, status_text=StatusText.INITIATED_DONE)

            # "APPROVER SITE WORKS"
            # await approver_raise_query(case_number=case_number)

            "Original Site"
            await original_site_query_modify(page=page, case_number=case_number)

            "getting the amounts to be used in db"
            incentive_received_data = await _scrape_amount(page=page, case_number=case_number)
            #  [('SURINDAR LAL CHANDRA', 'डाटा एंट्री ऑपरेटर', 'DME55173655', '22.1'), ('Yamini Verma', 'डाटा एंट्री ऑपरेटर', 'DME55173200', '22.1'), ('KUMARI JYOTI LAHRE', 'डाटा एंट्री ऑपरेटर', 'DME55173036', '22.1'), ('SUMITRA KURREY', 'डाटा एंट्री ऑपरेटर', 'J55173309', '22.1'), ('Mr. Kiran Kumar Sahu', 'डाटा एंट्री ऑपरेटर', 'J55172798', '22.1'), ('K. Damodar', 'चतुर्थ वर्ग एवं सफाई कर्मचारी', '05170020110', '1.54'), ('L.Kondiya', 'चतुर्थ वर्ग एवं सफाई कर्मचारी', '05170020107', '1.54'), ('Bal Chainya', 'चतुर्थ वर्ग एवं सफाई कर्मचारी', '05170020105', '1.54'), ('Smt. Shushila Bai', 'चतुर्थ वर्ग एवं सफाई कर्मचारी', '05170020103', '1.54'), ('Ankaiya', 'चतुर्थ वर्ग एवं सफाई कर्मचारी', '05170020093', '1.54'), ('Smt. Achamma Bai', 'चतुर्थ वर्ग एवं सफाई कर्मचारी', '05170020090', '1.54'), ('Smt. Chandrika Bai', 'चतुर्थ वर्ग एवं सफाई कर्मचारी', '05170020088', '1.54'), ('Mongra Tandon', 'चतुर्थ वर्ग एवं सफाई कर्मचारी', '04170150044', '1.54'), ('reman singh kanwar', 'चतुर्थ वर्ग एवं सफाई कर्मचारी', '07170250453', '1.54'), ('Mithlesh Kumar Markam', 'चतुर्थ वर्ग एवं सफाई कर्मचारी', '07170250399', '1.54'), ('V. Kondiya', 'चतुर्थ वर्ग एवं सफाई कर्मचारी', '05170020092', '1.54'), ('Panch Ram Nirmalkar', 'चतुर्थ वर्ग एवं सफाई कर्मचारी', '05170020091', '1.54'), ('Shri Shyam Das Mahant', 'चतुर्थ वर्ग एवं सफाई कर्मचारी', '05170020087', '1.54'), ('Shri Amar Das', 'चतुर्थ वर्ग एवं सफाई कर्मचारी', '05170020086', '1.54'), ('Shri Shiv Kumar Sarthi', 'चतुर्थ वर्ग एवं सफाई कर्मचारी', '05170020084', '1.54'), ('Shri Krishna Kumar Tripathi', 'चतुर्थ वर्ग एवं सफाई कर्मचारी', '05170020082', '1.54'), ('mohendra kumar janardan', 'चतुर्थ वर्ग एवं सफाई कर्मचारी', '04170150084', '1.54'), ('RAVINDRA KUMAR PANDEY', 'चतुर्थ वर्ग एवं सफाई कर्मचारी', '04170150035', '1.54'), ('Shri S.N.Shriwas', 'चतुर्थ वर्ग एवं सफाई कर्मचारी', '05170020100', '1.54'), ('Shri Sunil Kumar Channe', 'चतुर्थ वर्ग एवं सफाई कर्मचारी', '05170020099', '1.54'), ('Shri H.P.Tiwari', 'चतुर्थ वर्ग एवं सफाई कर्मचारी', '05170020098', '1.54'), ('Shri C.P. Chandra', 'चतुर्थ वर्ग एवं सफाई कर्मचारी', '05170020097', '1.54'), ('S SHANKAR RAO', 'चतुर्थ वर्ग एवं सफाई कर्मचारी', '05170020189', '1.54'), ('PUSHPENDRA KUMAR', 'चतुर्थ वर्ग एवं सफाई कर्मचारी', '05170020184', '1.54'), ('Smt.Lata Chandra', 'चतुर्थ वर्ग एवं सफाई कर्मचारी', '05170020137', '1.54'), ('Shri Maheshwar Prasad Tandon', 'चतुर्थ वर्ग एवं सफाई कर्मचारी', '05170020131', '1.54'), ('Anand Singh Rathiya', 'चतुर्थ वर्ग एवं सफाई कर्मचारी', 'N5517857', '1.54'), ('Smt. Sant Ram Kashyap', 'चतुर्थ वर्ग एवं सफाई कर्मचारी', '05170020095', '1.54'), ('Komal Dewangan', 'चतुर्थ वर्ग एवं सफाई कर्मचारी', 'N5517589', '1.54'), ('PRATIBHA', 'चतुर्थ वर्ग एवं सफाई कर्मचारी', 'N5517615', '1.54'), ('BAJANTI', 'चतुर्थ वर्ग एवं सफाई कर्मचारी', 'N5517614', '1.54'), ('SUNITA RATHORE', 'चतुर्थ वर्ग एवं सफाई कर्मचारी', 'N5517613', '1.54'), ('AAKANKSHA', 'चतुर्थ वर्ग एवं सफाई कर्मचारी', 'N5517610', '1.54'), ('Mrs. ANAMIKA', 'चतुर्थ वर्ग एवं सफाई कर्मचारी', 'N5517055', '1.54'), ('Mrs. BASANTI RATHIYA', 'चतुर्थ वर्ग एवं सफाई कर्मचारी', 'N5517028', '1.54'), ('KU. PRIYA', 'चतुर्थ वर्ग एवं सफाई कर्मचारी', 'N5517136', '1.54'), ('MAHENDRA KUMAR', 'चतुर्थ वर्ग एवं सफाई कर्मचारी', 'N55171003', '1.54'), ('Mr. BALDEV SINGH', 'चतुर्थ वर्ग एवं सफाई कर्मचारी', 'N5517085', '1.54'), ('Mr .AVINASH KUMAR BANJARE', 'चतुर्थ वर्ग एवं सफाई कर्मचारी', 'N5517084', '1.54'), ('Mr. MONESHWAR CHANDRA RATHIYA', 'चतुर्थ वर्ग एवं सफाई कर्मचारी', 'N5517083', '1.54'), ('KUMARI SHREMATI', 'चतुर्थ वर्ग एवं सफाई कर्मचारी', 'N5517056', '1.54'), ('RAMANAND', 'चतुर्थ वर्ग एवं सफाई कर्मचारी', 'N5517042', '1.54'), ('RAJESH', 'चतुर्थ वर्ग एवं सफाई कर्मचारी', 'N5517032', '1.54'), ('Narendra Kumar Kanwar', 'नर्सिंग एवं पैरामेडिकल स्टाफ', '04170140534', '2.95'), ('Manish Singh', 'नर्सिंग एवं पैरामेडिकल स्टाफ', '05170011626', '2.95'), ('Jyoti Porte', 'नर्सिंग एवं पैरामेडिकल स्टाफ', 'N5517990', '2.95'), ('Bhanu Priya Chauhan', 'नर्सिंग एवं पैरामेडिकल स्टाफ', 'N5517989', '2.95'), ('KIRAN CHANDRA', 'नर्सिंग एवं पैरामेडिकल स्टाफ', 'N5517979', '2.95'), ('TULSI YADAV', 'नर्सिंग एवं पैरामेडिकल स्टाफ', '0617009003', '2.95'), ('NEELIMA NISHAD', 'नर्सिंग एवं पैरामेडिकल स्टाफ', '05170050074', '2.95'), ('ASHA KINDO', 'नर्सिंग एवं पैरामेडिकल स्टाफ', '03170050126', '2.95'), ('Anita Yadav', 'नर्सिंग एवं पैरामेडिकल स्टाफ', '22170040038', '2.95'), ('Smt. Bhavana Bansiyar', 'नर्सिंग एवं पैरामेडिकल स्टाफ', '15170200090', '2.95'), ('Ku. Har Bai Khunte', 'नर्सिंग एवं पैरामेडिकल स्टाफ', '14170011702', '2.95'), ('vinita tigga', 'नर्सिंग एवं पैरामेडिकल स्टाफ', '12170050188', '2.95'), ('KU. ANJANA KUJUR', 'नर्सिंग एवं पैरामेडिकल स्टाफ', '11170260055', '2.95'), ('SAROJ RATHOR', 'नर्सिंग एवं पैरामेडिकल स्टाफ', '09170160077', '2.95'), ('Neelam Kanwar', 'नर्सिंग एवं पैरामेडिकल स्टाफ', '08170040064', '2.95'), ('Smt.Nirmala Miri', 'नर्सिंग एवं पैरामेडिकल स्टाफ', '07170250081', '2.95'), ('SMT NAMITA PATEL', 'नर्सिंग एवं पैरामेडिकल स्टाफ', '05170020176', '2.95'), ('ku Kiran Sahu', 'नर्सिंग एवं पैरामेडिकल स्टाफ', '05170020161', '2.95'), ('Ku.Marget Bishwash', 'नर्सिंग एवं पैरामेडिकल स्टाफ', '05170020153', '2.95'), ('Ku.Jyoti Gual', 'नर्सिंग एवं पैरामेडिकल स्टाफ', '05170020143', '2.95'), ('Smt Bharti Markam', 'नर्सिंग एवं पैरामेडिकल स्टाफ', '05170020142', '2.95'), ('Ku. Sandhya Nair', 'नर्सिंग एवं पैरामेडिकल स्टाफ', '05170020124', '2.95'), ('PREETI SONWANI', 'नर्सिंग एवं पैरामेडिकल स्टाफ', '05170011771', '2.95'), ('MANJULATA KANWAR', 'नर्सिंग एवं पैरामेडिकल स्टाफ', '04170150162', '2.95'), ('SUSHILATA RATHIYA', 'नर्सिंग एवं पैरामेडिकल स्टाफ', '04170140436', '2.95'), ('DIPTI LAKRA', 'नर्सिंग एवं पैरामेडिकल स्टाफ', '04170140331', '2.95'), ('Smt.DIVYA ALKA TOPPO', 'नर्सिंग एवं पैरामेडिकल स्टाफ', '04170060068', '2.95'), ('Sanjaya Kumar Tiwari', 'नर्सिंग एवं पैरामेडिकल स्टाफ', 'N5517515', '2.95'), ('PAWAN KUMAR PATEL', 'नर्सिंग एवं पैरामेडिकल स्टाफ', 'N5517809', '2.95'), ('VANDANA SAMUEL', 'नर्सिंग एवं पैरामेडिकल स्टाफ', '66170020614', '2.95'), ('CHANDRAKANTA SHRIVASTAV', 'नर्सिंग एवं पैरामेडिकल स्टाफ', '66170020507', '2.95'), ('PUSHPA BAG', 'नर्सिंग एवं पैरामेडिकल स्टाफ', '66170020500', '2.95'), ('Deepa sahu', 'नर्सिंग एवं पैरामेडिकल स्टाफ', '66170020472', '2.95'), ('Priti Masih Upma Kumar', 'नर्सिंग एवं पैरामेडिकल स्टाफ', '15170200071', '2.95'), ('SMT. PHULKUMARI TOPPO', 'नर्सिंग एवं पैरामेडिकल स्टाफ', '66170020668', '2.95'), ('Smt. Sarika Tirpude Ramteke', 'नर्सिंग एवं पैरामेडिकल स्टाफ', '66170020665', '2.95'), ('PRIYA VERMA', 'नर्सिंग एवं पैरामेडिकल स्टाफ', '66170020641', '2.95'), ('Victoria Gardia', 'नर्सिंग एवं पैरामेडिकल स्टाफ', '66170010273', '2.95'), ('Smt.Seema Rani Lajras', 'नर्सिंग एवं पैरामेडिकल स्टाफ', '05170020063', '2.95'), ('Rajni Kiran Kispotta', 'नर्सिंग एवं पैरामेडिकल स्टाफ', '11170010433', '2.95'), ('Sanjay Das Manikpuri', 'नर्सिंग एवं पैरामेडिकल स्टाफ', 'N5517599', '2.95'), ('Neelu Chaudhary', 'नर्सिंग एवं पैरामेडिकल स्टाफ', 'N5517492', '2.95'), ('Radhe Shyam Kashyap', 'नर्सिंग एवं पैरामेडिकल स्टाफ', 'AI55170022', '2.95'), ('Amarjeet Kour', 'नर्सिंग एवं पैरामेडिकल स्टाफ', '19170040056', '2.95'), ('Smt. Varnita Zilkar', 'नर्सिंग एवं पैरामेडिकल स्टाफ', '15170200084', '2.95'), ('Dr. Priyanka Ekka', 'सभी फिजिशियन / सर्जन', '05530010030', '198.9'), ('Dr Ankita Kapoor', 'सभी फिजिशियन / सर्जन', '05530010028', '198.9'), ('DR ARUNIKA SISODIYA', 'सभी फिजिशियन / सर्जन', '05170020182', '198.9'), ('Dr. V.Agrawal', 'सभी फिजिशियन / सर्जन', '05170020043', '198.9'), ('Dr.Sumit Gupta', 'सभी फिजिशियन / सर्जन', '05170020187', '198.9'), ('MAHENDRA KUMAR', 'पैथोलॉजी ,रेडियोलॉजी , माइक्रोबायोलॉजी , बायोकेमेस्ट्री , निश्चेतना (टेक्निशियन )', '05170011797', '2.49'), ('dinesh kumar patel', 'पैथोलॉजी ,रेडियोलॉजी , माइक्रोबायोलॉजी , बायोकेमेस्ट्री , निश्चेतना (टेक्निशियन )', '09530010377', '2.49'), ('Smt.Bhagwati Koshle', 'पैथोलॉजी ,रेडियोलॉजी , माइक्रोबायोलॉजी , बायोकेमेस्ट्री , निश्चेतना (टेक्निशियन )', '09170010257', '2.49'), ('ARUN KUMAR KANWAR', 'पैथोलॉजी ,रेडियोलॉजी , माइक्रोबायोलॉजी , बायोकेमेस्ट्री , निश्चेतना (टेक्निशियन )', '07170021387', '2.49'), ('SMT DURGESHWARI KARSH', 'पैथोलॉजी ,रेडियोलॉजी , माइक्रोबायोलॉजी , बायोकेमेस्ट्री , निश्चेतना (टेक्निशियन )', '06170050083', '2.49'), ('SHIVSHANKAR SINGH KANWAR', 'पैथोलॉजी ,रेडियोलॉजी , माइक्रोबायोलॉजी , बायोकेमेस्ट्री , निश्चेतना (टेक्निशियन )', '05170060099', '2.49'), ('Shri C.L. Dixena', 'पैथोलॉजी ,रेडियोलॉजी , माइक्रोबायोलॉजी , बायोकेमेस्ट्री , निश्चेतना (टेक्निशियन )', '05170020117', '2.49'), ('Shri Sushil kumar miri', 'पैथोलॉजी ,रेडियोलॉजी , माइक्रोबायोलॉजी , बायोकेमेस्ट्री , निश्चेतना (टेक्निशियन )', '03170010366', '2.49'), ('Santosh Kumar Singh', 'पैथोलॉजी ,रेडियोलॉजी , माइक्रोबायोलॉजी , बायोकेमेस्ट्री , निश्चेतना (टेक्निशियन )', 'AI55170030', '2.49'), ('Smt. Reena Verma', 'पैथोलॉजी ,रेडियोलॉजी , माइक्रोबायोलॉजी , बायोकेमेस्ट्री , निश्चेतना (टेक्निशियन )', 'AI55170024', '2.49'), ('Smt. Pinky Singh', 'पैथोलॉजी ,रेडियोलॉजी , माइक्रोबायोलॉजी , बायोकेमेस्ट्री , निश्चेतना (टेक्निशियन )', '06170100054', '2.49'), ('Shri Dildar Sahish', 'पैथोलॉजी ,रेडियोलॉजी , माइक्रोबायोलॉजी , बायोकेमेस्ट्री , निश्चेतना (टेक्निशियन )', '05170020054', '2.49'), ('Haricharan Jangde', 'पैथोलॉजी ,रेडियोलॉजी , माइक्रोबायोलॉजी , बायोकेमेस्ट्री , निश्चेतना (टेक्निशियन )', '05170010396', '2.49'), ('Pratima Sahu', 'पैथोलॉजी ,रेडियोलॉजी , माइक्रोबायोलॉजी , बायोकेमेस्ट्री , निश्चेतना (टेक्निशियन )', 'N5517593', '2.49'), ('Geeta patel', 'पैथोलॉजी ,रेडियोलॉजी , माइक्रोबायोलॉजी , बायोकेमेस्ट्री , निश्चेतना (टेक्निशियन )', 'J55172863', '2.49'), ('ACHCHHE KUMAR PATLEY', 'पैथोलॉजी ,रेडियोलॉजी , माइक्रोबायोलॉजी , बायोकेमेस्ट्री , निश्चेतना (टेक्निशियन )', '04170150173', '2.49'), ('Dr. manoj Kumar', 'पैथोलॉजी ,रेडियोलॉजी , माइक्रोबायोलॉजी , बायोकेमेस्ट्री , ( फैकल्टी एंड रेसीडेंट )', '05530010029', '15.47'), ('Dr. Sumit Gupta', 'पैथोलॉजी ,रेडियोलॉजी , माइक्रोबायोलॉजी , बायोकेमेस्ट्री , ( फैकल्टी एंड रेसीडेंट )', 'DME55173185', '15.47'), ('Dr. Ashutosh Kumar', 'पैथोलॉजी ,रेडियोलॉजी , माइक्रोबायोलॉजी , बायोकेमेस्ट्री , ( फैकल्टी एंड रेसीडेंट )', 'DME55173188', '15.47'), ('Shende Pranali kisandas', 'पैथोलॉजी ,रेडियोलॉजी , माइक्रोबायोलॉजी , बायोकेमेस्ट्री , ( फैकल्टी एंड रेसीडेंट )', 'DME55172441', '15.47'), ('Vibha Tandon', 'पैथोलॉजी ,रेडियोलॉजी , माइक्रोबायोलॉजी , बायोकेमेस्ट्री , ( फैकल्टी एंड रेसीडेंट )', '05530010013', '15.47'), ('HANISH KUMAR CHOWDA', 'पैथोलॉजी ,रेडियोलॉजी , माइक्रोबायोलॉजी , बायोकेमेस्ट्री , ( फैकल्टी एंड रेसीडेंट )', '04170140518', '15.47'), ('Reena Nayak', 'पैथोलॉजी ,रेडियोलॉजी , माइक्रोबायोलॉजी , बायोकेमेस्ट्री , ( फैकल्टी एंड रेसीडेंट )', 'DME55172439', '15.47'), ('Dr. Veenapani Mire', 'पैथोलॉजी ,रेडियोलॉजी , माइक्रोबायोलॉजी , बायोकेमेस्ट्री , ( फैकल्टी एंड रेसीडेंट )', '05530010079', '15.47'), ('Rajesh Kumar', 'पैथोलॉजी ,रेडियोलॉजी , माइक्रोबायोलॉजी , बायोकेमेस्ट्री , ( फैकल्टी एंड रेसीडेंट )', '05530010065', '15.47'), ('Dushyant Chandra', 'पैथोलॉजी ,रेडियोलॉजी , माइक्रोबायोलॉजी , बायोकेमेस्ट्री , ( फैकल्टी एंड रेसीडेंट )', '05530010031', '15.47'), ('Dr. Deepa Janghel', 'पैथोलॉजी ,रेडियोलॉजी , माइक्रोबायोलॉजी , बायोकेमेस्ट्री , ( फैकल्टी एंड रेसीडेंट )', '05530010026', '15.47'), ('Dr. Awadh Sahu', 'पैथोलॉजी ,रेडियोलॉजी , माइक्रोबायोलॉजी , बायोकेमेस्ट्री , ( फैकल्टी एंड रेसीडेंट )', 'DME55173201', '15.47'), ('DR RAKESH KUMAR AGRAWAL', 'पैथोलॉजी ,रेडियोलॉजी , माइक्रोबायोलॉजी , बायोकेमेस्ट्री , ( फैकल्टी एंड रेसीडेंट )', '05170040053', '15.47'), ('GHANSHYAM SINGH JATRA', 'पैथोलॉजी ,रेडियोलॉजी , माइक्रोबायोलॉजी , बायोकेमेस्ट्री , ( फैकल्टी एंड रेसीडेंट )', '05170011289', '15.47'), ('AVINASH MESHRAM', 'अधिष्ठाता अस्पताल अधीक्षक ,सहायक अधीक्षक नोडल अधिकारी एवं सहायक नोडल अधिकारी , अस्पताल सलाहकार', '66170010023', '3.16'), ('RAVIKANT JATWAR', 'अधिष्ठाता अस्पताल अधीक्षक ,सहायक अधीक्षक नोडल अधिकारी एवं सहायक नोडल अधिकारी , अस्पताल सलाहकार', '04170140099', '3.16'), ('Dr. Rakesh Kumar Verma', 'अधिष्ठाता अस्पताल अधीक्षक ,सहायक अधीक्षक नोडल अधिकारी एवं सहायक नोडल अधिकारी , अस्पताल सलाहकार', '02530010055', '3.16'), ('Dr. Aditya Siodiya', 'अधिष्ठाता अस्पताल अधीक्षक ,सहायक अधीक्षक नोडल अधिकारी एवं सहायक नोडल अधिकारी , अस्पताल सलाहकार', '09530010327', '3.16'), ('DR.ANMOL MADHUR MINZ', 'अधिष्ठाता अस्पताल अधीक्षक ,सहायक अधीक्षक नोडल अधिकारी एवं सहायक नोडल अधिकारी , अस्पताल सलाहकार', '04170140656', '3.16'), ('Dr. Durga Shankar Patel', 'अधिष्ठाता अस्पताल अधीक्षक ,सहायक अधीक्षक नोडल अधिकारी एवं सहायक नोडल अधिकारी , अस्पताल सलाहकार', '66170010344', '3.16'), ('Dr. Gopal Singh Kanwer', 'अधिष्ठाता अस्पताल अधीक्षक ,सहायक अधीक्षक नोडल अधिकारी एवं सहायक नोडल अधिकारी , अस्पताल सलाहकार', '11170010478', '3.16')]


            processed_inc_names = process_scraped_data(incentive_received_data)
            print()
            print('printing processes incentive received data')
            # print(processed_inc_names)

            main_master_c_num_with_name_amt_cat_e_code.extend(processed_inc_names)
            # print(main_master_c_num_with_name_amt_cat_e_code) # ['CASE/PS6/HOSP22G146659/CB7008568', '2210', 'Ophthalmology', 'SICS with non-foldable IOL', '2023-10-07', 'Lal Khan', ('Dr. Gopal Singh Kanwer', 'अधिष्ठाता अस्पताल अधीक्षक ,सहायक अधीक्षक नोडल अधिकारी एवं सहायक नोडल अधिकारी , अस्पताल सलाहकार', '11170010478', '3.16'), ('Dr. Durga Shankar Patel', 'अधिष्ठाता अस्पताल अधीक्षक ,सहायक अधीक्षक नोडल अधिकारी एवं सहायक नोडल अधिकारी , अस्पताल सलाहकार', '66170010344', '3.16'), ('DR.ANMOL MADHUR MINZ', 'अधिष्ठाता अस्पताल अधीक्षक ,सहायक अधीक्षक नोडल अधिकारी एवं सहायक नोडल अधिकारी , अस्पताल सलाहकार', '04170140656', '3.16'), ('Dr. Aditya Siodiya', 'अधिष्ठाता अस्पताल अधीक्षक ,सहायक अधीक्षक नोडल अधिकारी एवं सहायक नोडल अधिकारी , अस्पताल सलाहकार', '09530010327', '3.16'), ('Dr. Rakesh Kumar Verma', 'अधिष्ठाता अस्पताल अधीक्षक ,सहायक अधीक्षक नोडल अधिकारी एवं सहायक नोडल अधिकारी , अस्पताल सलाहकार', '02530010055', '3.16'), ('RAVIKANT JATWAR', 'अधिष्ठाता अस्पताल अधीक्षक ,सहायक अधीक्षक नोडल अधिकारी एवं सहायक नोडल अधिकारी , अस्पताल सलाहकार', '04170140099', '3.16'), ('AVINASH MESHRAM', 'अधिष्ठाता अस्पताल अधीक्षक ,सहायक अधीक्षक नोडल अधिकारी एवं सहायक नोडल अधिकारी , अस्पताल सलाहकार', '66170010023', '3.16'), ('GHANSHYAM SINGH JATRA', 'पैथोलॉजी ,रेडियोलॉजी , माइक्रोबायोलॉजी , बायोकेमेस्ट्री , ( फैकल्टी एंड रेसीडेंट )', '05170011289', '15.47'), ('DR RAKESH KUMAR AGRAWAL', 'पैथोलॉजी ,रेडियोलॉजी , माइक्रोबायोलॉजी , बायोकेमेस्ट्री , ( फैकल्टी एंड रेसीडेंट )', '05170040053', '15.47'), ('Dr. Awadh Sahu', 'पैथोलॉजी ,रेडियोलॉजी , माइक्रोबायोलॉजी , बायोकेमेस्ट्री , ( फैकल्टी एंड रेसीडेंट )', 'DME55173201', '15.47'), ('Dr. Deepa Janghel', 'पैथोलॉजी ,रेडियोलॉजी , माइक्रोबायोलॉजी , बायोकेमेस्ट्री , ( फैकल्टी एंड रेसीडेंट )', '05530010026', '15.47'), ('Dushyant Chandra', 'पैथोलॉजी ,रेडियोलॉजी , माइक्रोबायोलॉजी , बायोकेमेस्ट्री , ( फैकल्टी एंड रेसीडेंट )', '05530010031', '15.47'), ('Rajesh Kumar', 'पैथोलॉजी ,रेडियोलॉजी , माइक्रोबायोलॉजी , बायोकेमेस्ट्री , ( फैकल्टी एंड रेसीडेंट )', '05530010065', '15.47'), ('Dr. Veenapani Mire', 'पैथोलॉजी ,रेडियोलॉजी , माइक्रोबायोलॉजी , बायोकेमेस्ट्री , ( फैकल्टी एंड रेसीडेंट )', '05530010079', '15.47'), ('Reena Nayak', 'पैथोलॉजी ,रेडियोलॉजी , माइक्रोबायोलॉजी , बायोकेमेस्ट्री , ( फैकल्टी एंड रेसीडेंट )', 'DME55172439', '15.47'), ('HANISH KUMAR CHOWDA', 'पैथोलॉजी ,रेडियोलॉजी , माइक्रोबायोलॉजी , बायोकेमेस्ट्री , ( फैकल्टी एंड रेसीडेंट )', '04170140518', '15.47'), ('Vibha Tandon', 'पैथोलॉजी ,रेडियोलॉजी , माइक्रोबायोलॉजी , बायोकेमेस्ट्री , ( फैकल्टी एंड रेसीडेंट )', '05530010013', '15.47'), ('Shende Pranali kisandas', 'पैथोलॉजी ,रेडियोलॉजी , माइक्रोबायोलॉजी , बायोकेमेस्ट्री , ( फैकल्टी एंड रेसीडेंट )', 'DME55172441', '15.47'), ('Dr. Ashutosh Kumar', 'पैथोलॉजी ,रेडियोलॉजी , माइक्रोबायोलॉजी , बायोकेमेस्ट्री , ( फैकल्टी एंड रेसीडेंट )', 'DME55173188', '15.47'), ('Dr. Sumit Gupta', 'पैथोलॉजी ,रेडियोलॉजी , माइक्रोबायोलॉजी , बायोकेमेस्ट्री , ( फैकल्टी एंड रेसीडेंट )', 'DME55173185', '15.47'), ('Dr. manoj Kumar', 'पैथोलॉजी ,रेडियोलॉजी , माइक्रोबायोलॉजी , बायोकेमेस्ट्री , ( फैकल्टी एंड रेसीडेंट )', '05530010029', '15.47'), ('ACHCHHE KUMAR PATLEY', 'पैथोलॉजी ,रेडियोलॉजी , माइक्रोबायोलॉजी , बायोकेमेस्ट्री , निश्चेतना (टेक्निशियन )', '04170150173', '2.49'), ('Geeta patel', 'पैथोलॉजी ,रेडियोलॉजी , माइक्रोबायोलॉजी , बायोकेमेस्ट्री , निश्चेतना (टेक्निशियन )', 'J55172863', '2.49'), ('Pratima Sahu', 'पैथोलॉजी ,रेडियोलॉजी , माइक्रोबायोलॉजी , बायोकेमेस्ट्री , निश्चेतना (टेक्निशियन )', 'N5517593', '2.49'), ('Haricharan Jangde', 'पैथोलॉजी ,रेडियोलॉजी , माइक्रोबायोलॉजी , बायोकेमेस्ट्री , निश्चेतना (टेक्निशियन )', '05170010396', '2.49'), ('Shri Dildar Sahish', 'पैथोलॉजी ,रेडियोलॉजी , माइक्रोबायोलॉजी , बायोकेमेस्ट्री , निश्चेतना (टेक्निशियन )', '05170020054', '2.49'), ('Smt. Pinky Singh', 'पैथोलॉजी ,रेडियोलॉजी , माइक्रोबायोलॉजी , बायोकेमेस्ट्री , निश्चेतना (टेक्निशियन )', '06170100054', '2.49'), ('Smt. Reena Verma', 'पैथोलॉजी ,रेडियोलॉजी , माइक्रोबायोलॉजी , बायोकेमेस्ट्री , निश्चेतना (टेक्निशियन )', 'AI55170024', '2.49'), ('Santosh Kumar Singh', 'पैथोलॉजी ,रेडियोलॉजी , माइक्रोबायोलॉजी , बायोकेमेस्ट्री , निश्चेतना (टेक्निशियन )', 'AI55170030', '2.49'), ('Shri Sushil kumar miri', 'पैथोलॉजी ,रेडियोलॉजी , माइक्रोबायोलॉजी , बायोकेमेस्ट्री , निश्चेतना (टेक्निशियन )', '03170010366', '2.49'), ('Shri C.L. Dixena', 'पैथोलॉजी ,रेडियोलॉजी , माइक्रोबायोलॉजी , बायोकेमेस्ट्री , निश्चेतना (टेक्निशियन )', '05170020117', '2.49'), ('SHIVSHANKAR SINGH KANWAR', 'पैथोलॉजी ,रेडियोलॉजी , माइक्रोबायोलॉजी , बायोकेमेस्ट्री , निश्चेतना (टेक्निशियन )', '05170060099', '2.49'), ('SMT DURGESHWARI KARSH', 'पैथोलॉजी ,रेडियोलॉजी , माइक्रोबायोलॉजी , बायोकेमेस्ट्री , निश्चेतना (टेक्निशियन )', '06170050083', '2.49'), ('ARUN KUMAR KANWAR', 'पैथोलॉजी ,रेडियोलॉजी , माइक्रोबायोलॉजी , बायोकेमेस्ट्री , निश्चेतना (टेक्निशियन )', '07170021387', '2.49'), ('Smt.Bhagwati Koshle', 'पैथोलॉजी ,रेडियोलॉजी , माइक्रोबायोलॉजी , बायोकेमेस्ट्री , निश्चेतना (टेक्निशियन )', '09170010257', '2.49'), ('dinesh kumar patel', 'पैथोलॉजी ,रेडियोलॉजी , माइक्रोबायोलॉजी , बायोकेमेस्ट्री , निश्चेतना (टेक्निशियन )', '09530010377', '2.49'), ('MAHENDRA KUMAR', 'पैथोलॉजी ,रेडियोलॉजी , माइक्रोबायोलॉजी , बायोकेमेस्ट्री , निश्चेतना (टेक्निशियन )', '05170011797', '2.49'), ('Dr.Sumit Gupta', 'सभी फिजिशियन / सर्जन', '05170020187', '198.9'), ('Dr. V.Agrawal', 'सभी फिजिशियन / सर्जन', '05170020043', '198.9'), ('DR ARUNIKA SISODIYA', 'सभी फिजिशियन / सर्जन', '05170020182', '198.9'), ('Dr Ankita Kapoor', 'सभी फिजिशियन / सर्जन', '05530010028', '198.9'), ('Dr. Priyanka Ekka', 'सभी फिजिशियन / सर्जन', '05530010030', '198.9'), ('Smt. Varnita Zilkar', 'नर्सिंग एवं पैरामेडिकल स्टाफ', '15170200084', '2.95'), ('Amarjeet Kour', 'नर्सिंग एवं पैरामेडिकल स्टाफ', '19170040056', '2.95'), ('Radhe Shyam Kashyap', 'नर्सिंग एवं पैरामेडिकल स्टाफ', 'AI55170022', '2.95'), ('Neelu Chaudhary', 'नर्सिंग एवं पैरामेडिकल स्टाफ', 'N5517492', '2.95'), ('Sanjay Das Manikpuri', 'नर्सिंग एवं पैरामेडिकल स्टाफ', 'N5517599', '2.95'), ('Rajni Kiran Kispotta', 'नर्सिंग एवं पैरामेडिकल स्टाफ', '11170010433', '2.95'), ('Smt.Seema Rani Lajras', 'नर्सिंग एवं पैरामेडिकल स्टाफ', '05170020063', '2.95'), ('Victoria Gardia', 'नर्सिंग एवं पैरामेडिकल स्टाफ', '66170010273', '2.95'), ('PRIYA VERMA', 'नर्सिंग एवं पैरामेडिकल स्टाफ', '66170020641', '2.95'), ('Smt. Sarika Tirpude Ramteke', 'नर्सिंग एवं पैरामेडिकल स्टाफ', '66170020665', '2.95'), ('SMT. PHULKUMARI TOPPO', 'नर्सिंग एवं पैरामेडिकल स्टाफ', '66170020668', '2.95'), ('Priti Masih Upma Kumar', 'नर्सिंग एवं पैरामेडिकल स्टाफ', '15170200071', '2.95'), ('Deepa sahu', 'नर्सिंग एवं पैरामेडिकल स्टाफ', '66170020472', '2.95'), ('PUSHPA BAG', 'नर्सिंग एवं पैरामेडिकल स्टाफ', '66170020500', '2.95'), ('CHANDRAKANTA SHRIVASTAV', 'नर्सिंग एवं पैरामेडिकल स्टाफ', '66170020507', '2.95'), ('VANDANA SAMUEL', 'नर्सिंग एवं पैरामेडिकल स्टाफ', '66170020614', '2.95'), ('PAWAN KUMAR PATEL', 'नर्सिंग एवं पैरामेडिकल स्टाफ', 'N5517809', '2.95'), ('Sanjaya Kumar Tiwari', 'नर्सिंग एवं पैरामेडिकल स्टाफ', 'N5517515', '2.95'), ('Smt.DIVYA ALKA TOPPO', 'नर्सिंग एवं पैरामेडिकल स्टाफ', '04170060068', '2.95'), ('DIPTI LAKRA', 'नर्सिंग एवं पैरामेडिकल स्टाफ', '04170140331', '2.95'), ('SUSHILATA RATHIYA', 'नर्सिंग एवं पैरामेडिकल स्टाफ', '04170140436', '2.95'), ('MANJULATA KANWAR', 'नर्सिंग एवं पैरामेडिकल स्टाफ', '04170150162', '2.95'), ('PREETI SONWANI', 'नर्सिंग एवं पैरामेडिकल स्टाफ', '05170011771', '2.95'), ('Ku. Sandhya Nair', 'नर्सिंग एवं पैरामेडिकल स्टाफ', '05170020124', '2.95'), ('Smt Bharti Markam', 'नर्सिंग एवं पैरामेडिकल स्टाफ', '05170020142', '2.95'), ('Ku.Jyoti Gual', 'नर्सिंग एवं पैरामेडिकल स्टाफ', '05170020143', '2.95'), ('Ku.Marget Bishwash', 'नर्सिंग एवं पैरामेडिकल स्टाफ', '05170020153', '2.95'), ('ku Kiran Sahu', 'नर्सिंग एवं पैरामेडिकल स्टाफ', '05170020161', '2.95'), ('SMT NAMITA PATEL', 'नर्सिंग एवं पैरामेडिकल स्टाफ', '05170020176', '2.95'), ('Smt.Nirmala Miri', 'नर्सिंग एवं पैरामेडिकल स्टाफ', '07170250081', '2.95'), ('Neelam Kanwar', 'नर्सिंग एवं पैरामेडिकल स्टाफ', '08170040064', '2.95'), ('SAROJ RATHOR', 'नर्सिंग एवं पैरामेडिकल स्टाफ', '09170160077', '2.95'), ('KU. ANJANA KUJUR', 'नर्सिंग एवं पैरामेडिकल स्टाफ', '11170260055', '2.95'), ('vinita tigga', 'नर्सिंग एवं पैरामेडिकल स्टाफ', '12170050188', '2.95'), ('Ku. Har Bai Khunte', 'नर्सिंग एवं पैरामेडिकल स्टाफ', '14170011702', '2.95'), ('Smt. Bhavana Bansiyar', 'नर्सिंग एवं पैरामेडिकल स्टाफ', '15170200090', '2.95'), ('Anita Yadav', 'नर्सिंग एवं पैरामेडिकल स्टाफ', '22170040038', '2.95'), ('ASHA KINDO', 'नर्सिंग एवं पैरामेडिकल स्टाफ', '03170050126', '2.95'), ('NEELIMA NISHAD', 'नर्सिंग एवं पैरामेडिकल स्टाफ', '05170050074', '2.95'), ('TULSI YADAV', 'नर्सिंग एवं पैरामेडिकल स्टाफ', '0617009003', '2.95'), ('KIRAN CHANDRA', 'नर्सिंग एवं पैरामेडिकल स्टाफ', 'N5517979', '2.95'), ('Bhanu Priya Chauhan', 'नर्सिंग एवं पैरामेडिकल स्टाफ', 'N5517989', '2.95'), ('Jyoti Porte', 'नर्सिंग एवं पैरामेडिकल स्टाफ', 'N5517990', '2.95'), ('Manish Singh', 'नर्सिंग एवं पैरामेडिकल स्टाफ', '05170011626', '2.95'), ('Narendra Kumar Kanwar', 'नर्सिंग एवं पैरामेडिकल स्टाफ', '04170140534', '2.95'), ('RAJESH', 'चतुर्थ वर्ग एवं सफाई कर्मचारी', 'N5517032', '1.54'), ('RAMANAND', 'चतुर्थ वर्ग एवं सफाई कर्मचारी', 'N5517042', '1.54'), ('KUMARI SHREMATI', 'चतुर्थ वर्ग एवं सफाई कर्मचारी', 'N5517056', '1.54'), ('Mr. MONESHWAR CHANDRA RATHIYA', 'चतुर्थ वर्ग एवं सफाई कर्मचारी', 'N5517083', '1.54'), ('Mr .AVINASH KUMAR BANJARE', 'चतुर्थ वर्ग एवं सफाई कर्मचारी', 'N5517084', '1.54'), ('Mr. BALDEV SINGH', 'चतुर्थ वर्ग एवं सफाई कर्मचारी', 'N5517085', '1.54'), ('MAHENDRA KUMAR', 'चतुर्थ वर्ग एवं सफाई कर्मचारी', 'N55171003', '1.54'), ('KU. PRIYA', 'चतुर्थ वर्ग एवं सफाई कर्मचारी', 'N5517136', '1.54'), ('Mrs. BASANTI RATHIYA', 'चतुर्थ वर्ग एवं सफाई कर्मचारी', 'N5517028', '1.54'), ('Mrs. ANAMIKA', 'चतुर्थ वर्ग एवं सफाई कर्मचारी', 'N5517055', '1.54'), ('AAKANKSHA', 'चतुर्थ वर्ग एवं सफाई कर्मचारी', 'N5517610', '1.54'), ('SUNITA RATHORE', 'चतुर्थ वर्ग एवं सफाई कर्मचारी', 'N5517613', '1.54'), ('BAJANTI', 'चतुर्थ वर्ग एवं सफाई कर्मचारी', 'N5517614', '1.54'), ('PRATIBHA', 'चतुर्थ वर्ग एवं सफाई कर्मचारी', 'N5517615', '1.54'), ('Komal Dewangan', 'चतुर्थ वर्ग एवं सफाई कर्मचारी', 'N5517589', '1.54'), ('Smt. Sant Ram Kashyap', 'चतुर्थ वर्ग एवं सफाई कर्मचारी', '05170020095', '1.54'), ('Anand Singh Rathiya', 'चतुर्थ वर्ग एवं सफाई कर्मचारी', 'N5517857', '1.54'), ('Shri Maheshwar Prasad Tandon', 'चतुर्थ वर्ग एवं सफाई कर्मचारी', '05170020131', '1.54'), ('Smt.Lata Chandra', 'चतुर्थ वर्ग एवं सफाई कर्मचारी', '05170020137', '1.54'), ('PUSHPENDRA KUMAR', 'चतुर्थ वर्ग एवं सफाई कर्मचारी', '05170020184', '1.54'), ('S SHANKAR RAO', 'चतुर्थ वर्ग एवं सफाई कर्मचारी', '05170020189', '1.54'), ('Shri C.P. Chandra', 'चतुर्थ वर्ग एवं सफाई कर्मचारी', '05170020097', '1.54'), ('Shri H.P.Tiwari', 'चतुर्थ वर्ग एवं सफाई कर्मचारी', '05170020098', '1.54'), ('Shri Sunil Kumar Channe', 'चतुर्थ वर्ग एवं सफाई कर्मचारी', '05170020099', '1.54'), ('Shri S.N.Shriwas', 'चतुर्थ वर्ग एवं सफाई कर्मचारी', '05170020100', '1.54'), ('RAVINDRA KUMAR PANDEY', 'चतुर्थ वर्ग एवं सफाई कर्मचारी', '04170150035', '1.54'), ('mohendra kumar janardan', 'चतुर्थ वर्ग एवं सफाई कर्मचारी', '04170150084', '1.54'), ('Shri Krishna Kumar Tripathi', 'चतुर्थ वर्ग एवं सफाई कर्मचारी', '05170020082', '1.54'), ('Shri Shiv Kumar Sarthi', 'चतुर्थ वर्ग एवं सफाई कर्मचारी', '05170020084', '1.54'), ('Shri Amar Das', 'चतुर्थ वर्ग एवं सफाई कर्मचारी', '05170020086', '1.54'), ('Shri Shyam Das Mahant', 'चतुर्थ वर्ग एवं सफाई कर्मचारी', '05170020087', '1.54'), ('Panch Ram Nirmalkar', 'चतुर्थ वर्ग एवं सफाई कर्मचारी', '05170020091', '1.54'), ('V. Kondiya', 'चतुर्थ वर्ग एवं सफाई कर्मचारी', '05170020092', '1.54'), ('Mithlesh Kumar Markam', 'चतुर्थ वर्ग एवं सफाई कर्मचारी', '07170250399', '1.54'), ('reman singh kanwar', 'चतुर्थ वर्ग एवं सफाई कर्मचारी', '07170250453', '1.54'), ('Mongra Tandon', 'चतुर्थ वर्ग एवं सफाई कर्मचारी', '04170150044', '1.54'), ('Smt. Chandrika Bai', 'चतुर्थ वर्ग एवं सफाई कर्मचारी', '05170020088', '1.54'), ('Smt. Achamma Bai', 'चतुर्थ वर्ग एवं सफाई कर्मचारी', '05170020090', '1.54'), ('Ankaiya', 'चतुर्थ वर्ग एवं सफाई कर्मचारी', '05170020093', '1.54'), ('Smt. Shushila Bai', 'चतुर्थ वर्ग एवं सफाई कर्मचारी', '05170020103', '1.54'), ('Bal Chainya', 'चतुर्थ वर्ग एवं सफाई कर्मचारी', '05170020105', '1.54'), ('L.Kondiya', 'चतुर्थ वर्ग एवं सफाई कर्मचारी', '05170020107', '1.54'), ('K. Damodar', 'चतुर्थ वर्ग एवं सफाई कर्मचारी', '05170020110', '1.54'), ('Mr. Kiran Kumar Sahu', 'डाटा एंट्री ऑपरेटर', 'J55172798', '22.1'), ('SUMITRA KURREY', 'डाटा एंट्री ऑपरेटर', 'J55173309', '22.1'), ('KUMARI JYOTI LAHRE', 'डाटा एंट्री ऑपरेटर', 'DME55173036', '22.1'), ('Yamini Verma', 'डाटा एंट्री ऑपरेटर', 'DME55173200', '22.1'), ('SURINDAR LAL CHANDRA', 'डाटा एंट्री ऑपरेटर', 'DME55173655', '22.1')]
            'The above main_master_c_num_with_name_amt_cat_e_code will be sent for saving in excel and db after processing'

            'DISPLAY THE NAMES AND AMOUNT'
            indent_json_print(main_master_c_num_with_name_amt_cat_e_code)

            'Excel save'
            ExcelMethods().excel_save_new(final_lol=main_master_c_num_with_name_amt_cat_e_code)
            sql_update_for_3(full_list_of_list=main_master_c_num_with_name_amt_cat_e_code)
            ColourPrint.print_green(message_box('Excel and DB saved'))

            Status().set_status(case_number=case_number, status_text=StatusText.EXCEL_DB_SAVED)

        "Updating the incentive auto 2 from db3"
        update_sql_in_auto_2_by_emp_code()

        "updating auto2 globally"
        await main_aiohttp_optimized()

        # await asyncio.sleep(100)
def get_start_index(incentive_cases_list):
    last_initiated_json_case_number, last_initiated_status_json_value = Status().get_last_initiated_done()

    for i, item in enumerate(incentive_cases_list):
        if item.startswith(last_initiated_json_case_number):
            print(f'Found {last_initiated_json_case_number} at index:', i)
            if last_initiated_status_json_value == StatusText.EXCEL_DB_SAVED:
                return i + 1
            else:
                return i

    # If not found, start from beginning
    return 0

def stuck_display():
    status_json = r'G:\Other computers\My Computer\pythonProject\svnsssy\new_svnsssy\incentive_case_stuck_status.json'
    with open(status_json, 'r') as file:
        data = json.load(file)
        incomplete = []

        for case_number, status_text in data.items():
            if status_text != StatusText.EXCEL_DB_SAVED:
                # print('case_data', case_number)
                incomplete.append(case_number)
                err_msg = "The below mentioned(s) incentive cases were not completed.\nMark them incentive query in APPROVER site and than re-initiate\n\n" + "\n".join(incomplete) + "\n\nNote: This incentive case must be first case of entry. "
                error_tk_box(error_title='Incomplete Initiated',
                             error_message= err_msg)
                raise PendingDeprecationWarning (err_msg, f'https://dkbssy.cg.nic.in/secure/incentivemodule/incentivemoduleApViewDME.aspx?ci={case_number}')

def stuck_list():
    status_json = r'G:\Other computers\My Computer\pythonProject\svnsssy\new_svnsssy\incentive_case_stuck_status.json'
    with open(status_json, 'r') as file:
        data = json.load(file)
        incomplete = []

        for case_number, status_text in data.items():
            if status_text != StatusText.EXCEL_DB_SAVED:
                # print('case_data', case_number)
                incomplete.append(case_number)
    return incomplete

def stuck_display_multi():
    incomplete = stuck_list()

    incomplete_urls = '\n'.join([f'https://dkbssy.cg.nic.in/secure/incentivemodule/incentivemoduleApViewDME.aspx?ci={case_number}' for case_number in incomplete])

    if incomplete:
        err_msg = "The below mentioned(s) incentive cases were not completed.\nMark them incentive query in APPROVER site and than re-initiate\n\n" + "\n".join(incomplete) + "\n\nNote: This incentive case must be first case of entry."
        error_tk_box(error_title='Incomplete Initiated',
                     error_message= err_msg)
        raise PendingDeprecationWarning (incomplete_urls)

async def stuck_resolver(case_number, status_text):  # INCOMPLETE NOT USED
    status_json = r'G:\Other computers\My Computer\pythonProject\svnsssy\new_svnsssy\incentive_case_stuck_status.json'
    with open(status_json, 'r') as file:
        data = json.load(file)

    ordered_status_steps = [
        StatusText.INITIATED_DONE,
        StatusText.INITIATE_BUT_QUERY_PAGE_NOT_LOADED,
        StatusText.QUERY_CLEAR_CLICKED_BUT_TIMEOUT,
        StatusText.QUERIED_CLEAR_COMPLETED,
        StatusText.REINITIATED_PAGE_NOT_LOADED,
        StatusText.REINITIATED_CLICKED_BUT_TIMEOUT,
        StatusText.REINITIATED_COMPLETED,
        StatusText.PRINT_NOT_DISPLAYED,
        StatusText.FULL_WEB_WORK_COMPLETED,
        StatusText.EXCEL_DB_SAVED
    ]

    ColourPrint.print_yellow(f"Resuming case {case_number} from status {status_text}")

    async with async_playwright() as p:
        browser = await p.chromium.connect_over_cdp("http://localhost:9222")
        context = browser.contexts[0]
        page = await context.new_page()
        page.set_default_timeout(120_000)
        page.set_default_navigation_timeout(120_000)

        try:
            # Start from the current status and proceed forward
            current_index = ordered_status_steps.index(status_text)
            for next_status in ordered_status_steps[current_index + 1:]:
                if next_status == StatusText.INITIATE_BUT_QUERY_PAGE_NOT_LOADED:
                    # means load query approver site
                    'approver site methods required'
                    url = f"https://dkbssy.cg.nic.in/secure/incentivemodule/incentivemoduleApViewDME.aspx?ci={case_number}"
                    await page.goto(url)
                    Status().set_status(case_number=case_number,
                                        status_text=StatusText.QUERY_CLEAR_CLICKED_BUT_TIMEOUT)

                elif next_status == StatusText.QUERY_CLEAR_CLICKED_BUT_TIMEOUT:
                    # means the page has not loaded in approver
                    pass


        except Exception as e:
            ColourPrint.print_bg_red(f"Error resolving case {case_number}: {str(e)}")
            # Status remains at last successful step

        finally:
            await page.close()


def update_sql_in_auto_2():
    """
    Used to update the auto-2 the names returned by sql table by matching the name and row numbers of auto-2
    :return: None
    """
    excel_meths_obj = ExcelMethods()
    excel_path = r'G:\My Drive\GdrivePC\Hospital\RSBY\New\Incentive_auto2.xlsx'
    dict_name_row_number = excel_meths_obj.retrieve_employee_id()  # function ran
    workbook_master = openpyxl.load_workbook(excel_path)
    worksheet_for_getting_id = workbook_master["Sheet3"]

    # updating from database
    results = excel_meths_obj.sqlite_process_3()
    for employee_name, total_count_till_now, total_incentive_till_now in results:
        chosen_row_number = dict_name_row_number[employee_name]
        worksheet_for_getting_id.cell(row=chosen_row_number, column=5).value = int(total_count_till_now)
        worksheet_for_getting_id.cell(row=chosen_row_number, column=6).value = float(total_incentive_till_now)
        # print(employee_name, total_count_till_now, total_incentive_till_now)

    ColourPrint.print_yellow('Database 3 Done')
    workbook_master.save(excel_path)
    workbook_master.close()

def update_sql_in_auto_2_by_emp_code():
    excel_path = r'G:\My Drive\GdrivePC\Hospital\RSBY\New\Incentive_auto2.xlsx'
    dict_ecode_row_number = get_employee_code_row_number(excel_path)  # getting the ecode and row number dict
    # print(dict_ecode_row_number)
    workbook_master = openpyxl.load_workbook(excel_path)
    worksheet_for_getting_id = workbook_master["Sheet3"]

    # updating from database
    results = sqlite_process_3_by_ecode()

    # data process for repeated double e_codes
    for_updating_excel_data = to_update_excel_after_sum_data_of_duplicated_name(DUPLICATE_NAMED_CODES=DUPLICATE_NAMED_CODES, sql_returned_data=results)

    # print(results)
    for employee_name, ecode, total_count_till_now, total_incentive_till_now in for_updating_excel_data:
        # print(type(ecode))
        # print(employee_name, ecode, total_count_till_now, total_incentive_till_now)
        chosen_row_number = dict_ecode_row_number[ecode]
        worksheet_for_getting_id.cell(row=chosen_row_number, column=5).value = int(total_count_till_now)
        worksheet_for_getting_id.cell(row=chosen_row_number, column=6).value = float(total_incentive_till_now)

    ColourPrint.print_yellow('Database 3 Done by ecode')
    workbook_master.save(excel_path)
    workbook_master.close()

    # print(dict_ecode_row_number)
    # print(len(dict_ecode_row_number))

def to_update_excel_after_sum_data_of_duplicated_name(DUPLICATE_NAMED_CODES, sql_returned_data)->list[list]:
    """
    Used for adding the same person double emp_codes datas
    :param sql_returned_data: name,ecode,onc_count_inc_amount data
    :param DUPLICATE_NAMED_CODES: e_codes of same person
    :return: final lol of names with data
    """
    # making he master ode for the other ecode
    ecode_to_master_mapped = {}
    name_of_master = {}
    for name, ecode_list in DUPLICATE_NAMED_CODES.items():
        for ecode in ecode_list:
            master_ecode = ecode_list[0]
            ecode_to_master_mapped[ecode] = master_ecode
            name_of_master[master_ecode] = name
            # print(master_ecode)
    # print(ecode_to_master_mapped)  # {'05170020188': '05170020188', 'N0517466': '05170020188', '01234': '01234', '56789': '01234', 'abcde': '01234'}
    # print(name_of_master)  # {'05170020188': 'Vandana Kanwar', '01234': 'DEMO'}
    # print()

    # summing the duplicate ecode
    summed ={}
    for name_sql, ecode_sql, count_sql, amount_sql in sql_returned_data:
        if ecode_sql in ecode_to_master_mapped:
            master_for_sql = ecode_to_master_mapped[ecode_sql]

            if master_for_sql not in summed:
                summed[master_for_sql] = [count_sql,amount_sql]
            else:
                summed[master_for_sql][0] += count_sql
                summed[master_for_sql][1] += amount_sql
    # print(summed)

    # formatting the combined data in format required for Excel
    combined_data = []
    for ecode_comb, (count_comb, amount_comb) in summed.items():
        name_comb = name_of_master[ecode_comb]
        combined_data.append([name_comb, ecode_comb, count_comb, amount_comb])
    # print(combined_data)

    final_list = []
    # removing all the items of duplicate e_codes
    for name_sql, ecode_sql, count_sql, amount_sql in sql_returned_data:
        if ecode_sql in ecode_to_master_mapped:
            continue
        final_list.append([name_sql, ecode_sql, count_sql, amount_sql])
    # print(final_list)
    # adding the duplicated after merging the other codes in final list
    # final_list.append(combined_data)
    # print(final_list)
    final_list.extend(combined_data)
    # print(final_list)

    return final_list



class StatusText(str, Enum):
    INITIATED_DONE = "initiated_done"
    INITIATE_BUT_QUERY_PAGE_NOT_LOADED = "initiateDoneButQueryPageNotLoaded"
    QUERY_CLEAR_CLICKED_BUT_TIMEOUT = "queryClearClickedButTimeout"
    QUERIED_CLEAR_COMPLETED = "queriedClearCompleted"
    REINITIATED_PAGE_NOT_LOADED = "reinitiatedPageNotLoaded"
    REINITIATED_CLICKED_BUT_TIMEOUT = "reinitiatedClickedButTimeout"
    REINITIATED_COMPLETED = "reinitiatedCompleted"
    PRINT_NOT_DISPLAYED = "printNotDisplayed"
    FULL_WEB_WORK_COMPLETED = "fullWebWorkCompleted"
    EXCEL_DB_SAVED = "excel_db_saved"

class Status:
    status_json = r'G:\Other computers\My Computer\pythonProject\svnsssy\new_svnsssy\incentive_case_stuck_status.json'

    def get_status(self, case_number):
        try:
            with open(self.status_json, 'r') as file:
                data = json.load(file)

            # Return status if case_number exists in the JSON data
            return data.get(case_number, 'Status not found')

        except FileNotFoundError:
            return "Status file not found"
        except json.JSONDecodeError:
            return "Invalid JSON file"

    def set_status(self, case_number, status_text):
        with open(self.status_json, 'r') as file:
            data = json.load(file)

        data[case_number] = status_text

        with open(self.status_json, 'w') as file:
            json.dump(data, file, indent=4)

    def delete_status(self, case_number):
        try:
            with open(self.status_json, 'r') as file:
                data = json.load(file)

            if case_number in data:
                del data[case_number]
                with open(self.status_json, 'w') as file:
                    json.dump(data, file, indent=4)
                return f"[INFO] Deleted status for case_number '{case_number}'"
            else:
                return f"[INFO] Case number '{case_number}' not found in status file."

        except FileNotFoundError:
            return "Status file not found"
        except json.JSONDecodeError:
            return "Invalid JSON file"

    def get_last_initiated_done(self):
        with open(self.status_json, 'r') as file:
            data = json.load(file)
        last_initiated = list(data.items())[-1]
        # print(last_initiated)
        return last_initiated


async def last_print(case_number):
    status = ["initiated_done", "initiateDoneButQueryPageNotLoaded", "queryClearClickedButTimeout", "queriedClearCompleted",
             "reinitiatedPageNotLoaded", "reinitiatedClickedButTimeout", "reinitiatedCompleted", "printNotDisplayed",
             "fullWebWorkCompleted", "excel_db_saved"]

    print_page_url = f'https://dkbssy.cg.nic.in/secure/incentivemodule/incentivemoduleViewPrintDME.aspx?ci={case_number}'



def process_scraped_data(scraped_data):
    # [('SURINDAR LAL CHANDRA', 'डाटा एंट्री ऑपरेटर', 'DME55173655', '22.1'), ('Yamini Verma', 'डाटा एंट्री ऑपरेटर', 'DME55173200', '22.1'), ('KUMARI JYOTI LAHRE', 'डाटा एंट्री ऑपरेटर', 'DME55173036', '22.1'), ('SUMITRA KURREY', 'डाटा एंट्री ऑपरेटर', 'J55173309', '22.1'), ('Mr. Kiran Kumar Sahu', 'डाटा एंट्री ऑपरेटर', 'J55172798', '22.1'), ('K. Damodar', 'चतुर्थ वर्ग एवं सफाई कर्मचारी', '05170020110', '1.54'), ('L.Kondiya', 'चतुर्थ वर्ग एवं सफाई कर्मचारी', '05170020107', '1.54'), ('Bal Chainya', 'चतुर्थ वर्ग एवं सफाई कर्मचारी', '05170020105', '1.54'), ('Smt. Shushila Bai', 'चतुर्थ वर्ग एवं सफाई कर्मचारी', '05170020103', '1.54'), ('Ankaiya', 'चतुर्थ वर्ग एवं सफाई कर्मचारी', '05170020093', '1.54'), ('Smt. Achamma Bai', 'चतुर्थ वर्ग एवं सफाई कर्मचारी', '05170020090', '1.54'), ('Smt. Chandrika Bai', 'चतुर्थ वर्ग एवं सफाई कर्मचारी', '05170020088', '1.54'), ('Mongra Tandon', 'चतुर्थ वर्ग एवं सफाई कर्मचारी', '04170150044', '1.54'), ('reman singh kanwar', 'चतुर्थ वर्ग एवं सफाई कर्मचारी', '07170250453', '1.54'), ('Mithlesh Kumar Markam', 'चतुर्थ वर्ग एवं सफाई कर्मचारी', '07170250399', '1.54'), ('V. Kondiya', 'चतुर्थ वर्ग एवं सफाई कर्मचारी', '05170020092', '1.54'), ('Panch Ram Nirmalkar', 'चतुर्थ वर्ग एवं सफाई कर्मचारी', '05170020091', '1.54'), ('Shri Shyam Das Mahant', 'चतुर्थ वर्ग एवं सफाई कर्मचारी', '05170020087', '1.54'), ('Shri Amar Das', 'चतुर्थ वर्ग एवं सफाई कर्मचारी', '05170020086', '1.54'), ('Shri Shiv Kumar Sarthi', 'चतुर्थ वर्ग एवं सफाई कर्मचारी', '05170020084', '1.54'), ('Shri Krishna Kumar Tripathi', 'चतुर्थ वर्ग एवं सफाई कर्मचारी', '05170020082', '1.54'), ('mohendra kumar janardan', 'चतुर्थ वर्ग एवं सफाई कर्मचारी', '04170150084', '1.54'), ('RAVINDRA KUMAR PANDEY', 'चतुर्थ वर्ग एवं सफाई कर्मचारी', '04170150035', '1.54'), ('Shri S.N.Shriwas', 'चतुर्थ वर्ग एवं सफाई कर्मचारी', '05170020100', '1.54'), ('Shri Sunil Kumar Channe', 'चतुर्थ वर्ग एवं सफाई कर्मचारी', '05170020099', '1.54'), ('Shri H.P.Tiwari', 'चतुर्थ वर्ग एवं सफाई कर्मचारी', '05170020098', '1.54'), ('Shri C.P. Chandra', 'चतुर्थ वर्ग एवं सफाई कर्मचारी', '05170020097', '1.54'), ('S SHANKAR RAO', 'चतुर्थ वर्ग एवं सफाई कर्मचारी', '05170020189', '1.54'), ('PUSHPENDRA KUMAR', 'चतुर्थ वर्ग एवं सफाई कर्मचारी', '05170020184', '1.54'), ('Smt.Lata Chandra', 'चतुर्थ वर्ग एवं सफाई कर्मचारी', '05170020137', '1.54'), ('Shri Maheshwar Prasad Tandon', 'चतुर्थ वर्ग एवं सफाई कर्मचारी', '05170020131', '1.54'), ('Anand Singh Rathiya', 'चतुर्थ वर्ग एवं सफाई कर्मचारी', 'N5517857', '1.54'), ('Smt. Sant Ram Kashyap', 'चतुर्थ वर्ग एवं सफाई कर्मचारी', '05170020095', '1.54'), ('Komal Dewangan', 'चतुर्थ वर्ग एवं सफाई कर्मचारी', 'N5517589', '1.54'), ('PRATIBHA', 'चतुर्थ वर्ग एवं सफाई कर्मचारी', 'N5517615', '1.54'), ('BAJANTI', 'चतुर्थ वर्ग एवं सफाई कर्मचारी', 'N5517614', '1.54'), ('SUNITA RATHORE', 'चतुर्थ वर्ग एवं सफाई कर्मचारी', 'N5517613', '1.54'), ('AAKANKSHA', 'चतुर्थ वर्ग एवं सफाई कर्मचारी', 'N5517610', '1.54'), ('Mrs. ANAMIKA', 'चतुर्थ वर्ग एवं सफाई कर्मचारी', 'N5517055', '1.54'), ('Mrs. BASANTI RATHIYA', 'चतुर्थ वर्ग एवं सफाई कर्मचारी', 'N5517028', '1.54'), ('KU. PRIYA', 'चतुर्थ वर्ग एवं सफाई कर्मचारी', 'N5517136', '1.54'), ('MAHENDRA KUMAR', 'चतुर्थ वर्ग एवं सफाई कर्मचारी', 'N55171003', '1.54'), ('Mr. BALDEV SINGH', 'चतुर्थ वर्ग एवं सफाई कर्मचारी', 'N5517085', '1.54'), ('Mr .AVINASH KUMAR BANJARE', 'चतुर्थ वर्ग एवं सफाई कर्मचारी', 'N5517084', '1.54'), ('Mr. MONESHWAR CHANDRA RATHIYA', 'चतुर्थ वर्ग एवं सफाई कर्मचारी', 'N5517083', '1.54'), ('KUMARI SHREMATI', 'चतुर्थ वर्ग एवं सफाई कर्मचारी', 'N5517056', '1.54'), ('RAMANAND', 'चतुर्थ वर्ग एवं सफाई कर्मचारी', 'N5517042', '1.54'), ('RAJESH', 'चतुर्थ वर्ग एवं सफाई कर्मचारी', 'N5517032', '1.54'), ('Narendra Kumar Kanwar', 'नर्सिंग एवं पैरामेडिकल स्टाफ', '04170140534', '2.95'), ('Manish Singh', 'नर्सिंग एवं पैरामेडिकल स्टाफ', '05170011626', '2.95'), ('Jyoti Porte', 'नर्सिंग एवं पैरामेडिकल स्टाफ', 'N5517990', '2.95'), ('Bhanu Priya Chauhan', 'नर्सिंग एवं पैरामेडिकल स्टाफ', 'N5517989', '2.95'), ('KIRAN CHANDRA', 'नर्सिंग एवं पैरामेडिकल स्टाफ', 'N5517979', '2.95'), ('TULSI YADAV', 'नर्सिंग एवं पैरामेडिकल स्टाफ', '0617009003', '2.95'), ('NEELIMA NISHAD', 'नर्सिंग एवं पैरामेडिकल स्टाफ', '05170050074', '2.95'), ('ASHA KINDO', 'नर्सिंग एवं पैरामेडिकल स्टाफ', '03170050126', '2.95'), ('Anita Yadav', 'नर्सिंग एवं पैरामेडिकल स्टाफ', '22170040038', '2.95'), ('Smt. Bhavana Bansiyar', 'नर्सिंग एवं पैरामेडिकल स्टाफ', '15170200090', '2.95'), ('Ku. Har Bai Khunte', 'नर्सिंग एवं पैरामेडिकल स्टाफ', '14170011702', '2.95'), ('vinita tigga', 'नर्सिंग एवं पैरामेडिकल स्टाफ', '12170050188', '2.95'), ('KU. ANJANA KUJUR', 'नर्सिंग एवं पैरामेडिकल स्टाफ', '11170260055', '2.95'), ('SAROJ RATHOR', 'नर्सिंग एवं पैरामेडिकल स्टाफ', '09170160077', '2.95'), ('Neelam Kanwar', 'नर्सिंग एवं पैरामेडिकल स्टाफ', '08170040064', '2.95'), ('Smt.Nirmala Miri', 'नर्सिंग एवं पैरामेडिकल स्टाफ', '07170250081', '2.95'), ('SMT NAMITA PATEL', 'नर्सिंग एवं पैरामेडिकल स्टाफ', '05170020176', '2.95'), ('ku Kiran Sahu', 'नर्सिंग एवं पैरामेडिकल स्टाफ', '05170020161', '2.95'), ('Ku.Marget Bishwash', 'नर्सिंग एवं पैरामेडिकल स्टाफ', '05170020153', '2.95'), ('Ku.Jyoti Gual', 'नर्सिंग एवं पैरामेडिकल स्टाफ', '05170020143', '2.95'), ('Smt Bharti Markam', 'नर्सिंग एवं पैरामेडिकल स्टाफ', '05170020142', '2.95'), ('Ku. Sandhya Nair', 'नर्सिंग एवं पैरामेडिकल स्टाफ', '05170020124', '2.95'), ('PREETI SONWANI', 'नर्सिंग एवं पैरामेडिकल स्टाफ', '05170011771', '2.95'), ('MANJULATA KANWAR', 'नर्सिंग एवं पैरामेडिकल स्टाफ', '04170150162', '2.95'), ('SUSHILATA RATHIYA', 'नर्सिंग एवं पैरामेडिकल स्टाफ', '04170140436', '2.95'), ('DIPTI LAKRA', 'नर्सिंग एवं पैरामेडिकल स्टाफ', '04170140331', '2.95'), ('Smt.DIVYA ALKA TOPPO', 'नर्सिंग एवं पैरामेडिकल स्टाफ', '04170060068', '2.95'), ('Sanjaya Kumar Tiwari', 'नर्सिंग एवं पैरामेडिकल स्टाफ', 'N5517515', '2.95'), ('PAWAN KUMAR PATEL', 'नर्सिंग एवं पैरामेडिकल स्टाफ', 'N5517809', '2.95'), ('VANDANA SAMUEL', 'नर्सिंग एवं पैरामेडिकल स्टाफ', '66170020614', '2.95'), ('CHANDRAKANTA SHRIVASTAV', 'नर्सिंग एवं पैरामेडिकल स्टाफ', '66170020507', '2.95'), ('PUSHPA BAG', 'नर्सिंग एवं पैरामेडिकल स्टाफ', '66170020500', '2.95'), ('Deepa sahu', 'नर्सिंग एवं पैरामेडिकल स्टाफ', '66170020472', '2.95'), ('Priti Masih Upma Kumar', 'नर्सिंग एवं पैरामेडिकल स्टाफ', '15170200071', '2.95'), ('SMT. PHULKUMARI TOPPO', 'नर्सिंग एवं पैरामेडिकल स्टाफ', '66170020668', '2.95'), ('Smt. Sarika Tirpude Ramteke', 'नर्सिंग एवं पैरामेडिकल स्टाफ', '66170020665', '2.95'), ('PRIYA VERMA', 'नर्सिंग एवं पैरामेडिकल स्टाफ', '66170020641', '2.95'), ('Victoria Gardia', 'नर्सिंग एवं पैरामेडिकल स्टाफ', '66170010273', '2.95'), ('Smt.Seema Rani Lajras', 'नर्सिंग एवं पैरामेडिकल स्टाफ', '05170020063', '2.95'), ('Rajni Kiran Kispotta', 'नर्सिंग एवं पैरामेडिकल स्टाफ', '11170010433', '2.95'), ('Sanjay Das Manikpuri', 'नर्सिंग एवं पैरामेडिकल स्टाफ', 'N5517599', '2.95'), ('Neelu Chaudhary', 'नर्सिंग एवं पैरामेडिकल स्टाफ', 'N5517492', '2.95'), ('Radhe Shyam Kashyap', 'नर्सिंग एवं पैरामेडिकल स्टाफ', 'AI55170022', '2.95'), ('Amarjeet Kour', 'नर्सिंग एवं पैरामेडिकल स्टाफ', '19170040056', '2.95'), ('Smt. Varnita Zilkar', 'नर्सिंग एवं पैरामेडिकल स्टाफ', '15170200084', '2.95'), ('Dr. Durga Shankar Patel', 'एनेस्थीसिया', '66170010344', '221'), ('Dr. Priyanka Ekka', 'सभी सीनियर एवं जूनियर रेसिडेंट', '05530010030', '221'), ('Dr Ankita Kapoor', 'सभी फिजिशियन / सर्जन', '05530010028', '248.63'), ('DR ARUNIKA SISODIYA', 'सभी फिजिशियन / सर्जन', '05170020182', '248.63'), ('Dr. V.Agrawal', 'सभी फिजिशियन / सर्जन', '05170020043', '248.63'), ('Dr.Sumit Gupta', 'सभी फिजिशियन / सर्जन', '05170020187', '248.63'), ('MAHENDRA KUMAR', 'पैथोलॉजी ,रेडियोलॉजी , माइक्रोबायोलॉजी , बायोकेमेस्ट्री , निश्चेतना (टेक्निशियन )', '05170011797', '2.49'), ('dinesh kumar patel', 'पैथोलॉजी ,रेडियोलॉजी , माइक्रोबायोलॉजी , बायोकेमेस्ट्री , निश्चेतना (टेक्निशियन )', '09530010377', '2.49'), ('Smt.Bhagwati Koshle', 'पैथोलॉजी ,रेडियोलॉजी , माइक्रोबायोलॉजी , बायोकेमेस्ट्री , निश्चेतना (टेक्निशियन )', '09170010257', '2.49'), ('ARUN KUMAR KANWAR', 'पैथोलॉजी ,रेडियोलॉजी , माइक्रोबायोलॉजी , बायोकेमेस्ट्री , निश्चेतना (टेक्निशियन )', '07170021387', '2.49'), ('SMT DURGESHWARI KARSH', 'पैथोलॉजी ,रेडियोलॉजी , माइक्रोबायोलॉजी , बायोकेमेस्ट्री , निश्चेतना (टेक्निशियन )', '06170050083', '2.49'), ('SHIVSHANKAR SINGH KANWAR', 'पैथोलॉजी ,रेडियोलॉजी , माइक्रोबायोलॉजी , बायोकेमेस्ट्री , निश्चेतना (टेक्निशियन )', '05170060099', '2.49'), ('Shri C.L. Dixena', 'पैथोलॉजी ,रेडियोलॉजी , माइक्रोबायोलॉजी , बायोकेमेस्ट्री , निश्चेतना (टेक्निशियन )', '05170020117', '2.49'), ('Shri Sushil kumar miri', 'पैथोलॉजी ,रेडियोलॉजी , माइक्रोबायोलॉजी , बायोकेमेस्ट्री , निश्चेतना (टेक्निशियन )', '03170010366', '2.49'), ('Santosh Kumar Singh', 'पैथोलॉजी ,रेडियोलॉजी , माइक्रोबायोलॉजी , बायोकेमेस्ट्री , निश्चेतना (टेक्निशियन )', 'AI55170030', '2.49'), ('Smt. Reena Verma', 'पैथोलॉजी ,रेडियोलॉजी , माइक्रोबायोलॉजी , बायोकेमेस्ट्री , निश्चेतना (टेक्निशियन )', 'AI55170024', '2.49'), ('Smt. Pinky Singh', 'पैथोलॉजी ,रेडियोलॉजी , माइक्रोबायोलॉजी , बायोकेमेस्ट्री , निश्चेतना (टेक्निशियन )', '06170100054', '2.49'), ('Shri Dildar Sahish', 'पैथोलॉजी ,रेडियोलॉजी , माइक्रोबायोलॉजी , बायोकेमेस्ट्री , निश्चेतना (टेक्निशियन )', '05170020054', '2.49'), ('Haricharan Jangde', 'पैथोलॉजी ,रेडियोलॉजी , माइक्रोबायोलॉजी , बायोकेमेस्ट्री , निश्चेतना (टेक्निशियन )', '05170010396', '2.49'), ('Pratima Sahu', 'पैथोलॉजी ,रेडियोलॉजी , माइक्रोबायोलॉजी , बायोकेमेस्ट्री , निश्चेतना (टेक्निशियन )', 'N5517593', '2.49'), ('Geeta patel', 'पैथोलॉजी ,रेडियोलॉजी , माइक्रोबायोलॉजी , बायोकेमेस्ट्री , निश्चेतना (टेक्निशियन )', 'J55172863', '2.49'), ('ACHCHHE KUMAR PATLEY', 'पैथोलॉजी ,रेडियोलॉजी , माइक्रोबायोलॉजी , बायोकेमेस्ट्री , निश्चेतना (टेक्निशियन )', '04170150173', '2.49'), ('Dr. manoj Kumar', 'पैथोलॉजी ,रेडियोलॉजी , माइक्रोबायोलॉजी , बायोकेमेस्ट्री , ( फैकल्टी एंड रेसीडेंट )', '05530010029', '15.47'), ('Dr. Sumit Gupta', 'पैथोलॉजी ,रेडियोलॉजी , माइक्रोबायोलॉजी , बायोकेमेस्ट्री , ( फैकल्टी एंड रेसीडेंट )', 'DME55173185', '15.47'), ('Dr. Ashutosh Kumar', 'पैथोलॉजी ,रेडियोलॉजी , माइक्रोबायोलॉजी , बायोकेमेस्ट्री , ( फैकल्टी एंड रेसीडेंट )', 'DME55173188', '15.47'), ('Shende Pranali kisandas', 'पैथोलॉजी ,रेडियोलॉजी , माइक्रोबायोलॉजी , बायोकेमेस्ट्री , ( फैकल्टी एंड रेसीडेंट )', 'DME55172441', '15.47'), ('Vibha Tandon', 'पैथोलॉजी ,रेडियोलॉजी , माइक्रोबायोलॉजी , बायोकेमेस्ट्री , ( फैकल्टी एंड रेसीडेंट )', '05530010013', '15.47'), ('HANISH KUMAR CHOWDA', 'पैथोलॉजी ,रेडियोलॉजी , माइक्रोबायोलॉजी , बायोकेमेस्ट्री , ( फैकल्टी एंड रेसीडेंट )', '04170140518', '15.47'), ('Reena Nayak', 'पैथोलॉजी ,रेडियोलॉजी , माइक्रोबायोलॉजी , बायोकेमेस्ट्री , ( फैकल्टी एंड रेसीडेंट )', 'DME55172439', '15.47'), ('Dr. Veenapani Mire', 'पैथोलॉजी ,रेडियोलॉजी , माइक्रोबायोलॉजी , बायोकेमेस्ट्री , ( फैकल्टी एंड रेसीडेंट )', '05530010079', '15.47'), ('Rajesh Kumar', 'पैथोलॉजी ,रेडियोलॉजी , माइक्रोबायोलॉजी , बायोकेमेस्ट्री , ( फैकल्टी एंड रेसीडेंट )', '05530010065', '15.47'), ('Dushyant Chandra', 'पैथोलॉजी ,रेडियोलॉजी , माइक्रोबायोलॉजी , बायोकेमेस्ट्री , ( फैकल्टी एंड रेसीडेंट )', '05530010031', '15.47'), ('Dr. Deepa Janghel', 'पैथोलॉजी ,रेडियोलॉजी , माइक्रोबायोलॉजी , बायोकेमेस्ट्री , ( फैकल्टी एंड रेसीडेंट )', '05530010026', '15.47'), ('Dr. Awadh Sahu', 'पैथोलॉजी ,रेडियोलॉजी , माइक्रोबायोलॉजी , बायोकेमेस्ट्री , ( फैकल्टी एंड रेसीडेंट )', 'DME55173201', '15.47'), ('DR RAKESH KUMAR AGRAWAL', 'पैथोलॉजी ,रेडियोलॉजी , माइक्रोबायोलॉजी , बायोकेमेस्ट्री , ( फैकल्टी एंड रेसीडेंट )', '05170040053', '15.47'), ('GHANSHYAM SINGH JATRA', 'पैथोलॉजी ,रेडियोलॉजी , माइक्रोबायोलॉजी , बायोकेमेस्ट्री , ( फैकल्टी एंड रेसीडेंट )', '05170011289', '15.47'), ('AVINASH MESHRAM', 'अधिष्ठाता अस्पताल अधीक्षक ,सहायक अधीक्षक नोडल अधिकारी एवं सहायक नोडल अधिकारी , अस्पताल सलाहकार', '66170010023', '3.68'), ('RAVIKANT JATWAR', 'अधिष्ठाता अस्पताल अधीक्षक ,सहायक अधीक्षक नोडल अधिकारी एवं सहायक नोडल अधिकारी , अस्पताल सलाहकार', '04170140099', '3.68'), ('Dr. Rakesh Kumar Verma', 'अधिष्ठाता अस्पताल अधीक्षक ,सहायक अधीक्षक नोडल अधिकारी एवं सहायक नोडल अधिकारी , अस्पताल सलाहकार', '02530010055', '3.68'), ('Dr. Aditya Siodiya', 'अधिष्ठाता अस्पताल अधीक्षक ,सहायक अधीक्षक नोडल अधिकारी एवं सहायक नोडल अधिकारी , अस्पताल सलाहकार', '09530010327', '3.68'), ('DR.ANMOL MADHUR MINZ', 'अधिष्ठाता अस्पताल अधीक्षक ,सहायक अधीक्षक नोडल अधिकारी एवं सहायक नोडल अधिकारी , अस्पताल सलाहकार', '04170140656', '3.68'), ('Dr. Gopal Singh Kanwer', 'अधिष्ठाता अस्पताल अधीक्षक ,सहायक अधीक्षक नोडल अधिकारी एवं सहायक नोडल अधिकारी , अस्पताल सलाहकार', '11170010478', '3.68')]
    incentive_received_data = scraped_data[::-1]
    # [('Dr. Gopal Singh Kanwer', 'अधिष्ठाता अस्पताल अधीक्षक ,सहायक अधीक्षक नोडल अधिकारी एवं सहायक नोडल अधिकारी , अस्पताल सलाहकार', '11170010478', '3.16'), ('Dr. Durga Shankar Patel', 'अधिष्ठाता अस्पताल अधीक्षक ,सहायक अधीक्षक नोडल अधिकारी एवं सहायक नोडल अधिकारी , अस्पताल सलाहकार', '66170010344', '3.16'), ('DR.ANMOL MADHUR MINZ', 'अधिष्ठाता अस्पताल अधीक्षक ,सहायक अधीक्षक नोडल अधिकारी एवं सहायक नोडल अधिकारी , अस्पताल सलाहकार', '04170140656', '3.16'), ('Dr. Aditya Siodiya', 'अधिष्ठाता अस्पताल अधीक्षक ,सहायक अधीक्षक नोडल अधिकारी एवं सहायक नोडल अधिकारी , अस्पताल सलाहकार', '09530010327', '3.16'), ('Dr. Rakesh Kumar Verma', 'अधिष्ठाता अस्पताल अधीक्षक ,सहायक अधीक्षक नोडल अधिकारी एवं सहायक नोडल अधिकारी , अस्पताल सलाहकार', '02530010055', '3.16'), ('RAVIKANT JATWAR', 'अधिष्ठाता अस्पताल अधीक्षक ,सहायक अधीक्षक नोडल अधिकारी एवं सहायक नोडल अधिकारी , अस्पताल सलाहकार', '04170140099', '3.16'), ('AVINASH MESHRAM', 'अधिष्ठाता अस्पताल अधीक्षक ,सहायक अधीक्षक नोडल अधिकारी एवं सहायक नोडल अधिकारी , अस्पताल सलाहकार', '66170010023', '3.16'), ('GHANSHYAM SINGH JATRA', 'पैथोलॉजी ,रेडियोलॉजी , माइक्रोबायोलॉजी , बायोकेमेस्ट्री , ( फैकल्टी एंड रेसीडेंट )', '05170011289', '15.47'), ('DR RAKESH KUMAR AGRAWAL', 'पैथोलॉजी ,रेडियोलॉजी , माइक्रोबायोलॉजी , बायोकेमेस्ट्री , ( फैकल्टी एंड रेसीडेंट )', '05170040053', '15.47'), ('Dr. Awadh Sahu', 'पैथोलॉजी ,रेडियोलॉजी , माइक्रोबायोलॉजी , बायोकेमेस्ट्री , ( फैकल्टी एंड रेसीडेंट )', 'DME55173201', '15.47'), ('Dr. Deepa Janghel', 'पैथोलॉजी ,रेडियोलॉजी , माइक्रोबायोलॉजी , बायोकेमेस्ट्री , ( फैकल्टी एंड रेसीडेंट )', '05530010026', '15.47'), ('Dushyant Chandra', 'पैथोलॉजी ,रेडियोलॉजी , माइक्रोबायोलॉजी , बायोकेमेस्ट्री , ( फैकल्टी एंड रेसीडेंट )', '05530010031', '15.47'), ('Rajesh Kumar', 'पैथोलॉजी ,रेडियोलॉजी , माइक्रोबायोलॉजी , बायोकेमेस्ट्री , ( फैकल्टी एंड रेसीडेंट )', '05530010065', '15.47'), ('Dr. Veenapani Mire', 'पैथोलॉजी ,रेडियोलॉजी , माइक्रोबायोलॉजी , बायोकेमेस्ट्री , ( फैकल्टी एंड रेसीडेंट )', '05530010079', '15.47'), ('Reena Nayak', 'पैथोलॉजी ,रेडियोलॉजी , माइक्रोबायोलॉजी , बायोकेमेस्ट्री , ( फैकल्टी एंड रेसीडेंट )', 'DME55172439', '15.47'), ('HANISH KUMAR CHOWDA', 'पैथोलॉजी ,रेडियोलॉजी , माइक्रोबायोलॉजी , बायोकेमेस्ट्री , ( फैकल्टी एंड रेसीडेंट )', '04170140518', '15.47'), ('Vibha Tandon', 'पैथोलॉजी ,रेडियोलॉजी , माइक्रोबायोलॉजी , बायोकेमेस्ट्री , ( फैकल्टी एंड रेसीडेंट )', '05530010013', '15.47'), ('Shende Pranali kisandas', 'पैथोलॉजी ,रेडियोलॉजी , माइक्रोबायोलॉजी , बायोकेमेस्ट्री , ( फैकल्टी एंड रेसीडेंट )', 'DME55172441', '15.47'), ('Dr. Ashutosh Kumar', 'पैथोलॉजी ,रेडियोलॉजी , माइक्रोबायोलॉजी , बायोकेमेस्ट्री , ( फैकल्टी एंड रेसीडेंट )', 'DME55173188', '15.47'), ('Dr. Sumit Gupta', 'पैथोलॉजी ,रेडियोलॉजी , माइक्रोबायोलॉजी , बायोकेमेस्ट्री , ( फैकल्टी एंड रेसीडेंट )', 'DME55173185', '15.47'), ('Dr. manoj Kumar', 'पैथोलॉजी ,रेडियोलॉजी , माइक्रोबायोलॉजी , बायोकेमेस्ट्री , ( फैकल्टी एंड रेसीडेंट )', '05530010029', '15.47'), ('ACHCHHE KUMAR PATLEY', 'पैथोलॉजी ,रेडियोलॉजी , माइक्रोबायोलॉजी , बायोकेमेस्ट्री , निश्चेतना (टेक्निशियन )', '04170150173', '2.49'), ('Geeta patel', 'पैथोलॉजी ,रेडियोलॉजी , माइक्रोबायोलॉजी , बायोकेमेस्ट्री , निश्चेतना (टेक्निशियन )', 'J55172863', '2.49'), ('Pratima Sahu', 'पैथोलॉजी ,रेडियोलॉजी , माइक्रोबायोलॉजी , बायोकेमेस्ट्री , निश्चेतना (टेक्निशियन )', 'N5517593', '2.49'), ('Haricharan Jangde', 'पैथोलॉजी ,रेडियोलॉजी , माइक्रोबायोलॉजी , बायोकेमेस्ट्री , निश्चेतना (टेक्निशियन )', '05170010396', '2.49'), ('Shri Dildar Sahish', 'पैथोलॉजी ,रेडियोलॉजी , माइक्रोबायोलॉजी , बायोकेमेस्ट्री , निश्चेतना (टेक्निशियन )', '05170020054', '2.49'), ('Smt. Pinky Singh', 'पैथोलॉजी ,रेडियोलॉजी , माइक्रोबायोलॉजी , बायोकेमेस्ट्री , निश्चेतना (टेक्निशियन )', '06170100054', '2.49'), ('Smt. Reena Verma', 'पैथोलॉजी ,रेडियोलॉजी , माइक्रोबायोलॉजी , बायोकेमेस्ट्री , निश्चेतना (टेक्निशियन )', 'AI55170024', '2.49'), ('Santosh Kumar Singh', 'पैथोलॉजी ,रेडियोलॉजी , माइक्रोबायोलॉजी , बायोकेमेस्ट्री , निश्चेतना (टेक्निशियन )', 'AI55170030', '2.49'), ('Shri Sushil kumar miri', 'पैथोलॉजी ,रेडियोलॉजी , माइक्रोबायोलॉजी , बायोकेमेस्ट्री , निश्चेतना (टेक्निशियन )', '03170010366', '2.49'), ('Shri C.L. Dixena', 'पैथोलॉजी ,रेडियोलॉजी , माइक्रोबायोलॉजी , बायोकेमेस्ट्री , निश्चेतना (टेक्निशियन )', '05170020117', '2.49'), ('SHIVSHANKAR SINGH KANWAR', 'पैथोलॉजी ,रेडियोलॉजी , माइक्रोबायोलॉजी , बायोकेमेस्ट्री , निश्चेतना (टेक्निशियन )', '05170060099', '2.49'), ('SMT DURGESHWARI KARSH', 'पैथोलॉजी ,रेडियोलॉजी , माइक्रोबायोलॉजी , बायोकेमेस्ट्री , निश्चेतना (टेक्निशियन )', '06170050083', '2.49'), ('ARUN KUMAR KANWAR', 'पैथोलॉजी ,रेडियोलॉजी , माइक्रोबायोलॉजी , बायोकेमेस्ट्री , निश्चेतना (टेक्निशियन )', '07170021387', '2.49'), ('Smt.Bhagwati Koshle', 'पैथोलॉजी ,रेडियोलॉजी , माइक्रोबायोलॉजी , बायोकेमेस्ट्री , निश्चेतना (टेक्निशियन )', '09170010257', '2.49'), ('dinesh kumar patel', 'पैथोलॉजी ,रेडियोलॉजी , माइक्रोबायोलॉजी , बायोकेमेस्ट्री , निश्चेतना (टेक्निशियन )', '09530010377', '2.49'), ('MAHENDRA KUMAR', 'पैथोलॉजी ,रेडियोलॉजी , माइक्रोबायोलॉजी , बायोकेमेस्ट्री , निश्चेतना (टेक्निशियन )', '05170011797', '2.49'), ('Dr.Sumit Gupta', 'सभी फिजिशियन / सर्जन', '05170020187', '198.9'), ('Dr. V.Agrawal', 'सभी फिजिशियन / सर्जन', '05170020043', '198.9'), ('DR ARUNIKA SISODIYA', 'सभी फिजिशियन / सर्जन', '05170020182', '198.9'), ('Dr Ankita Kapoor', 'सभी फिजिशियन / सर्जन', '05530010028', '198.9'), ('Dr. Priyanka Ekka', 'सभी फिजिशियन / सर्जन', '05530010030', '198.9'), ('Smt. Varnita Zilkar', 'नर्सिंग एवं पैरामेडिकल स्टाफ', '15170200084', '2.95'), ('Amarjeet Kour', 'नर्सिंग एवं पैरामेडिकल स्टाफ', '19170040056', '2.95'), ('Radhe Shyam Kashyap', 'नर्सिंग एवं पैरामेडिकल स्टाफ', 'AI55170022', '2.95'), ('Neelu Chaudhary', 'नर्सिंग एवं पैरामेडिकल स्टाफ', 'N5517492', '2.95'), ('Sanjay Das Manikpuri', 'नर्सिंग एवं पैरामेडिकल स्टाफ', 'N5517599', '2.95'), ('Rajni Kiran Kispotta', 'नर्सिंग एवं पैरामेडिकल स्टाफ', '11170010433', '2.95'), ('Smt.Seema Rani Lajras', 'नर्सिंग एवं पैरामेडिकल स्टाफ', '05170020063', '2.95'), ('Victoria Gardia', 'नर्सिंग एवं पैरामेडिकल स्टाफ', '66170010273', '2.95'), ('PRIYA VERMA', 'नर्सिंग एवं पैरामेडिकल स्टाफ', '66170020641', '2.95'), ('Smt. Sarika Tirpude Ramteke', 'नर्सिंग एवं पैरामेडिकल स्टाफ', '66170020665', '2.95'), ('SMT. PHULKUMARI TOPPO', 'नर्सिंग एवं पैरामेडिकल स्टाफ', '66170020668', '2.95'), ('Priti Masih Upma Kumar', 'नर्सिंग एवं पैरामेडिकल स्टाफ', '15170200071', '2.95'), ('Deepa sahu', 'नर्सिंग एवं पैरामेडिकल स्टाफ', '66170020472', '2.95'), ('PUSHPA BAG', 'नर्सिंग एवं पैरामेडिकल स्टाफ', '66170020500', '2.95'), ('CHANDRAKANTA SHRIVASTAV', 'नर्सिंग एवं पैरामेडिकल स्टाफ', '66170020507', '2.95'), ('VANDANA SAMUEL', 'नर्सिंग एवं पैरामेडिकल स्टाफ', '66170020614', '2.95'), ('PAWAN KUMAR PATEL', 'नर्सिंग एवं पैरामेडिकल स्टाफ', 'N5517809', '2.95'), ('Sanjaya Kumar Tiwari', 'नर्सिंग एवं पैरामेडिकल स्टाफ', 'N5517515', '2.95'), ('Smt.DIVYA ALKA TOPPO', 'नर्सिंग एवं पैरामेडिकल स्टाफ', '04170060068', '2.95'), ('DIPTI LAKRA', 'नर्सिंग एवं पैरामेडिकल स्टाफ', '04170140331', '2.95'), ('SUSHILATA RATHIYA', 'नर्सिंग एवं पैरामेडिकल स्टाफ', '04170140436', '2.95'), ('MANJULATA KANWAR', 'नर्सिंग एवं पैरामेडिकल स्टाफ', '04170150162', '2.95'), ('PREETI SONWANI', 'नर्सिंग एवं पैरामेडिकल स्टाफ', '05170011771', '2.95'), ('Ku. Sandhya Nair', 'नर्सिंग एवं पैरामेडिकल स्टाफ', '05170020124', '2.95'), ('Smt Bharti Markam', 'नर्सिंग एवं पैरामेडिकल स्टाफ', '05170020142', '2.95'), ('Ku.Jyoti Gual', 'नर्सिंग एवं पैरामेडिकल स्टाफ', '05170020143', '2.95'), ('Ku.Marget Bishwash', 'नर्सिंग एवं पैरामेडिकल स्टाफ', '05170020153', '2.95'), ('ku Kiran Sahu', 'नर्सिंग एवं पैरामेडिकल स्टाफ', '05170020161', '2.95'), ('SMT NAMITA PATEL', 'नर्सिंग एवं पैरामेडिकल स्टाफ', '05170020176', '2.95'), ('Smt.Nirmala Miri', 'नर्सिंग एवं पैरामेडिकल स्टाफ', '07170250081', '2.95'), ('Neelam Kanwar', 'नर्सिंग एवं पैरामेडिकल स्टाफ', '08170040064', '2.95'), ('SAROJ RATHOR', 'नर्सिंग एवं पैरामेडिकल स्टाफ', '09170160077', '2.95'), ('KU. ANJANA KUJUR', 'नर्सिंग एवं पैरामेडिकल स्टाफ', '11170260055', '2.95'), ('vinita tigga', 'नर्सिंग एवं पैरामेडिकल स्टाफ', '12170050188', '2.95'), ('Ku. Har Bai Khunte', 'नर्सिंग एवं पैरामेडिकल स्टाफ', '14170011702', '2.95'), ('Smt. Bhavana Bansiyar', 'नर्सिंग एवं पैरामेडिकल स्टाफ', '15170200090', '2.95'), ('Anita Yadav', 'नर्सिंग एवं पैरामेडिकल स्टाफ', '22170040038', '2.95'), ('ASHA KINDO', 'नर्सिंग एवं पैरामेडिकल स्टाफ', '03170050126', '2.95'), ('NEELIMA NISHAD', 'नर्सिंग एवं पैरामेडिकल स्टाफ', '05170050074', '2.95'), ('TULSI YADAV', 'नर्सिंग एवं पैरामेडिकल स्टाफ', '0617009003', '2.95'), ('KIRAN CHANDRA', 'नर्सिंग एवं पैरामेडिकल स्टाफ', 'N5517979', '2.95'), ('Bhanu Priya Chauhan', 'नर्सिंग एवं पैरामेडिकल स्टाफ', 'N5517989', '2.95'), ('Jyoti Porte', 'नर्सिंग एवं पैरामेडिकल स्टाफ', 'N5517990', '2.95'), ('Manish Singh', 'नर्सिंग एवं पैरामेडिकल स्टाफ', '05170011626', '2.95'), ('Narendra Kumar Kanwar', 'नर्सिंग एवं पैरामेडिकल स्टाफ', '04170140534', '2.95'), ('RAJESH', 'चतुर्थ वर्ग एवं सफाई कर्मचारी', 'N5517032', '1.54'), ('RAMANAND', 'चतुर्थ वर्ग एवं सफाई कर्मचारी', 'N5517042', '1.54'), ('KUMARI SHREMATI', 'चतुर्थ वर्ग एवं सफाई कर्मचारी', 'N5517056', '1.54'), ('Mr. MONESHWAR CHANDRA RATHIYA', 'चतुर्थ वर्ग एवं सफाई कर्मचारी', 'N5517083', '1.54'), ('Mr .AVINASH KUMAR BANJARE', 'चतुर्थ वर्ग एवं सफाई कर्मचारी', 'N5517084', '1.54'), ('Mr. BALDEV SINGH', 'चतुर्थ वर्ग एवं सफाई कर्मचारी', 'N5517085', '1.54'), ('MAHENDRA KUMAR', 'चतुर्थ वर्ग एवं सफाई कर्मचारी', 'N55171003', '1.54'), ('KU. PRIYA', 'चतुर्थ वर्ग एवं सफाई कर्मचारी', 'N5517136', '1.54'), ('Mrs. BASANTI RATHIYA', 'चतुर्थ वर्ग एवं सफाई कर्मचारी', 'N5517028', '1.54'), ('Mrs. ANAMIKA', 'चतुर्थ वर्ग एवं सफाई कर्मचारी', 'N5517055', '1.54'), ('AAKANKSHA', 'चतुर्थ वर्ग एवं सफाई कर्मचारी', 'N5517610', '1.54'), ('SUNITA RATHORE', 'चतुर्थ वर्ग एवं सफाई कर्मचारी', 'N5517613', '1.54'), ('BAJANTI', 'चतुर्थ वर्ग एवं सफाई कर्मचारी', 'N5517614', '1.54'), ('PRATIBHA', 'चतुर्थ वर्ग एवं सफाई कर्मचारी', 'N5517615', '1.54'), ('Komal Dewangan', 'चतुर्थ वर्ग एवं सफाई कर्मचारी', 'N5517589', '1.54'), ('Smt. Sant Ram Kashyap', 'चतुर्थ वर्ग एवं सफाई कर्मचारी', '05170020095', '1.54'), ('Anand Singh Rathiya', 'चतुर्थ वर्ग एवं सफाई कर्मचारी', 'N5517857', '1.54'), ('Shri Maheshwar Prasad Tandon', 'चतुर्थ वर्ग एवं सफाई कर्मचारी', '05170020131', '1.54'), ('Smt.Lata Chandra', 'चतुर्थ वर्ग एवं सफाई कर्मचारी', '05170020137', '1.54'), ('PUSHPENDRA KUMAR', 'चतुर्थ वर्ग एवं सफाई कर्मचारी', '05170020184', '1.54'), ('S SHANKAR RAO', 'चतुर्थ वर्ग एवं सफाई कर्मचारी', '05170020189', '1.54'), ('Shri C.P. Chandra', 'चतुर्थ वर्ग एवं सफाई कर्मचारी', '05170020097', '1.54'), ('Shri H.P.Tiwari', 'चतुर्थ वर्ग एवं सफाई कर्मचारी', '05170020098', '1.54'), ('Shri Sunil Kumar Channe', 'चतुर्थ वर्ग एवं सफाई कर्मचारी', '05170020099', '1.54'), ('Shri S.N.Shriwas', 'चतुर्थ वर्ग एवं सफाई कर्मचारी', '05170020100', '1.54'), ('RAVINDRA KUMAR PANDEY', 'चतुर्थ वर्ग एवं सफाई कर्मचारी', '04170150035', '1.54'), ('mohendra kumar janardan', 'चतुर्थ वर्ग एवं सफाई कर्मचारी', '04170150084', '1.54'), ('Shri Krishna Kumar Tripathi', 'चतुर्थ वर्ग एवं सफाई कर्मचारी', '05170020082', '1.54'), ('Shri Shiv Kumar Sarthi', 'चतुर्थ वर्ग एवं सफाई कर्मचारी', '05170020084', '1.54'), ('Shri Amar Das', 'चतुर्थ वर्ग एवं सफाई कर्मचारी', '05170020086', '1.54'), ('Shri Shyam Das Mahant', 'चतुर्थ वर्ग एवं सफाई कर्मचारी', '05170020087', '1.54'), ('Panch Ram Nirmalkar', 'चतुर्थ वर्ग एवं सफाई कर्मचारी', '05170020091', '1.54'), ('V. Kondiya', 'चतुर्थ वर्ग एवं सफाई कर्मचारी', '05170020092', '1.54'), ('Mithlesh Kumar Markam', 'चतुर्थ वर्ग एवं सफाई कर्मचारी', '07170250399', '1.54'), ('reman singh kanwar', 'चतुर्थ वर्ग एवं सफाई कर्मचारी', '07170250453', '1.54'), ('Mongra Tandon', 'चतुर्थ वर्ग एवं सफाई कर्मचारी', '04170150044', '1.54'), ('Smt. Chandrika Bai', 'चतुर्थ वर्ग एवं सफाई कर्मचारी', '05170020088', '1.54'), ('Smt. Achamma Bai', 'चतुर्थ वर्ग एवं सफाई कर्मचारी', '05170020090', '1.54'), ('Ankaiya', 'चतुर्थ वर्ग एवं सफाई कर्मचारी', '05170020093', '1.54'), ('Smt. Shushila Bai', 'चतुर्थ वर्ग एवं सफाई कर्मचारी', '05170020103', '1.54'), ('Bal Chainya', 'चतुर्थ वर्ग एवं सफाई कर्मचारी', '05170020105', '1.54'), ('L.Kondiya', 'चतुर्थ वर्ग एवं सफाई कर्मचारी', '05170020107', '1.54'), ('K. Damodar', 'चतुर्थ वर्ग एवं सफाई कर्मचारी', '05170020110', '1.54'), ('Mr. Kiran Kumar Sahu', 'डाटा एंट्री ऑपरेटर', 'J55172798', '22.1'), ('SUMITRA KURREY', 'डाटा एंट्री ऑपरेटर', 'J55173309', '22.1'), ('KUMARI JYOTI LAHRE', 'डाटा एंट्री ऑपरेटर', 'DME55173036', '22.1'), ('Yamini Verma', 'डाटा एंट्री ऑपरेटर', 'DME55173200', '22.1'), ('SURINDAR LAL CHANDRA', 'डाटा एंट्री ऑपरेटर', 'DME55173655', '22.1')]

    _category_all_value_pair_website_based = {
        'अधिष्ठाता अस्पताल अधीक्षक ,सहायक अधीक्षक नोडल अधिकारी एवं सहायक नोडल अधिकारी , अस्पताल सलाहकार': '1',
        'पैथोलॉजी ,रेडियोलॉजी , माइक्रोबायोलॉजी , बायोकेमेस्ट्री , ( फैकल्टी एंड रेसीडेंट )': '2',
        'पैथोलॉजी ,रेडियोलॉजी , माइक्रोबायोलॉजी , बायोकेमेस्ट्री , निश्चेतना (टेक्निशियन )': '3',
        'सभी फिजिशियन / सर्जन': '4', 'सभी सीनियर एवं जूनियर रेसिडेंट': '5',
        'एनेस्थीसिया': '6', 'नर्सिंग एवं पैरामेडिकल स्टाफ': '7', 'चतुर्थ वर्ग एवं सफाई कर्मचारी': '8',
        'डाटा एंट्री ऑपरेटर': '9'
    }
    processed = []
    for indi_data in incentive_received_data:
        name, hindi_cat, e_id, amt = indi_data

        cat_num = _category_all_value_pair_website_based[hindi_cat]
        detail_str = f'{name}@{amt}#{cat_num}^{e_id}'
        # print(detail_str)
        processed.append(detail_str)
    return processed



async def _scrape_amount(page:Page, case_number):
    # get table
    # Locate all rows inside tbody

    Status().set_status(case_number=case_number, status_text=StatusText.PRINT_NOT_DISPLAYED)

    await page.wait_for_selector("//tbody[@id='incentiveTableData']/tr")
    rows = await page.locator("//tbody[@id='incentiveTableData']/tr").all()
    # print(rows)

    data_list = []
    for row in rows:
        # individual_data = []
        td_data = await row.locator("td").all_inner_texts()
        # td_data -> ['1', 'डाटा एंट्री ऑपरेटर', 'Data Entry Operator(DME)', 'SURINDAR LAL CHANDRA', 'DME55173655', '22.1']
        hindi_designation = td_data[1].strip()
        name = td_data[3].strip()
        emp_id = td_data[4].strip()
        amount = td_data[5].strip()
        # individual_data.append((name, hindi_designation, emp_id, amount))
        # print(name, hindi_designation, emp_id, amount)

        data_list.append((name, hindi_designation, emp_id, amount))
    # indent_json_print(data=data_list,indent=2)
    # print(data_list[::-1])
    # print('Length of names', len(data_list))

    Status().set_status(case_number=case_number, status_text=StatusText.FULL_WEB_WORK_COMPLETED)

    return data_list

async def approver_raise_query(case_number, set_timeout_is=120_000, port =9223):
    ColourPrint.print_green("Inside Async")
    approver_page_indexes = await get_desired_page_indexes_in_cdp_async_for_ASYNC(user_title_of_page='Shaheed Veer Narayan Singh Ayushman Swasthya Yojna', cdp_port=port)
    async with async_playwright() as p:
        cdp_for_main = port
        browser = await p.chromium.connect_over_cdp(f"http://localhost:{cdp_for_main}")
        context = browser.contexts[0]
        all_pages = context.pages

        page_index = approver_page_indexes[0]  # selecting the first index of matching page
        page = all_pages[page_index]  # selecting the first PAGE of matching page

        page.set_default_timeout(set_timeout_is)
        page.set_default_navigation_timeout(set_timeout_is)

        url = f"https://dkbssy.cg.nic.in/secure/incentivemodule/incentivemoduleApViewDME.aspx?ci={case_number}"
        await page.goto(url)

        Status().set_status(case_number=case_number, status_text=StatusText.INITIATE_BUT_QUERY_PAGE_NOT_LOADED)

        await page.select_option('#ctl00_ContentPlaceHolder1_StatusTxt', '3')
        await page.locator('#ctl00_ContentPlaceHolder1_queryText').fill('REINITIATE')
        # time.sleep(5)
        await page.locator("//input[@id='ctl00_ContentPlaceHolder1_statusSubmit']").click()
        # time.sleep(5)
        Status().set_status(case_number=case_number, status_text=StatusText.QUERY_CLEAR_CLICKED_BUT_TIMEOUT)
        # await page.locator("//button[normalize-space()='OK']").is_visible()
        await page.locator("//button[normalize-space()='OK']").click()
        ColourPrint.print_green('Async done')

        Status().set_status(case_number=case_number, status_text=StatusText.QUERIED_CLEAR_COMPLETED)


async def original_site_query_modify(page:Page, case_number):
    print('Original Site - Here in reinitiate')

    Status().set_status(case_number=case_number, status_text=StatusText.REINITIATED_PAGE_NOT_LOADED)

    await page.goto(f'https://dkbssy.cg.nic.in/secure/incentivemodule/incentivemoduleQuerryViewDME_Edit.aspx?ci={case_number}')
    # Set up handler before the alert appears
    # page.on("dialog", lambda dialog: dialog.accept())

    await page.locator('//*[@id="ctl00_ContentPlaceHolder1_button_submit"]').click()  # trigger alert

    Status().set_status(case_number=case_number, status_text=StatusText.REINITIATED_CLICKED_BUT_TIMEOUT)

    try:
        await page.locator("//button[normalize-space()='OK']").click()  # wait for OK button
    except TimeoutError as e:
        ColourPrint.print_pink('TIMEOUT AT SUBMIT CLICK')
        await page.locator('//*[@id="ctl00_ContentPlaceHolder1_button_submit"]').click()
        await page.locator("//button[normalize-space()='OK']").click()

    Status().set_status(case_number=case_number, status_text=StatusText.REINITIATED_COMPLETED)


async def scrape_incentive_case_details(page):
    case_number = await page.locator('//*[@id="ctl00_ContentPlaceHolder1_caseno"]').text_content()
    depart_name_web = await page.locator('//*[@id="ctl00_ContentPlaceHolder1_dis_main_name"]').text_content()
    incentive_amount = await page.locator('//*[@id="ctl00_ContentPlaceHolder1_settledamt"]').text_content()
    patient_name = await page.locator('//*[@id="ctl00_ContentPlaceHolder1_patientName"]').text_content()
    diagnosis = await page.locator('//*[@id="ctl00_ContentPlaceHolder1_procdName"]').text_content()
    pre_auth_date = await page.locator('//*[@id="ctl00_ContentPlaceHolder1_Label1"]').text_content()
    # return case_number, depart_name_web, incentive_amount, patient_name, diagnosis, pre_auth_date
    return case_number, incentive_amount,depart_name_web,diagnosis,pre_auth_date,patient_name


async def entry_proper(page:Page, incentive_cat_and_names):
    _web_category_xpath = '//*[@id="ctl00_ContentPlaceHolder1_empCategory"]'
    _waiting_for_category_xpath = '//*[@id="ctl00_ContentPlaceHolder1_empCategory"]/option[10]'
    _web_emp_name_xpath = '//*[@id="ctl00_ContentPlaceHolder1_empName"]'
    _waiting_for_emp_name_xpath = '//*[@id="ctl00_ContentPlaceHolder1_empName"]/option[7]'
    _add_staff_button_xpath = '//input[@value="Add Staff"]'


    _category_all_value_pair = {
        'अधिष्ठाता अस्पताल अधीक्षक ,सहायक अधीक्षक नोडल अधिकारी एवं सहायक नोडल अधिकारी , अस्पताल  सलाहकार  ': '1',
        'पैथोलॉजी ,रेडियोलॉजी , माइक्रोबायोलॉजी , बायोकेमेस्ट्री , ( फैकल्टी एंड रेसीडेंट )': '2',
        'पैथोलॉजी ,रेडियोलॉजी , माइक्रोबायोलॉजी , बायोकेमेस्ट्री , निश्चेतना (टेक्निशियन )': '3',
        'सभी फिजिशियन / सर्जन ': '4', 'सभी सीनियर एवं जूनियर रेसिडेंट ': '5',
        'एनेस्थीसिया': '6', 'नर्सिंग एवं पैरामेडिकल स्टाफ ': '7', 'चतुर्थ वर्ग एवं सफाई कर्मचारी': '8',
        'डाटा एंट्री ऑपरेटर': '9'
    }

    # options_in_web = None
    # entered_names_cat_amount = []
    number = 1
    for cat_and_names in incentive_cat_and_names:
        cat = cat_and_names[0]  # hindi category
        # print('[[[[[[[[[[[[',cat)
        category_value = _category_all_value_pair[cat]  # hindi category value
        names = cat_and_names[1:]
        len_names = len(names)
        number += len_names
        percentage = inc_percent_amt_calc(cat)
        incentive_amount = float(await page.locator('//*[@id="ctl00_ContentPlaceHolder1_settledamt"]').text_content())

        if len_names:  # USED FOR NON NAMES CATEGORY

            # selection of category
            await page.select_option(_web_category_xpath,category_value)
            # print(']]]]]]]]]]]]]]] selected the cat')

            # printing all options
            await page.wait_for_selector(_waiting_for_emp_name_xpath)
            options = await page.locator(_web_emp_name_xpath).all_inner_texts()
            options_in_web = [opt.strip() for opt in options[0].split('\n')]
            # print(len(options_in_web))
            # print(options_in_web)


            # getting the names after checking the web options
            checked_names_label, checked_names_value = selection_of_names(names, options_in_web)
            # adding the names after checking
            await page.select_option(_web_emp_name_xpath, label=checked_names_label, value=checked_names_value)

            await page.get_by_text('Add Staff').click()

            await waiting_for_table(cat=cat,len_names=len_names,number=number,percentage=percentage,incentive_amount=incentive_amount,page=page,_category_all_value_pair=_category_all_value_pair)


async def submit_upload(page, upload_file_path):
    _attachment_file_xpath = '//*[@id="ctl00_ContentPlaceHolder1_FileUpload1"]'
    _submit_button_xpath = '//*[@id="ctl00_ContentPlaceHolder1_button_submit"]'
    _ok_button_xp = "//button[normalize-space()='OK']"
    await page.set_input_files(_attachment_file_xpath, upload_file_path)
    # Handle alert automatically
    # page.on("dialog", lambda dialog: dialog.accept())
    await page.locator(_submit_button_xpath).click()
    await page.locator(_ok_button_xp).click()

async def check_box_and_submit(page):
    checkbox_1_xp = '//label[@for="ctl00_ContentPlaceHolder1_CheckBox1"]'
    checkbox_2_xp = '//label[@for="ctl00_ContentPlaceHolder1_CheckBox2"]'
    checkbox_3_xp = '//label[@for="ctl00_ContentPlaceHolder1_CheckBox3"]'
    checkbox_4_xp = '//label[@for="ctl00_ContentPlaceHolder1_CheckBox4"]'
    _submit_button_xpath = '//*[@id="ctl00_ContentPlaceHolder1_button_submit"]'
    _ok_button_xp = "//button[normalize-space()='OK']"

    await page.locator(checkbox_1_xp).click()
    await page.locator(checkbox_2_xp).click()
    await page.locator(checkbox_3_xp).click()
    await page.locator(checkbox_4_xp).click()
    await page.locator(_submit_button_xpath).click()
    await page.locator(_ok_button_xp).click()


async def handle_dialog(dialog):
    print(f"Dialog appeared with message: {dialog.message}")
    await dialog.accept()


async def waiting_for_table(cat, len_names, number, percentage, incentive_amount, page,_category_all_value_pair):
    ColourPrint.print_turquoise(f'------------------- {cat} - Names: {len_names} -------------------')
    # incentive_names_amount_cat = []
    for num in range(number + 1 - len_names, number + 1):
        _web_name_xpath = f'//*[@id="ctl00_ContentPlaceHolder1_GridView1"]/tbody/tr[{num}]/td[6]'
        name_in_web_table = await page.locator(_web_name_xpath).text_content()

        # name_amount = f"{name_in_web_table.strip()}@{percentage * float(incentive_amount) / len_names:.2f}#{_category_all_value_pair[cat]}"
        name_amount = f"{name_in_web_table.strip()}@->#{_category_all_value_pair[cat]}"
        print(name_amount)


def selection_of_names(cat_name_list, web_name_options):
    to_select_value = []
    to_select_label = []
    for name in cat_name_list:
        name = name.strip()
        if name == 'MAHENDRA KUMAR HOUSE':
            to_select_value.append('N55171003')
        elif name == 'MAHENDRA KUMAR EYE':
            to_select_value.append('05170011797')
        elif name == 'MAHENDRA KUMAR':
            mahe_err = f"Unexpected employee name: 'MAHENDRA KUMAR'. Choose either MAHENDRA KUMAR EYE or MAHENDRA KUMAR HOUSE"
            error_tk_box(error_message=mahe_err)
            raise NameError(mahe_err)
        elif name not in web_name_options:
            err_msg = f'The name -> {name} is not in the options for selecting the name. Check manually the name list.'
            error_tk_box(error_message=err_msg)
            raise NameError(err_msg)
        else:
            to_select_label.append(name)

    # print(to_select)
    return to_select_label, to_select_value

def check_duplicate_names(cat_and_names_list_of_list):
    seen = []
    dupli = []
    for cat_and_name in cat_and_names_list_of_list:
        for name in cat_and_name:
            if name not in seen:
                seen.append(name)
            else:
                dupli.append(name)
    if dupli:
        err_msg = f'The name(s)->\n\n{"\n".join(dupli)} \n\nare duplicates in Excel Sheet. Remove then restart.'
        error_tk_box(error_message=err_msg)
        # print('dupli', dupli)
        raise NameError(err_msg)

def get_cat_and_names_for_entry():
    work_book = openpyxl.load_workbook(r'G:\My Drive\GdrivePC\Hospital\RSBY\New\Incentive_auto_ver_3.xlsx')
    # wb = self.work_book
    ws_1_sheet = work_book['Sheet1']
    ws_1_data = ws_1_sheet.iter_cols(min_row=1, max_row=51, min_col=3, max_col=11, values_only=True)
    ws_1_data_list = list(ws_1_data)
    # print(ws_1_data_list)
    # Filter None values inside each column
    filtered_not_none_names = [[cell for cell in col if cell is not None] for col in ws_1_data_list]
    # indent_json_print(filtered_not_none_names)
    work_book.close()

    "checking for duplicate value"
    check_duplicate_names(filtered_not_none_names)

    return filtered_not_none_names
    """[
          [
            "अधिष्ठाता अस्पताल अधीक्षक ,सहायक अधीक्षक नोडल अधिकारी एवं सहायक नोडल अधिकारी , अस्पताल  सलाहकार  ",
            "DR.ANMOL MADHUR MINZ",
            "AVINASH MESHRAM",
            "RAVIKANT JATWAR",
            "Dr. Durga Shankar Patel",
            "Dr. Rakesh Kumar Verma",
            "Dr. Gopal Singh Kanwer",
            "Dr. Aditya Siodiya"
            ],[...] ]"""


async def radio_click(page):
    # checking the radio
    await page.get_by_role('radio', name="Name Wise").check()


async def selection_of_category(page, cat_value):
    _web_category_xpath = '//*[@id="ctl00_ContentPlaceHolder1_empCategory"]'
    await page.select_option(_web_category_xpath, cat_value)



async def check_already_initiated(page, case_number):
    search_box_xpath = '//input[@type="search"]'
    searched_case_number_xp = f"//a[normalize-space()='{case_number}']"
    await page.locator(search_box_xpath).fill(case_number)
    try:
        # await page.locator(searched_case_number_xp).click()
        await page.get_by_text(case_number).click()
    except TimeoutError:
        err_msg_already_done = f"The Case Number -> {case_number} not found. It might be already Initiated."
        error_tk_box(error_message=err_msg_already_done)
        raise TimeoutError(err_msg_already_done)



async def select_department_and_dates(page, depart_name, from_date, to_date, case_number):
    second_page_url = "https://dkbssy.cg.nic.in/secure/incentivemodule/incentivemoduledme.aspx"
    speciality_DD_xpath = '//*[@id="ctl00_ContentPlaceHolder1_speciality"]'
    from_date_xpath = '//*[@id="ctl00_ContentPlaceHolder1_fdate"]'
    to_date_xpath = '//*[@id="ctl00_ContentPlaceHolder1_tdate"]'
    search_button_xpath = '//*[@id="ctl00_ContentPlaceHolder1_SearchCases"]'

    # page.set_default_timeout(300000)

    await page.goto(second_page_url)
    await page.select_option(speciality_DD_xpath, depart_name)
    await page.locator(from_date_xpath).fill(from_date)
    await page.locator(to_date_xpath).fill(to_date)
    # Correct usage of expect_response
    async with page.expect_response(
            lambda response: 'incentivemoduledme.aspx' in response.url and response.status == 200) as resp_info:
        await page.locator(search_button_xpath).click()

    response = await resp_info.value
    data = await response.text()  # or await response.json() if JSON
    if f'{case_number}</a>' in data:
        Status().delete_status(case_number=case_number)

    # print('Table data response:', data)


def check_file_path(attach_file_path):
    attach_file_path = attach_file_path.strip('\n').strip().strip('"')
    ColourPrint.print_pink('The file path is: ', attach_file_path)
    user_checked_file = input('IF ABOVE FILE LOCATION IS CORRECT, Press 1 and than press Enter: ')
    if user_checked_file != '1':
        error_tk_box(error_title="Error",
                     error_message='THE FILE PATH IS NOT CORRECTLY CONFIRMED. Press 1 and than press Enter to confirm')
        raise ValueError('THE FILE PATH IS NOT CORRECTLY CONFIRMED. Press 1 and than press Enter to confirm')
    return attach_file_path

def indent_json_print(data, indent=2):
    print(json.dumps(obj=data, indent=indent, ensure_ascii=False))


def split_multiline_to_list(all_incentive_multiline:str):
    all_incentive_multiline_split = all_incentive_multiline.split('\n')

    filter_non_blank = [i for i in all_incentive_multiline_split if not i.strip()=="" ]
    # indent_json_print(filter_blank)
    return filter_non_blank

def retrieve_case_number(each_case_number_line_data):
    # "CASE/PS6/HOSP22G146659/CK7008808\tOphthalmology\tSICS with non-foldable IOL\t2550\t07-10-2023\tMayamati"
    case_number_is, amount = each_case_number_line_data.split('\t')[0], each_case_number_line_data.split('\t')[1]

    # print("case number is:", case_number_is)
    return case_number_is, amount

class TempIncentiveSave:
    TEMP_FILE_PATH = r"G:\Other computers\My Computer\pythonProject\svnsssy\new_svnsssy\temp_cases_before_db.json"
    def save_temp_before_db(self, all_cases_data):
        # TEMP_FILE = r"G:\Other computers\My Computer\pythonProject\svnsssy\new_svnsssy\temp_cases_before_db.json"
        """Always save scraped data before DB write."""
        with open(self.TEMP_FILE_PATH, "w", encoding="utf-8") as f:
            json.dump(all_cases_data, f, indent=2, ensure_ascii=False)

    def clean_temp_file(self):
        """Clear the temp JSON file but keep the file itself."""
        with open(self.TEMP_FILE_PATH, "w", encoding="utf-8") as f:
            json.dump([], f, indent=4)  # write an empty JSON list
        print("Temp JSON content cleared.")
    def get_temp_json_data(self):
        with open (self.TEMP_FILE_PATH, "r", encoding="utf8") as f:
            return json.load(f)


def main(all_incentive_multiline:str, set_timeout_is=120000):
    asyncio.run(_main(all_incentive_multiline, set_timeout_is=set_timeout_is))



if __name__ == '__main__':
    # print(get_cat_and_names_for_entry())
    # indent_json_print(get_cat_and_names_for_entry())
    # asyncio.run(approver_raise_query(port=9223, case_number='CASE/PS6/HOSP22G146659/CB7008568'))
    # s = Status().get_status('CASE/PS6/HOSP22G146659/CK7008808')
    # Status().set_status('CASE/PS6/HOSP22G146659/CK7008808', 'initiated_done')
    # print(s)
    # s = Status().get_status('CASE/PS6/HOSP22G146659/CK7008808')
    # print(s)
    # asyncio.run(stuck_resolver())
    # print(Status().get_last_initiated_done())
    # update_sql_in_auto_2()
    # save_temp_before_db([1,2,3,4])
    update_sql_in_auto_2_by_emp_code()
    pass