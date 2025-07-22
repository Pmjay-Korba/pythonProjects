import datetime, sqlite3, sys, time, openpyxl
from selenium import webdriver
from selenium.common import NoSuchElementException
from selenium.webdriver.common.by import By
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import Select
from selenium.common.exceptions import ElementClickInterceptedException, NoSuchElementException, TimeoutException
from openpyxl.styles import PatternFill

fill_pattern_yellow = PatternFill(patternType='solid', fgColor='FFFF00')
fill_pattern_blue = PatternFill(patternType='solid', fgColor='00FFFF')

options = webdriver.ChromeOptions()
options.add_experimental_option("detach", True)
driver = webdriver.Chrome(options=options)
wait = WebDriverWait(driver, 20)

# driver.get('https://dkbssy.cg.nic.in/secure/incentivemodule/IncentiveDetails_EmpCodeWiseDME.aspx')
driver.maximize_window()


def login():
    driver.get('https://dkbssy.cg.nic.in/secure/login.aspx')
    driver.maximize_window()
    driver.find_element(By.NAME, "username").send_keys("HOSP22G146659")
    driver.find_element(By.NAME, "password").send_keys("Pmjaykorba@1")
    driver.find_element(By.ID, 'txtCaptcha').click()
    time.sleep(7)  ### SLEEP GIVEN FOR CAPTCHA ENTRY
    sign_in = driver.find_element(By.NAME, "loginbutton")
    sign_in.click()
    new_url = 'https://dkbssy.cg.nic.in/secure/incentivemodule/IncentiveDetails_EmpCodeWiseDME.aspx'
    driver.execute_script(f'window.location.href = "{new_url}"')
    time.sleep(1)
    '''new'''
    driver.execute_script(f'window.location.href = "{new_url}"')


def emp_code(code):
    try:
        # driver.get('https://dkbssy.cg.nic.in/secure/incentivemodule/IncentiveDetails_EmpCodeWiseDME.aspx')
        driver.find_element(By.NAME, 'ctl00$ContentPlaceHolder1$TextBox1').send_keys(code)
        driver.find_element(By.NAME, 'ctl00$ContentPlaceHolder1$search').click()
        new_path = f'//*[@id="ctl00_ContentPlaceHolder1_TextBox1" and @value="{code}"]'
        wait.until(EC.presence_of_element_located((By.XPATH, new_path)))

        total_amount_pending = driver.find_element(By.XPATH, '//*[@id="ctl00_ContentPlaceHolder1_Label4"]')
        total_paid_cases = driver.find_element(By.XPATH, '//*[@id="ctl00_ContentPlaceHolder1_Label14"]')
        total_paid_amount = driver.find_element(By.XPATH, '//*[@id="ctl00_ContentPlaceHolder1_Label5"]')
        try:
            number_of_cases_pending = driver.find_element(By.XPATH,
                                                          '//*[@id="ctl00_ContentPlaceHolder1_GridView1"]/tbody/tr[2]/td[3]').text
        except NoSuchElementException:
            number_of_cases_pending = 0
        print(code, '\t', number_of_cases_pending, '\t', total_amount_pending.text, '\t', total_paid_cases.text, '\t',
              total_paid_amount.text)
        driver.find_element(By.NAME, 'ctl00$ContentPlaceHolder1$TextBox1').clear()
        return code, number_of_cases_pending, total_amount_pending.text, total_paid_cases.text, total_paid_amount.text
    except TimeoutException:
        print(f"\033[93mTime Out\033[0m")
        raise TimeoutException
    except ElementClickInterceptedException:
        print(f"\033[91mThe amount not updated for {code}. Check Manually or Update it again\033[0m")


login()

workbook_auto1 = openpyxl.load_workbook(r'G:\My Drive\GdrivePC\Hospital\RSBY\New\Incentive_auto_ver_3.xlsx')

workbook_master = openpyxl.load_workbook(r'G:\My Drive\GdrivePC\Hospital\RSBY\New\Incentive_auto2.xlsx')
worksheet_for_getting_names = workbook_auto1["Sheet3"]  # id = auto 2, name = auto 3
# print(worksheet["B1:B5"])

# master amount getter workbook load
# master_amount_getter_workbook = openpyxl.load_workbook(
#     r'G:\My Drive\GdrivePC\Hospital\RSBY\New\MASTER_AMOUNT_GETTER.xlsx')
# master_amount_getter_worksheet = master_amount_getter_workbook['Sheet1']

incentive_names_to_enter = []
column_values = worksheet_for_getting_names.iter_rows(min_col=2, max_col=2, min_row=2)
for row in column_values:
    row_data = [cell.value for cell in row]
    incentive_names_to_enter += row_data

incentive_names_to_enter = [i for i in incentive_names_to_enter if i is not None]
print(incentive_names_to_enter)

# retrieving the employee ID--------
worksheet_for_getting_id = workbook_master["Sheet3"]  # id = auto 2, name = auto 3

column_names_in_sheet3 = worksheet_for_getting_id.iter_rows(min_col=2, max_col=2)  # auto2

dict_name_row_number = {}
for namez in column_names_in_sheet3:
    # print(namez[0].value)
    name_is = namez[0].value
    row_number = namez[0].row
    # print({name_is:row_number})
    dict_name_row_number[name_is] = row_number  # gives the row containing name
# print(dict_name_row_number)
counting = 1
count_names_for_check = []
for name in incentive_names_to_enter:
    name = name.strip()
    # print(dict_name_row_number.keys())
    if name in dict_name_row_number.keys():
        count_names_for_check.append(dict_name_row_number[name])
        chosen_row_number = dict_name_row_number[name]
        # print(chosen_row_number)
        chosen_emp_code = worksheet_for_getting_id.cell(row=chosen_row_number, column=3).value
        # print(chosen_emp_code)
        old_pendng_cases = worksheet_for_getting_id.cell(row=chosen_row_number, column=9).value
        chosen_amount = worksheet_for_getting_id.cell(row=chosen_row_number, column=10).value
        print(counting)
        counting += 1
        print(name, chosen_emp_code, old_pendng_cases, chosen_amount)

        '''Function run for getting details of all amount'''
        e_code, pend_cases, pend_amount, total_cases, total_amount = emp_code(chosen_emp_code)

        worksheet_for_getting_id.cell(row=chosen_row_number, column=8).value = e_code
        worksheet_for_getting_id.cell(row=chosen_row_number, column=9).value = int(pend_cases)
        worksheet_for_getting_id.cell(row=chosen_row_number, column=9).fill = fill_pattern_yellow
        worksheet_for_getting_id.cell(row=chosen_row_number, column=10).value = int(pend_amount)
        worksheet_for_getting_id.cell(row=chosen_row_number, column=10).fill = fill_pattern_blue
        worksheet_for_getting_id.cell(row=chosen_row_number, column=11).value = int(total_cases)
        worksheet_for_getting_id.cell(row=chosen_row_number, column=12).value = int(total_amount)

        # entry for amount getterttx29
        date = datetime.datetime.now().date()
        time = datetime.datetime.now().time()
        # updating the master_amount_getter
        # master_amount_getter_worksheet.append(
        #     [date, time, name, int(pend_cases), int(pend_amount), int(total_cases), int(total_amount)])
        workbook_master.save(r'G:\My Drive\GdrivePC\Hospital\RSBY\New\Incentive_auto2.xlsx')
        # master_amount_getter_workbook.save(r'G:\My Drive\GdrivePC\Hospital\RSBY\New\MASTER_AMOUNT_GETTER.xlsx')

    else:
        print(f"\033[91mTHERE OCCURRED AN ERROR, Check {name} in incentive auto 2\033[0m")
        sys.exit()

print(len(count_names_for_check), len(incentive_names_to_enter))
workbook_master.save(r'G:\My Drive\GdrivePC\Hospital\RSBY\New\Incentive_auto2.xlsx')
# master_amount_getter_workbook.save(r'G:\My Drive\GdrivePC\Hospital\RSBY\New\MASTER_AMOUNT_GETTER.xlsx')
print('Excel done')

# Connect to the SQLite database
conn = sqlite3.connect(r'G:\My Drive\GdrivePC\Hospital\RSBY\New\incentiveDatabase.db')
cursor = conn.cursor()

# Execute the SQL query
cursor.execute("""
    SELECT employee_name, COUNT(incentive_amount) AS total_count, SUM(incentive_amount) AS total_incentive
    FROM incentive_entry_T
    GROUP BY employee_name
""")

# Fetch the result
results = cursor.fetchall()  # [(n1,a1),(n2,a2)..]

# Close the cursor and connection
cursor.close()
conn.close()
# print(results)
for employee_name, total_count_till_now, total_incentive_till_now in results:
    if employee_name == "CHANDRAKANTA SHRIVASTAV":
        employee_name = "CHANDRAKANTA SHRIVASTAV"

    chosen_row_number = dict_name_row_number[employee_name]
    worksheet_for_getting_id.cell(row=chosen_row_number, column=13).value = int(total_count_till_now)
    worksheet_for_getting_id.cell(row=chosen_row_number, column=14).value = float(total_incentive_till_now)
    print(employee_name, total_count_till_now, total_incentive_till_now)
print('Database done')
workbook_master.save(r'G:\My Drive\GdrivePC\Hospital\RSBY\New\Incentive_auto2.xlsx')
print("Excel Saved")
excel_file_path = r'G:\My Drive\GdrivePC\Hospital\RSBY\New\Incentive_auto2.xlsx'
# os.system(f'start excel "{excel_file_path}"')
