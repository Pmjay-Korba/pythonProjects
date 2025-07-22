import datetime

import openpyxl


def working_date_extractor():
    workbook_auto1 = openpyxl.load_workbook(r'G:\My Drive\GdrivePC\Hospital\RSBY\Incentive_auto_ver_3.xlsx')

    workbook_master = openpyxl.load_workbook(r'G:\My Drive\GdrivePC\Hospital\RSBY\Incentive_auto_ver_3.xlsx')
    worksheet_for_getting_names = workbook_auto1["Sheet1"]
    # print(worksheet["B1:B5"])

    incentive_names_to_enter = []
    column_values = worksheet_for_getting_names.iter_rows(min_col=3, max_col=12, min_row=2, max_row=26)
    for row in column_values:
        row_data = [cell.value for cell in row]
        incentive_names_to_enter += row_data

    incentive_names_to_enter = [i for i in incentive_names_to_enter if i is not None]
    print(incentive_names_to_enter)

    # retrieving the employee ID--------
    worksheet_for_getting_id = workbook_master["Sheet3"]

    column_names_in_sheet3 = worksheet_for_getting_id.iter_rows(min_col=2, max_col=2)

    dict_name_row_number = {}
    for namez in column_names_in_sheet3:
        # print(namez[0].value)
        name_is = namez[0].value
        row_number = namez[0].row
        # print({name_is:row_number})
        dict_name_row_number[name_is] = row_number  # gives the row containing name
    # print(dict_name_row_number)

    count_names_for_check = []
    name_and_joining_date_dict={}
    for name in incentive_names_to_enter:
        name = name.strip()
        # print(dict_name_row_number.keys())
        if name in dict_name_row_number.keys():
            count_names_for_check.append(dict_name_row_number[name])
            chosen_row_number = dict_name_row_number[name]
            # print(chosen_row_number)
            chosen_emp_date = worksheet_for_getting_id.cell(row=chosen_row_number, column=8).value
            # print('.....', chosen_emp_date)
            name_and_joining_date_dict[name]=chosen_emp_date
    return name_and_joining_date_dict


def extract_date(casear):
    # print(casear)  # list of incen
    # loading the data sheet for getting the date of incentive cases
    date_workbook = openpyxl.load_workbook(r'G:\My Drive\GdrivePC\Hospital\RSBY\obgy.xlsx')
    date_worksheet = date_workbook['all']
    column_of_case_number = list(date_worksheet.iter_rows(min_col=2, max_col=2))

    date_all = []
    for case_nox in casear:
        for case_no in column_of_case_number:
            if case_nox == case_no[0].value:
                caseNo_is = case_no[0].value
                row_numberOfCaseNo = case_no[0].row
                case_no_date_is = date_worksheet.cell(row=row_numberOfCaseNo, column=6).value
                date_all.append(case_no_date_is)
    # print(date_all)
    return date_all

# casear = '''CASE/PS4/HS22005049/CB3995950
# CASE/PS4/HS22005049/CK4051870
# '''.split('\n')
# print(working_date_extractor())
# print(extract_date(casear))

def checker_with_dict_output(casear):
    dict_output = {}
    working_date_extracted_dict = working_date_extractor()
    print()
    casewise_extracted_dates_list = extract_date(casear)
    for date in casewise_extracted_dates_list:
        out = []
        for key, val in working_date_extracted_dict.items():
            if val is not None: #val=joining date
                # print('....',val)
                # print(date - val)
                if val > date:
                    # print('y')
                    # print(key)
                    out.append(key)
                    # print(out)

        dict_output[str(date)]=out
    for k, v in dict_output.items():
        print(f'{k}:{v}')
    print('Press Enter to continue')
    # return dict_output


# checker_with_dict_output(casear)