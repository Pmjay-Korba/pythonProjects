import sys
import time
from datetime import datetime
import openpyxl
import win32com.client

from dkbssy.utils.colour_prints import ColourPrint
import os


def dates_of_incentive_cases(workbook_name, sheet_name) -> list[str | float | datetime]:
    """
    Getting the dates of the incentive cases from the department wise Excel
    :param workbook_name: department wise Excel workbook name
    :param sheet_name: sheet name where the incentive cases located. Typically, Sheet4
    :return: list of dates in datetime
    """
    workbook = openpyxl.load_workbook(workbook_name, data_only=True)
    worksheet = workbook[sheet_name]
    col_data = worksheet.iter_cols(min_row=2, min_col=4, max_col=4, values_only=True)
    # print('---===--==', list(col_data)[0])
    date_data_tuple =list(col_data)[0]
    # print('=-=-=-=-', date_data_tuple)
    date_data = [d_data for d_data in date_data_tuple if d_data is not None]
    # print('date_datas', date_data)
    return date_data


class IncentiveNameGenerator:
    def __init__(self, workbook_path):
        """
        Check if the Excel file is open, if not, prompt user to open it.
        """
        self.workbook_path = os.path.abspath(workbook_path)
        self.workbook_name = os.path.basename(workbook_path)
        self._open_excel_and_auto_update_links()
        self.workbook = openpyxl.load_workbook(self.workbook_path, data_only=True)

    def _open_excel_and_auto_update_links(self):
        """
        Open Excel and automatically update external links without prompt.
        """
        excel = win32com.client.Dispatch("Excel.Application")
        excel.Visible = True
        excel.DisplayAlerts = False  # suppress popups (like your screenshot)

        # UpdateLinks=3 → update both external and remote references
        wb = excel.Workbooks.Open(self.workbook_path, UpdateLinks=3)
        time.sleep(5)
        # Optionally:
        wb.Save()
        print('✅ Excel file saved.')

        time.sleep(1)  # Optional pause before closing
        wb.Close(SaveChanges=False)  # Already saved manually

        # excel.Quit()
        # print('✅ Excel closed completely.')

    def get_all_sheet_data(self, sheet_name) -> list[tuple[str | float | datetime | None, ...]]:
        """
        Getting all the data of a sheet
        :param sheet_name:  name for which data need to generated
        :return: list[tuple of each employee details]
        """
        work_sheet = self.workbook[sheet_name]
        row_data = list(work_sheet.iter_rows(max_col=26, values_only=True))
        return row_data


    def filter_unwanted_names(self, list_with_unwanted_names, unwanted_post_list):
        """
        Using this for removing names in counselor, Psychologist etc.
        :param list_with_unwanted_names: master list with all names including the unwanted
        :param unwanted_post_list: dict with name and corresponding category in which it is not required
        :return: unwanted removed master list
        """
        '''(3, 'Dr. Durga Shankar Patel', 66170010344, '6,1', 'Professor', 'ANAES', 'ANAES', datetime.datetime(2022, 3, 3, 0, 0), 2106, 28511, 9898, 474434, 12004, 502945, 10275, 431551.19, 432444.1703000001, None, None, None, None, None, None, None, None, None)]'''
        # unwanted_post_list = ['Counsellor', 'Psychologist', 'Health Assistant']
        # unwanted_post_list = []

        filtered_unwanted_list =[i for i in list_with_unwanted_names if i[4] not in unwanted_post_list]
        # ColourPrint.print_pink(filtered_unwanted_list)
        return filtered_unwanted_list


    def filter_category(self, sheet_name, category_number, unwanted_post_list, depart_name_for_gen=None, sort_by_amount=True) -> list[
        tuple]:
        """
        Filtering according to provided category
        :param sort_by_amount: sort by ascending order
        :param unwanted_post_list: unwanted posts
        :param depart_name_for_gen: department for which entry to be made
        :param sheet_name: Sheet Name
        :param category_number: category's number
        :return: filtered list
        """
        all_sheet_data_return = self.get_all_sheet_data(sheet_name)
        '''[('Column1', 'Employee Name', 'Employee Code', 'Category', 'Post Name', 'Posting', 'Depart', 'Cases', 'Pend Cases', 'Pend Amount', 'Paid Cases', 'Paid Amount', 'Web Tot cases', 'Web Tot Amount', 'DB1Tot Case entry', 'DB1Tot amount entry', 'DB2Tot Case 2', 'DB2Tot Amount 2', temp_db_amount, ' ', None, None, None, None, None, None), <- OLD COLUMN IN NEW DB1 DB2 DELETED AND DB3 INCLUDED
            (3, 'Dr. Durga Shankar Patel', 66170010344, '6,1', 'Professor', 'ANAES', 'ANAES', datetime.datetime(2022, 3, 3, 0, 0), 8154, 403446, 1750, 71397, 9904, 474843, 8229, 403718.3462428572, 7793, 399197.5933333333, None, None, None, None, None, None, None, None), ...]'''
        filtered_row_category_data = [row_data for row_data in all_sheet_data_return if
                                      str(category_number) in str(row_data[3])]
        ColourPrint.print_green(filtered_row_category_data)

        'filtering the unwanted category'
        filtered_row_category_data = self.filter_unwanted_names(list_with_unwanted_names=filtered_row_category_data,unwanted_post_list=unwanted_post_list)
        # ColourPrint.print_pink(filtered_row_category_data)

        # sys.exit('stopped')

        if sort_by_amount:
            # ColourPrint.print_yellow(filtered_row_category_data[18])
            filtered_row_category_data.sort(key=lambda x: x[16])  # by temp_db_amount
        if depart_name_for_gen is None:
            return filtered_row_category_data
        else:
            if category_number == 4 or category_number == 5:
                filtered = [item for item in filtered_row_category_data if depart_name_for_gen in item[6]]
                # print('////////////////////', category_number, filtered)
                return filtered
            else:
                return filtered_row_category_data

    def depart_name_for_entry(self, sheet_name, cell_value):
        """
        Getting the cell value of the cell_value i.e. department name
        :param sheet_name: Sheet name
        :param cell_value: Cell where the department name selection has been made in Sheet3 - 'B5'
        :return: Sheet cell actual value
        """
        worksheet = self.workbook[sheet_name]
        depart_name = worksheet[cell_value].value
        return depart_name

    # def filter_category_by_department(self, sheet_name, category_number, depart_name_for_gen):
    #     all_category_list = self.filter_category(sheet_name, category_number, None)
    #     filtered = [item for item in all_category_list if depart_name_for_gen in item[6]]
    #     return filtered

    def _working_periods(self, sheet_name):
        """
        Returns working periods and leaves
        :param sheet_name: Sheet name:- Sheet2
        :return: Generator object
        """
        worksheet = self.workbook[sheet_name]
        row_data = list(worksheet.iter_rows(values_only=True))
        header = row_data[0]
        dict_data = [dict(zip(header, row)) for row in row_data[1:]]

        return dict_data

    def _is_eligible(self, each_emp_data:dict):  # emp_data = working_employee_period_dict
        """
        Check the Joining and Relieving dates are nor 'NOT ELIGIBLE'
        :param each_emp_data: 'Sheet2' employee row data
        :return: Bool
        """
        is_eligible = False
        '''working_employee_period_dict = {'Name_After': 'chandrakanta kashyap', 'Name_with_spaces': 'chandrakanta kashyap', '
                            Post': 'Staff Nurse (Govt.)', 'Serial number': 194, 'JOINING DATE': datetime.datetime(2022, 7, 12, 0, 0), 
                            'RELIEVING DATE': datetime.datetime(2023, 12, 31, 0, 0), 'LEAVE START1': datetime.datetime(2022, 8, 1, 0, 0), 
                            'LEAVE END1': datetime.datetime(2023, 1, 31, 0, 0), 'LEAVE START2': None, 'LEAVE END2': None, 'LEAVE START3': None, 
                            'LEAVE END3': None, 'EmpID': 15170200070, 'Working Days': 334, 'Percentage': 64.60348162475822, None: None},'''
        if each_emp_data['JOINING DATE'] != 'NOT ELIGIBLE' and each_emp_data['RELIEVING DATE'] != 'NOT ELIGIBLE':
            is_eligible = True

        return is_eligible

    def _is_on_leave(self, date_of_incentive, each_emp_data:dict) -> bool:  # emp_data = working_employee_period_dict
        """
        returns if employee is on leave on incentive date
        :param date_of_incentive: date of incentive
        :param each_emp_data: each employee row data
        :return: True if on leave.
        special mention: EXIT PROGRAM IF LEAVES NOT UPDATED
        """
        _1_leave, _1_return = each_emp_data['LEAVE START1'] ,each_emp_data['LEAVE END1']
        _2_leave, _2_return = each_emp_data['LEAVE START2'] ,each_emp_data['LEAVE END2']
        _3_leave, _3_return = each_emp_data['LEAVE START3'] ,each_emp_data['LEAVE END3']

        leave_periods_pairs = _1_leave, _1_return, _2_leave, _2_return, _3_leave, _3_return

        for non_work_day in range(0, len(leave_periods_pairs), 2):
            leave_date = leave_periods_pairs[non_work_day]
            return_date = leave_periods_pairs[non_work_day + 1]
            if leave_date and return_date:  # non None data
                if leave_date < date_of_incentive < return_date:
                    # print('Name =', each_emp_data["Name_After"], 'Leave dates= ', leave_date, 'return date= ', return_date)
                    return  True

            elif (leave_date and not return_date) or (not leave_date and return_date):
                ColourPrint.print_bg_red(f'Leave data not updated for {each_emp_data["Name_After"]}')
                sys.exit('UPDATE DATA AND RE-RUN')
        return False

    def collect_names_by_each_category(self,
                                       sheet_name_is_sheet3,
                                       sheet_name_is_sheet2,
                                       category_number,
                                       incentive_dates_list_data,
                                       number_of_names: int | None = None,
                                       depart_name_for_gen=None,
                                       sort_by_amount=True):


        filtered_names_sheet3 = self.filter_category(sheet_name=sheet_name_is_sheet3, category_number=category_number, depart_name_for_gen=depart_name_for_gen, sort_by_amount=sort_by_amount)
        emp_working_periods_sheet2 = self._working_periods(sheet_name=sheet_name_is_sheet2)  #  generator object

        # number of names redefining
        if number_of_names is None:
            number_of_names = len(filtered_names_sheet3)
        print('Number of names >>>>>>>>>>>>> ', number_of_names)

        collected_names = []
        # print('----------------', len(incentive_dates_list_data))
        set_list_sorted = sorted(list(set(incentive_dates_list_data)))
        # print('===============', len(set_list_sorted), set_list_sorted)
        for inc_date in set_list_sorted:
            date_wise_names = [inc_date]  # list for collecting names

            ColourPrint.print_bg_red(str(inc_date))
            for employee_in_sheet3 in filtered_names_sheet3:

                # checking the number of names to be in a particular category
                if len(date_wise_names) == number_of_names + 1:  # +1 for correcting the date added at start in list
                    break

                for working_employee_period_dict in emp_working_periods_sheet2:
                    '''working_employee_period_dict = {'Name_After': 'chandrakanta kashyap', 'Name_with_spaces': 'chandrakanta kashyap', '
                    Post': 'Staff Nurse (Govt.)', 'Serial number': 194, 'JOINING DATE': datetime.datetime(2022, 7, 12, 0, 0), 
                    'RELIEVING DATE ': datetime.datetime(2023, 12, 31, 0, 0), 'LEAVE START1': datetime.datetime(2022, 8, 1, 0, 0), 
                    'LEAVE END1': datetime.datetime(2023, 1, 31, 0, 0), 'LEAVE START2': None, 'LEAVE END2': None, 'LEAVE START3': None, 
                    'LEAVE END3': None, 'EmpID': 15170200070, 'Working Days': 334, 'Percentage': 64.60348162475822, None: None},'''

                    if employee_in_sheet3[1] == working_employee_period_dict['Name_After']:  # name matching in sheet 3 and sheet 2
                        """
                        till above the sheet 3 and sheet 2 names and working periods have been matched and extracted.
                        next is matching the dates of incentive cases to the working periods
                        """

                        # ColourPrint.print_green(f'{working_employee_period_dict["Name_After"]=},{working_employee_period_dict["JOINING DATE"]=}, {working_employee_period_dict["RELIEVING DATE"]}')
                        if self._is_eligible(working_employee_period_dict):
                            if working_employee_period_dict['JOINING DATE'] < inc_date < working_employee_period_dict["RELIEVING DATE"]:
                                # ColourPrint.print_blue(f'Not Employed - {employee_in_sheet3[1]}, {working_employee_period_dict['Name_After']}')
                                """
                                now checking weather in leave or not
                                """
                                if not self._is_on_leave(date_of_incentive=inc_date, each_emp_data=working_employee_period_dict):
                                    date_wise_names.append(employee_in_sheet3[1])

                            # else:  # get the names not included here
                            #     print(employee_in_sheet3[1])
            collected_names.append(date_wise_names)

        # print(collected_names)
        return collected_names

    def collect_names_for_single_date(self,
                sheet_name_is_sheet3,
                sheet_name_is_sheet2,
                category_number,
                inc_date,
                unwanted_post_list,
                number_of_names: int | None = None,
                depart_name_for_gen=None,
                sort_by_amount=True):
        """
        Getting the names of the employees for the single date
        :param sheet_name_is_sheet3: sheet 3 for getting the employee data
        :param sheet_name_is_sheet2: sheet 2 to get the leaves data
        :param category_number: cat num
        :param inc_date: incentive date
        :param number_of_names: how many employee names to be included in a single case number incentive
        :param depart_name_for_gen: required for filtering the consultants and residents names of the department
        :param sort_by_amount: sorting as per the received amount
        :param unwanted_post_list: posts which can be optional like counselor
        :return: names for the specific date
        """

        filtered_names_sheet3 = self.filter_category(sheet_name=sheet_name_is_sheet3, category_number=category_number, depart_name_for_gen=depart_name_for_gen, sort_by_amount=sort_by_amount, unwanted_post_list=unwanted_post_list)
        emp_working_periods_sheet2 = self._working_periods(sheet_name=sheet_name_is_sheet2)  #  generator object

        # number of names redefining
        if number_of_names is None:
            number_of_names = len(filtered_names_sheet3)
        # print('Number of names >>>>>>>>>>>>> ', number_of_names)

        # collected_names = []
        # date_wise_names = [inc_date]
        date_wise_names = []

        for employee_in_sheet3 in filtered_names_sheet3:

            # checking the number of names to be in a particular category
            # if len(date_wise_names) == number_of_names + 1:  # +1 for correcting the date added at start in list
            if len(date_wise_names) == number_of_names:  # +1 for correcting the date added at start in list
                break

            for working_employee_period_dict in emp_working_periods_sheet2:
                '''working_employee_period_dict = {'Name_After': 'chandrakanta kashyap', 'Name_with_spaces': 'chandrakanta kashyap', '
                Post': 'Staff Nurse (Govt.)', 'Serial number': 194, 'JOINING DATE': datetime.datetime(2022, 7, 12, 0, 0), 
                'RELIEVING DATE ': datetime.datetime(2023, 12, 31, 0, 0), 'LEAVE START1': datetime.datetime(2022, 8, 1, 0, 0), 
                'LEAVE END1': datetime.datetime(2023, 1, 31, 0, 0), 'LEAVE START2': None, 'LEAVE END2': None, 'LEAVE START3': None, 
                'LEAVE END3': None, 'EmpID': 15170200070, 'Working Days': 334, 'Percentage': 64.60348162475822, None: None},'''

                if employee_in_sheet3[1] == working_employee_period_dict['Name_After']:  # name matching in sheet 3 and sheet 2
                    """
                    till above the sheet 3 and sheet 2 names and working periods have been matched and extracted.
                    next is matching the dates of incentive cases to the working periods
                    """

                    # ColourPrint.print_green(f'{working_employee_period_dict["Name_After"]=},{working_employee_period_dict["JOINING DATE"]=}, {working_employee_period_dict["RELIEVING DATE"]=}')
                    if self._is_eligible(working_employee_period_dict):
                        # print(type(inc_date), inc_date)
                        # print(working_employee_period_dict['JOINING DATE'], inc_date, working_employee_period_dict["RELIEVING DATE"])
                        if working_employee_period_dict['JOINING DATE'] < inc_date < working_employee_period_dict["RELIEVING DATE"]:
                            # ColourPrint.print_blue(f'{employee_in_sheet3[1] =}, {working_employee_period_dict['Name_After']=}')
                            """
                            now checking weather in leave or not
                            """
                            if not self._is_on_leave(date_of_incentive=inc_date, each_emp_data=working_employee_period_dict):
                                date_wise_names.append(employee_in_sheet3[1])

                        # else:  # get the names not included here
                        #     print(employee_in_sheet3[1])

        # print(collected_names)
        return date_wise_names

def retrieving_all_incentive_case_data(workbook_path):
    """
    Getting the all row data of entry Excel workbook e.g. eye 2 2 of sheet 4
    :param workbook_path: path of workbook
    :return: list of all data of sheet 4
    """
    wb = openpyxl.load_workbook(workbook_path)
    ws = wb['Sheet4']
    row_datas = list(ws.iter_rows(min_row=2, values_only=True))
    # print(list(row_datas))
    wb.close()
    return row_datas

def retrieving_all_incentive_names_data(workbook_path):
    """Getting the all row data of entry Excel workbook e.g. eye 2 2 sheet 2
    :param workbook_path: path of workbook
    :return: Name list to make entry allocated by program
    """
    wb = openpyxl.load_workbook(workbook_path)
    ws = wb['Sheet2']
    row_datas = list(ws.iter_rows(min_row=2, values_only=True))
    # print(list(row_datas))
    wb.close()
    return row_datas

def inc_percent_amt_calculate(cat):
    category_percentages = {
        1: 0.01,
        2: 0.14,
        3: 0.06,
        4: 0.45,
        5: 0.10,
        6: 0.10,
        7: 0.06,
        8: 0.03,
        9: 0.05
    }
    return category_percentages.get(cat)

def count_of_name_in_each_category(retrieving_all_incentive_names_datas):
    """
    Getting the all row data of entry Excel workbook e.g. eye 2 2
    :param retrieving_all_incentive_names_datas:
    :return: count of names in each category
    """
     # retrieving_all_incentive_names_datas = [(1, 'DR.ANMOL MADHUR MINZ', 'Associate Professor', 'vfèk"Bkrk vLirky vèkh{kd ]lgk;d vèkh{kd uksMy vfèkdkjh', '04170140656', 1), (2, 'AVINASH MESHRAM', 'Dean',...]

    dict_cat_count = {}
    count = 1
    # print(55444)
    # print('ppppppppppppppppp',retrieving_all_incentive_names_datas)
    for name_data in retrieving_all_incentive_names_datas:
        cat = name_data[5]
        if cat is not None:
            # print('=========',name_data)
            # cat = name_data[5]
            # print('=========',cat)
            if cat not in dict_cat_count:
                dict_cat_count[cat] = count
            else:
                dict_cat_count[cat] +=1

    return dict_cat_count


def amount_distribution_of_all_case_number_of_sheet(retrieved_all_incentive_case_datas, retrieving_all_incentive_names_datas):
    """
    Generating the incentive amount received by each employee according to category and employee count in each category.
    Used to feed the SQL temp_database
    :param retrieved_all_incentive_case_datas: Getting the all row data of entry Excel workbook e.g. eye 2 2 of sheet 4
    :param retrieving_all_incentive_names_datas: Getting the all row data of entry Excel workbook e.g. eye 2 2 of sheet 2
    :return: name, emp_code, inc_category, divided_amount
    """
    # print(retrieved_all_incentive_case_datas)
    master_list_with_amount = []  # containing the case_number: each case number emp datas
    for case_num_data in retrieved_all_incentive_case_datas:  #[(3253, 'CASE/PS6/HOSP22G146659/CK7130670', 'Ophthalmology', 'SICS with non-foldable IOL', 2210, datetime.datetime(2023, 11, 1, 0, 0), 'Kartik Ram'), (3254, 'CASE/PS6/HOSP22G...)]
        # print(case_num_data)
        case_number = case_num_data[1]  # UPDATED (1966, 'CASE/PS6/HOSP22G146659/CK7130736', 2210, datetime.datetime(2023, 11, 1, 0, 0), 'Mahetarin Bai', 'Ophthalmology', 'SICS with non-foldable IOL')
        # ColourPrint.print_blue('/////////////////////////')
        total_amount = case_num_data[2]
        # print(case_number, total_amount)

        name_counts_in_each_cats = count_of_name_in_each_category(retrieving_all_incentive_names_datas)
        print(f'Counts of name in each category [{case_number}]', name_counts_in_each_cats)

        each_case_number_emp_data_with_amount = [case_number]
        for name_data in retrieving_all_incentive_names_datas:  #[(1, 'DR.ANMOL MADHUR MINZ', 'Associate Professor', 'vfèk"Bkrk vLirky vèkh{kd ]lgk;d vèkh{kd uksMy vfèkdkjh', '04170140656', 1), (2, 'AVINASH MESHRAM', 'Dean',...]
            name = name_data[1]
            emp_code = name_data[4]
            inc_category = name_data[5]
            # print(inc_category, f'--------{name}...........................>>')

            if inc_category is not None:
                percentage = inc_percent_amt_calculate(inc_category)
                divided_percentage = percentage/name_counts_in_each_cats[inc_category]
                # print('pppppp', divided_percentage, total_amount)

                try:
                    divided_amount = round(divided_percentage * total_amount, 4)
                except:
                    raise ValueError(f'Error in name: {name}, total amount: {total_amount}, percentage: {percentage}')

            # each_emp_data_with_amount =
            # print(name, emp_code, inc_category, percentage, divided_percentage, divided_amount)

                each_case_number_emp_data_with_amount.append((name, emp_code,inc_category,divided_amount))

        # print(each_case_number_emp_data_with_amount)
        master_list_with_amount.append(each_case_number_emp_data_with_amount)
    return master_list_with_amount

def update_incentive2_from_sql(emp_id_and_total_amount_list : list[tuple]):
    """
    Used to update the incentive 2 Excel
    :param emp_id_and_total_amount_list: employee data and future amounts
    :return: None
    """
    # [('J55172798', 510095.26200000005), ('J55173309', 509304.76200000005), ('N5517569', 17326.6266), ('04170090031', 21402.25117), ('05170020166', 21054.794599999997), ('05170020167', 11296.874),...]
    wb_path = r"G:\My Drive\GdrivePC\Hospital\RSBY\New\Incentive_auto2.xlsx"
    wb = openpyxl.load_workbook(wb_path)
    ws = wb['Sheet3']
    rows_data = list(ws.iter_rows())

    for emp_id, amount in emp_id_and_total_amount_list:
        for emp_details_tuple in rows_data:
            if emp_id == str(emp_details_tuple[2].value):
                # ColourPrint.print_yellow('==============')
                # print(emp_details_tuple[2].value, 'type: ', type(emp_details_tuple[2].value))
                emp_details_tuple[16].value = amount  # Column 'Q'
                break
    wb.save(wb_path)
    wb.close()








#
# INCENTIVE_NAME_ADDING_WORKBOOK_NAME = r"G:\My Drive\GdrivePC\Hospital\RSBY\New\Upload Formats\eye\eye 3 3.xlsx"
# # print((count_of_name_in_each_category()))
# ret_case_num_data = retrieving_all_incentive_case_data(INCENTIVE_NAME_ADDING_WORKBOOK_NAME)
# ret_name_data = retrieving_all_incentive_names_data(INCENTIVE_NAME_ADDING_WORKBOOK_NAME)
# # print(list(ret_name_data))
# # print('-----------------')
# # print(len(amount_distribution_of_all_case_number_of_sheet(ret_case_num_data, ret_name_data)))
# data_for_db_amount_of_all_case_number_and_employee_wise = amount_distribution_of_all_case_number_of_sheet(ret_case_num_data, ret_name_data)
# update_incentive2_from_sql('l')